"""One-off extraction: manager's "Dark Store Data" workbook -> config.xlsx.

Temporary migration tool. Reads the messy source workbook and writes a clean
`config.xlsx` (sheets locations / brands / coverage) that `cli sync` consumes.
After this runs, `config.xlsx` is the source of truth and this script is disposable.

Import rule: keep the reliable fields (merchant_id, city, state, region, lat/lon),
leave ambiguous metadata blank (grid_/missing pincodes -> blank, zone -> blank).

Usage:
    python scripts/extract_config.py "Dark Store Data'.xlsx" config.xlsx
"""
import re
import sys

from openpyxl import Workbook, load_workbook

SRC = sys.argv[1] if len(sys.argv) > 1 else "Dark Store Data'.xlsx"
OUT = sys.argv[2] if len(sys.argv) > 2 else "config.xlsx"

TENANT = "Dobra"
BRAND = "dobra"

_PIN = re.compile(r"^\d{6}$")


def _mid(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _pincode(v) -> str:
    """Keep only real 6-digit pincodes; drop grid_ placeholders and blanks."""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    s = str(v).strip()
    return s if _PIN.match(s) else ""


def extract_locations(wb) -> list[dict]:
    ws = wb["All Stores"]
    rows = list(ws.iter_rows(values_only=True))
    # row 0 = title banner, row 1 = header, row 2+ = data
    # cols: 0 S.No, 1 Merchant ID, 2 Store Type, 3 City, 4 State, 5 Region,
    #       6 Latitude, 7 Longitude, 8 Pincode, 9 Zone, 10 Rank, ...
    out, seen = [], set()
    dropped = 0
    for r in rows[2:]:
        if not r or all(c is None for c in r):
            continue
        mid = _mid(r[1])
        lat, lon = r[6], r[7]
        if not mid or lat in (None, "") or lon in (None, ""):
            dropped += 1
            continue
        if mid in seen:  # border-overlap duplicates (same store, two regions)
            continue
        seen.add(mid)
        out.append({
            "merchant_id": mid,
            "city": str(r[3]).strip().lower() if r[3] else "",
            "state": str(r[4]).strip() if r[4] else "",
            "region": str(r[5]).strip() if r[5] else "",
            "zone": "",
            "pincode": _pincode(r[8]),
            "lat": float(lat),
            "lon": float(lon),
            "active": "yes",
        })
    print(f"locations: {len(out)} stores  ({dropped} rows dropped for missing id/coords, "
          f"{sum(1 for x in out if not x['pincode'])} with blank pincode)")
    return out


def extract_keywords(wb) -> list[str]:
    ws = wb["Dobra Keyword Performence"]
    rows = ws.iter_rows(min_row=2, values_only=True)  # header on row 1
    seen, order = set(), []
    for r in rows:
        kw = r[7] if len(r) > 7 else None  # 'keyword' column
        if kw and str(kw).strip() and str(kw).strip() not in seen:
            seen.add(str(kw).strip())
            order.append(str(kw).strip())
    print(f"keywords: {len(order)} distinct -> {order}")
    return order


def main():
    wb = load_workbook(SRC, data_only=True, read_only=True)
    locations = extract_locations(wb)
    keywords = extract_keywords(wb)
    cities = sorted({l["city"] for l in locations if l["city"]})

    out = Workbook()
    ws = out.active
    ws.title = "locations"
    ws.append(["merchant_id", "city", "state", "region", "zone", "pincode", "lat", "lon", "active"])
    for l in locations:
        ws.append([l["merchant_id"], l["city"], l["state"], l["region"],
                   l["zone"], l["pincode"], l["lat"], l["lon"], l["active"]])

    b = out.create_sheet("brands")
    b.append(["tenant", "brand", "relationship", "keywords", "aliases"])
    b.append([TENANT, BRAND, "own", ", ".join(keywords), BRAND])

    c = out.create_sheet("coverage")
    c.append(["tenant", "city", "zone"])
    for city in cities:
        c.append([TENANT, city, ""])

    out.save(OUT)
    print(f"\nWrote {OUT}: {len(locations)} locations, 1 brand ({len(keywords)} keywords), "
          f"{len(cities)} coverage cities.")


if __name__ == "__main__":
    main()
