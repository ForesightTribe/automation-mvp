"""Export every Zepto PRIVATE (seller-panel) table for one tenant to one workbook.

The Zepto counterpart of `export_to_excel.py`, which does the same for Blinkit.
Public/search data is not included — that lives in the shared `search_*` and
`sku_snapshots` tables and is not per-platform-private.

One sheet per table, plus a Contents sheet listing row counts and date ranges so
you can see at a glance which days are complete. Read-only: nothing is written
back to the database.

Run from backend/:
    python -m scripts.zepto_export_private -t <tenant_id>
    python -m scripts.zepto_export_private -t <tenant_id> --from 2026-08-29 --to 2026-08-30
    python -m scripts.zepto_export_private -t <tenant_id> -o zepto_aug.xlsx

Dates filter each table on its own date column (see `_SHEETS`). `zepto_po_items`
has no date of its own, so it is filtered through its parent PO's `po_date`.
"""
import argparse
import asyncio
import json
import uuid
from datetime import date, datetime
from pathlib import Path

import openpyxl
from dotenv import load_dotenv
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import text

load_dotenv(Path(__file__).parent.parent / ".env")

from app.core.database import AsyncSessionLocal  # noqa: E402

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(color="FFFFFF", bold=True)

# Columns that are plumbing, not data — same set the Blinkit export strips.
_STRIP = {"id", "upsert_key", "tenant_id", "scrape_job_id", "platform"}

# (sheet name, table, date column, order-by). Order mirrors how the data is
# read on the dashboard: sales first, then ads, then supply chain.
_SHEETS: list[tuple[str, str, str, str]] = [
    ("Sales Summary",      "zepto_seller_sales_summary",      "date",         "date"),
    ("Sales per SKU",      "zepto_seller_sales",              "period_start", "period_start, product_variant_id"),
    ("Sales per SKU-City", "zepto_seller_product_city_daily", "date",         "date, city_id, product_variant_id"),
    ("Ads Campaigns",      "zepto_ad_campaign_daily",         "date",         "date, campaign_id"),
    ("Ads Keywords",       "zepto_ad_keyword_daily",          "date",         "date, keyword"),
    ("Ads Products",       "zepto_ad_product_daily",          "date",         "date, product_variant_id"),
    ("Ads Breakdown",      "zepto_ad_breakdown_daily",        "date",         "date, dimension, name"),
    ("POs",                "zepto_po",                        "po_date",      "po_date, po_id"),
    ("PO Line Items",      "zepto_po_items",                  None,           "po_id, sku_name"),
    ("ASNs",               "zepto_asn",                       "asn_date",     "asn_date, asn_no"),
    ("GRNs",               "zepto_grn",                       "grn_date",     "grn_date, grn_no"),
]


def _to_header(key: str) -> str:
    return key.replace("_", " ").title()


def _clean(val):
    if isinstance(val, datetime):
        return val.replace(tzinfo=None)
    if isinstance(val, (dict, list)):
        return json.dumps(val, default=str)
    if isinstance(val, uuid.UUID):
        return str(val)
    return val


def _write_sheet(ws, rows: list[dict]) -> None:
    if not rows:
        ws.append(["(no data for this window)"])
        return

    keys = [k for k in rows[0] if k not in _STRIP]

    for col, key in enumerate(keys, start=1):
        cell = ws.cell(row=1, column=col, value=_to_header(key))
        cell.fill, cell.font = _HEADER_FILL, _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for r, row in enumerate(rows, start=2):
        for col, key in enumerate(keys, start=1):
            ws.cell(row=r, column=col, value=_clean(row.get(key)))

    for col, key in enumerate(keys, start=1):
        width = max(
            len(_to_header(key)),
            max((len(str(row.get(key) or "")) for row in rows), default=0),
        )
        ws.column_dimensions[get_column_letter(col)].width = min(width + 2, 55)
    ws.freeze_panes = "A2"


def _build_query(table: str, date_col: str | None, order_by: str,
                 start: str | None, end: str | None) -> str:
    """SELECT for one sheet, with the date window applied where one exists."""
    if table == "zepto_po_items" and (start or end):
        # No date column of its own — scope through the parent PO.
        conds = ["i.tenant_id = :t"]
        if start:
            conds.append("p.po_date >= :start")
        if end:
            conds.append("p.po_date <= :end")
        return (
            f"select i.* from zepto_po_items i "
            f"join zepto_po p on p.po_id = i.po_id and p.tenant_id = i.tenant_id "
            f"where {' and '.join(conds)} order by i.{order_by}"
        )

    conds = ["tenant_id = :t"]
    if date_col and start:
        conds.append(f"{date_col} >= :start")
    if date_col and end:
        conds.append(f"{date_col} <= :end")
    return f"select * from {table} where {' and '.join(conds)} order by {order_by}"


async def export(tenant_id: str, output: str, start: str | None, end: str | None) -> None:
    params: dict = {"t": uuid.UUID(tenant_id)}
    if start:
        params["start"] = date.fromisoformat(start)
    if end:
        params["end"] = date.fromisoformat(end)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    contents = wb.create_sheet("Contents")
    summary: list[dict] = []

    async with AsyncSessionLocal() as db:
        for name, table, date_col, order_by in _SHEETS:
            sql = _build_query(table, date_col, order_by, start, end)
            rows = [dict(r) for r in (await db.execute(text(sql), params)).mappings()]
            _write_sheet(wb.create_sheet(name), rows)

            dates = [r[date_col] for r in rows if date_col and r.get(date_col)]
            summary.append({
                "sheet": name,
                "table": table,
                "rows": len(rows),
                "from": min(dates) if dates else None,
                "to": max(dates) if dates else None,
            })
            print(f"  {name:20} {len(rows):>6} rows")

    _write_sheet(contents, summary)
    wb.move_sheet("Contents", offset=-len(_SHEETS))
    wb.save(output)
    print(f"\nSaved {output}")


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("-t", "--tenant", required=True, help="Tenant ID")
    ap.add_argument("--from", dest="start", help="Start date YYYY-MM-DD")
    ap.add_argument("--to", dest="end", help="End date YYYY-MM-DD")
    ap.add_argument("-o", "--output", help="Output .xlsx (default: zepto_private_<dates>.xlsx)")
    args = ap.parse_args()

    out = args.output
    if not out:
        window = f"_{args.start or 'all'}_to_{args.end or 'all'}" if (args.start or args.end) else ""
        out = f"zepto_private{window}.xlsx"

    print(f"Exporting Zepto private data for {args.tenant}")
    asyncio.run(export(args.tenant, out, args.start, args.end))


if __name__ == "__main__":
    main()
