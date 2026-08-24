"""Export loaded Zepto scrape data from Postgres to Excel — for eyeballing.

The pipeline's output lives in the database; this is a read-only view of it in the
shape the client workbooks used, so a run can be checked without writing SQL.

It reads what `cli scrape load` already pushed. It does NOT scrape, and it is not
part of the pipeline — delete it the day the dashboard shows this data.

Run from backend/:
    python -m scripts.zepto_export_from_db
    python -m scripts.zepto_export_from_db --city bengaluru
    python -m scripts.zepto_export_from_db --mp blinkit
"""
import asyncio
import sys
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import text

from app.core.database import AsyncSessionLocal

BACKEND = Path(__file__).parent.parent
MP = "blinkit" if "--mp" in sys.argv and "blinkit" in sys.argv else "zepto"
CITY = None
if "--city" in sys.argv:
    i = sys.argv.index("--city")
    if i + 1 < len(sys.argv):
        CITY = sys.argv[i + 1]

HDR_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
HDR_FILL = PatternFill("solid", fgColor="2E4057")
OWN_FILL = PatternFill("solid", fgColor="C6EFCE")
ALT_FILL = PatternFill("solid", fgColor="F0F4F8")


def head(ws, cols):
    ws.append(cols)
    for c in ws[1]:
        c.font, c.fill = HDR_FONT, HDR_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"


def widths(ws, cap=46):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        longest = max((len(str(c.value)) for c in col[:300] if c.value), default=0)
        ws.column_dimensions[letter].width = min(max(longest + 2, 9), cap)


async def main():
    where = "l.mp_slug = :mp" + (" and l.city = :city" if CITY else "")
    params = {"mp": MP}
    if CITY:
        params["city"] = CITY

    async with AsyncSessionLocal() as db:
        rows = (await db.execute(text(
            f"""select l.pincode, l.zone, s.lat, s.lon, l.merchant_id,
                       l.keyword, l.position, l.brand_slug, l.is_brand,
                       l.product_name, l.price, l.mrp, l.discount_pct,
                       l.in_stock, l.inventory, l.pack_raw, l.scraped_at
                from search_listings l
                join search_snapshots s on s.id = l.snapshot_id
                where {where}
                order by l.pincode, l.zone, l.keyword, l.position"""),
            params)).all()

        snaps = (await db.execute(text(
            f"""select s.keyword, count(*) n,
                       count(distinct s.merchant_id) stores,
                       min(s.brand_rank) best_rank,
                       round(avg(s.brand_rank)::numeric, 1) avg_rank,
                       round(avg(s.brand_sov)::numeric, 1) avg_sov
                from search_snapshots s
                where s.mp_slug = :mp""" + (" and s.city = :city" if CITY else "") +
            " group by s.keyword order by n desc"), params)).all()

    if not rows:
        print(f"No {MP} rows in the database"
              + (f" for city '{CITY}'" if CITY else "")
              + " — run `cli scrape load` first.")
        return

    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "All Products"
    head(ws, ["pincode", "area", "lat", "lon", "store_id", "keyword", "rank",
              "brand", "is_own", "product_name", "price", "mrp", "discount_pct",
              "in_stock", "inventory", "pack", "scraped_at"])
    for i, r in enumerate(rows, start=2):
        ws.append([r[0], r[1], r[2], r[3], r[4], r[5], r[6], r[7],
                   "YES" if r[8] else "", r[9], r[10], r[11], r[12],
                   "yes" if r[13] else "no", r[14], r[15],
                   r[16].strftime("%d-%b %H:%M") if r[16] else ""])
        if r[8]:
            for c in range(1, 18):
                ws.cell(row=i, column=c).fill = OWN_FILL
    ws.auto_filter.ref = ws.dimensions
    widths(ws)

    ws2 = wb.create_sheet("By Keyword")
    head(ws2, ["keyword", "snapshots", "stores", "best rank", "avg rank", "avg SoV %"])
    for i, r in enumerate(snaps, start=2):
        ws2.append(list(r))
        if i % 2 == 0:
            for c in range(1, 7):
                ws2.cell(row=i, column=c).fill = ALT_FILL
    widths(ws2)

    ws3 = wb.create_sheet("By Store")
    head(ws3, ["pincode", "area", "store_id", "products", "own rows", "keywords"])
    by_store: dict = {}
    for r in rows:
        k = (r[0], r[1], r[4])
        d = by_store.setdefault(k, {"n": 0, "own": 0, "kw": set()})
        d["n"] += 1
        d["own"] += 1 if r[8] else 0
        d["kw"].add(r[5])
    for i, (k, d) in enumerate(sorted(by_store.items()), start=2):
        ws3.append([k[0], k[1], k[2], d["n"], d["own"], len(d["kw"])])
        if i % 2 == 0:
            for c in range(1, 7):
                ws3.cell(row=i, column=c).fill = ALT_FILL
    widths(ws3)

    ts = datetime.now().strftime("%Y%m%d_%H%M")
    name = f"{MP}_from_db_{CITY or 'all'}_{ts}.xlsx"
    out = BACKEND / name
    wb.save(out)

    own = sum(1 for r in rows if r[8])
    print(f"{MP}" + (f" / {CITY}" if CITY else "")
          + f"   {len(rows):,} product rows   {len(by_store)} stores   "
            f"{own:,} own-brand rows")
    print(f"Saved -> {out}")


asyncio.run(main())
