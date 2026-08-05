"""
Keyword competitive scan — resumable.

Zepto's search API enforces a VOLUME limit, roughly 300-350 searches per window,
not just a rate limit. Slowing down only postpones the wall; it does not avoid it.
A 169-store x 9-keyword sweep needs ~1,500 searches, so it can never finish in one
session however gently it is paced.

So this version is built to be run repeatedly instead of once:

  - every store's results are appended to a checkpoint the moment it completes,
    so a block can never lose work already done
  - on start it reads the checkpoint and skips store/keyword pairs already
    collected, resuming where it stopped
  - when blocked it exits cleanly and tells you to re-run later, rather than
    burning long cooldowns against a window that is not reopening
  - a small non-zero result (7 products for a niche term) is a real answer, not
    a failure; only zero-across-retries counts as blocked

Run from backend/ — same command every time until it reports complete:
    python -m scripts.zepto_keyword_scan
    python -m scripts.zepto_keyword_scan --status     # progress, no scraping
    python -m scripts.zepto_keyword_scan --export     # build the workbook now
"""
import asyncio
import csv
import json
import re
import sys
from pathlib import Path
from datetime import datetime
from collections import Counter, defaultdict

sys.path.insert(0, str(Path(__file__).parent.parent))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from playwright.async_api import async_playwright

# ── Config ────────────────────────────────────────────────────────────────────
CLIENT_BRAND = "Brik Oven"
# Sentinel brand for a search that legitimately returned nothing. Recorded so the
# pair counts as done; filtered out of every export sheet.
NO_RESULTS = "(no results)"
CITY = "Bengaluru"

# The competitor set the client asked to be tracked. Everything on the SERP is
# still scraped and still lands in 'All Products' — position is only meaningful
# as a rank out of the full 30, and re-scraping to add a brand later would cost
# another full run. The narrowing happens at export instead.
#
# Left = canonical label used in the report. Right = normalised fragments to
# match Zepto's own brand string against. Fragments are matched on a normalised
# form (lowercased, apostrophes dropped, spaces collapsed) so "The Baker's Dozen"
# and "Bakers Dozen" both land.
#
# 'health factory' is deliberately the full two words: Zepto also carries an
# unrelated 'Healthy Master', which a bare 'health' would wrongly pull in.
COMPETITORS = {
    "The Health Factory": ["health factory"],
    "Suchali":            ["suchali"],
    "Theobroma":          ["theobroma"],
    "The Baker's Dozen":  ["bakers dozen"],
    "Baker's Loaf":       ["bakers loaf"],
}

# Report order: client first, then competitors as the client listed them.
TRACKED_ORDER = [CLIENT_BRAND] + list(COMPETITORS)

# One fill per tracked brand so a row is identifiable at a glance.
TRACK_FILL = {
    CLIENT_BRAND:         "C6EFCE",   # green  - the client
    "The Health Factory": "FCD9D9",   # red
    "Suchali":            "FFF3C4",   # yellow
    "Theobroma":          "E4DAF6",   # purple
    "The Baker's Dozen":  "FFE3C2",   # orange
    "Baker's Loaf":       "D6E9F8",   # blue
}


def _norm(s):
    """Lowercase, drop apostrophes, collapse whitespace — so brand strings
    compare regardless of punctuation styling."""
    return re.sub(r"\s+", " ", str(s or "").lower().replace("'", "")
                  .replace("’", "")).strip()


def tracked_brand(brand):
    """Canonical tracked name for a Zepto brand string, or None if untracked."""
    n = _norm(brand)
    if not n:
        return None
    if _norm(CLIENT_BRAND) in n:
        return CLIENT_BRAND
    for canon, frags in COMPETITORS.items():
        if any(f in n for f in frags):
            return canon
    return None

# 'sourdough bread loaf' removed: it returned 1-9 products and often 0, which the
# old code read as throttling — it caused most of the pointless identity rebuilds.
KEYWORDS = [
    "sourdough",
    "sourdough bread",
    "sourdough bread loaf",
    "rosemary sourdough",
    "sour cream",
    "mozzarella",
    "ricotta",
    "cheesecake",
    "buratta",
]

RESULT_CAP = 30
MAX_PAGES = 3
# Zepto allows roughly 320 searches per window. If that window is ROLLING
# (320 per rolling hour) then pacing below the sustainable rate avoids blocks
# entirely, which finishes sooner than sprinting and then waiting out 15-60 min
# recoveries:
#     320 searches / 60 min = 5.3 per min = 11.3 s per search
# 12 s sits just under that line. Store gap folded in, since the per-search
# spacing is now what matters.
# Measured: 5 s -> 5 stores before a block; 12 s -> 34 stores. Slower pacing
# buys real throughput, so 20 s is deliberately conservative — 3 searches/min,
# well under the ~5.3/min that the ~320-per-window ceiling implies. It cannot
# beat a volume cap, but it stops us reaching it early and keeps each recovery
# short. The run is unattended and checkpointed, so wall-clock cost is cheap.
# Measured, all on this IP:
#     0.4 s (150/min) -> blocked after   1 search
#     5   s ( 12/min) -> blocked after  21
#    12   s (  5/min) -> blocked after 137
# Rate matters, not just volume — bursting blocks almost immediately.
#
# 2026-08-04: tried 6 s, reverted to 12 s. The 5/12 s figures were measured before
# the `layout: null` bug was found, so it was worth re-testing whether that bug had
# inflated them. It had not:
#     6 s ( 10/min) -> blocked after  47 searches
# The block was genuine — one store blocked on `sourdough`, which returns 30 and
# never pages past the end, so it cannot be the parse bug.
# Throughput, block penalty included, is what settles it:
#     6 s ->  47 searches per (282 s + 900 s hard block)  = 2.4/min
#    12 s -> 110 searches per (1320 s + 900 s clean rest) = 3.0/min
# Faster pacing is SLOWER end to end, because a hard block costs more than a
# scheduled pause — recovery after a block yields ~3 searches per 5-min cycle,
# while a clean pause resets the window properly. Do not re-litigate without new
# evidence; 12 s is measured, not guessed.
SEARCH_GAP_S = 12.0       # buffer after every search
STORE_GAP_S = 3.0         # extra breathing room between stores
TRIES = 2                 # retries per search; more just feeds an active block
STOP_AFTER_BAD_STORES = 2   # consecutive fully-failed stores -> the window is shut
MAX_STORE_ATTEMPTS = 4      # re-queue cap per store, so a persistently failing
                            # store cannot spin the queue forever

# Blocks are unavoidable: the window allows ~300-350 searches and a full sweep
# needs ~1,350, so it WILL be hit several times. Rather than exiting, wait the
# window out and carry on — one command, unattended. Waits escalate because a
# window that did not reopen after 15 min will not reopen after another 15.
# PROACTIVE break — the important one. The window allows ~320 searches; resting
# BEFORE that point avoids the block entirely, which matters because a hard block
# appears to deepen the penalty: after one, recovery yielded only ~3 searches per
# 5-minute cycle, while a clean pause resets the window properly.
# 250 leaves headroom under the observed ceiling.
# Set BELOW the measured block point, not at a guessed ceiling. At 12 s pacing
# this IP blocked after 137 searches, so 250 would never have fired — the wall
# would arrive first. 110 leaves margin under 137. Each run logs where blocks
# actually land (see BLOCK POINTS in the summary), so this number should be
# re-tuned from evidence rather than left as a guess.
PAUSE_EVERY = 110            # searches between scheduled breaks
PAUSE_S = 900                # 15 min rest, taken before the limit is reached

# Reactive recovery, for when a block happens anyway.
RECOVERY_WAITS_S = [900, 1800, 2700, 3600]   # 15, 30, 45, 60 min
PROBE_EVERY_S = 300          # while waiting, test every 5 min with ONE search
MAX_CONSECUTIVE_STALLS = 6   # give up only after ~4h of failed recovery

BACKEND = Path(__file__).parent.parent
# Both are reassigned by main() when --tag is passed, so an experimental run
# gets its own checkpoint and its own workbook.
TAG = ""
CKPT = BACKEND / "keyword_scan_checkpoint.csv"

_BASE = "https://www.zepto.com"
_BFF = "https://bff-gateway.zepto.com"
_GP = "/lms/api/v2/get_page"
_SEARCH = "/user-search-service/api/v3/search"
_DROP = {"host", "content-length", "connection", "accept-encoding",
         ":method", ":path", ":authority", ":scheme"}

FIELDS = ["store_id", "store_name", "pincode", "area", "lat", "lng", "keyword",
          "position", "brand", "is_client", "product_name", "pack_size",
          "mrp_rs", "selling_price_rs", "discount_pct", "discount_rs",
          "in_stock", "available_qty", "rating", "rating_count", "category",
          "match_bucket", "weight_g", "product_id", "variant_id", "scraped_at"]

HDR_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
HDR_FILL = PatternFill("solid", fgColor="2E4057")
HDR_ALIGN = Alignment(horizontal="center", vertical="center")
ALT_FILL = PatternFill("solid", fgColor="F0F4F8")
CLIENT_FILL = PatternFill("solid", fgColor="C6EFCE")
CLIENT_FONT = Font(name="Calibri", size=10, bold=True, color="0B6B3A")
TOP3_FILL = PatternFill("solid", fgColor="FFF2CC")
OOS_FONT = Font(name="Calibri", size=10, color="C00000")
TITLE_FILL = PatternFill("solid", fgColor="1A3C5E")


# ── Stores ────────────────────────────────────────────────────────────────────

def load_stores():
    f = sorted(BACKEND.glob("zepto_FINAL_master_filled_*.xlsx")) or \
        sorted(BACKEND.glob("zepto_FINAL_master_*.xlsx"))
    if not f:
        raise SystemExit("No FINAL master found.")
    wb = openpyxl.load_workbook(f[-1], read_only=True, data_only=True)
    rr = list(wb["All Stores"].iter_rows(values_only=True))
    wb.close()
    hdr = [str(x).strip().lower() if x else "" for x in rr[2]]
    out, seen = [], set()
    for r in rr[3:]:
        if not r or not r[0]:
            continue
        d = dict(zip(hdr, r))
        if str(d.get("city")) != CITY:
            continue
        sid = str(d.get("store_id") or "").strip()
        if not sid or sid in seen:
            continue
        try:
            la, ln = float(d["lat"]), float(d["lng"])
        except (TypeError, ValueError):
            continue
        seen.add(sid)
        sec = [x.strip() for x in str(d.get("secondary") or "").split(",")
               if x.strip() and x.strip().lower() != "none"]
        out.append({"store_id": sid,
                    "store_name": str(d.get("store_name") or "").strip(),
                    "pincode": str(d.get("source_pincode") or "").strip(),
                    "area": str(d.get("area_suggestion") or "").strip(),
                    "secondary": sec,
                    "lat": la, "lng": ln})
    return out, f[-1].name


# ── Checkpoint ────────────────────────────────────────────────────────────────

def load_ckpt():
    """(rows, {(store_id, keyword)} already collected)."""
    if not CKPT.exists():
        return [], set()
    rows, done = [], set()
    with CKPT.open(newline="", encoding="utf-8") as fh:
        for r in csv.DictReader(fh):
            rows.append(r)
            done.add((r["store_id"], r["keyword"]))
    return rows, done


def append_ckpt(rows):
    """Append immediately so a later block cannot lose completed work."""
    new = not CKPT.exists()
    with CKPT.open("a", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=FIELDS)
        if new:
            w.writeheader()
        for r in rows:
            w.writerow({k: r.get(k, "") for k in FIELDS})


# ── API ───────────────────────────────────────────────────────────────────────

async def capture(page, part, nav, settle=8000):
    cap = {"h": None, "body": None}

    async def on_req(req):
        if part in req.url and cap["h"] is None:
            cap["h"] = dict(req.headers)
            try:
                cap["body"] = req.post_data
            except Exception:
                pass

    page.on("request", on_req)
    try:
        await nav()
        w = 0
        while cap["h"] is None and w < settle:
            await page.wait_for_timeout(250)
            w += 250
    except Exception:
        pass
    finally:
        page.remove_listener("request", on_req)
    if not cap["h"]:
        return {}, None
    return {k: v for k, v in cap["h"].items() if k.lower() not in _DROP}, cap["body"]


async def open_session(browser, state):
    old = state.get("ctx")
    if old is not None:
        try:
            await old.close()
        except Exception:
            pass
    ctx = await browser.new_context(
        user_agent=("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                    "(KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36"))
    page = await ctx.new_page()
    state["ctx"], state["page"] = ctx, page
    gp, _ = await capture(page, _GP,
                          lambda: page.goto(_BASE, timeout=30000,
                                            wait_until="domcontentloaded"))
    if gp:
        state["gp"] = gp
    await page.wait_for_timeout(1500)
    se, raw = await capture(page, _SEARCH,
                            lambda: page.goto(f"{_BASE}/search?query=bread",
                                              timeout=30000,
                                              wait_until="domcontentloaded"),
                            settle=9000)
    if se:
        state["se"] = se
        if raw:
            try:
                state["body"] = json.loads(raw)
            except Exception:
                pass
    # Only the search headers matter. get_page headers are captured opportunistically
    # (the homepage fires that request itself) but the scan no longer depends on
    # them, so a throttled get_page must not fail an otherwise-healthy session.
    return bool(se)


def gp_url(lat, lng):
    return (f"{_BFF}{_GP}?latitude={lat}&longitude={lng}"
            f"&page_type=HOME&version=v2&show_new_eta_banner=true"
            f"&page_size=3&enforce_platform_type=WEB")


async def resolve_store(state, lat, lng):
    for a in range(1, 3):
        try:
            r = await state["ctx"].request.get(gp_url(lat, lng),
                                               headers=state["gp"], timeout=15000)
            if r.status == 200:
                svc = (await r.json()).get("storeServiceableResponse") or {}
                return svc.get("storeId"), [str(x) for x in
                                            (svc.get("secondaryStoreIds") or [])]
        except Exception:
            pass
        await asyncio.sleep(0.6 * a)
    return None, []


def _rs(p):
    return round((p or 0) / 100, 2)


def extract(data, at):
    rows = []
    # 'or []' not a .get default: on a page past the end of the result set Zepto
    # sends "layout": null, and the key IS present, so .get("layout", []) hands
    # back None and iterating it raises.
    for w in (data.get("layout") or []):
        if w.get("widgetId") != "PRODUCT_GRID":
            continue
        for it in (w.get("data", {}).get("resolver", {})
                   .get("data", {}).get("items") or []):
            pr = it.get("productResponse")
            if not pr:
                continue
            prod = pr.get("product", {}) or {}
            pv = pr.get("productVariant", {}) or {}
            rat = pv.get("ratingSummary", {}) or {}
            meta = pr.get("meta", {}) or {}
            brand = (prod.get("brand") or "").strip()
            mrp = pv.get("mrp") or pr.get("mrp") or 0
            sp = pr.get("discountedSellingPrice") or pr.get("sellingPrice") or 0
            rows.append({
                "position": (it.get("position") or 0) + 1,
                "brand": brand or "(unbranded)",
                "is_client": "YES" if CLIENT_BRAND.lower() in brand.lower() else "",
                "product_name": prod.get("name", ""),
                "pack_size": pv.get("formattedPacksize", ""),
                "mrp_rs": _rs(mrp), "selling_price_rs": _rs(sp),
                "discount_pct": pr.get("discountPercent", 0),
                "discount_rs": _rs(pr.get("discountAmount")),
                "in_stock": "" if pr.get("outOfStock") else "YES",
                "available_qty": pr.get("availableQuantity", 0),
                "rating": rat.get("averageRating", ""),
                "rating_count": rat.get("totalRatings", 0),
                "category": pr.get("primaryCategoryName", ""),
                "match_bucket": meta.get("query_matching_bucket", ""),
                "weight_g": pv.get("weightInGms", ""),
                "product_id": prod.get("id", ""), "variant_id": pv.get("id", ""),
                "scraped_at": at,
            })
    return rows


async def search(state, sid, all_ids, kw):
    """(rows, blocked). A small non-zero result is a real answer, not a failure."""
    got, seen, blocked = [], set(), False
    for page_no in range(MAX_PAGES):
        if len(got) >= RESULT_CAP:
            break
        body = dict(state["body"])
        body.update(query=kw, pageNumber=page_no, mode="SHOW_ALL_RESULTS")
        page_rows = None
        for a in range(1, TRIES + 1):
            h = dict(state["se"])
            h["store_id"] = h["storeid"] = sid
            h["store_ids"] = all_ids
            h["store_etas"] = json.dumps({s: -1 for s in all_ids.split(",")})
            try:
                r = await state["ctx"].request.post(f"{_BFF}{_SEARCH}", headers=h,
                                                    data=json.dumps(body),
                                                    timeout=20000)
                if r.status in (299, 401, 403, 429):
                    blocked = True
                    return got[:RESULT_CAP], True
                if r.status == 200:
                    page_rows = extract(await r.json(), state["at"])
                    break
            except Exception:
                pass
            await asyncio.sleep(1.0 * a)
        if page_rows is None:
            # Only a throttle STATUS means blocked, and that already returned
            # above. Getting here means the page never parsed — a transport or
            # payload problem. If earlier pages already yielded rows, keep them:
            # failing to read page 2 says nothing about the 26 rows in hand, and
            # discarding the store here is what made healthy stores look
            # rate limited and sent them round the retry queue.
            if got:
                break
            blocked = True
            break
        if not page_rows:
            break                      # genuinely nothing more for this term
        for x in page_rows:
            k = x["variant_id"] or (x["product_name"], x["position"])
            if k not in seen:
                seen.add(k)
                got.append(x)
    got.sort(key=lambda x: x["position"])
    return got[:RESULT_CAP], blocked


# ── Export ────────────────────────────────────────────────────────────────────

def widths(ws, mx=44):
    for col in ws.columns:
        cl = get_column_letter(col[0].column)
        ml = max((len(str(c.value)) for c in col[:400] if c.value), default=0)
        ws.column_dimensions[cl].width = min(max(ml + 2, 9), mx)


def head(ws, cols):
    ws.append(cols)
    for c in ws[1]:
        c.font, c.fill, c.alignment = HDR_FONT, HDR_FILL, HDR_ALIGN
    ws.freeze_panes = "A2"


def export():
    rows, _ = load_ckpt()
    if not rows:
        print("Checkpoint is empty — nothing to export.")
        return
    empties = [r for r in rows if r.get("brand") == NO_RESULTS]
    rows = [r for r in rows if r.get("brand") != NO_RESULTS]
    if empties:
        from collections import Counter as _C
        by_kw = _C(r["keyword"] for r in empties)
        print(f"searches that returned nothing: {len(empties)}  "
              + ", ".join(f"{k} x{v}" for k, v in by_kw.most_common()))
    for r in rows:
        for k in ("position", "available_qty", "rating_count"):
            try:
                r[k] = int(float(r[k]))
            except (TypeError, ValueError):
                r[k] = ""
        for k in ("mrp_rs", "selling_price_rs", "discount_pct", "discount_rs"):
            try:
                r[k] = float(r[k])
            except (TypeError, ValueError):
                r[k] = ""
    client = [r for r in rows if r["is_client"] == "YES"]
    stores = {r["store_id"] for r in rows}
    kws = [k for k in KEYWORDS if any(r["keyword"] == k for r in rows)]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "All Products"
    head(ws, FIELDS)
    for i, r in enumerate(sorted(rows, key=lambda x: (x["store_name"], x["keyword"],
                                                      x["position"] or 999)), start=2):
        ws.append([r.get(c, "") for c in FIELDS])
        if r["is_client"] == "YES":
            for c in range(1, len(FIELDS) + 1):
                ws.cell(row=i, column=c).fill = CLIENT_FILL
                ws.cell(row=i, column=c).font = CLIENT_FONT
        elif isinstance(r["position"], int) and r["position"] <= 3:
            ws.cell(row=i, column=FIELDS.index("position") + 1).fill = TOP3_FILL
        if r["in_stock"] != "YES":
            ws.cell(row=i, column=FIELDS.index("in_stock") + 1).font = OOS_FONT
    ws.auto_filter.ref = ws.dimensions
    widths(ws)

    ws2 = wb.create_sheet(f"{CLIENT_BRAND} Only")
    head(ws2, FIELDS)
    for i, r in enumerate(sorted(client, key=lambda x: (x["keyword"],
                                                        x["position"] or 999)), start=2):
        ws2.append([r.get(c, "") for c in FIELDS])
        for c in range(1, len(FIELDS) + 1):
            ws2.cell(row=i, column=c).fill = CLIENT_FILL
    ws2.auto_filter.ref = ws2.dimensions
    widths(ws2)

    # ── Tracked brands only ───────────────────────────────────────────────────
    # The client asked to see Brik Oven against five named competitors rather
    # than the whole SERP. 'All Products' above keeps everything; these two
    # sheets are the narrowed view.
    tr_rows = []
    for r in rows:
        t = tracked_brand(r["brand"])
        if t:
            r = dict(r)
            r["tracked_brand"] = t
            tr_rows.append(r)

    TFIELDS = ["tracked_brand"] + FIELDS
    ws2b = wb.create_sheet("Tracked Brands")
    head(ws2b, TFIELDS)
    order = {b: i for i, b in enumerate(TRACKED_ORDER)}
    for i, r in enumerate(sorted(tr_rows, key=lambda x: (x["keyword"],
                                                         x["store_name"],
                                                         x["position"] or 999)),
                          start=2):
        ws2b.append([r.get(c, "") for c in TFIELDS])
        fill = PatternFill("solid", fgColor=TRACK_FILL[r["tracked_brand"]])
        for c in range(1, len(TFIELDS) + 1):
            ws2b.cell(row=i, column=c).fill = fill
        if r["tracked_brand"] == CLIENT_BRAND:
            ws2b.cell(row=i, column=1).font = CLIENT_FONT
        if r["in_stock"] != "YES":
            ws2b.cell(row=i, column=TFIELDS.index("in_stock") + 1).font = OOS_FONT
    ws2b.auto_filter.ref = ws2b.dimensions
    widths(ws2b)

    # ── Head to head ──────────────────────────────────────────────────────────
    # One row per store x keyword, with each tracked brand's best rank side by
    # side. This is the sheet that actually answers "are we beating them here".
    ws2c = wb.create_sheet("Head to Head")
    head(ws2c, ["store_name", "pincode", "keyword", "results"]
         + TRACKED_ORDER + ["best rank", "winner", "Brik Oven beats"])
    cell = defaultdict(dict)          # (store, kw) -> brand -> best rank
    meta = {}
    for r in tr_rows:
        k = (r["store_name"], r["keyword"])
        p = r["position"] if isinstance(r["position"], int) else None
        if p is None:
            continue
        b = r["tracked_brand"]
        if b not in cell[k] or p < cell[k][b]:
            cell[k][b] = p
    for r in rows:
        k = (r["store_name"], r["keyword"])
        meta.setdefault(k, {"pincode": r["pincode"], "n": 0})
        meta[k]["n"] += 1

    hh = 0
    for i, (k, brands) in enumerate(
            sorted(cell.items(), key=lambda kv: (kv[0][1], kv[0][0])), start=2):
        store, kw = k
        ranks = [(b, p) for b, p in brands.items()]
        best_b, best_p = min(ranks, key=lambda x: x[1])
        mine = brands.get(CLIENT_BRAND)
        # "n of m" rather than a bare count: 0 competitors present and 0 beaten
        # are very different stories, and a lone "0" reads like a loss.
        rivals = [p for b, p in ranks if b != CLIENT_BRAND]
        beats = ("absent" if mine is None
                 else f"{sum(1 for p in rivals if mine < p)} of {len(rivals)}")
        ws2c.append([store, meta.get(k, {}).get("pincode", ""), kw,
                     meta.get(k, {}).get("n", ""),
                     *[brands.get(b, "") for b in TRACKED_ORDER],
                     best_p, best_b, beats])
        # colour the winning brand's cell, and the client's own cell always
        wcol = 4 + TRACKED_ORDER.index(best_b) + 1
        ws2c.cell(row=i, column=wcol).fill = PatternFill(
            "solid", fgColor=TRACK_FILL[best_b])
        ws2c.cell(row=i, column=wcol).font = Font(name="Calibri", size=10, bold=True)
        if mine is not None:
            ws2c.cell(row=i, column=5).fill = PatternFill("solid",
                                                          fgColor=TRACK_FILL[CLIENT_BRAND])
        if best_b == CLIENT_BRAND:
            hh += 1
    ws2c.auto_filter.ref = ws2c.dimensions
    ws2c.freeze_panes = "D2"
    widths(ws2c)

    ws3 = wb.create_sheet("By Keyword")
    head(ws3, ["keyword", "stores", "results", f"{CLIENT_BRAND} SKUs",
               "best rank", "avg rank", "SoV %", "stores stocking", "top competitor"])
    for i, kw in enumerate(kws, start=2):
        rs = [r for r in rows if r["keyword"] == kw]
        cl = [r for r in rs if r["is_client"] == "YES"]
        st = {r["store_id"] for r in rs}
        stocking = {r["store_id"] for r in cl}
        comp = Counter(r["brand"] for r in rs if r["is_client"] != "YES")
        ranks = [r["position"] for r in cl if isinstance(r["position"], int)]
        ws3.append([kw, len(st), len(rs), len(cl),
                    min(ranks) if ranks else "",
                    round(sum(ranks) / len(ranks), 1) if ranks else "",
                    round(len(cl) / len(rs) * 100, 1) if rs else 0,
                    len(stocking),
                    comp.most_common(1)[0][0] if comp else ""])
        if i % 2 == 0:
            for c in range(1, 10):
                ws3.cell(row=i, column=c).fill = ALT_FILL
    widths(ws3)

    ws4 = wb.create_sheet("Brand Ranking")
    head(ws4, ["brand", "is_client", "appearances", "SoV %", "avg rank",
               "best rank", "avg price", "min price", "max price"])
    bybrand = defaultdict(list)
    for r in rows:
        bybrand[r["brand"]].append(r)
    tot = len(rows) or 1
    for i, (b, rs) in enumerate(sorted(bybrand.items(), key=lambda kv: -len(kv[1])),
                                start=2):
        ranks = [r["position"] for r in rs if isinstance(r["position"], int)]
        pz = [r["selling_price_rs"] for r in rs
              if isinstance(r["selling_price_rs"], float) and r["selling_price_rs"]]
        isc = rs[0]["is_client"] == "YES"
        ws4.append([b, "YES" if isc else "", len(rs), round(len(rs) / tot * 100, 2),
                    round(sum(ranks) / len(ranks), 1) if ranks else "",
                    min(ranks) if ranks else "",
                    round(sum(pz) / len(pz), 2) if pz else "",
                    min(pz) if pz else "", max(pz) if pz else ""])
        if isc:
            for c in range(1, 10):
                ws4.cell(row=i, column=c).fill = CLIENT_FILL
                ws4.cell(row=i, column=c).font = CLIENT_FONT
        elif i % 2 == 0:
            for c in range(1, 10):
                ws4.cell(row=i, column=c).fill = ALT_FILL
    widths(ws4)

    ws5 = wb.create_sheet("Summary")
    head(ws5, ["Metric", "Value"])
    ranks = [r["position"] for r in client if isinstance(r["position"], int)]
    pz = [r["selling_price_rs"] for r in client
          if isinstance(r["selling_price_rs"], float) and r["selling_price_rs"]]
    stats = [("Client brand", CLIENT_BRAND), ("City", CITY),
             ("Stores covered", len(stores)),
             ("Keywords covered", len(kws)),
             ("Searches completed", len({(r["store_id"], r["keyword"]) for r in rows})),
             ("Product rows", len(rows)),
             (f"{CLIENT_BRAND} rows", len(client)),
             (f"{CLIENT_BRAND} SoV",
              f"{len(client)/len(rows)*100:.1f}%" if rows else "—"),
             ("Distinct brands", len(bybrand)),
             (f"{CLIENT_BRAND} best rank", min(ranks) if ranks else "—"),
             (f"{CLIENT_BRAND} avg rank",
              f"{sum(ranks)/len(ranks):.1f}" if ranks else "—"),
             (f"{CLIENT_BRAND} avg price",
              f"Rs{sum(pz)/len(pz):.2f}" if pz else "—"),
             ("Exported at", datetime.now().strftime("%d-%b-%Y %H:%M")),
             ("", ""),
             ("— Tracked brands —", "appearances / best / avg rank")]
    for b in TRACKED_ORDER:
        rs = [r for r in tr_rows if r["tracked_brand"] == b]
        rk = [r["position"] for r in rs if isinstance(r["position"], int)]
        if not rs:
            v = "not found on any SERP"
        elif rk:
            v = f"{len(rs)} / #{min(rk)} / {sum(rk)/len(rk):.1f}"
        else:
            v = f"{len(rs)} / — / —"
        stats.append((b, v))
    stats += [("", ""),
              ("Head-to-head rows", len(cell)),
              (f"{CLIENT_BRAND} ranks best", f"{hh} of {len(cell)}"
               + (f"  ({hh/len(cell)*100:.0f}%)" if cell else ""))]
    for i, (m, v) in enumerate(stats, start=2):
        ws5.cell(row=i, column=1, value=m)
        ws5.cell(row=i, column=2, value=v)
        if i % 2 == 0:
            for c in (1, 2):
                ws5.cell(row=i, column=c).fill = ALT_FILL
    widths(ws5)

    T = f"{CLIENT_BRAND} keyword scan — {CITY} — {len(stores)} stores, {len(kws)} keywords"
    S = (f"{len(rows)} product rows  |  {len(client)} {CLIENT_BRAND} rows  |  "
         f"green = {CLIENT_BRAND}, amber = top-3 rank")
    for w in (ws, ws2, ws2b, ws2c, ws3, ws4, ws5):
        w.insert_rows(1, amount=2)
        mc = w.max_column
        w["A1"], w["A2"] = T, S
        w["A1"].font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
        w["A1"].fill = TITLE_FILL
        w["A1"].alignment = Alignment(horizontal="left", vertical="center")
        w.row_dimensions[1].height = 22
        w["A2"].font = Font(name="Calibri", size=10, italic=True, color="444444")
        w["A2"].fill = ALT_FILL
        w.merge_cells(start_row=1, start_column=1, end_row=1, end_column=mc)
        w.merge_cells(start_row=2, start_column=1, end_row=2, end_column=mc)

    ts = datetime.now().strftime("%Y%m%d_%H%M")
    suffix = f"_{TAG}" if TAG else ""
    out = BACKEND / f"zepto_keyword_scan_{CITY.lower()}{suffix}_{ts}.xlsx"
    wb.save(out)
    print(f"Rows {len(rows)} | stores {len(stores)} | {CLIENT_BRAND} rows {len(client)}")
    print(f"Tracked-brand rows {len(tr_rows)} of {len(rows)}  "
          f"| head-to-head {len(cell)} store x keyword"
          + (f" | {CLIENT_BRAND} best in {hh} ({hh/len(cell)*100:.0f}%)"
             if cell else ""))
    for b in TRACKED_ORDER:
        n = sum(1 for r in tr_rows if r["tracked_brand"] == b)
        print(f"    {b:22s} {n:5d}" + ("   <- not on any SERP yet" if not n else ""))
    print(f"Saved -> {out}")


def status():
    stores, src = load_stores()
    rows, done = load_ckpt()
    want = len(stores) * len(KEYWORDS)
    per = Counter(s for s, _ in done)
    full = [s for s in stores if per.get(s["store_id"], 0) >= len(KEYWORDS)]
    print(f"Store source : {src}")
    print(f"Stores       : {len(stores)}   keywords: {len(KEYWORDS)}")
    print(f"Searches     : {len(done)}/{want}  ({len(done)/want*100:.1f}%)")
    print(f"Stores done  : {len(full)}/{len(stores)}")
    print(f"Rows banked  : {len(rows)}")
    print(f"Checkpoint   : {CKPT.name}")
    if len(done) < want:
        print(f"\nRe-run to continue:  python -m scripts.zepto_keyword_scan")
    else:
        print(f"\nCOMPLETE — build the workbook:")
        print(f"   python -m scripts.zepto_keyword_scan --export")


# ── Main ──────────────────────────────────────────────────────────────────────

async def probe_open(state):
    """One cheap search to test whether the window has reopened."""
    body = dict(state["body"])
    body.update(query="bread", pageNumber=0, mode="SHOW_ALL_RESULTS")
    sid = state.get("probe_store") or ""
    h = dict(state["se"])
    if sid:
        h["store_id"] = h["storeid"] = h["store_ids"] = sid
        h["store_etas"] = json.dumps({sid: -1})
    try:
        r = await state["ctx"].request.post(f"{_BFF}{_SEARCH}", headers=h,
                                            data=json.dumps(body), timeout=20000)
        if r.status != 200:
            return False
        d = await r.json()
        for w in d.get("layout", []):
            if w.get("widgetId") == "PRODUCT_GRID":
                if w.get("data", {}).get("resolver", {}).get("data", {}).get("items"):
                    return True
        return False
    except Exception:
        return False


async def wait_for_window(browser, state, stall):
    """Sleep out a block, probing periodically. True once search works again."""
    wait = RECOVERY_WAITS_S[min(stall, len(RECOVERY_WAITS_S) - 1)]
    print()
    print(f"        [window shut — waiting up to {wait//60} min, "
          f"probing every {PROBE_EVERY_S//60} min]")
    waited = 0
    while waited < wait:
        await asyncio.sleep(PROBE_EVERY_S)
        waited += PROBE_EVERY_S
        await open_session(browser, state)     # fresh session for the probe
        if await probe_open(state):
            print(f"        [window reopened after {waited//60} min — resuming]")
            return True
        print(f"        [still shut at {waited//60} min]")
    return False


async def run():
    stores, src = load_stores()
    rows_done, done = load_ckpt()
    at = datetime.now().strftime("%d-%b-%Y %H:%M")
    want = len(stores) * len(KEYWORDS)

    todo = [(s, [k for k in KEYWORDS if (s["store_id"], k) not in done])
            for s in stores]
    todo = [(s, ks) for s, ks in todo if ks]

    print(f"Store source : {src}")
    print(f"Client       : {CLIENT_BRAND}   city: {CITY}")
    print(f"Stores       : {len(stores)}   keywords: {len(KEYWORDS)}")
    print(f"Already done : {len(done)}/{want} searches")
    print(f"This run     : {len(todo)} stores remaining")
    print("=" * 96)
    if not todo:
        print("Nothing left — everything already collected.")
        print("   python -m scripts.zepto_keyword_scan --export")
        return

    banked = 0
    bad_streak = 0
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        state = {"gp": {}, "se": {}, "body": {}, "ctx": None, "page": None, "at": at,
                 "searches": 0, "block_points": []}
        print("Opening session...")
        if not await open_session(browser, state):
            print("Header capture failed — try again shortly.")
            await browser.close()
            return
        print()

        pending = []          # stores deferred by a block, retried after recovery
        attempts = {}         # store_id -> re-queue count, so a store that keeps
                              # blocking cannot loop forever
        stalls = 0
        queue = list(todo)
        idx = 0
        while queue or pending:
            if not queue:
                queue, pending = pending, []
            st, kws = queue.pop(0)
            idx += 1
            # No get_page here. The search API binds to a store purely through the
            # store_id/storeid/store_ids/store_etas headers — lat/lng in the body is
            # ignored — and the catalog already holds every store_id and its
            # secondaries. Calling get_page per store bought only a SAME/DIFF
            # re-confirmation, at the cost of doubling the request count and
            # spending a SECOND, independently-rate-limited budget. When that budget
            # ran out the store was skipped even though search was perfectly healthy.
            sid, sec = st["store_id"], st.get("secondary") or []
            print(f"[{idx}/{len(todo)}] {st['store_name'][:28]:<29} "
                  f"{st['pincode']:<7} {sid[:8]}")

            store_rows, blocked_here, done_kws = [], False, []
            all_ids = ",".join([sid] + sec)

            # Scheduled rest before the window can close on us.
            if state["searches"] >= PAUSE_EVERY:
                done_n = len(load_ckpt()[1])
                print(f"        [scheduled rest {PAUSE_S//60} min after "
                      f"{state['searches']} searches — {done_n} banked]")
                await asyncio.sleep(PAUSE_S)
                state["searches"] = 0
                await open_session(browser, state)   # fresh session after the rest
                print("        [resuming]")

            for kw in kws:
                res, blocked = await search(state, sid, all_ids, kw)
                state["searches"] += 1
                if blocked:
                    blocked_here = True
                    state["block_points"].append(state["searches"])
                    print(f"        {kw:<20} BLOCKED "
                          f"(after {state['searches']} searches this window)")
                    break
                cl = [r for r in res if r["is_client"] == "YES"]
                best = min((r["position"] for r in cl), default=None)
                done_kws.append(kw)
                if not res:
                    # A genuinely empty result ('sourdough bread loaf' returns 0 at
                    # many stores) still has to be RECORDED, or the pair never
                    # enters the checkpoint and is retried on every future run —
                    # an infinite loop that also masks the real finding, which is
                    # that the term returns nothing there. One sentinel row marks
                    # it done; export drops these.
                    store_rows.append({
                        "store_id": st["store_id"], "store_name": st["store_name"],
                        "pincode": st["pincode"], "area": st["area"],
                        "lat": st["lat"], "lng": st["lng"], "keyword": kw,
                        "brand": NO_RESULTS, "is_client": "", "position": "",
                        "product_name": "", "scraped_at": state["at"],
                    })
                tag = (f"{CLIENT_BRAND} #{best} ({len(cl)} SKU)" if cl
                       else f"no {CLIENT_BRAND}")
                print(f"        {kw:<20} {len(res):>2} products | {tag}")
                for r in res:
                    store_rows.append({**r, "keyword": kw,
                                       "store_id": st["store_id"],
                                       "store_name": st["store_name"],
                                       "pincode": st["pincode"], "area": st["area"],
                                       "lat": st["lat"], "lng": st["lng"]})
                await asyncio.sleep(SEARCH_GAP_S)

            # bank whatever this store produced, immediately
            if store_rows:
                append_ckpt(store_rows)
                banked += len(store_rows)

            # A store blocked PARTWAY through leaves the rest of its keywords
            # uncollected. Re-queue exactly those, so one run finishes the job —
            # a production scrape cannot depend on someone noticing an incomplete
            # total and running the command again.
            remaining = [k for k in kws if k not in done_kws]
            if remaining and blocked_here:
                n = attempts.get(st["store_id"], 0) + 1
                attempts[st["store_id"]] = n
                if n < MAX_STORE_ATTEMPTS:
                    pending.append((st, remaining))
                    print(f"        [{len(remaining)} keywords re-queued "
                          f"(attempt {n}/{MAX_STORE_ATTEMPTS})]")
                else:
                    print(f"        [giving up on {len(remaining)} keywords after "
                          f"{n} attempts]")

            if blocked_here and not store_rows:
                bad_streak += 1
                if bad_streak >= STOP_AFTER_BAD_STORES:
                    state["probe_store"] = sid
                    if await wait_for_window(browser, state, stalls):
                        stalls = 0
                        state["searches"] = 0     # the window reopened
                        bad_streak = 0
                        pending.append((st, kws))     # redo this store
                        continue
                    stalls += 1
                    if stalls >= MAX_CONSECUTIVE_STALLS:
                        print()
                        print("=" * 96)
                        print("Window has not reopened after repeated waits — stopping.")
                        print(f"{banked} rows banked to {CKPT.name}; nothing lost.")
                        print("Re-run the same command later to continue.")
                        break
                    bad_streak = 0
                    pending.append((st, kws))
                    continue
            else:
                bad_streak = 0
            await asyncio.sleep(STORE_GAP_S)

        await browser.close()

    rows_done, done = load_ckpt()
    print()
    if state.get("block_points"):
        bp = state["block_points"]
        print(f"BLOCK POINTS this run : {bp}")
        print(f"  earliest {min(bp)}, median {sorted(bp)[len(bp)//2]} searches "
              f"into a window")
        print(f"  -> set PAUSE_EVERY a little BELOW {min(bp)} for the next run")
    else:
        print(f"No blocks — PAUSE_EVERY={PAUSE_EVERY} is holding. It could go higher.")
    print()
    print(f"Searches complete : {len(done)}/{want}  ({len(done)/want*100:.1f}%)")
    print(f"Rows banked       : {len(rows_done)}")
    if len(done) >= want:
        print("\nALL DONE — build the workbook:")
        print("   python -m scripts.zepto_keyword_scan --export")
    else:
        print("\nRe-run later to continue:")
        print("   python -m scripts.zepto_keyword_scan")


def main():
    # --tag NAME runs against its own checkpoint and its own workbook, so a
    # timing experiment cannot touch a completed run's data. Without it the
    # paths are exactly as before, so the default behaviour is unchanged.
    #     python -m scripts.zepto_keyword_scan --tag test1
    #     python -m scripts.zepto_keyword_scan --tag test1 --export
    global CKPT, TAG
    if "--tag" in sys.argv:
        i = sys.argv.index("--tag")
        if i + 1 >= len(sys.argv):
            raise SystemExit("--tag needs a name, e.g. --tag test1")
        TAG = sys.argv[i + 1].strip().replace(" ", "_")
        CKPT = BACKEND / f"keyword_scan_checkpoint_{TAG}.csv"
        print(f"[tag {TAG}]  checkpoint: {CKPT.name}")

    if "--status" in sys.argv:
        status()
    elif "--export" in sys.argv:
        export()
    else:
        asyncio.run(run())


if __name__ == "__main__":
    main()
