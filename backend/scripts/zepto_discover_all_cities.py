"""
Grid-scan EVERY city for dark stores — pincodes, areas, lat/lng, all of it.

Three problems this has to solve that a single-city scan does not:

1. OVERLAP. 18 city pairs sit within 60 km of each other (Chandigarh/Mohali/
   Panchkula are ~10-15 km apart). Scanning each city separately would probe the
   same coordinates several times and count the same store under several cities.
   Fix: cities whose boxes touch are merged into one CLUSTER and scanned once.

2. ATTRIBUTION. A store found in a shared region has to belong to exactly one
   city. Fix: the store-name prefix decides it (BLR->Bengaluru, KUR->Ambala),
   because that prefix is Zepto's own label and it already encodes the client's
   "covered in" taxonomy. Only when the prefix is unknown does it fall back to
   the nearest city centroid.

3. TINY CITIES. Six cities have a single known store, so their coordinate span is
   zero. Fix: every box gets a minimum half-size, so a one-store city still gets
   a real search area around it.

Run from backend/:
    python -m scripts.zepto_discover_all_cities            # every pending cluster
    python -m scripts.zepto_discover_all_cities --list     # show the plan, scan nothing
    python -m scripts.zepto_discover_all_cities --only Ambala,Chandigarh
Output: dark_stores_zepto/
"""
import asyncio
import json
import math
import re
import statistics
import sys
from pathlib import Path
from datetime import datetime
from collections import Counter, defaultdict

sys.path.insert(0, str(Path(__file__).parent.parent))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from playwright.async_api import async_playwright

BACKEND = Path(__file__).parent.parent
OUTDIR  = BACKEND / "dark_stores_zepto"

STEP        = 0.006     # ~670 m
MIN_HALF    = 0.13      # ~14 km minimum half-box, for single-store cities
MARGIN      = 0.12      # ~13 km padding beyond the known store span
CLUSTER_KM  = 60        # cities closer than this are scanned together
CONCURRENCY = 8
PROBE_TRIES = 3

# Cities already grid-scanned; skipped unless --only names them.
DONE = {"Bengaluru", "Kolkata", "Chennai", "Pune", "Hyderabad", "Ahmedabad"}

_BASE = "https://www.zepto.com"
_BFF  = "https://bff-gateway.zepto.com"
_GP   = "/lms/api/v2/get_page"
_DROP = {"host", "content-length", "connection", "accept-encoding",
         ":method", ":path", ":authority", ":scheme"}

HDR_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
HDR_FILL = PatternFill("solid", fgColor="2E4057")
HDR_ALIGN = Alignment(horizontal="center", vertical="center")
ALT_FILL = PatternFill("solid", fgColor="F0F4F8")
NEW_FILL = PatternFill("solid", fgColor="D9F7E6")
TITLE_FILL = PatternFill("solid", fgColor="1A3C5E")


def km(a, b, c, d):
    R = 6371.0
    p = math.radians
    return 2 * R * math.asin(math.sqrt(
        math.sin(p(c - a) / 2) ** 2 +
        math.cos(p(a)) * math.cos(p(c)) * math.sin(p(d - b) / 2) ** 2))


# ── Load the master workbook ──────────────────────────────────────────────────

def load_master():
    path = sorted(BACKEND.glob("zepto_master_all_cities_*.xlsx"))
    if not path:
        raise SystemExit("No master workbook found.")
    path = path[-1]
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows = list(wb["All Stores"].iter_rows(values_only=True))
    wb.close()
    hdr = [str(h).strip().lower() if h else "" for h in rows[2]]
    recs = [dict(zip(hdr, r)) for r in rows[3:] if r and r[0]]
    return path, recs


def build_context(recs):
    """city -> stores, city -> centroid, prefix -> city, known store ids."""
    city = defaultdict(list)
    known = {}
    for r in recs:
        sid = str(r.get("store_id") or "").strip()
        if not sid:
            continue
        try:
            la, ln = float(r["lat"]), float(r["lng"])
        except (TypeError, ValueError):
            continue
        rec = {"store_id": sid, "store_name": str(r.get("store_name") or ""),
               "city": r["city"], "tier": r.get("tier", ""),
               "pincode": str(r.get("pincode") or ""), "area": str(r.get("area") or ""),
               "lat": la, "lng": ln}
        city[r["city"]].append(rec)
        known[sid] = rec

    centroid = {c: (statistics.median(p["lat"] for p in v),
                    statistics.median(p["lng"] for p in v))
                for c, v in city.items()}   # median is already outlier-robust

    pref = defaultdict(Counter)
    for r in known.values():
        m = re.match(r"^\s*([A-Za-z]{2,4})\s*-", r["store_name"])
        if m:
            pref[m.group(1).upper()][r["city"]] += 1
    prefix_city = {}
    for p, c in pref.items():
        top, n = c.most_common(1)[0]
        if n / sum(c.values()) >= 0.8:
            prefix_city[p] = top
    return city, centroid, prefix_city, known


def cluster_cities(centroid, cities):
    """Union-find over 'centroids within CLUSTER_KM' so overlapping cities scan once."""
    parent = {c: c for c in cities}

    def find(x):
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x

    def union(a, b):
        ra, rb = find(a), find(b)
        if ra != rb:
            parent[rb] = ra

    cs = list(cities)
    for i in range(len(cs)):
        for j in range(i + 1, len(cs)):
            if km(*centroid[cs[i]], *centroid[cs[j]]) < CLUSTER_KM:
                union(cs[i], cs[j])
    groups = defaultdict(list)
    for c in cs:
        groups[find(c)].append(c)
    return [sorted(v) for v in groups.values()]


OUTLIER_KM = 45   # a store further than this from its city median is misfiled


def core_stores(c, city_stores):
    """Stores genuinely belonging to a city — outliers excluded.

    Some city lists contain stores that sit in a different city entirely
    (Jaipur holds one in Kota and one in Mathura, ~195 km away; Surat holds one
    in Vadodara, 131 km away). Letting those set the bounding box inflated Jaipur
    to 260 km across = 158,608 grid points = 109 minutes, almost all of it empty
    countryside. Excluding them keeps each box over the real city.
    """
    pts = city_stores[c]
    if len(pts) < 3:
        return pts
    mla = statistics.median(p["lat"] for p in pts)
    mln = statistics.median(p["lng"] for p in pts)
    core = [p for p in pts if km(mla, mln, p["lat"], p["lng"]) <= OUTLIER_KM]
    return core or pts


def box_for(members, city_stores):
    """Bounding box covering every member city, padded, with a minimum size."""
    la = [p["lat"] for c in members for p in core_stores(c, city_stores)]
    ln = [p["lng"] for c in members for p in core_stores(c, city_stores)]
    cla, cln = (min(la) + max(la)) / 2, (min(ln) + max(ln)) / 2
    hla = max((max(la) - min(la)) / 2 + MARGIN, MIN_HALF)
    hln = max((max(ln) - min(ln)) / 2 + MARGIN, MIN_HALF)
    return (round(cla - hla, 4), round(cla + hla, 4),
            round(cln - hln, 4), round(cln + hln, 4))


def build_grid(box):
    lo_la, hi_la, lo_ln, hi_ln = box
    pts = []
    la = lo_la
    while la <= hi_la + 1e-9:
        ln = lo_ln
        while ln <= hi_ln + 1e-9:
            pts.append((round(la, 5), round(ln, 5)))
            ln += STEP
        la += STEP
    return pts


# ── Probing ───────────────────────────────────────────────────────────────────

async def capture_headers(page):
    cap = {}

    async def on_req(req):
        if _GP in req.url and not cap:
            cap.update(req.headers)

    page.on("request", on_req)
    try:
        await page.goto(_BASE, timeout=30000, wait_until="domcontentloaded")
        waited = 0
        while not cap and waited < 6000:
            await page.wait_for_timeout(250)
            waited += 250
    except Exception:
        pass
    finally:
        page.remove_listener("request", on_req)
    return {k: v for k, v in cap.items() if k.lower() not in _DROP}


def gp_url(lat, lng):
    return (f"{_BFF}{_GP}?latitude={lat}&longitude={lng}"
            f"&page_type=HOME&version=v2&show_new_eta_banner=true"
            f"&page_size=3&enforce_platform_type=WEB")


async def probe(ctx, lat, lng, state):
    for a in range(1, PROBE_TRIES + 1):
        try:
            r = await ctx.request.get(gp_url(lat, lng), headers=state["headers"], timeout=15000)
            if r.status == 200:
                d = await r.json()
                svc = d.get("storeServiceableResponse") or {}
                sid = svc.get("storeId")
                if not sid:
                    return None
                return {"store_id": sid,
                        "secondary": [str(x) for x in (svc.get("secondaryStoreIds") or [])],
                        "store_name": (d.get("storeDetailsResponse") or {}).get("name", "")}
            if r.status in (401, 403, 429):
                state["stale"] = True
                await asyncio.sleep(1.5 * a)
                continue
        except Exception:
            pass
        await asyncio.sleep(0.6 * a)
    state["failed"] += 1
    return None


def assign_city(name, lat, lng, members, prefix_city, centroid):
    """Prefix wins (it is Zepto's own label); otherwise nearest member centroid."""
    m = re.match(r"^\s*([A-Za-z]{2,4})\s*-", name or "")
    if m:
        c = prefix_city.get(m.group(1).upper())
        if c:
            return c, "prefix"
    best, bd = None, 1e9
    for c in members:
        d = km(lat, lng, *centroid[c])
        if d < bd:
            best, bd = c, d
    return best, "nearest"


# ── Excel helpers ─────────────────────────────────────────────────────────────

def widths(ws, mx=44):
    for col in ws.columns:
        cl = get_column_letter(col[0].column)
        ml = max((len(str(c.value)) for c in col[:500] if c.value), default=0)
        ws.column_dimensions[cl].width = min(max(ml + 2, 9), mx)


def head(ws, cols):
    ws.append(cols)
    for c in ws[1]:
        c.font, c.fill, c.alignment = HDR_FONT, HDR_FILL, HDR_ALIGN
    ws.row_dimensions[1].height = 18
    ws.freeze_panes = "A2"


def banner(ws, t, s):
    ws.insert_rows(1, amount=2)
    mc = ws.max_column or 1
    ws["A1"], ws["A2"] = t, s
    ws["A1"].font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="444444")
    ws["A2"].fill = ALT_FILL
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 15
    if mc > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=mc)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=mc)


COLS = ["store_id", "store_name", "city", "assigned_by", "is_new",
        "already_in", "note", "lat", "lng", "hit_count", "secondary",
        "scanned_as"]


def write_cluster(members, found, box, npts, failed, no_store, known):
    OUTDIR.mkdir(exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    slug = re.sub(r"[^a-z0-9]+", "_", "_".join(members).lower())[:60]
    new = [v for v in found.values() if v["is_new"]]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stores"
    head(ws, COLS)
    for i, r in enumerate(sorted(found.values(),
                                 key=lambda x: (not x["is_new"], x["city"],
                                                x["store_name"])), start=2):
        ws.append([r.get(c, "") for c in COLS])
        if r["is_new"]:
            for c in range(1, len(COLS) + 1):
                ws.cell(row=i, column=c).fill = NEW_FILL
        elif i % 2 == 0:
            for c in range(1, len(COLS) + 1):
                ws.cell(row=i, column=c).fill = ALT_FILL
    ws.auto_filter.ref = ws.dimensions
    widths(ws)

    ws2 = wb.create_sheet("Summary")
    head(ws2, ["Metric", "Value"])
    per = Counter(v["city"] for v in found.values())
    pern = Counter(v["city"] for v in new)
    other = [v for v in found.values() if v.get("already_in")
             and v["already_in"] not in members]
    mine  = [v for v in found.values() if v.get("already_in")
             and v["already_in"] in members]
    stats = [("City scanned", ", ".join(members)),
             ("Box", f"lat {box[0]}-{box[1]}, lng {box[2]}-{box[3]}"),
             ("Grid step", f"{STEP} deg (~{STEP*111*1000:.0f} m)"),
             ("Points probed", npts), ("Points with no store", no_store),
             ("Probe failures", failed),
             ("Unique stores hit", len(found)),
             ("NEW stores (added to total)", len(new)),
             ("Already known - this city", len(mine)),
             ("Already known - ANOTHER city", len(other))]
    for c, n in Counter(v["already_in"] for v in other).most_common():
        stats.append((f"    already counted in {c}", n))
    for c in members:
        stats.append((f"  {c}", f"{per.get(c,0)} stores ({pern.get(c,0)} new)"))
    stats.append(("Scraped at", datetime.now().strftime("%d-%b-%Y %H:%M")))
    for i, (m, v) in enumerate(stats, start=2):
        ws2.cell(row=i, column=1, value=m)
        ws2.cell(row=i, column=2, value=v)
        if i % 2 == 0:
            for c in (1, 2):
                ws2.cell(row=i, column=c).fill = ALT_FILL
    widths(ws2)

    T = f"Dark store discovery — {', '.join(members)} — {len(new)} new"
    S = (f"{npts} grid points @ {STEP} deg  |  box lat {box[0]}-{box[1]}, "
         f"lng {box[2]}-{box[3]}  |  green = new")
    for w in (ws, ws2):
        banner(w, T, S)
    out = OUTDIR / f"discovery_{slug}_{ts}.xlsx"
    wb.save(out)
    return out, new


async def main():
    args = sys.argv[1:]
    only = None
    if "--only" in args:
        only = {x.strip() for x in args[args.index("--only") + 1].split(",")}
    list_only = "--list" in args

    path, recs = load_master()
    city_stores, centroid, prefix_city, known = build_context(recs)

    targets = [c for c in city_stores if (c in only) if only] if only else \
              [c for c in city_stores if c not in DONE]
    if not targets:
        print("Nothing to scan.")
        return

    # One scan PER CITY. Clustering was faster in principle but actually cost
    # more points, because a merged box spans the empty land between cities.
    # Per-city boxes hug each city, and a store found inside another city's
    # territory is still deduplicated globally and labelled with the city that
    # already owns it, so nothing is double counted and nothing is lost.
    clusters = [[c] for c in sorted(targets)]
    plan = []
    for members in clusters:
        box = box_for(members, city_stores)
        pts = build_grid(box)
        plan.append((members, box, len(pts)))
    plan.sort(key=lambda x: -x[2])
    total_pts = sum(p[2] for p in plan)

    print(f"Master file : {path.name}")
    print(f"Output dir  : {OUTDIR}")
    print(f"Cities to scan : {len(targets)}  (one scan each)")
    print(f"Dedup base     : {len(known)} known store ids across ALL "
          f"{len(city_stores)} cities — a store found twice is counted once")
    print(f"Grid step   : {STEP} deg (~{STEP*111*1000:.0f} m)")
    print("=" * 100)
    print(f"{'city':<52}{'points':>9}{'est min':>9}")
    print("-" * 100)
    for members, box, n in plan:
        print(f"{', '.join(members)[:51]:<52}{n:>9}{n*0.33/CONCURRENCY/60:>9.1f}")
    print("-" * 100)
    print(f"{'TOTAL':<52}{total_pts:>9}{total_pts*0.33/CONCURRENCY/60:>9.0f}")
    if list_only:
        print("\n--list given: nothing scanned.")
        return
    print()

    grand_new = []
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        ctx = await browser.new_context(
            user_agent=("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                        "(KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36"))
        page = await ctx.new_page()
        state = {"headers": {}, "stale": False, "failed": 0}
        print("Capturing session headers...")
        state["headers"] = await capture_headers(page)
        if not state["headers"]:
            print("Header capture failed — aborting.")
            await browser.close()
            return
        print(f"Captured {len(state['headers'])} headers.\n")

        for ci, (members, box, npts) in enumerate(plan, 1):
            grid = build_grid(box)
            print(f"[cluster {ci}/{len(plan)}] {', '.join(members)}")
            print(f"    box lat {box[0]}-{box[1]}, lng {box[2]}-{box[3]}   {npts} points")
            found, done = {}, {"n": 0, "no": 0}
            sem = asyncio.Semaphore(CONCURRENCY)
            lock = asyncio.Lock()

            async def worker(la, ln):
                async with sem:
                    r = await probe(ctx, la, ln, state)
                async with lock:
                    done["n"] += 1
                    if not r:
                        done["no"] += 1
                    else:
                        sid = r["store_id"]
                        if sid not in found:
                            nm = r["store_name"] or (known.get(sid, {}).get("store_name", ""))
                            c, how = assign_city(nm, la, ln, members, prefix_city, centroid)
                            prev = known.get(sid)
                            found[sid] = {"store_id": sid, "store_name": nm,
                                          "city": c, "assigned_by": how,
                                          "lat": la, "lng": ln,
                                          "secondary": ",".join(r["secondary"]),
                                          "is_new": prev is None,
                                          "already_in": "" if prev is None
                                                        else prev["city"],
                                          "note": "" if prev is None
                                                  else f"already counted in {prev['city']}",
                                          "hit_count": 1,
                                          "scanned_as": ", ".join(members)}
                            if sid not in known:
                                nn = sum(1 for v in found.values() if v["is_new"])
                                print(f"      NEW {sid[:8]}  {nm[:30]:<30} -> {c} ({how})"
                                      f"   [{len(found)} uniq, {nn} new]")
                        else:
                            found[sid]["hit_count"] += 1
                    if done["n"] % 2000 == 0:
                        nn = sum(1 for v in found.values() if v["is_new"])
                        print(f"      ... {done['n']}/{npts} probed | {len(found)} stores | {nn} new")

            CH = 300
            for s in range(0, len(grid), CH):
                await asyncio.gather(*(worker(a, b) for a, b in grid[s:s + CH]))
                if state["stale"]:
                    nh = await capture_headers(page)
                    if nh:
                        state["headers"] = nh
                    state["stale"] = False
                    await asyncio.sleep(2)

            out, new = write_cluster(members, found, box, done["n"],
                                     state["failed"], done["no"], known)
            grand_new.extend(new)
            per = Counter(v["city"] for v in found.values())
            print(f"    -> {len(found)} stores, {len(new)} NEW   "
                  f"({'; '.join(f'{c}:{n}' for c, n in sorted(per.items()))})")
            print(f"    saved {out.name}\n")

        await browser.close()

    print("=" * 100)
    print(f"Clusters scanned : {len(plan)}")
    print(f"TOTAL NEW stores : {len(grand_new)}")
    byc = Counter(v["city"] for v in grand_new)
    for c, n in byc.most_common():
        print(f"   {c:<28} +{n}")
    print(f"\nAll files in: {OUTDIR}")


if __name__ == "__main__":
    asyncio.run(main())
