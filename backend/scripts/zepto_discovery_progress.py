"""
Progress across the 58-city discovery run.

Reports FRESH per-city counts — the stores a scan actually saw and assigned to
that city. Deliberately does NOT report a running national total: until all 58
cities are scanned, any such figure mixes fresh and stale data and overstates
how much is actually verified.

"FRESH" is not old+new, because old+new is wrong in two directions:
  - a scan can miss a previously-known store (closed, or outside the box)
  - a store can move city on re-attribution (HBL-Vidhya Nagar: Belagavi -> Hubballi)

Run from backend/:
    python -m scripts.zepto_discovery_progress
"""
import re
import sys
from pathlib import Path
from collections import Counter

sys.path.insert(0, str(Path(__file__).parent.parent))

import openpyxl

BACKEND = Path(__file__).parent.parent
OUTDIR = BACKEND / "dark_stores_zepto"

ALL_CITIES = [
    "Delhi NCR", "Belagavi", "Mumbai", "Kolkata", "Bengaluru", "Chennai",
    "Dehradun", "Hyderabad", "Pune", "Jalandhar", "Surat", "Ambala",
    "Ahmedabad", "Hosur", "Jaipur", "Lucknow", "Kochi", "Mohali", "Chandigarh",
    "Prayagraj", "Kanpur", "Coimbatore", "Indore", "Nagpur", "Meerut", "Agra",
    "Rewari", "Ludhiana", "Nashik", "Vadodara", "Mysuru", "Vijayawada",
    "Varanasi", "Amritsar", "Gorakhpur", "Bareilly", "Warangal", "Madurai",
    "Kota", "Patiala", "Panipat", "Rajkot", "Udaipur", "Karimnagar", "Guntur",
    "Vellore", "Hubballi", "Palakkad", "Panchkula", "Puducherry", "Bhiwadi",
    "Chhatrapati Sambhaji Nagar", "Davangere", "Hisar", "Karnal", "Mehsana",
    "Sonipat", "Tumkuru",
]


def load_prev():
    """(count per city, grand total, store-id set per city) from the master."""
    p = sorted(BACKEND.glob("zepto_master_all_cities_*.xlsx"))
    if not p:
        return Counter(), 0, {}
    wb = openpyxl.load_workbook(p[-1], read_only=True, data_only=True)
    rows = list(wb["All Stores"].iter_rows(values_only=True))
    wb.close()
    hdr = [str(x).strip().lower() if x else "" for x in rows[2]]
    cnt, ids, total = Counter(), {}, 0
    for r in rows[3:]:
        if not r or not r[0]:
            continue
        d = dict(zip(hdr, r))
        cnt[d["city"]] += 1
        ids.setdefault(d["city"], set()).add(str(d["store_id"]).strip())
        total += 1
    return cnt, total, ids


def latest_files():
    """city slug -> newest file (Excel lock files ignored)."""
    best = {}
    for f in OUTDIR.glob("darkstores_*.xlsx"):
        if f.name.startswith("~$"):
            continue
        m = re.match(r"darkstores_(.+)_(\d{8}_\d{4})\.xlsx$", f.name)
        if not m:
            continue
        slug, ts = m.group(1), m.group(2)
        if slug not in best or ts > best[slug][0]:
            best[slug] = (ts, f)
    return {k: v[1] for k, v in best.items()}


def read(f):
    wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
    rr = list(wb["Stores"].iter_rows(values_only=True))
    wb.close()
    hdr = [str(x).strip().lower() if x else "" for x in rr[2]]
    return [dict(zip(hdr, r)) for r in rr[3:] if r and r[0]]


def main():
    prev, prev_total, prev_ids = load_prev()
    files = latest_files()
    slug_of = {c: re.sub(r"[^a-z0-9]+", "_", c.lower())[:40] for c in ALL_CITIES}

    done, pending = [], []
    all_new = set()
    fresh_sub = old_sub = 0

    print(f"{'#':<4}{'city':<28}{'OLD':>6}{'pin':>6}{'grid':>6}{'NEW':>5}{'FRESH':>8}{'miss':>6}")
    print("-" * 71)
    for i, c in enumerate(ALL_CITIES, 1):
        f = files.get(slug_of[c])
        if not f:
            pending.append(c)
            print(f"{i:<4}{c[:27]:<28}{prev.get(c,0):>6}{'-':>6}{'-':>6}{'-':>5}"
                  f"{'pending':>8}{'':>6}")
            continue
        rows = read(f)
        mine = [r for r in rows if r.get("city") == c]
        new = [r for r in mine if str(r.get("is_new")) == "True"]
        bp = sum(1 for r in rows if r.get("found_by") == "pincode")
        bg = sum(1 for r in rows if r.get("found_by") == "grid")
        hit = {str(r["store_id"]).strip() for r in rows}
        miss = len(prev_ids.get(c, set()) - hit)
        all_new.update(str(r["store_id"]) for r in new)
        fresh_sub += len(mine)
        old_sub += prev.get(c, 0)
        done.append(c)
        print(f"{i:<4}{c[:27]:<28}{prev.get(c,0):>6}{bp:>6}{bg:>6}{len(new):>5}"
              f"{len(mine):>8}{miss:>6}")
    print("-" * 71)

    print()
    print(f"SCANNED {len(done)}/58 cities")
    print(f"   their OLD subtotal    : {old_sub}")
    print(f"   their FRESH subtotal  : {fresh_sub}")
    print(f"   NEW stores confirmed  : {len(all_new)}")
    print()
    print(f"   {len(pending)} cities still carry OLD numbers, so there is no")
    print("   meaningful national total yet — merge the files once all 58 are done.")
    print()
    print("   FRESH = stores this scan saw and assigned to that city.")
    print("   miss  = previously-known stores the scan did NOT hit: closed, or")
    print("           just outside the box. Unverified either way.")

    # Only safe to skip a pending city when its box lies ~entirely inside an
    # already-scanned box — only then can it not possibly reveal a new store.
    skippable = []
    if pending and done:
        try:
            from scripts.zepto_discover_city_full import (load_master, build_context,
                                                          box_for, build_grid)
            _, recs = load_master()
            cs = build_context(recs)[0]
            dboxes = [box_for(c, cs) for c in done if c in cs]
            for c in pending:
                if c not in cs:
                    continue
                g = build_grid(box_for(c, cs))
                inside = sum(1 for la, ln in g
                             if any(b[0] <= la <= b[1] and b[2] <= ln <= b[3]
                                    for b in dboxes))
                if inside / len(g) >= 0.99:
                    skippable.append(c)
        except Exception:
            pass

    if skippable:
        print()
        print(f"Fully covered by a completed scan — safe to SKIP ({len(skippable)}):")
        print("   " + ", ".join(skippable))

    nxt = [c for c in pending if c not in skippable]
    if pending:
        print()
        print(f"Pending ({len(pending)}): next up = {nxt[0] if nxt else '-'}")
        print("   " + ", ".join(pending[:12]) + (" ..." if len(pending) > 12 else ""))
        if nxt:
            q = '"' + nxt[0] + '"' if " " in nxt[0] else nxt[0]
            print()
            print(f"   python -m scripts.zepto_discover_city_full --only {q}")


if __name__ == "__main__":
    main()
