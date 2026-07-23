"""
Scrape Blinkit public search for a keyword across all Pune PIN codes and export to Excel.

Usage (from backend/):
    python -m scripts.pune_keyword_scrape
    python -m scripts.pune_keyword_scrape --keyword "gluten free snack" --brand dobra
    python -m scripts.pune_keyword_scrape --keyword "tapioca chips" --brand dobra --out results.xlsx

Requires: openpyxl  (pip install openpyxl)
"""
import argparse
import asyncio
import sys
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

TENANT_ID = "a870fd8d-7373-47ec-ad69-5dd08ce35542"

# All Pune areas with PIN codes and approximate centroid lat/lon
PUNE_AREAS = [
    ("Pune GPO / Camp",     "411001", 18.5196, 73.8791),
    ("Camp",                "411001", 18.5140, 73.8783),
    ("Shivajinagar",        "411005", 18.5362, 73.8427),
    ("Deccan Gymkhana",     "411004", 18.5178, 73.8317),
    ("Swargate",            "411042", 18.5007, 73.8588),
    ("Kothrud",             "411038", 18.5074, 73.8077),
    ("Karve Nagar",         "411052", 18.4986, 73.8199),
    ("Erandwane",           "411004", 18.5214, 73.8351),
    ("Hadapsar",            "411028", 18.5018, 73.9260),
    ("Magarpatta",          "411028", 18.5168, 73.9302),
    ("Wanowrie",            "411040", 18.4868, 73.9002),
    ("Kondhwa",             "411048", 18.4681, 73.8936),
    ("NIBM",                "411048", 18.4699, 73.8968),
    ("Bibwewadi",           "411037", 18.4790, 73.8598),
    ("Dhankawadi",          "411043", 18.4595, 73.8612),
    ("Sahakar Nagar",       "411009", 18.4897, 73.8379),
    ("Parvati",             "411009", 18.4941, 73.8397),
    ("Sinhagad Road",       "411051", 18.4757, 73.8176),
    ("Vadgaon Budruk",      "411041", 18.4681, 73.8077),
    ("Warje",               "411058", 18.4965, 73.7964),
    ("Bavdhan",             "411021", 18.5224, 73.7866),
    ("Baner",               "411045", 18.5590, 73.7868),
    ("Balewadi",            "411045", 18.5780, 73.7769),
    ("Aundh",               "411007", 18.5588, 73.8080),
    ("Pashan",              "411021", 18.5300, 73.7900),
    ("Sus",                 "411021", 18.5380, 73.7565),
    ("Viman Nagar",         "411014", 18.5679, 73.9143),
    ("Kharadi",             "411014", 18.5594, 73.9435),
    ("Wagholi",             "412207", 18.5799, 73.9852),
    ("Lohegaon",            "411047", 18.5971, 73.9271),
    ("Dhanori",             "411015", 18.5970, 73.8996),
    ("Vishrantwadi",        "411015", 18.5797, 73.8870),
    ("Yerawada",            "411006", 18.5554, 73.9007),
    ("Koregaon Park",       "411001", 18.5363, 73.9014),
    ("Kalyani Nagar",       "411006", 18.5490, 73.9092),
    ("Bhosari",             "411026", 18.6345, 73.8529),
    ("Moshi",               "412105", 18.6658, 73.8523),
    ("Chinchwad",           "411033", 18.6279, 73.7969),
    ("Pimpri",              "411017", 18.6279, 73.7969),
    ("Akurdi",              "411035", 18.6475, 73.7646),
    ("Nigdi",               "411044", 18.6577, 73.7646),
    ("Hinjewadi",           "411057", 18.5912, 73.7387),
    ("Talegaon Dabhade",    "410507", 18.7276, 73.6744),
    ("Alandi",              "412105", 18.6814, 73.8897),
]


async def scrape_area(keyword: str, brand: str, area: str, pincode: str, lat: float, lon: float) -> dict:
    from scraper.platforms.blinkit.public_data import scraper as bl_scraper, parser as bl_parser
    try:
        raw = await bl_scraper.scrape(
            keyword=keyword,
            brand_slug=brand,
            city_slug="pune",
            zone=area,
            pincode=pincode,
            lat=lat,
            lon=lon,
            aliases=None,
        )
        result = bl_parser.parse(raw)
        return {"area": area, "pincode": pincode, "success": True, **result}
    except Exception as e:
        return {
            "area": area, "pincode": pincode, "success": False, "error": str(e),
            "total_results": 0, "brand_product_count": 0,
            "brand_rank": None, "brand_sov_pct": 0,
            "brand_products": [], "competitors": [],
        }


async def run(keyword: str, brand: str, out_path: Path, workers: int = 3) -> None:
    print(f"\nKeyword : {keyword!r}")
    print(f"Brand   : {brand!r}")
    print(f"Areas   : {len(PUNE_AREAS)}")
    print(f"Workers : {workers}\n")

    sem = asyncio.Semaphore(workers)

    async def bounded(area, pincode, lat, lon):
        async with sem:
            print(f"  → {area} ({pincode})", end="", flush=True)
            result = await scrape_area(keyword, brand, area, pincode, lat, lon)
            if result["success"]:
                rank = f"#{result['brand_rank']}" if result["brand_rank"] else "—"
                print(f"  total={result['total_results']}  brand={result['brand_product_count']}  rank={rank}  sov={result['brand_sov_pct']:.1f}%")
            else:
                print(f"  ERROR: {result['error'][:80]}")
            return result

    tasks = [bounded(a, p, lat, lon) for (a, p, lat, lon) in PUNE_AREAS]
    results = await asyncio.gather(*tasks)

    _write_excel(keyword, brand, results, out_path)
    print(f"\nSaved → {out_path.resolve()}")


def _write_excel(keyword: str, brand: str, results: list[dict], out_path: Path) -> None:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("\nERROR: openpyxl not installed. Run: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Summary"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    ok_fill = PatternFill("solid", fgColor="E2EFDA")
    err_fill = PatternFill("solid", fgColor="FCE4EC")

    headers_s = ["Area", "PIN Code", "Total Results", f"{brand.title()} Products",
                 f"{brand.title()} Rank", f"{brand.title()} SoV %", "Status"]
    ws1.append(headers_s)
    for i, h in enumerate(headers_s, 1):
        c = ws1.cell(1, i)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center")

    for r in results:
        row = [
            r["area"],
            r["pincode"],
            r.get("total_results", 0),
            r.get("brand_product_count", 0),
            f"#{r['brand_rank']}" if r.get("brand_rank") else "—",
            round(r.get("brand_sov_pct", 0), 2),
            "OK" if r["success"] else f"ERR: {r.get('error', '')[:60]}",
        ]
        ws1.append(row)
        fill = ok_fill if r["success"] else err_fill
        for col in range(1, len(row) + 1):
            ws1.cell(ws1.max_row, col).fill = fill

    # column widths
    for col, w in zip(range(1, 8), [22, 10, 14, 16, 14, 14, 40]):
        ws1.column_dimensions[get_column_letter(col)].width = w

    # ── Sheet 2: All Products ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("All Products")
    headers_p = ["Area", "PIN Code", "Position", "Product Name", "Brand",
                 "Price (₹)", "MRP (₹)", "Unit", "In Stock", "Rating",
                 "Category L1", "Category L2", "Is Own Brand"]
    ws2.append(headers_p)
    for i in range(1, len(headers_p) + 1):
        c = ws2.cell(1, i)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center")

    brand_fill = PatternFill("solid", fgColor="FFF9C4")
    for r in results:
        for p in r.get("listings", []):
            is_own = p.get("is_brand", False)
            cat = p.get("category") or {}
            row = [
                r["area"], r["pincode"],
                p.get("position"), p.get("name"), p.get("brand"),
                p.get("price"), p.get("mrp"), p.get("unit"),
                "Yes" if p.get("in_stock") else "No",
                p.get("rating"),
                cat.get("l1") if isinstance(cat, dict) else None,
                cat.get("l2") if isinstance(cat, dict) else None,
                "Yes" if is_own else "No",
            ]
            ws2.append(row)
            if is_own:
                for col in range(1, len(row) + 1):
                    ws2.cell(ws2.max_row, col).fill = brand_fill

    for col, w in zip(range(1, len(headers_p) + 1),
                      [22, 10, 10, 40, 20, 10, 10, 12, 10, 8, 20, 20, 14]):
        ws2.column_dimensions[get_column_letter(col)].width = w

    # ── Sheet 3: Competitors ──────────────────────────────────────────────────
    ws3 = wb.create_sheet("Competitors")
    headers_c = ["Area", "PIN Code", "Competitor Brand", "Products in Results"]
    ws3.append(headers_c)
    for i in range(1, len(headers_c) + 1):
        c = ws3.cell(1, i)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center")

    for r in results:
        for comp in r.get("competitors", []):
            ws3.append([r["area"], r["pincode"],
                        comp.get("name"), comp.get("count_in_results")])

    for col, w in zip(range(1, 5), [22, 10, 30, 18]):
        ws3.column_dimensions[get_column_letter(col)].width = w

    # metadata row at top of summary
    ws1.insert_rows(1)
    ws1.merge_cells("A1:G1")
    meta = ws1.cell(1, 1)
    meta.value = (f"Blinkit | keyword: '{keyword}' | brand: '{brand}' | "
                  f"city: Pune | scraped: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    meta.font = Font(bold=True, italic=True)
    meta.alignment = Alignment(horizontal="center")

    wb.save(out_path)




if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Scrape Blinkit keyword across Pune PINs → Excel")
    parser.add_argument("--keyword", "-k", default="gluten free snack", help="Search keyword")
    parser.add_argument("--brand", "-b", default="dobra", help="Brand slug")
    parser.add_argument("--out", "-o", default=None, help="Output Excel file path")
    parser.add_argument("--workers", "-w", type=int, default=3,
                        help="Parallel browser workers (default 3, lower = more polite)")
    args = parser.parse_args()

    keyword = args.keyword
    brand = args.brand
    out_path = Path(args.out) if args.out else Path(
        f"pune_{keyword.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    )

    if hasattr(asyncio, "WindowsProactorEventLoopPolicy"):
        asyncio.set_event_loop_policy(asyncio.WindowsProactorEventLoopPolicy())

    asyncio.run(run(keyword, brand, out_path, args.workers))
