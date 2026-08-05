"""
Fill the blank pincode / area cells in the FINAL master.

Grid-discovered stores have no pincode or area, because a grid probe asks
"which store serves this coordinate?" — no pincode is ever part of that
transaction. Phase-1 stores get both for free, since the pincode is the input.

This resolves those coordinates via OpenStreetMap Nominatim (usage policy:
identifying User-Agent, max 1 request/second — both honoured).

A `pincode_source` column records provenance:
    scraper          - the pincode that was searched to find the store
    reverse_geocode  - derived from the store's coordinate
Those are not the same thing: the first is the pincode Zepto matched, the second
is the pincode the coordinate physically sits in. Anything unresolved is written
as UNRESOLVED rather than guessed.

Run from backend/:
    python -m scripts.zepto_fill_missing_pincodes
Output: zepto_FINAL_master_filled_TIMESTAMP.xlsx
"""
import sys
import time
from pathlib import Path
from datetime import datetime
from collections import Counter

sys.path.insert(0, str(Path(__file__).parent.parent))

import httpx
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BACKEND = Path(__file__).parent.parent
NOMINATIM = "https://nominatim.openstreetmap.org/reverse"
UA = "zepto-darkstore-research/1.0 (contact: soumi97bag@gmail.com)"
RATE_LIMIT_S = 1.1
TRIES = 3

HDR_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
HDR_FILL = PatternFill("solid", fgColor="2E4057")
HDR_ALIGN = Alignment(horizontal="center", vertical="center")
ALT_FILL = PatternFill("solid", fgColor="F0F4F8")
NEW_FILL = PatternFill("solid", fgColor="D9F7E6")
GEO_FILL = PatternFill("solid", fgColor="FFF2CC")
BAD_FILL = PatternFill("solid", fgColor="FDE8E8")
TITLE_FILL = PatternFill("solid", fgColor="1A3C5E")


def blank(v):
    return not str(v or "").strip() or str(v).strip().lower() == "none"


def reverse_geocode(client, lat, lng):
    for a in range(1, TRIES + 1):
        try:
            r = client.get(NOMINATIM,
                           params={"lat": lat, "lon": lng, "format": "jsonv2",
                                   "addressdetails": 1, "zoom": 18},
                           headers={"User-Agent": UA, "Accept-Language": "en"},
                           timeout=20)
            if r.status_code == 200:
                ad = (r.json() or {}).get("address") or {}
                area = next((ad[k] for k in
                             ("neighbourhood", "suburb", "village", "town",
                              "city_district", "residential", "quarter", "hamlet")
                             if ad.get(k)), "")
                return {"pincode": (ad.get("postcode") or "").replace(" ", ""),
                        "area": area,
                        "state": ad.get("state") or ""}
            if r.status_code in (429, 503):
                time.sleep(2.0 * a)
                continue
        except Exception:
            pass
        time.sleep(1.0 * a)
    return None


def widths(ws, mx=46):
    for col in ws.columns:
        cl = get_column_letter(col[0].column)
        ml = max((len(str(c.value)) for c in col[:500] if c.value), default=0)
        ws.column_dimensions[cl].width = min(max(ml + 2, 9), mx)


def main():
    src = sorted(BACKEND.glob("zepto_FINAL_master_*.xlsx"))
    src = [p for p in src if "filled" not in p.name]
    if not src:
        raise SystemExit("No zepto_FINAL_master_*.xlsx found — run the merge first.")
    src = src[-1]

    wb = openpyxl.load_workbook(src, data_only=True)
    rr = list(wb["All Stores"].iter_rows(values_only=True))
    hdr = [str(x).strip().lower() if x else "" for x in rr[2]]
    rows = [dict(zip(hdr, r)) for r in rr[3:] if r and r[0]]
    other_sheets = {n: [list(x) for x in wb[n].iter_rows(values_only=True)]
                    for n in wb.sheetnames if n != "All Stores"}
    wb.close()

    todo = [r for r in rows if blank(r.get("source_pincode"))]
    print(f"Source : {src.name}")
    print(f"Rows   : {len(rows)}   blank pincode: {len(todo)}")
    print(f"Est    : ~{len(todo)*RATE_LIMIT_S/60:.1f} min\n")

    for r in rows:
        r["pincode_source"] = "" if blank(r.get("source_pincode")) else "scraper"

    ok = fail = 0
    with httpx.Client() as client:
        for i, r in enumerate(todo, 1):
            try:
                la, ln = float(r["lat"]), float(r["lng"])
            except (TypeError, ValueError):
                r["source_pincode"] = "UNRESOLVED"
                r["pincode_source"] = "failed"
                fail += 1
                continue
            g = reverse_geocode(client, la, ln)
            nm = str(r.get("store_name") or "")[:30]
            if g and g["pincode"]:
                r["source_pincode"] = g["pincode"]
                if blank(r.get("area_suggestion")):
                    r["area_suggestion"] = g["area"] or nm
                if blank(r.get("state")):
                    r["state"] = g["state"]
                r["pincode_source"] = "reverse_geocode"
                ok += 1
                print(f"  [{i:>3}/{len(todo)}] {nm:<31} -> {g['pincode']}  {g['area'][:24]}")
            else:
                r["source_pincode"] = "UNRESOLVED"
                r["pincode_source"] = "failed"
                fail += 1
                print(f"  [{i:>3}/{len(todo)}] {nm:<31} -> UNRESOLVED")
            time.sleep(RATE_LIMIT_S)

    print(f"\nGeocoded ok: {ok}   failed: {fail}")

    COLS = ["store_id", "store_name", "city", "tier", "source_pincode",
            "pincode_source", "area_suggestion", "state", "lat", "lng",
            "found_by", "is_new", "serviceable", "assigned_by", "secondary"]

    out_wb = openpyxl.Workbook()
    ws = out_wb.active
    ws.title = "All Stores"
    ws.append(COLS)
    for c in ws[1]:
        c.font, c.fill, c.alignment = HDR_FONT, HDR_FILL, HDR_ALIGN
    ws.freeze_panes = "A2"
    order = {"Tier 1": 0, "Tier 2": 1, "Tier 3": 2, "": 3}
    for i, r in enumerate(sorted(rows, key=lambda x: (order.get(x.get("tier"), 3),
                                                      str(x.get("city")),
                                                      str(x.get("source_pincode")),
                                                      str(x.get("store_name")))), start=2):
        ws.append([r.get(c, "") for c in COLS])
        if str(r.get("source_pincode")) == "UNRESOLVED":
            fill = BAD_FILL
        elif r.get("pincode_source") == "reverse_geocode":
            fill = GEO_FILL
        elif str(r.get("is_new")) == "True":
            fill = NEW_FILL
        else:
            fill = ALT_FILL if i % 2 == 0 else None
        if fill:
            for c in range(1, len(COLS) + 1):
                ws.cell(row=i, column=c).fill = fill
    ws.auto_filter.ref = ws.dimensions
    widths(ws)

    for name, data in other_sheets.items():
        w = out_wb.create_sheet(name)
        for row in data:
            w.append(row)
        widths(w)

    at = datetime.now().strftime("%d-%b-%Y %H:%M")
    T = f"Zepto Dark Stores — FINAL — {len(rows)} unique stores across 58 cities"
    S = (f"Generated: {at}  |  pincode: {sum(1 for r in rows if r['pincode_source']=='scraper')} "
         f"from scraper, {ok} reverse-geocoded, {fail} unresolved  |  "
         f"yellow = geocoded, green = new")
    ws.insert_rows(1, amount=2)
    mc = ws.max_column
    ws["A1"], ws["A2"] = T, S
    ws["A1"].font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="444444")
    ws["A2"].fill = ALT_FILL
    ws.row_dimensions[2].height = 15
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=mc)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=mc)

    ts = datetime.now().strftime("%Y%m%d_%H%M")
    out = BACKEND / f"zepto_FINAL_master_filled_{ts}.xlsx"
    out_wb.save(out)
    src_n = sum(1 for r in rows if r["pincode_source"] == "scraper")
    print(f"\npincode_source: scraper {src_n} | reverse_geocode {ok} | failed {fail}")
    print(f"unique pincodes: {len({str(r['source_pincode']) for r in rows if not blank(r.get('source_pincode')) and r['source_pincode']!='UNRESOLVED'})}")
    print(f"Saved -> {out}")


if __name__ == "__main__":
    main()
