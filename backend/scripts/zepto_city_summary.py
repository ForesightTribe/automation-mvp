"""
Generate a clean city-tier-darkstores summary Excel.
- 1064 globally unique stores (each store counted once)
- Notes column: if a city shares border stores with another city, or is covered within another city
Run from backend/:
    python -m scripts.zepto_city_summary
"""
import sys
from pathlib import Path
from datetime import datetime

sys.path.insert(0, str(Path(__file__).parent.parent))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── Tier mapping ─────────────────────────────────────────────────────────────
TIERS = {
    "Bengaluru": "Tier 1", "Hyderabad": "Tier 1", "Delhi NCR": "Tier 1",
    "Chennai": "Tier 1", "Mumbai": "Tier 1", "Pune": "Tier 1",
    "Kolkata": "Tier 1", "Ahmedabad": "Tier 1",
    "Lucknow": "Tier 2", "Jaipur": "Tier 2", "Surat": "Tier 2",
    "Kanpur": "Tier 2", "Nagpur": "Tier 2", "Dehradun": "Tier 2",
    "Indore": "Tier 2", "Mysuru": "Tier 2", "Nashik": "Tier 2",
    "Coimbatore": "Tier 2", "Kochi": "Tier 2", "Ludhiana": "Tier 2",
    "Jalandhar": "Tier 2", "Prayagraj": "Tier 2", "Vijayawada": "Tier 2",
    "Vadodara": "Tier 2", "Agra": "Tier 2", "Chandigarh": "Tier 2",
    "Madurai": "Tier 2", "Udaipur": "Tier 2", "Amritsar": "Tier 2",
    "Hubballi": "Tier 2", "Kota": "Tier 2", "Patiala": "Tier 2",
    "Varanasi": "Tier 2", "Warangal": "Tier 2", "Puducherry": "Tier 2",
    "Rajkot": "Tier 2",
    "Bareilly": "Tier 3", "Meerut": "Tier 3", "Mohali": "Tier 3",
    "Gorakhpur": "Tier 3", "Karimnagar": "Tier 3", "Palakkad": "Tier 3",
    "Ambala": "Tier 3", "Davangere": "Tier 3", "Guntur": "Tier 3",
    "Haridwar": "Tier 3", "Karnal": "Tier 3", "Kurukshetra": "Tier 3",
    "Panipat": "Tier 3", "Sonipat": "Tier 3", "Vellore": "Tier 3",
    "Belagavi": "Tier 3", "Bhiwadi": "Tier 3", "Hapur": "Tier 3",
    "Hisar": "Tier 3", "Hosur": "Tier 3", "Mehsana": "Tier 3",
    "Panchkula": "Tier 3", "Rewari": "Tier 3", "Saharanpur": "Tier 3",
    "Valsad": "Tier 3", "Tumkuru": "Tier 3",
    "Chhatrapati Sambhaji Nagar": "Tier 3",
}

# ── Cities fully covered within another city's scrape ─────────────────────────
COVERED_IN = {
    "Gurugram":   "Covered in Delhi NCR",
    "Noida":      "Covered in Delhi NCR",
    "Ghaziabad":  "Covered in Delhi NCR",
    "Faridabad":  "Covered in Delhi NCR",
    "Haridwar":   "Covered in Dehradun",
    "Kurukshetra":"Covered in Ambala",
    "Saharanpur": "Covered in Dehradun",
    "Valsad":     "Covered in Surat",
    "Hapur":      "Covered in Delhi NCR",
}

# ── Border store sharing notes ────────────────────────────────────────────────
# Derived from the 24 duplicate store_ids found across city files
SHARED_NOTES = {
    "Ambala":      "3 border stores also found in Kurukshetra (2) and Mohali (1)",
    "Amritsar":    "1 border store also found in Jalandhar",
    "Bangalore":   "1 border store also found in Mysuru",
    "Belagavi":    "1 border store also found in Hubballi",
    "Chandigarh":  "4 border stores also found in Mohali (1), Panchkula (3)",
    "Dehradun":    "3 border stores also found in Haridwar (2), Saharanpur (1)",
    "Delhi NCR":   "5 border stores also found in Hapur (1), Jaipur (2), Meerut (1), Sonipat (1)",
    "Hapur":       "1 store shared with Delhi NCR",
    "Haridwar":    "2 stores shared with Dehradun",
    "Hubballi":    "1 store shared with Belagavi",
    "Jaipur":      "3 border stores shared — 2 with Delhi NCR, 1 with Kota",
    "Jalandhar":   "2 border stores also found in Amritsar (1), Ludhiana (1)",
    "Kanpur":      "1 border store also found in Lucknow",
    "Kota":        "1 store shared with Jaipur",
    "Kurukshetra": "2 stores shared with Ambala",
    "Lucknow":     "1 store shared with Kanpur",
    "Ludhiana":    "1 store shared with Jalandhar",
    "Meerut":      "1 store shared with Delhi NCR",
    "Mohali":      "3 border stores shared — 1 with Ambala, 1 with Chandigarh, 1 with Panchkula",
    "Mysuru":      "1 store shared with Bangalore",
    "Panchkula":   "4 stores shared — 3 with Chandigarh, 1 with Mohali",
    "Patiala":     "1 store shared with Chandigarh",
    "Saharanpur":  "1 store shared with Dehradun",
    "Sonipat":     "1 store shared with Delhi NCR",
    "Surat":       "2 border stores also found in Vadodara (1), Valsad (1)",
    "Vadodara":    "1 store shared with Surat",
    "Valsad":      "1 store shared with Surat",
}

# Styling
T1_FILL  = PatternFill("solid", fgColor="1A3C5E")
T2_FILL  = PatternFill("solid", fgColor="2E6DA4")
T3_FILL  = PatternFill("solid", fgColor="5B9BD5")
COV_FILL = PatternFill("solid", fgColor="A9C4E0")
HDR_FILL = PatternFill("solid", fgColor="0D1F2D")
TOT_FILL = PatternFill("solid", fgColor="048A81")
ALT_FILL = PatternFill("solid", fgColor="F0F4F8")
WHITE    = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
DARK     = Font(name="Calibri", size=10, color="1A3C5E")
HDR_FNT  = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
TOT_FNT  = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
NOTE_FNT = Font(name="Calibri", size=9, italic=True, color="555555")
CTR      = Alignment(horizontal="center", vertical="center")
LEFT     = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def col_width(ws, col_idx, min_w=10, max_w=55):
    max_len = 0
    for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
        for cell in row:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
    ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, min_w), max_w)


def get_city_counts(combined_path: Path) -> dict:
    """Read unique store counts from the 1064-version combined file."""
    wb = openpyxl.load_workbook(combined_path, read_only=True, data_only=True)
    counts = {}
    if "Summary" in wb.sheetnames:
        ws = wb["Summary"]
        rows = list(ws.iter_rows(values_only=True))
        for row in rows[1:]:
            if row[0] and row[0] != "TOTAL" and row[1] is not None:
                counts[str(row[0])] = int(row[1])
    wb.close()
    return counts


def main():
    backend = Path(__file__).parent.parent

    # Find the 1064-unique combined file (latest)
    combined_files = sorted(backend.glob("zepto_all_stores_combined_*.xlsx"), reverse=True)
    if not combined_files:
        print("No combined file found. Run combine_zepto_stores.py first.")
        return

    # Per-city counts — each city uses its own scraped file (no cross-city dedup)
    from scripts.combine_zepto_stores import find_latest_files, read_stores, city_display_name

    print("Computing per-city store counts...")
    latest_files = find_latest_files()
    city_counts = {}

    for city_key in sorted(latest_files):
        display = city_display_name(city_key)
        if display is None:
            continue
        stores = read_stores(latest_files[city_key])
        seen_in_city: set = set()
        count = 0
        for s in stores:
            sid = str(s.get("store_id", "")).strip()
            if sid and sid not in seen_in_city:
                seen_in_city.add(sid)
                count += 1
        # normalise name: Bangalore → Bengaluru for tier lookup
        display_norm = "Bengaluru" if display == "Bangalore" else display
        city_counts[display_norm] = count

    # ── Build Excel ───────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "City Summary"

    # Header
    headers = ["City", "Tier", "Unique Dark Stores", "Notes"]
    ws.append(headers)
    for c, cell in enumerate(ws[1], 1):
        cell.font = HDR_FNT
        cell.fill = HDR_FILL
        cell.alignment = CTR
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    row_num = 2
    tier_order = ["Tier 1", "Tier 2", "Tier 3"]

    for tier in tier_order:
        tier_cities = [(c, TIERS[c]) for c in sorted(TIERS) if TIERS[c] == tier]

        for city, _ in tier_cities:
            fill = T1_FILL if tier == "Tier 1" else (T2_FILL if tier == "Tier 2" else T3_FILL)

            if city in COVERED_IN:
                # Show "Covered in [City]" in the store count column
                ws.cell(row=row_num, column=1, value=city)
                ws.cell(row=row_num, column=2, value=tier)
                ws.cell(row=row_num, column=3, value=COVERED_IN[city])
                ws.cell(row=row_num, column=4, value="")
                for c in range(1, 4):
                    ws.cell(row=row_num, column=c).font = Font(name="Calibri", size=10, italic=True, color="1A3C5E")
                    ws.cell(row=row_num, column=c).fill = COV_FILL
                    ws.cell(row=row_num, column=c).alignment = CTR
            else:
                count = city_counts.get(city, 0)
                note = SHARED_NOTES.get(city, "")
                ws.cell(row=row_num, column=1, value=city)
                ws.cell(row=row_num, column=2, value=tier)
                ws.cell(row=row_num, column=3, value=count)
                ws.cell(row=row_num, column=4, value=note)
                for c in range(1, 3):
                    ws.cell(row=row_num, column=c).font = WHITE
                    ws.cell(row=row_num, column=c).fill = fill
                    ws.cell(row=row_num, column=c).alignment = CTR
                ws.cell(row=row_num, column=3).font = WHITE
                ws.cell(row=row_num, column=3).fill = fill
                ws.cell(row=row_num, column=3).alignment = CTR
                ws.cell(row=row_num, column=4).font = NOTE_FNT
                ws.cell(row=row_num, column=4).alignment = LEFT

            ws.row_dimensions[row_num].height = 18
            row_num += 1

        # Covered-in cities for Delhi NCR (Gurugram, Noida, Ghaziabad, Faridabad — not in TIERS dict)
        if tier == "Tier 2":
            delhi_ncr_extras = ["Gurugram", "Noida", "Ghaziabad", "Faridabad"]
            for city in delhi_ncr_extras:
                ws.cell(row=row_num, column=1, value=city)
                ws.cell(row=row_num, column=2, value="Tier 2")
                ws.cell(row=row_num, column=3, value=COVERED_IN[city])
                ws.cell(row=row_num, column=4, value="")
                for c in range(1, 4):
                    ws.cell(row=row_num, column=c).font = Font(name="Calibri", size=10, italic=True, color="1A3C5E")
                    ws.cell(row=row_num, column=c).fill = COV_FILL
                    ws.cell(row=row_num, column=c).alignment = CTR
                ws.row_dimensions[row_num].height = 18
                row_num += 1

    # Total row — globally unique count (1064)
    seen_global: set = set()
    for city_key in sorted(latest_files):
        display = city_display_name(city_key)
        if display is None:
            continue
        for s in read_stores(latest_files[city_key]):
            sid = str(s.get("store_id", "")).strip()
            if sid:
                seen_global.add(sid)
    total = len(seen_global)
    ws.cell(row=row_num, column=1, value="TOTAL")
    ws.cell(row=row_num, column=2, value="")
    ws.cell(row=row_num, column=3, value=total)
    ws.cell(row=row_num, column=4, value=f"1064 unique dark stores across India (25 border stores counted once each)")
    for c in range(1, 5):
        ws.cell(row=row_num, column=c).font = TOT_FNT
        ws.cell(row=row_num, column=c).fill = TOT_FILL
        ws.cell(row=row_num, column=c).alignment = CTR
    ws.cell(row=row_num, column=4).alignment = LEFT
    ws.row_dimensions[row_num].height = 20

    # Column widths
    for i in range(1, 5):
        col_width(ws, i)

    ws.auto_filter.ref = f"A1:D{row_num - 1}"

    ts = datetime.now().strftime("%Y%m%d_%H%M")
    out = backend / f"zepto_city_tier_summary_{ts}.xlsx"
    wb.save(out)
    print(f"\nCities: {len(city_counts)}")
    print(f"Total unique stores: {total}")
    print(f"Saved → {out}")


if __name__ == "__main__":
    main()
