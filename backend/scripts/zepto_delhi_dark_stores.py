"""
Scrape Zepto dark stores in Delhi NCR — delivery areas + pincodes + grid scan.

How it works:
  1. Open zepto.com in Playwright once (WAF token obtained automatically by real browser)
  2. Register init_script: overrides window.fetch + XMLHttpRequest to inject lat/lng
     from localStorage before every get_page call — survives page reloads
  3. For each coordinate (delivery areas + pincodes + 1km grid):
     - Write target lat/lng to localStorage
     - Reload → zepto.com fires get_page with our injected coordinates
     - Capture store_id + store_name via page.on("response")
  4. After all probes: compute centroid of all hits per store → estimated_lat/lng
  5. Save deduplicated results to Excel

No cookies to paste, no tokens to copy.
Run from backend/:
    python -m scripts.zepto_delhi_dark_stores
"""
import asyncio
import json
import re
import sys
import itertools
from pathlib import Path
from collections import Counter, defaultdict

sys.path.insert(0, str(Path(__file__).parent.parent))

import httpx
import openpyxl
from playwright.async_api import async_playwright

from scraper.platforms.zepto.dark_store.endpoints import (
    NCR_BBOX, DELHI_NCR_PINCODES,
)

ZEPTO_DELIVERY_AREAS = [
    # Delhi
    "Karol Bagh Delhi", "Dilshad Garden Delhi", "Mayur Vihar Delhi",
    "Model Town Delhi", "Pitampura Delhi", "South Extension Delhi",
    "Rohini Sector 3 Delhi", "Rohini Sector 15 Delhi", "Sector 10 Dwarka Delhi",
    "Paschim Vihar Delhi", "Uttam Nagar Delhi", "Shakti Nagar Delhi",
    "Kalkaji Delhi", "Kirti Nagar Delhi", "RK Puram Delhi",
    "Sector 12 Dwarka Delhi", "New Friends Colony Delhi", "Vikaspuri Delhi",
    "Malviya Nagar Delhi", "Vasant Kunj Delhi", "Mahavir Enclave Delhi",
    "Janakpuri Delhi", "Anand Vihar Delhi", "IP Extension Delhi",
    "East of Kailash Delhi", "Rajouri Garden Delhi", "Wazirpur Delhi",
    # Ghaziabad
    "Indirapuram Ghaziabad", "Vaishali Ghaziabad", "Gaur City Ghaziabad",
    "Rajnagar Extension Ghaziabad", "Raj Nagar Ghaziabad",
    # Gurgaon
    "Sector 12 Gurugram", "Sector 66 Gurugram", "Udyog Vihar Gurugram",
    "Sector 71 Gurugram", "Sector 53 Gurugram", "Sector 39 Gurugram",
    "Sun City Gurugram", "Sector 63 Gurugram", "Sector 27 Gurugram",
    "Ardee City Gurugram", "Sector 69 Gurugram",
    # Noida
    "Sector 63 Noida", "Sector 142 Noida", "Sector 117 Noida",
    "Sector 10 Noida", "Sector 46 Noida", "Sector 73 Noida",
]

GRID_WORKERS = 8  # parallel browser pages for grid scan

_BFF = "https://bff-gateway.zepto.com"

# JS injected before any page script — overrides fetch + XHR to inject lat/lng
_INIT_SCRIPT = """
(() => {
    const injectCoords = (url) => {
        if (typeof url !== 'string' || !url.includes('/lms/api/v2/get_page')) return url;
        const lat = localStorage.getItem('__probe_lat');
        const lng = localStorage.getItem('__probe_lng');
        if (!lat || !lng) return url;
        try {
            const u = new URL(url, location.href);
            u.searchParams.set('latitude', lat);
            u.searchParams.set('longitude', lng);
            return u.toString();
        } catch(e) { return url; }
    };

    const origFetch = window.fetch;
    window.fetch = function(url, opts) {
        return origFetch.apply(this, [injectCoords(url), opts]);
    };

    const origOpen = XMLHttpRequest.prototype.open;
    XMLHttpRequest.prototype.open = function(method, url, ...rest) {
        return origOpen.call(this, method, injectCoords(url), ...rest);
    };
})();
"""


def make_grid(lat_min, lat_max, lng_min, lng_max, step_km=1.0):
    step = step_km / 111.0
    lats, lngs = [], []
    lat = lat_min
    while lat <= lat_max:
        lats.append(round(lat, 5))
        lat += step
    lng = lng_min
    while lng <= lng_max:
        lngs.append(round(lng, 5))
        lng += step
    return list(itertools.product(lats, lngs))


async def get_coords(query: str, ctx, postal_code_only: bool = False) -> dict | None:
    """Resolve query to lat/lng via Zepto autocomplete (no WAF on these endpoints).
    postal_code_only=True: only use predictions with types=['postal_code'] (for pincode searches).
    postal_code_only=False: use first prediction (for named area searches like 'Karol Bagh Delhi').
    """
    headers = {
        "Accept": "application/json",
        "Origin": "https://www.zepto.com",
        "Referer": "https://www.zepto.com/",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "app_sub_platform": "WEB",
        "app_version": "16.20.0",
        "platform": "WEB",
        "tenant": "ZEPTO",
    }
    for q in [query, query.split(",")[0].strip(), f"{query} India"]:
        try:
            r = await ctx.request.get(
                f"{_BFF}/api/v1/maps/place/autocomplete/",
                params={"place_name": q},
                headers=headers,
                timeout=10000,
            )
            if r.status != 200:
                continue
            data = await r.json()
            preds = data.get("data", {}).get("predictions") or data.get("predictions") or []
            if not preds:
                continue
            if postal_code_only:
                pred = next((p for p in preds if "postal_code" in p.get("types", [])), None)
                if not pred:
                    continue
            else:
                pred = preds[0]
            r2 = await ctx.request.get(
                f"{_BFF}/api/v1/maps/place/details/",
                params={"place_id": pred["place_id"]},
                headers=headers,
                timeout=10000,
            )
            if r2.status != 200:
                continue
            d2 = await r2.json()
            result = d2.get("data", {}).get("result") or d2.get("result") or {}
            geo = result.get("geometry", {}).get("location", {})
            lat, lng = geo.get("lat"), geo.get("lng")
            if not lat or not lng:
                continue

            pincode = state = city = None
            for comp in result.get("address_components", []):
                types = comp.get("types", [])
                if "postal_code" in types:
                    pincode = comp.get("long_name")
                if "locality" in types or "administrative_area_level_2" in types:
                    city = city or comp.get("long_name")
                if "administrative_area_level_1" in types:
                    state = state or comp.get("long_name")

            return {
                "lat": lat, "lng": lng,
                "pincode": pincode or "", "city": city or "Delhi", "state": state or "Delhi",
            }
        except Exception as e:
            print(f"    [coord error] {q}: {e}")
            continue
    return None


async def reverse_geocode(lat: float, lng: float, client: httpx.AsyncClient) -> dict:
    """Reverse geocode lat/lng → area + pincode via OSM Nominatim."""
    try:
        r = await client.get(
            "https://nominatim.openstreetmap.org/reverse",
            params={"lat": lat, "lon": lng, "format": "json", "addressdetails": 1},
            headers={"User-Agent": "zepto-store-scraper/1.0"},
            timeout=10,
        )
        if r.status_code != 200:
            return {}
        data = r.json()
        addr = data.get("address", {})
        pincode = addr.get("postcode", "")
        area = (
            addr.get("suburb")
            or addr.get("neighbourhood")
            or addr.get("quarter")
            or addr.get("village")
            or addr.get("town")
            or ""
        )
        state = addr.get("state", "")
        city = addr.get("city") or addr.get("town") or addr.get("county") or ""
        return {"pincode": pincode, "area": area, "city": city, "state": state}
    except Exception:
        return {}


async def geocode_store_name(store_name: str, city: str, client: httpx.AsyncClient) -> dict:
    """
    Geocode store name using Zepto's own autocomplete API — gives the same
    coordinates the TL gets when manually searching on zepto.com.
    "DEL-Paharganj" → query "Paharganj Delhi" → Zepto autocomplete → lat/lng
    """
    if not store_name:
        return {}
    area = re.sub(r'^[A-Z]{2,5}-', '', store_name)
    area = re.sub(r'\s+(New|Old|\d+[A-Za-z]?|Phase\s*\d+|Block\s*\d+)$', '', area, flags=re.IGNORECASE).strip()
    if not area:
        return {}
    headers = {
        "Accept": "application/json",
        "Origin": "https://www.zepto.com",
        "Referer": "https://www.zepto.com/",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "app_sub_platform": "WEB", "app_version": "16.20.0",
        "platform": "WEB", "tenant": "ZEPTO",
    }
    for query in [f"{area} {city}", area]:
        try:
            r = await client.get(
                f"{_BFF}/api/v1/maps/place/autocomplete/",
                params={"place_name": query}, headers=headers, timeout=10,
            )
            if r.status_code != 200:
                continue
            preds = r.json().get("data", {}).get("predictions") or r.json().get("predictions") or []
            if not preds:
                continue
            r2 = await client.get(
                f"{_BFF}/api/v1/maps/place/details/",
                params={"place_id": preds[0]["place_id"]}, headers=headers, timeout=10,
            )
            if r2.status_code != 200:
                continue
            result = r2.json().get("data", {}).get("result") or r2.json().get("result") or {}
            geo = result.get("geometry", {}).get("location", {})
            lat, lng = geo.get("lat"), geo.get("lng")
            if lat and lng:
                return {"estimated_lat": lat, "estimated_lng": lng}
        except Exception:
            continue
    return {}


async def probe_store_at(lat: float, lng: float, page) -> dict:
    """
    Set localStorage coords, reload page.
    init_script injects lat/lng into every get_page call → captures natural response.
    """
    result = {}
    done = asyncio.Event()

    async def on_response(response):
        if "/lms/api/v2/get_page" in response.url and not done.is_set():
            try:
                data = await response.json()
                svc = data.get("storeServiceableResponse", {})
                sid = svc.get("storeId")
                if sid:
                    result["store_id"] = str(sid)
                    result["secondary_ids"] = [str(s) for s in svc.get("secondaryStoreIds", [])]
                    result["store_name"] = data.get("storeDetailsResponse", {}).get("name", "")
                done.set()
            except Exception:
                pass

    try:
        await page.evaluate(f"""() => {{
            localStorage.setItem('__probe_lat', '{lat}');
            localStorage.setItem('__probe_lng', '{lng}');
        }}""")
    except Exception:
        pass

    page.on("response", on_response)
    try:
        await page.reload(timeout=20000, wait_until="domcontentloaded")
        try:
            await asyncio.wait_for(done.wait(), timeout=8)
        except asyncio.TimeoutError:
            pass
    except Exception:
        pass
    finally:
        page.remove_listener("response", on_response)

    return result


async def main():
    print("Zepto Delhi NCR Dark Store Scraper")
    print("=" * 55)

    from datetime import datetime
    ts = datetime.now().strftime("%Y%m%d_%H%M")

    seen_primary: set[str] = set()
    primary_rows: list[dict] = []
    secondary_rows: list[dict] = []
    probe_hits: dict[str, list] = defaultdict(list)
    lock = asyncio.Lock()

    # ── Resume from checkpoint if --resume <file> is given ──────────────────────
    resume_file = None
    args = sys.argv[1:]
    for i, arg in enumerate(args):
        if arg == "--resume" and i + 1 < len(args):
            resume_file = args[i + 1]
            break
        if arg.startswith("--resume="):
            resume_file = arg.split("=", 1)[1]
            break

    if not resume_file:
        async with async_playwright() as p:
            browser = await p.chromium.launch(headless=True)
            ctx = await browser.new_context(
                user_agent=(
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36"
                ),
            )

            warmup_page = await ctx.new_page()
            await warmup_page.add_init_script(_INIT_SCRIPT)
            print("Warming up browser (WAF challenge)...")
            try:
                await warmup_page.goto("https://www.zepto.com", timeout=30000, wait_until="domcontentloaded")
                await warmup_page.wait_for_timeout(3000)
                print("Browser ready.\n")
            except Exception as e:
                print(f"Warmup: {e}\n")

            probes: list[dict] = []

            # ── Step 1: Delivery area names ──────────────────────────────────────
            print(f"Step 1: Resolving {len(ZEPTO_DELIVERY_AREAS)} delivery area coordinates...")
            for area in ZEPTO_DELIVERY_AREAS:
                coords = await get_coords(area, ctx)
                if coords:
                    probes.append({
                        "area": area, "lat": coords["lat"], "lng": coords["lng"],
                        "pincode": coords["pincode"], "city": coords["city"],
                        "state": coords["state"], "found_by": "autosuggestion",
                    })
                    print(f"  ✓ {area} → ({coords['lat']:.4f}, {coords['lng']:.4f})")
                else:
                    print(f"  ✗ {area} → not resolved")
                await asyncio.sleep(0.4)

            # ── Step 1b: Pincode sweep ───────────────────────────────────────────
            print(f"\nStep 1b: Pincode sweep ({len(DELHI_NCR_PINCODES)} pincodes)...")
            for pincode, loc_name in DELHI_NCR_PINCODES.items():
                coords = await get_coords(pincode, ctx, postal_code_only=True)
                if coords:
                    probes.append({
                        "area": loc_name, "lat": coords["lat"], "lng": coords["lng"],
                        "pincode": pincode, "city": coords["city"],
                        "state": coords["state"], "found_by": "pincode",
                    })
                await asyncio.sleep(0.3)

            named_probes = probes[:]
            print(f"\nStep 2: Probing {len(named_probes)} named coordinates...")
            for i, item in enumerate(named_probes, 1):
                lat, lng = item["lat"], item["lng"]
                r = await probe_store_at(lat, lng, warmup_page)
                sid = r.get("store_id")
                label = item["area"]
                if sid:
                    probe_hits[sid].append((lat, lng))
                    if sid not in seen_primary:
                        seen_primary.add(sid)
                        secondary = ",".join(r.get("secondary_ids", []))
                        primary_rows.append({
                            "store_id": sid,
                            "store_name": r.get("store_name", ""),
                            "city": item.get("city", "Delhi"),
                            "state": item.get("state", "Delhi"),
                            "pincode": item.get("pincode", ""),
                            "area": item.get("area", ""),
                            "lat": lat, "lng": lng,
                            "hit_count": 0,
                            "secondary_store_id": secondary,
                            "found_by": item["found_by"],
                        })
                        for ssid in r.get("secondary_ids", []):
                            secondary_rows.append({
                                "primary_store_id": sid, "secondary_store_id": ssid,
                                "probe_lat": lat, "probe_lng": lng,
                            })
                        print(f"  [{i}/{len(named_probes)}] NEW {sid} ← {label}")
                else:
                    print(f"  [{i}/{len(named_probes)}] no store ← {label}")

            # ── Step 3: Grid scan ────────────────────────────────────────────────
            grid_coords = make_grid(**NCR_BBOX, step_km=1.0)
            print(f"\nStep 3: Grid scan — {len(grid_coords)} points @ 1.0 km step, {GRID_WORKERS} parallel workers")

            worker_pages = [warmup_page]
            for _ in range(GRID_WORKERS - 1):
                wp = await ctx.new_page()
                await wp.add_init_script(_INIT_SCRIPT)
                await wp.goto("https://www.zepto.com", timeout=30000, wait_until="domcontentloaded")
                await wp.wait_for_timeout(1000)
                worker_pages.append(wp)

            queue: asyncio.Queue = asyncio.Queue()
            for idx, (lat, lng) in enumerate(grid_coords):
                await queue.put((idx + 1, lat, lng))
            total_grid = len(grid_coords)

            async def grid_worker(page, worker_id):
                while True:
                    try:
                        idx, lat, lng = queue.get_nowait()
                    except asyncio.QueueEmpty:
                        break
                    r = await probe_store_at(lat, lng, page)
                    sid = r.get("store_id")
                    if sid:
                        async with lock:
                            probe_hits[sid].append((lat, lng))
                            if sid not in seen_primary:
                                seen_primary.add(sid)
                                secondary = ",".join(r.get("secondary_ids", []))
                                primary_rows.append({
                                    "store_id": sid,
                                    "store_name": r.get("store_name", ""),
                                    "city": "Delhi", "state": "Delhi",
                                    "pincode": "", "area": "",
                                    "lat": "", "lng": "",
                                    "_fallback_lat": lat, "_fallback_lng": lng,
                                    "hit_count": 0,
                                    "secondary_store_id": secondary,
                                    "found_by": "grid",
                                })
                                for ssid in r.get("secondary_ids", []):
                                    secondary_rows.append({
                                        "primary_store_id": sid, "secondary_store_id": ssid,
                                        "probe_lat": lat, "probe_lng": lng,
                                    })
                                print(f"  [grid {idx}/{total_grid}] NEW {sid} ← ({lat:.4f},{lng:.4f}) w{worker_id}")
                    queue.task_done()

            await asyncio.gather(*[grid_worker(wp, i) for i, wp in enumerate(worker_pages, 1)])

            for wp in worker_pages:
                try:
                    await wp.close()
                except Exception:
                    pass
            try:
                await browser.close()
            except Exception:
                pass

        # Save checkpoint so geocoding can be retried without re-running grid
        ckpt_path = Path(__file__).parent.parent / f"zepto_delhi_checkpoint_{ts}.json"
        with open(ckpt_path, "w", encoding="utf-8") as f:
            json.dump({
                "primary_rows": primary_rows,
                "secondary_rows": secondary_rows,
                "probe_hits": {sid: list(hits) for sid, hits in probe_hits.items()},
            }, f)
        print(f"Checkpoint saved → {ckpt_path}")

    else:
        # Load from checkpoint
        print(f"Resuming from checkpoint: {resume_file}")
        with open(resume_file, "r", encoding="utf-8") as f:
            ckpt = json.load(f)
        primary_rows = ckpt["primary_rows"]
        secondary_rows = ckpt.get("secondary_rows", [])
        for sid, hits in ckpt["probe_hits"].items():
            probe_hits[sid] = [tuple(h) for h in hits]
        seen_primary = {r["store_id"] for r in primary_rows}
        print(f"Loaded {len(primary_rows)} stores, {sum(len(v) for v in probe_hits.values())} probe hits")
        print("Skipping Steps 1-3 — going straight to geocoding.\n")

    # ── Summary ──────────────────────────────────────────────────────────────────
    print(f"\n{'='*55}")
    print(f"Total unique Delhi NCR dark stores: {len(primary_rows)}")

    if not primary_rows:
        print("No stores found. Try running again — WAF may have blocked warm-up.")
        return

    # ── Centroid + geocoding pass ────────────────────────────────────────────────
    print(f"\nGeocoding {len(primary_rows)} stores (centroid → estimated lat/lng + pincode/area)...")
    async with httpx.AsyncClient(follow_redirects=True) as client:
        for i, row in enumerate(primary_rows, 1):
            sid = row["store_id"]
            hits = probe_hits[sid]
            row["hit_count"] = len(hits)

            # Grid-only stores: get exact lat/lng from Zepto place/details via store name
            if not row.get("lat") and row.get("store_name"):
                geo = await geocode_store_name(row["store_name"], "Delhi", client)
                row["lat"] = geo.get("estimated_lat", "")
                row["lng"] = geo.get("estimated_lng", "")
                await asyncio.sleep(1.1)

            # Final fallback: centroid of grid hits
            if not row.get("lat") and hits:
                row["lat"] = round(sum(h[0] for h in hits) / len(hits), 6)
                row["lng"] = round(sum(h[1] for h in hits) / len(hits), 6)

            # Clean up internal fallback fields
            row.pop("_fallback_lat", None)
            row.pop("_fallback_lng", None)

            lat_for_geo = row.get("lat") or ""
            lng_for_geo = row.get("lng") or ""
            if not row["pincode"] or not row["area"]:
                if lat_for_geo and lng_for_geo:
                    geo = await reverse_geocode(lat_for_geo, lng_for_geo, client)
                    if not row["pincode"] and geo.get("pincode"):
                        row["pincode"] = geo["pincode"]
                    if not row["area"] and geo.get("area"):
                        row["area"] = geo["area"]
                    if geo.get("state"):
                        row["state"] = geo["state"]
                    if geo.get("city"):
                        row["city"] = geo["city"]
                    await asyncio.sleep(1.1)

            print(f"  [{i}/{len(primary_rows)}] {row.get('store_name') or sid[:8]} "
                  f"→ lat={row.get('lat','')}, lng={row.get('lng','')}, pin={row.get('pincode','')}")
    print("Geocoding done.\n")

    out = Path(__file__).parent.parent / f"zepto_delhi_dark_stores_{ts}.xlsx"
    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "primary_stores"
    hdr1 = ["store_id", "store_name", "city", "state", "pincode", "area",
            "lat", "lng", "hit_count", "secondary_store_id", "found_by"]
    ws1.append(hdr1)
    for row in sorted(primary_rows, key=lambda x: x["pincode"] or "999999"):
        ws1.append([row[h] for h in hdr1])

    if secondary_rows:
        ws2 = wb.create_sheet("secondary_stores")
        ws2.append(["primary_store_id", "secondary_store_id", "probe_lat", "probe_lng"])
        for row in secondary_rows:
            ws2.append([row["primary_store_id"], row["secondary_store_id"],
                        row["probe_lat"], row["probe_lng"]])

    ws3 = wb.create_sheet("summary")
    ws3.append(["found_by", "count"])
    for method, count in sorted(Counter(r["found_by"] for r in primary_rows).items()):
        ws3.append([method, count])
        print(f"  {method}: {count} stores")

    wb.save(out)
    print(f"\nSaved → {out}")


if __name__ == "__main__":
    asyncio.run(main())
