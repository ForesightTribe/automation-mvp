"""
Bengaluru dark store GRID SCAN — 1 km lattice, exhaustive store discovery.

Why a grid: the pincode method is only as good as Zepto's autocomplete. Where
suggestions are sparse, stores go undiscovered. A grid asks get_page "which store
serves this exact point?" at every node, so nothing can hide between samples.

Speed: probing via page.reload() costs ~3s/point (2+ hours at 1 km). Instead we
capture the session headers from the page's OWN get_page request once, then
replay them through ctx.request — ~0.33s/point, and correct (verified against
known store coordinates). Headers are re-captured automatically if the session
goes stale (429/403).

Run from backend/:
    python -m scripts.zepto_bengaluru_grid_scan
Output: zepto_bengaluru_grid_scan_TIMESTAMP.xlsx
"""
import asyncio
import sys
from pathlib import Path
from datetime import datetime

sys.path.insert(0, str(Path(__file__).parent.parent))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from playwright.async_api import async_playwright

# ── Grid definition ───────────────────────────────────────────────────────────
# Whole Bengaluru Urban + Rural. The previous 12.72-13.25 box was measurably too
# small on two edges: the northernmost store sat 1.9 km from the north edge and
# the southernmost sat exactly ON the south edge — both cut off. East and west
# were already saturated (nearest store 6 km and 14.6 km from those edges, with
# zero stores in the edge bands), so they are widened only slightly.
#   north -> Doddaballapur / Vijayapura / Chikkaballapur corridor
#   south -> Anekal / Attibele / Chandapura
LAT_MIN, LAT_MAX = 12.58, 13.45
LNG_MIN, LNG_MAX = 77.25, 77.95

# ~670 m. Finer than 1 km because a 1 km lattice provably still hides stores in
# dense clusters: BLR-Bellandur 6 is a known store that was never the primary
# server at any 1 km node. Tighter sampling should surface that class of store.
STEP = 0.006

CONCURRENCY = 8         # raised from 5 — the 1 km run had 0 probe failures
PROBE_TRIES = 3

COMBINED   = Path(__file__).parent.parent / "zepto_all_stores_combined_20260731_1147.xlsx"
CITY_SHEET = "Bangalore"

_BASE = "https://www.zepto.com"
_BFF  = "https://bff-gateway.zepto.com"
_GP   = "/lms/api/v2/get_page"

HDR_FONT  = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
HDR_FILL  = PatternFill("solid", fgColor="2E4057")
HDR_ALIGN = Alignment(horizontal="center", vertical="center")
ALT_FILL  = PatternFill("solid", fgColor="F0F4F8")
NEW_FILL  = PatternFill("solid", fgColor="D9F7E6")
SEC_FILL  = PatternFill("solid", fgColor="FFF4CE")
TITLE_FILL = PatternFill("solid", fgColor="1A3C5E")

_DROP_HEADERS = {"host", "content-length", "connection", "accept-encoding",
                 ":method", ":path", ":authority", ":scheme"}


# ── Input ─────────────────────────────────────────────────────────────────────

def load_known() -> tuple[set, set, dict, dict]:
    """(bengaluru_primary, bengaluru_secondary, {id: name}, {id: owning_city}).

    The owning-city map spans EVERY city sheet, not just Bangalore. The 1 km run
    reported 36 new stores, but two of them (HOS-Hosur, BLR-Mico layout network-2)
    were already recorded under Hosur — the grid box overlaps that city. Checking
    only the Bangalore sheet cannot catch that; this map can.
    """
    wb = openpyxl.load_workbook(COMBINED, read_only=True, data_only=True)
    primary, secondary, names, owner = set(), set(), {}, {}

    for sheet in wb.sheetnames:
        if sheet in ("All Stores", "Summary"):
            continue
        rows = list(wb[sheet].iter_rows(values_only=True))
        if not rows:
            continue
        hdr = [str(h).strip().lower() if h else "" for h in rows[0]]
        if "store_id" not in hdr:
            continue
        for r in rows[1:]:
            if not r or not r[0]:
                continue
            d = dict(zip(hdr, r))
            sid = str(d.get("store_id") or "").strip()
            if not sid:
                continue
            owner.setdefault(sid, sheet)
            nm = str(d.get("store_name") or "").strip()
            if nm and sid not in names:
                names[sid] = nm
            if sheet == CITY_SHEET:
                primary.add(sid)
                for x in str(d.get("secondary_store_id") or "").split(","):
                    x = x.strip()
                    if x:
                        secondary.add(x)
    wb.close()
    return primary, secondary, names, owner


def build_grid() -> list[tuple[float, float]]:
    pts = []
    lat = LAT_MIN
    while lat <= LAT_MAX + 1e-9:
        lng = LNG_MIN
        while lng <= LNG_MAX + 1e-9:
            pts.append((round(lat, 5), round(lng, 5)))
            lng += STEP
        lat += STEP
    return pts


# ── Session ───────────────────────────────────────────────────────────────────

async def capture_headers(page) -> dict:
    """Load the homepage and lift the session headers off its own get_page call."""
    captured: dict[str, str] = {}

    async def on_req(req):
        if _GP in req.url and not captured:
            captured.update(req.headers)

    page.on("request", on_req)
    try:
        await page.goto(_BASE, timeout=30000, wait_until="domcontentloaded")
        await page.wait_for_timeout(4000)
    except Exception:
        pass
    finally:
        page.remove_listener("request", on_req)

    return {k: v for k, v in captured.items() if k.lower() not in _DROP_HEADERS}


def gp_url(lat: float, lng: float) -> str:
    return (f"{_BFF}{_GP}?latitude={lat}&longitude={lng}"
            f"&page_type=HOME&version=v2&show_new_eta_banner=true"
            f"&page_size=3&enforce_platform_type=WEB")


# ── Probe ─────────────────────────────────────────────────────────────────────

async def probe(ctx, headers: dict, lat: float, lng: float, state: dict) -> dict | None:
    """Resolve the store serving (lat, lng). Returns None where Zepto doesn't serve."""
    for attempt in range(1, PROBE_TRIES + 1):
        try:
            r = await ctx.request.get(gp_url(lat, lng),
                                      headers=state["headers"], timeout=15000)
            if r.status == 200:
                d = await r.json()
                svc = d.get("storeServiceableResponse") or {}
                sid = svc.get("storeId")
                if not sid:
                    return None            # genuinely unserviceable point
                return {
                    "store_id":   sid,
                    "secondary":  [str(x) for x in (svc.get("secondaryStoreIds") or [])],
                    "store_name": (d.get("storeDetailsResponse") or {}).get("name", ""),
                }
            if r.status in (401, 403, 429):
                state["stale"] = True       # ask the driver to refresh the session
                await asyncio.sleep(1.5 * attempt)
                continue
        except Exception:
            pass
        await asyncio.sleep(0.6 * attempt)
    state["failed"] += 1
    return None


# ── Excel helpers ─────────────────────────────────────────────────────────────

def widths(ws):
    for col in ws.columns:
        cl = get_column_letter(col[0].column)
        ml = max((len(str(c.value)) for c in col if c.value), default=0)
        ws.column_dimensions[cl].width = min(max(ml + 2, 8), 46)


def head(ws, cols):
    ws.append(cols)
    for cell in ws[1]:
        cell.font, cell.fill, cell.alignment = HDR_FONT, HDR_FILL, HDR_ALIGN
    ws.row_dimensions[1].height = 18
    ws.freeze_panes = "A2"


def banner(ws, title, sub):
    ws.insert_rows(1, amount=2)
    mc = ws.max_column or 1
    ws["A1"], ws["A2"] = title, sub
    ws["A1"].font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="444444")
    ws["A2"].fill = ALT_FILL
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 15
    if mc > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=mc)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=mc)


# ── Main ──────────────────────────────────────────────────────────────────────

async def main():
    ts         = datetime.now().strftime("%Y%m%d_%H%M")
    scraped_at = datetime.now().strftime("%d-%b-%Y %H:%M")

    known_primary, known_secondary, known_names, owner = load_known()
    grid = build_grid()

    print(f"Bengaluru Dark Store GRID SCAN - {ts}")
    print(f"Box     : lat {LAT_MIN}-{LAT_MAX}, lng {LNG_MIN}-{LNG_MAX}")
    print(f"Step    : {STEP} deg (~{STEP*111:.2f} km)")
    print(f"Points  : {len(grid)}   (concurrency {CONCURRENCY})")
    print(f"Known   : {len(known_primary)} Bengaluru primary, {len(known_secondary)} secondary")
    print(f"          {len(owner)} store ids known across ALL cities (dedup base)")
    print("=" * 78)

    found: dict[str, dict] = {}
    sec_seen: dict[str, set] = {}
    state = {"headers": {}, "stale": False, "failed": 0}
    done = {"n": 0, "no_store": 0}

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        ctx = await browser.new_context(
            user_agent=("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                        "(KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36"))
        page = await ctx.new_page()

        print("Capturing session headers...")
        state["headers"] = await capture_headers(page)
        if not state["headers"]:
            print("FAILED to capture session headers — aborting.")
            await browser.close()
            return
        print(f"Captured {len(state['headers'])} headers. Scanning...\n")

        sem = asyncio.Semaphore(CONCURRENCY)
        lock = asyncio.Lock()

        async def worker(idx: int, lat: float, lng: float):
            async with sem:
                r = await probe(ctx, state["headers"], lat, lng, state)
            async with lock:
                done["n"] += 1
                n = done["n"]
                if not r:
                    done["no_store"] += 1
                else:
                    sid = r["store_id"]
                    for s in r["secondary"]:
                        sec_seen.setdefault(s, set()).add((lat, lng))
                    if sid not in found:
                        is_new = sid not in known_primary
                        # Truly new = absent from EVERY city sheet, not just Bangalore.
                        other_city = owner.get(sid, "")
                        truly_new = sid not in owner
                        found[sid] = {
                            "store_id": sid,
                            "store_name": r["store_name"] or known_names.get(sid, ""),
                            "lat": lat, "lng": lng,
                            "secondary": ",".join(r["secondary"]),
                            "is_new": is_new,
                            "truly_new": truly_new,
                            "known_city": "" if truly_new else other_city,
                            "hit_count": 1,
                        }
                        n_new = sum(1 for v in found.values() if v["truly_new"])
                        if truly_new:
                            tag = "NEW  "
                        elif is_new:
                            tag = f"dup:{other_city[:6]}"
                        else:
                            tag = "known"
                        print(f"[{n:>4}/{len(grid)}] ({lat},{lng}) {tag:<10} {sid[:8]} "
                              f"{(r['store_name'] or '')[:28]:<28} | uniq {len(found)} new {n_new}")
                    else:
                        found[sid]["hit_count"] += 1
                if n % 250 == 0:
                    n_new = sum(1 for v in found.values() if v["is_new"])
                    print(f"    ... {n}/{len(grid)} probed | unique {len(found)} | new {n_new} "
                          f"| no-store {done['no_store']}")

        # Run in chunks so a stale session can be refreshed between them.
        CHUNK = 200
        for start in range(0, len(grid), CHUNK):
            batch = grid[start:start + CHUNK]
            await asyncio.gather(*(worker(start + j, la, ln)
                                   for j, (la, ln) in enumerate(batch)))
            if state["stale"]:
                print("    [session stale — re-capturing headers]")
                new_h = await capture_headers(page)
                if new_h:
                    state["headers"] = new_h
                state["stale"] = False
                await asyncio.sleep(2)

        await browser.close()

    sec_only = {s: c for s, c in sec_seen.items() if s not in found}
    new_primary = [v for v in found.values() if v["is_new"]]
    brand_new   = [v for v in found.values() if v["truly_new"]]
    cross_city  = [v for v in found.values() if v["is_new"] and not v["truly_new"]]
    all_known_before = known_primary | known_secondary
    new_sec_only = [s for s in sec_only if s not in all_known_before and s not in owner]
    total_after = len(known_primary | known_secondary | set(found) | set(sec_only))
    missed = [s for s in known_primary if s not in found]

    print(f"\n{'='*78}")
    print(f"Grid points probed          : {done['n']}")
    print(f"  no store at point         : {done['no_store']}")
    print(f"  probe failures            : {state['failed']}")
    print(f"Unique stores hit           : {len(found)}")
    print(f"  already in Bengaluru 134  : {len(found) - len(new_primary)}")
    print(f"  new vs Bengaluru 134      : {len(new_primary)}")
    print(f"    of which BRAND NEW      : {len(brand_new)}   <- new to the whole dataset")
    print(f"    of which another city's : {len(cross_city)}  <- border spill, not new")
    for v in cross_city:
        print(f"        {v['store_id'][:8]}  {v['store_name'][:32]:<32} belongs to {v['known_city']}")
    print(f"Secondary-only stores       : {len(sec_only)}  ({len(new_sec_only)} brand new)")
    print(f"Known-134 not hit by grid   : {len(missed)}")
    print(f"\nBengaluru confirmed stores  : {len(found) - len(cross_city) + len(missed)}")
    print(f"TOTAL unique store IDs after this scan: {total_after}")

    # ── Excel ────────────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    b_title = (f"Bengaluru Grid Scan - {len(found)} stores hit, "
               f"{len(brand_new)} brand new")
    b_sub   = (f"Generated: {scraped_at}  |  {len(grid)} points @ {STEP} deg "
               f"(~{STEP*111:.2f} km)  |  total unique store IDs now: {total_after}")

    ws1 = wb.active
    ws1.title = "Grid Stores"
    c1 = ["store_id", "store_name", "lat", "lng", "is_new", "truly_new",
          "known_city", "hit_count", "secondary"]
    head(ws1, c1)
    for r_idx, row in enumerate(
            sorted(found.values(), key=lambda x: (not x["truly_new"], x["store_name"])), start=2):
        ws1.append([row.get(c, "") for c in c1])
        if row["truly_new"]:
            for c in range(1, len(c1) + 1):
                ws1.cell(row=r_idx, column=c).fill = NEW_FILL
        elif row["is_new"]:
            for c in range(1, len(c1) + 1):
                ws1.cell(row=r_idx, column=c).fill = SEC_FILL   # border spill
        elif r_idx % 2 == 0:
            for c in range(1, len(c1) + 1):
                ws1.cell(row=r_idx, column=c).fill = ALT_FILL
    ws1.auto_filter.ref = ws1.dimensions
    widths(ws1)

    # Only genuinely-new stores here — border spill is excluded so this sheet is
    # a clean "add these to the dataset" list.
    ws2 = wb.create_sheet("NEW Stores")
    head(ws2, c1)
    for r_idx, row in enumerate(sorted(brand_new, key=lambda x: x["store_name"]), start=2):
        ws2.append([row.get(c, "") for c in c1])
        for c in range(1, len(c1) + 1):
            ws2.cell(row=r_idx, column=c).fill = NEW_FILL
    widths(ws2)

    # Stores the grid found that belong to a neighbouring city's sheet.
    ws5 = wb.create_sheet("Border Spill")
    head(ws5, c1)
    for r_idx, row in enumerate(sorted(cross_city, key=lambda x: x["store_name"]), start=2):
        ws5.append([row.get(c, "") for c in c1])
        for c in range(1, len(c1) + 1):
            ws5.cell(row=r_idx, column=c).fill = SEC_FILL
    widths(ws5)

    ws3 = wb.create_sheet("Secondary Only")
    c3 = ["store_id", "times_referenced", "sample_lat", "sample_lng", "brand_new"]
    head(ws3, c3)
    for r_idx, (s, coords) in enumerate(
            sorted(sec_only.items(), key=lambda kv: -len(kv[1])), start=2):
        c0 = sorted(coords)[0]
        ws3.append([s, len(coords), c0[0], c0[1],
                    "YES" if s not in all_known_before else "no"])
        for c in range(1, len(c3) + 1):
            ws3.cell(row=r_idx, column=c).fill = SEC_FILL
    widths(ws3)

    ws4 = wb.create_sheet("Summary")
    head(ws4, ["Metric", "Value"])
    stats = [
        ("Grid box", f"lat {LAT_MIN}-{LAT_MAX}, lng {LNG_MIN}-{LNG_MAX}"),
        ("Grid step", f"{STEP} deg (~{STEP*111:.2f} km)"),
        ("Grid points probed", done["n"]),
        ("Points with no store", done["no_store"]),
        ("Probe failures", state["failed"]),
        ("Unique stores hit by grid", len(found)),
        ("Already among the 134", len(found) - len(new_primary)),
        ("New vs Bengaluru 134", len(new_primary)),
        ("  of which BRAND NEW to dataset", len(brand_new)),
        ("  of which border spill (other city)", len(cross_city)),
        ("Secondary-only stores", len(sec_only)),
        ("Known-134 not hit by grid", len(missed)),
        ("Bengaluru confirmed stores", len(found) - len(cross_city) + len(missed)),
        ("TOTAL unique store IDs now", total_after),
        ("Scraped at", scraped_at),
    ]
    for r_idx, (m, v) in enumerate(stats, start=2):
        ws4.cell(row=r_idx, column=1, value=m)
        ws4.cell(row=r_idx, column=2, value=v)
        if r_idx % 2 == 0:
            for c in range(1, 3):
                ws4.cell(row=r_idx, column=c).fill = ALT_FILL
    widths(ws4)

    for ws in (ws1, ws2, ws3, ws4, ws5):
        banner(ws, b_title, b_sub)

    out = Path(__file__).parent.parent / f"zepto_bengaluru_grid_scan_{ts}.xlsx"
    wb.save(out)
    print(f"\nSaved -> {out}")


if __name__ == "__main__":
    asyncio.run(main())
