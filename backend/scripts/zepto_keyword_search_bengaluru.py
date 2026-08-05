"""
Zepto keyword product scraper — Bengaluru
Navigates to zepto.com/search?query=KEYWORD, intercepts the search API response,
and extracts product position, price, brand, and store data into Excel.

Client: Dobra
Run from backend/:
    python -m scripts.zepto_keyword_search_bengaluru
Output: zepto_keyword_search_bengaluru_TIMESTAMP.xlsx
"""
import asyncio
import sys
from pathlib import Path
from datetime import datetime
from urllib.parse import quote

sys.path.insert(0, str(Path(__file__).parent.parent))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from playwright.async_api import async_playwright

# ── Config ────────────────────────────────────────────────────────────────────
CLIENT_BRAND    = "Brik Oven"      # your client's brand name
BENGALURU_LAT   = 13.0035068
BENGALURU_LNG   = 77.5890953

# Keywords to search — add / remove as needed
KEYWORDS = [
    "brik oven",
    "sourdough bread",
    "whole wheat bread",
    "multigrain bread",
    "sourdough",
    "brick oven bread",
]

_SEARCH_PATH = "/user-search-service/api/v3/search"
_BASE        = "https://www.zepto.com"

# Re-use the same init script that injects lat/lng into get_page calls
_INIT_SCRIPT = """
(() => {
    const injectCoords = (url) => {
        if (typeof url !== 'string' || !url.includes('/lms/api/v2/get_page')) return url;
        const lat = localStorage.getItem('__probe_lat');
        const lng = localStorage.getItem('__probe_lng');
        if (!lat || !lng) return url;
        try {
            const u = new URL(url, location.href);
            u.searchParams.set('latitude', lat);
            u.searchParams.set('longitude', lng);
            return u.toString();
        } catch(e) { return url; }
    };
    const origFetch = window.fetch;
    window.fetch = function(url, opts) {
        return origFetch.apply(this, [injectCoords(url), opts]);
    };
    const origOpen = XMLHttpRequest.prototype.open;
    XMLHttpRequest.prototype.open = function(method, url, ...rest) {
        return origOpen.call(this, method, injectCoords(url), ...rest);
    };
})();
"""

COLUMNS = [
    "keyword", "position", "product_name", "brand",
    "mrp_rs", "selling_price_rs", "discount_pct",
    "pack_size", "rating", "rating_count", "in_stock",
    "category", "query_match_bucket", "is_client_brand",
    "product_id", "variant_id", "store_id", "scraped_at",
]

HDR_FONT   = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
HDR_FILL   = PatternFill("solid", fgColor="2E4057")
HDR_ALIGN  = Alignment(horizontal="center", vertical="center")
ALT_FILL   = PatternFill("solid", fgColor="F0F4F8")
CLI_FILL   = PatternFill("solid", fgColor="D9F7E6")   # green highlight for client rows
KWTL_FILL  = PatternFill("solid", fgColor="048A81")
KWTL_FONT  = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
TITLE_FILL = PatternFill("solid", fgColor="1A3C5E")


def _paise_to_rs(paise) -> float:
    return round((paise or 0) / 100, 2)


def extract_products(data: dict, keyword: str, scraped_at: str) -> list[dict]:
    """Pull all product rows from a single search API response."""
    rows = []
    for widget in data.get("layout", []):
        if widget.get("widgetId") != "PRODUCT_GRID":
            continue
        items = (
            widget.get("data", {})
            .get("resolver", {})
            .get("data", {})
            .get("items") or []
        )
        for item in items:
            pr = item.get("productResponse")
            if not pr:
                continue
            prod = pr.get("product", {}) or {}
            pv   = pr.get("productVariant", {}) or {}
            rat  = pv.get("ratingSummary", {}) or {}
            meta = pr.get("meta", {}) or {}

            brand = prod.get("brand", "")
            mrp   = pv.get("mrp") or pr.get("mrp") or 0
            sp    = pr.get("discountedSellingPrice") or pr.get("sellingPrice") or 0

            rows.append({
                "keyword":          keyword,
                "position":         (item.get("position") or 0) + 1,  # 1-based
                "product_name":     prod.get("name", ""),
                "brand":            brand,
                "mrp_rs":           _paise_to_rs(mrp),
                "selling_price_rs": _paise_to_rs(sp),
                "discount_pct":     pr.get("discountPercent", 0),
                "pack_size":        pv.get("formattedPacksize", ""),
                "rating":           rat.get("averageRating", ""),
                "rating_count":     rat.get("totalRatings", 0),
                "in_stock":         not pr.get("outOfStock", False),
                "category":         pr.get("primaryCategoryName", ""),
                "query_match_bucket": meta.get("query_matching_bucket", ""),
                "is_client_brand":  "YES" if CLIENT_BRAND.lower() in brand.lower() else "NO",
                "product_id":       prod.get("id", ""),
                "variant_id":       pv.get("id", ""),
                "store_id":         pr.get("storeId", ""),
                "scraped_at":       scraped_at,
            })
    return rows


async def search_keyword(keyword: str, page, scraped_at: str) -> list[dict]:
    """Navigate to the Zepto search page and collect all product results."""
    collected: list[dict] = []
    seen_page_keys: set = set()

    async def on_response(response):
        if _SEARCH_PATH not in response.url:
            return
        try:
            data = await response.json()
            # Skip pure autocomplete responses (no product grid)
            has_grid = any(
                w.get("widgetId") == "PRODUCT_GRID"
                for w in data.get("layout", [])
            )
            if not has_grid:
                return
            page_key = data.get("currentPage", 0)
            if page_key in seen_page_keys:
                return
            seen_page_keys.add(page_key)
            rows = extract_products(data, keyword, scraped_at)
            collected.extend(rows)
            end = data.get("hasReachedEnd", True)
            print(f"    page {page_key}: {len(rows)} products  ({'end' if end else 'more...'})")
        except Exception:
            pass

    page.on("response", on_response)
    try:
        url = f"{_BASE}/search?query={quote(keyword)}"
        await page.goto(url, timeout=30000, wait_until="domcontentloaded")
        await page.wait_for_timeout(4000)

        # Scroll to load paginated results (up to 5 extra pages)
        for _ in range(5):
            await page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
            await page.wait_for_timeout(2000)
            # Stop if the last captured response said hasReachedEnd
            # (we track this via seen_page_keys vs total pages)
            last_count = len(collected)
            await page.wait_for_timeout(500)
            if len(collected) == last_count:
                break

    except Exception as e:
        print(f"    [error] {keyword}: {e}")
    finally:
        page.remove_listener("response", on_response)

    return collected


def apply_banner(ws, title: str, subtitle: str):
    ws.insert_rows(1, amount=2)
    mc = ws.max_column or 1
    ws["A1"] = title
    ws["A2"] = subtitle
    ws["A1"].font  = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    ws["A1"].fill  = TITLE_FILL
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22
    ws["A2"].font  = Font(name="Calibri", size=10, italic=True, color="444444")
    ws["A2"].fill  = ALT_FILL
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 15
    if mc > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=mc)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=mc)


def set_col_widths(ws, min_w=8, max_w=45):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max(
            (len(str(cell.value)) for cell in col if cell.value), default=0
        )
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_w), max_w)


async def main():
    ts         = datetime.now().strftime("%Y%m%d_%H%M")
    scraped_at = datetime.now().strftime("%d-%b-%Y %H:%M")
    print(f"Zepto Keyword Scraper — Bengaluru — {ts}")
    print(f"Client brand : {CLIENT_BRAND}")
    print(f"Keywords     : {len(KEYWORDS)}")
    print("=" * 60)

    all_rows: list[dict] = []
    store_id = None

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        ctx = await browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/125.0.0.0 Safari/537.36"
            )
        )
        page = await ctx.new_page()
        await page.add_init_script(_INIT_SCRIPT)

        # ── Warm-up: open zepto.com with Bengaluru coords ──────────────────
        print("Warming up browser (setting Bengaluru location)...")

        async def capture_store(response):
            nonlocal store_id
            if "/lms/api/v2/get_page" in response.url and not store_id:
                try:
                    data = await response.json()
                    sid = (data.get("storeServiceableResponse") or {}).get("storeId")
                    if sid:
                        store_id = sid
                except Exception:
                    pass

        page.on("response", capture_store)
        await page.goto(_BASE, timeout=30000, wait_until="domcontentloaded")
        await page.evaluate(f"""() => {{
            localStorage.setItem('__probe_lat', '{BENGALURU_LAT}');
            localStorage.setItem('__probe_lng', '{BENGALURU_LNG}');
        }}""")
        await page.reload(timeout=25000, wait_until="domcontentloaded")
        await page.wait_for_timeout(3000)
        page.remove_listener("response", capture_store)
        print(f"Store ID     : {store_id or 'not captured (will still search)'}")
        print()

        # ── Search each keyword ─────────────────────────────────────────────
        for i, kw in enumerate(KEYWORDS, 1):
            print(f"[{i}/{len(KEYWORDS)}] '{kw}'")
            rows = await search_keyword(kw, page, scraped_at)
            print(f"  → {len(rows)} products total")

            client_rows = [r for r in rows if r["is_client_brand"] == "YES"]
            if client_rows:
                best = min(client_rows, key=lambda x: x["position"])
                print(f"  ★ {CLIENT_BRAND} best rank: #{best['position']} — {best['product_name'][:50]}")
            else:
                print(f"  (no {CLIENT_BRAND} products found for this keyword)")

            # Print top 5 regardless
            for r in rows[:5]:
                flag = " ★" if r["is_client_brand"] == "YES" else ""
                print(f"    #{r['position']:>2}  {r['product_name'][:48]:<48}  ₹{r['selling_price_rs']}{flag}")

            all_rows.extend(rows)
            await asyncio.sleep(1.5)

        await browser.close()

    print(f"\n{'='*60}")
    print(f"Total rows: {len(all_rows)}")
    client_count = sum(1 for r in all_rows if r["is_client_brand"] == "YES")
    print(f"{CLIENT_BRAND} rows: {client_count}")

    # ── Build Excel ─────────────────────────────────────────────────────────
    banner_title = f"Zepto Keyword Search — Bengaluru — {CLIENT_BRAND}"
    banner_sub   = f"Generated: {scraped_at}  |  Store: {store_id or 'unknown'}  |  Coords: {BENGALURU_LAT}, {BENGALURU_LNG}"

    wb = openpyxl.Workbook()

    # Sheet 1: All results
    ws_all = wb.active
    ws_all.title = "All Results"
    ws_all.append(COLUMNS)
    for cell in ws_all[1]:
        cell.font  = HDR_FONT
        cell.fill  = HDR_FILL
        cell.alignment = HDR_ALIGN
    ws_all.row_dimensions[1].height = 18
    ws_all.freeze_panes = "A2"

    for r_idx, row in enumerate(all_rows, start=2):
        ws_all.append([row.get(c, "") for c in COLUMNS])
        fill = CLI_FILL if row["is_client_brand"] == "YES" else (ALT_FILL if r_idx % 2 == 0 else None)
        if fill:
            for c in range(1, len(COLUMNS) + 1):
                ws_all.cell(row=r_idx, column=c).fill = fill
    ws_all.auto_filter.ref = ws_all.dimensions
    set_col_widths(ws_all)

    # Sheet 2: Summary by keyword
    ws_sum = wb.create_sheet("By Keyword")
    sum_cols = ["Keyword", "Total Products", f"{CLIENT_BRAND} Products",
                f"Best {CLIENT_BRAND} Rank", f"Best {CLIENT_BRAND} Product"]
    ws_sum.append(sum_cols)
    for cell in ws_sum[1]:
        cell.font  = HDR_FONT
        cell.fill  = HDR_FILL
        cell.alignment = HDR_ALIGN
    ws_sum.row_dimensions[1].height = 18
    ws_sum.freeze_panes = "A2"

    for r_idx, kw in enumerate(KEYWORDS, start=2):
        kw_rows = [r for r in all_rows if r["keyword"] == kw]
        cli     = [r for r in kw_rows if r["is_client_brand"] == "YES"]
        best    = min(cli, key=lambda x: x["position"]) if cli else None
        ws_sum.append([
            kw,
            len(kw_rows),
            len(cli),
            best["position"] if best else "—",
            best["product_name"] if best else "—",
        ])
        if r_idx % 2 == 0:
            for c in range(1, 6):
                ws_sum.cell(row=r_idx, column=c).fill = ALT_FILL
    set_col_widths(ws_sum)

    # Sheet 3: Client brand rows only
    ws_cli = wb.create_sheet(f"{CLIENT_BRAND} Only")
    ws_cli.append(COLUMNS)
    for cell in ws_cli[1]:
        cell.font  = HDR_FONT
        cell.fill  = HDR_FILL
        cell.alignment = HDR_ALIGN
    ws_cli.row_dimensions[1].height = 18
    ws_cli.freeze_panes = "A2"

    cli_rows = sorted(
        (r for r in all_rows if r["is_client_brand"] == "YES"),
        key=lambda x: (x["keyword"], x["position"])
    )
    for r_idx, row in enumerate(cli_rows, start=2):
        ws_cli.append([row.get(c, "") for c in COLUMNS])
        ws_cli.cell(row=r_idx, column=1)  # already green by default for client
        if r_idx % 2 == 0:
            for c in range(1, len(COLUMNS) + 1):
                ws_cli.cell(row=r_idx, column=c).fill = CLI_FILL
    ws_cli.auto_filter.ref = ws_cli.dimensions
    set_col_widths(ws_cli)

    # Apply banner to all sheets
    for ws in [ws_all, ws_sum, ws_cli]:
        apply_banner(ws, banner_title, banner_sub)

    out = Path(__file__).parent.parent / f"zepto_keyword_search_bengaluru_{ts}.xlsx"
    wb.save(out)
    print(f"Saved → {out}")


if __name__ == "__main__":
    asyncio.run(main())
