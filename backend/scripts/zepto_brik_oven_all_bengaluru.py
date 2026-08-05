"""
Brik Oven public product data across EVERY Bengaluru dark store.

Chain (both hops through Zepto's own APIs):

    lat/lng  ──►  get_page  ──►  storeId  ──►  search  ──►  products
                  (coordinate sent to the API; the serving store comes back)

Every store's lat/lng from the master workbook is sent to get_page, and the store
that coordinate resolves to is used for the search — so the store is never assumed,
it is confirmed live and recorded per row.

Speed: both hops replay session headers captured once from the page's own requests
(~0.3 s/call) instead of navigating the browser per store (~4 s). That is the
difference between a 25-minute sweep and a 2-minute one.

Run from backend/:
    python -m scripts.zepto_brik_oven_all_bengaluru
Output: zepto_brik_oven_bengaluru_TIMESTAMP.xlsx
"""
import asyncio
import json
import sys
from pathlib import Path
from datetime import datetime

sys.path.insert(0, str(Path(__file__).parent.parent))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from playwright.async_api import async_playwright

BRAND   = "Brik Oven"
KEYWORD = "brik oven"
CITY_SHEET = "Bengaluru"

BACKEND = Path(__file__).parent.parent
_BASE = "https://www.zepto.com"
_BFF  = "https://bff-gateway.zepto.com"
_GP   = "/lms/api/v2/get_page"
_SEARCH = "/user-search-service/api/v3/search"

TRIES = 3
REFRESH_EVERY = 40      # rotate the session proactively, before it can expire
_DROP = {"host", "content-length", "connection", "accept-encoding",
         ":method", ":path", ":authority", ":scheme"}

HDR_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
HDR_FILL = PatternFill("solid", fgColor="2E4057")
HDR_ALIGN = Alignment(horizontal="center", vertical="center")
ALT_FILL = PatternFill("solid", fgColor="F0F4F8")
OK_FILL  = PatternFill("solid", fgColor="D9F7E6")
NO_FILL  = PatternFill("solid", fgColor="FDE8E8")
ERR_FILL = PatternFill("solid", fgColor="FFF4CE")
TITLE_FILL = PatternFill("solid", fgColor="1A3C5E")


def latest(pat):
    f = sorted(BACKEND.glob(pat))
    return f[-1] if f else None


def load_stores() -> list[dict]:
    """Bengaluru stores from the master workbook — all 168, with lat/lng."""
    path = latest("zepto_master_all_cities_*.xlsx")
    if not path:
        raise SystemExit("No master workbook found. Run zepto_build_master_all_cities first.")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if CITY_SHEET not in wb.sheetnames:
        wb.close()
        raise SystemExit(f"Sheet '{CITY_SHEET}' not in {path.name}")
    rows = list(wb[CITY_SHEET].iter_rows(values_only=True))
    wb.close()
    hdr = [str(h).strip().lower() if h else "" for h in rows[0]]
    out, seen = [], set()
    for r in rows[1:]:
        if not r or not r[0]:
            continue
        d = dict(zip(hdr, r))
        sid = str(d.get("store_id") or "").strip()
        if not sid or sid in seen:
            continue
        try:
            la, ln = float(d.get("lat") or 0), float(d.get("lng") or 0)
        except (TypeError, ValueError):
            continue
        if not la or not ln:
            continue
        seen.add(sid)
        out.append({
            "store_id": sid,
            "store_name": str(d.get("store_name") or "").strip(),
            "pincode": str(d.get("pincode") or "").strip(),
            "area": str(d.get("area") or "").strip(),
            "lat": la, "lng": ln,
            "discovery": str(d.get("discovery") or "").strip(),
        })
    print(f"Store source: {path.name}")
    return out


async def capture(page, url_part, do_nav, settle_ms: int = 5000) -> tuple[dict, str | None]:
    """Capture the session headers (and body) from the page's own request.

    The request fires AFTER domcontentloaded, so the listener has to stay attached
    through a settle window — removing it the moment navigation returns captures
    nothing.
    """
    cap = {"h": None, "body": None}

    async def on_req(req):
        if url_part in req.url and cap["h"] is None:
            cap["h"] = dict(req.headers)
            try:
                cap["body"] = req.post_data
            except Exception:
                pass

    page.on("request", on_req)
    try:
        await do_nav()
        waited = 0
        while cap["h"] is None and waited < settle_ms:
            await page.wait_for_timeout(250)
            waited += 250
    except Exception:
        pass
    finally:
        page.remove_listener("request", on_req)
    if not cap["h"]:
        return {}, None
    return {k: v for k, v in cap["h"].items() if k.lower() not in _DROP}, cap["body"]


def gp_url(lat, lng):
    return (f"{_BFF}{_GP}?latitude={lat}&longitude={lng}"
            f"&page_type=HOME&version=v2&show_new_eta_banner=true"
            f"&page_size=3&enforce_platform_type=WEB")


async def refresh_session(page, state) -> bool:
    """Re-capture BOTH header sets. Refreshing only get_page was the bug that let
    an expired search session fail every remaining store."""
    gp, _ = await capture(
        page, _GP,
        lambda: page.goto(_BASE, timeout=30000, wait_until="domcontentloaded"))
    if gp:
        state["gp"] = gp
    await page.wait_for_timeout(1200)
    se, raw = await capture(
        page, _SEARCH,
        lambda: page.goto(f"{_BASE}/search?query={KEYWORD.replace(' ', '%20')}",
                          timeout=30000, wait_until="domcontentloaded"),
        settle_ms=8000)
    if se:
        state["se"] = se
        if raw:
            try:
                b = json.loads(raw)
                b["query"] = KEYWORD
                state["body"] = b          # carries a fresh userSessionId
            except Exception:
                pass
    return bool(gp and se)


async def resolve_store(ctx, state, lat, lng):
    """HOP 1 — send the coordinate to get_page, read back the serving store."""
    for a in range(1, TRIES + 1):
        try:
            r = await ctx.request.get(gp_url(lat, lng), headers=state["gp"], timeout=15000)
            if r.status == 200:
                svc = (await r.json()).get("storeServiceableResponse") or {}
                return svc.get("storeId"), [str(x) for x in (svc.get("secondaryStoreIds") or [])]
            if r.status in (401, 403, 429):
                state["stale"] = True
                await asyncio.sleep(1.2 * a)
                continue
        except Exception:
            pass
        await asyncio.sleep(0.5 * a)
    return None, []


def _rs(p):
    return round((p or 0) / 100, 2)


def extract(data: dict) -> tuple[list[dict], set, int]:
    """All Brik Oven products from a search response, plus store ids seen and total."""
    prods, ids, total = [], set(), 0
    for w in data.get("layout", []):
        if w.get("widgetId") != "PRODUCT_GRID":
            continue
        for it in (w.get("data", {}).get("resolver", {}).get("data", {}).get("items") or []):
            pr = it.get("productResponse")
            if not pr:
                continue
            total += 1
            if pr.get("storeId"):
                ids.add(pr["storeId"])
            prod = pr.get("product", {}) or {}
            if BRAND.lower() not in (prod.get("brand") or "").lower():
                continue
            pv  = pr.get("productVariant", {}) or {}
            rat = pv.get("ratingSummary", {}) or {}
            meta = pr.get("meta", {}) or {}
            imgs = pv.get("images") or []
            mrp = pv.get("mrp") or pr.get("mrp") or 0
            sp  = pr.get("discountedSellingPrice") or pr.get("sellingPrice") or 0
            prods.append({
                "position":       (it.get("position") or 0) + 1,
                "product_name":   prod.get("name", ""),
                "brand":          prod.get("brand", ""),
                "pack_size":      pv.get("formattedPacksize", ""),
                "mrp_rs":         _rs(mrp),
                "selling_price_rs": _rs(sp),
                "discount_pct":   pr.get("discountPercent", 0),
                "discount_rs":    _rs(pr.get("discountAmount")),
                "super_saver_rs": _rs(pr.get("superSaverSellingPrice")),
                "in_stock":       not pr.get("outOfStock", False),
                "available_qty":  pr.get("availableQuantity", 0),
                "max_order_qty":  pv.get("maxAllowedQuantity", ""),
                "rating":         rat.get("averageRating", ""),
                "rating_count":   rat.get("totalRatings", 0),
                "category":       pr.get("primaryCategoryName", ""),
                "match_bucket":   meta.get("query_matching_bucket", ""),
                "weight_g":       pv.get("weightInGms", ""),
                "shelf_life_hrs": pv.get("shelfLifeInHours", ""),
                "country":        prod.get("countryOfOrigin", ""),
                "manufacturer":   prod.get("manufacturerName") or "",
                "fssai":          pv.get("fssaiLicense", ""),
                "product_id":     prod.get("id", ""),
                "variant_id":     pv.get("id", ""),
                "image":          (imgs[0].get("path") if imgs else ""),
                "resp_store_id":  pr.get("storeId", ""),
            })
    return prods, ids, total


async def search_store(ctx, state, sid, all_ids):
    """HOP 2 — search KEYWORD at this store.

    A 200 carrying zero products is NOT a real answer — every Bengaluru store
    returns something for this keyword — so it is retried and, if it persists,
    reported as a failed measurement rather than "not stocked". `search_stale`
    tells the caller to re-capture the SEARCH headers specifically; refreshing
    only the get_page headers leaves an expired search session expired forever.
    """
    body = dict(state["body"])
    last_status = None
    for a in range(1, TRIES + 1):
        h = dict(state["se"])
        h["store_id"] = sid
        h["storeid"] = sid
        h["store_ids"] = all_ids
        h["store_etas"] = json.dumps({s: -1 for s in all_ids.split(",")})
        try:
            r = await ctx.request.post(f"{_BFF}{_SEARCH}", headers=h,
                                       data=json.dumps(body), timeout=20000)
            last_status = r.status
            if r.status == 200:
                prods, ids, total = extract(await r.json())
                if total > 0:
                    return prods, ids, total, True
                # 200 but empty page -> session is being soft-throttled
                state["search_stale"] = True
            elif r.status in (401, 403, 429):
                state["search_stale"] = True
        except Exception:
            pass
        await asyncio.sleep(0.8 * a)
    state["search_stale"] = True
    state["last_status"] = last_status
    return [], set(), 0, False


def widths(ws, mx=46):
    for col in ws.columns:
        cl = get_column_letter(col[0].column)
        ml = max((len(str(c.value)) for c in col if c.value), default=0)
        ws.column_dimensions[cl].width = min(max(ml + 2, 9), mx)


def head(ws, cols):
    ws.append(cols)
    for c in ws[1]:
        c.font, c.fill, c.alignment = HDR_FONT, HDR_FILL, HDR_ALIGN
    ws.row_dimensions[1].height = 18
    ws.freeze_panes = "A2"


def banner(ws, t, s):
    ws.insert_rows(1, amount=2)
    mc = ws.max_column or 1
    ws["A1"], ws["A2"] = t, s
    ws["A1"].font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="444444")
    ws["A2"].fill = ALT_FILL
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 15
    if mc > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=mc)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=mc)


async def main():
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    at = datetime.now().strftime("%d-%b-%Y %H:%M")

    stores = load_stores()
    print(f"\n{BRAND} public data across Bengaluru dark stores — {ts}")
    print(f"Dark stores : {len(stores)}")
    print(f"Keyword     : \"{KEYWORD}\"")
    print(f"Method      : lat/lng -> get_page -> storeId -> search")
    print("=" * 108)

    detail, summary = [], []

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        ctx = await browser.new_context(
            user_agent=("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                        "(KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36"))
        page = await ctx.new_page()
        state = {"gp": {}, "se": {}, "body": {}, "stale": False,
                 "search_stale": False, "last_status": None}

        print("Capturing session headers...")
        state["gp"], _ = await capture(
            page, _GP,
            lambda: page.goto(_BASE, timeout=30000, wait_until="domcontentloaded"))
        print(f"  get_page headers : {len(state['gp'])}")
        await page.wait_for_timeout(1500)
        state["se"], raw = await capture(
            page, _SEARCH,
            lambda: page.goto(f"{_BASE}/search?query=brik%20oven",
                              timeout=30000, wait_until="domcontentloaded"),
            settle_ms=8000)
        print(f"  search headers   : {len(state['se'])}")
        if not state["gp"] or not state["se"]:
            print("Header capture failed — aborting.")
            print("  (retry usually fixes it; Zepto occasionally serves a WAF "
                  "challenge on a cold session)")
            await browser.close()
            return
        state["body"] = json.loads(raw) if raw else {
            "query": KEYWORD, "pageNumber": 0, "mode": "SHOW_ALL_RESULTS"}
        state["body"]["query"] = KEYWORD
        print()

        for i, st in enumerate(stores, 1):
            # HOP 1 — coordinate through the API
            api_sid, sec = await resolve_store(ctx, state, st["lat"], st["lng"])
            if not api_sid:
                print(f"[{i:>3}/{len(stores)}] {st['store_name'][:30]:<30} "
                      f"get_page failed — store not resolved from lat/lng")
                summary.append({**st, "api_store_id": "", "id_match": "",
                                "status": "GETPAGE_FAILED", "found": "",
                                "product_count": "", "best_rank": "",
                                "min_price": "", "max_price": "", "total_results": ""})
                continue

            match = "SAME" if api_sid == st["store_id"] else "DIFF"
            prods, ids, total, ok = await search_store(
                ctx, state, api_sid, ",".join([api_sid] + sec))

            hdr = (f"[{i:>3}/{len(stores)}] {st['store_name'][:30]:<30} "
                   f"{st['pincode']:<7} {api_sid[:8]} [{match}]")
            if not ok:
                print(f"{hdr}  search failed")
                status, found = "SEARCH_FAILED", ""
            elif prods:
                print(f"{hdr}  {len(prods)} products")
                for pr in sorted(prods, key=lambda x: x["position"]):
                    stock = "in stock" if pr["in_stock"] else "OUT OF STOCK"
                    print(f"        #{pr['position']:<3} {pr['product_name'][:44]:<44} "
                          f"{pr['pack_size'][:14]:<14} MRP Rs{pr['mrp_rs']:<7.2f} "
                          f"SP Rs{pr['selling_price_rs']:<7.2f} -{pr['discount_pct']:<3}% "
                          f"{stock:<12} {pr['rating']}({pr['rating_count']})")
                status, found = "OK", "YES"
            else:
                print(f"{hdr}  no {BRAND} (page had {total} products)")
                status, found = "OK", "NO"

            prices = [pr["selling_price_rs"] for pr in prods if pr["selling_price_rs"]]
            summary.append({
                **st,
                "api_store_id": api_sid,
                "id_match": match,
                "status": status,
                "found": found,
                "product_count": len(prods) if ok else "",
                "best_rank": min((pr["position"] for pr in prods), default="") if prods else "",
                "min_price": min(prices) if prices else "",
                "max_price": max(prices) if prices else "",
                "total_results": total if ok else "",
            })
            await asyncio.sleep(0.25)

            for pr in prods:
                detail.append({
                    "store_name": st["store_name"], "pincode": st["pincode"],
                    "area": st["area"], "lat": st["lat"], "lng": st["lng"],
                    "api_store_id": api_sid, **pr,
                })

            # Refresh the session when it goes stale, and proactively every
            # REFRESH_EVERY stores so it never gets the chance to expire mid-sweep.
            if state["search_stale"] or state["stale"] or i % REFRESH_EVERY == 0:
                why = ("search stale" if state["search_stale"]
                       else "get_page stale" if state["stale"] else "routine")
                print(f"        [refreshing session — {why}]")
                await asyncio.sleep(3)
                await refresh_session(page, state)
                state["stale"] = state["search_stale"] = False

        await browser.close()

    ok_rows  = [s for s in summary if s["status"] == "OK"]
    stocked  = [s for s in ok_rows if s["found"] == "YES"]
    failed   = [s for s in summary if s["status"] != "OK"]
    prices   = [d["selling_price_rs"] for d in detail if d["selling_price_rs"]]

    print(f"\n{'='*108}")
    print(f"Dark stores checked      : {len(stores)}")
    print(f"Measured successfully    : {len(ok_rows)}   failed: {len(failed)}")
    print(f"Stocking {BRAND:<14}: {len(stocked)}")
    if ok_rows:
        print(f"Coverage                 : {len(stocked)/len(ok_rows)*100:.1f}% of measured")
    print(f"Product rows             : {len(detail)}")
    print(f"Unique products          : {len({d['product_name'] for d in detail})}")
    if prices:
        print(f"Price range              : Rs{min(prices):.2f} - Rs{max(prices):.2f}"
              f"   avg Rs{sum(prices)/len(prices):.2f}")

    # ── Excel ────────────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    cov = f"{len(stocked)}/{len(ok_rows)}" if ok_rows else "n/a"
    T = f"{BRAND} — Bengaluru dark stores — stocked in {cov}"
    S = (f"Generated: {at}  |  keyword \"{KEYWORD}\"  |  {len(stores)} dark stores  |  "
         f"lat/lng -> get_page -> storeId -> search  |  {len(detail)} product rows")

    DET = ["store_name", "pincode", "area", "lat", "lng", "api_store_id",
           "position", "product_name", "brand", "pack_size",
           "mrp_rs", "selling_price_rs", "discount_pct", "discount_rs", "super_saver_rs",
           "in_stock", "available_qty", "max_order_qty", "rating", "rating_count",
           "category", "match_bucket", "weight_g", "shelf_life_hrs", "country",
           "manufacturer", "fssai", "product_id", "variant_id", "image"]
    ws1 = wb.active
    ws1.title = "Product Detail"
    head(ws1, DET)
    for i, r in enumerate(sorted(detail, key=lambda x: (x["store_name"], x["position"])), start=2):
        ws1.append([r.get(c, "") for c in DET])
        if i % 2 == 0:
            for c in range(1, len(DET) + 1):
                ws1.cell(row=i, column=c).fill = ALT_FILL
    ws1.auto_filter.ref = ws1.dimensions
    widths(ws1)

    SUM = ["store_name", "pincode", "area", "lat", "lng", "store_id", "api_store_id",
           "id_match", "status", "found", "product_count", "best_rank",
           "min_price", "max_price", "total_results", "discovery"]
    ws2 = wb.create_sheet("Store Coverage")
    head(ws2, SUM)
    ordr = {"YES": 0, "NO": 1, "": 2}
    for i, r in enumerate(sorted(summary, key=lambda x: (ordr.get(x["found"], 3),
                                                         x["best_rank"] if x["best_rank"] != "" else 999)),
                          start=2):
        ws2.append([r.get(c, "") for c in SUM])
        fill = ERR_FILL if r["status"] != "OK" else (OK_FILL if r["found"] == "YES" else NO_FILL)
        for c in range(1, len(SUM) + 1):
            ws2.cell(row=i, column=c).fill = fill
    ws2.auto_filter.ref = ws2.dimensions
    widths(ws2)

    ws3 = wb.create_sheet("By Product")
    P = ["product_name", "pack_size", "stores_stocking", "in_stock_stores",
         "avg_rank", "best_rank", "min_price_rs", "max_price_rs", "avg_price_rs",
         "mrp_rs", "avg_discount_pct", "rating", "rating_count"]
    head(ws3, P)
    byp = {}
    for d in detail:
        byp.setdefault(d["product_name"], []).append(d)
    for i, (nm, rs) in enumerate(sorted(byp.items(), key=lambda kv: -len(kv[1])), start=2):
        pz = [r["selling_price_rs"] for r in rs if r["selling_price_rs"]]
        rk = [r["position"] for r in rs if r["position"]]
        dc = [r["discount_pct"] for r in rs if r["discount_pct"] is not None]
        ws3.append([nm, rs[0]["pack_size"], len(rs), sum(1 for r in rs if r["in_stock"]),
                    round(sum(rk)/len(rk), 1) if rk else "", min(rk) if rk else "",
                    min(pz) if pz else "", max(pz) if pz else "",
                    round(sum(pz)/len(pz), 2) if pz else "", rs[0]["mrp_rs"],
                    round(sum(dc)/len(dc), 1) if dc else "",
                    rs[0]["rating"], rs[0]["rating_count"]])
        if i % 2 == 0:
            for c in range(1, len(P) + 1):
                ws3.cell(row=i, column=c).fill = ALT_FILL
    widths(ws3)

    ws4 = wb.create_sheet("Summary")
    head(ws4, ["Metric", "Value"])
    stats = [
        ("Brand", BRAND), ("Keyword", KEYWORD), ("City", "Bengaluru"),
        ("Dark stores checked", len(stores)),
        ("Measured successfully", len(ok_rows)),
        ("Failed to measure", len(failed)),
        ("Stores stocking brand", len(stocked)),
        ("Stores NOT stocking", len(ok_rows) - len(stocked)),
        ("Coverage % of measured",
         f"{len(stocked)/len(ok_rows)*100:.1f}%" if ok_rows else "—"),
        ("Product rows", len(detail)),
        ("Unique products", len(byp)),
        ("Min price", f"Rs{min(prices):.2f}" if prices else "—"),
        ("Max price", f"Rs{max(prices):.2f}" if prices else "—"),
        ("Avg price", f"Rs{sum(prices)/len(prices):.2f}" if prices else "—"),
        ("Best rank seen", min((d["position"] for d in detail), default="—")),
        ("Avg rank", f"{sum(d['position'] for d in detail)/len(detail):.1f}" if detail else "—"),
        ("Scraped at", at),
    ]
    for i, (m, v) in enumerate(stats, start=2):
        ws4.cell(row=i, column=1, value=m)
        ws4.cell(row=i, column=2, value=v)
        if i % 2 == 0:
            for c in (1, 2):
                ws4.cell(row=i, column=c).fill = ALT_FILL
    widths(ws4)

    for w in (ws1, ws2, ws3, ws4):
        banner(w, T, S)

    out = BACKEND / f"zepto_brik_oven_bengaluru_{ts}.xlsx"
    wb.save(out)
    print(f"\nSaved -> {out}")


if __name__ == "__main__":
    asyncio.run(main())
