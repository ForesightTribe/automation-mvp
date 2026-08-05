"""
Brik Oven coverage checker across all Bengaluru dark stores.
Reads every Bengaluru store from the combined file, sets lat/lng for each,
searches "brik oven" on Zepto, and records availability + position per store.

Run from backend/:
    python -m scripts.zepto_brik_oven_coverage_bengaluru
Output: zepto_brik_oven_coverage_bengaluru_TIMESTAMP.xlsx
"""
import asyncio
import json
import sys
from pathlib import Path
from datetime import datetime
from urllib.parse import quote

sys.path.insert(0, str(Path(__file__).parent.parent))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from playwright.async_api import async_playwright

BRAND       = "Brik Oven"
KEYWORD     = "brik oven"
CITY_SHEET  = "Bangalore"
COMBINED    = Path(__file__).parent.parent / "zepto_all_stores_combined_20260731_1147.xlsx"
_BASE       = "https://www.zepto.com"
_SEARCH_PATH = "/user-search-service/api/v3/search"

_INIT_SCRIPT = """
(() => {
    const injectCoords = (url) => {
        if (typeof url !== 'string' || !url.includes('/lms/api/v2/get_page')) return url;
        const lat = localStorage.getItem('__probe_lat');
        const lng = localStorage.getItem('__probe_lng');
        if (!lat || !lng) return url;
        try {
            const u = new URL(url, location.href);
            u.searchParams.set('latitude', lat);
            u.searchParams.set('longitude', lng);
            return u.toString();
        } catch(e) { return url; }
    };
    const origFetch = window.fetch;
    window.fetch = function(url, opts) {
        return origFetch.apply(this, [injectCoords(url), opts]);
    };
    const origOpen = XMLHttpRequest.prototype.open;
    XMLHttpRequest.prototype.open = function(method, url, ...rest) {
        return origOpen.call(this, method, injectCoords(url), ...rest);
    };
})();
"""

HDR_FONT  = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
HDR_FILL  = PatternFill("solid", fgColor="2E4057")
HDR_ALIGN = Alignment(horizontal="center", vertical="center")
ALT_FILL  = PatternFill("solid", fgColor="F0F4F8")
YES_FILL  = PatternFill("solid", fgColor="D9F7E6")   # green — store carries Brik Oven
NO_FILL   = PatternFill("solid", fgColor="FDE8E8")   # red   — not found
TOT_FILL  = PatternFill("solid", fgColor="048A81")
TOT_FONT  = Font(name="Calibri", size=10, bold=True, color="FFFFFF")


def load_bengaluru_stores() -> list[dict]:
    wb = openpyxl.load_workbook(COMBINED, read_only=True, data_only=True)
    ws = wb[CITY_SHEET]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    header = [str(h).strip().lower() for h in rows[0]]
    stores = []
    seen_ids = set()
    for row in rows[1:]:
        rec = dict(zip(header, row))
        sid = str(rec.get("store_id", "")).strip()
        if not sid or sid in seen_ids:
            continue
        seen_ids.add(sid)
        try:
            lat = float(rec.get("lat") or 0)
            lng = float(rec.get("lng") or 0)
        except (TypeError, ValueError):
            continue
        if not lat or not lng:
            continue
        stores.append({
            "store_id":   sid,
            "store_name": str(rec.get("store_name", "")).strip(),
            "area":       str(rec.get("area", "")).strip(),
            "lat":        lat,
            "lng":        lng,
            "secondary":  str(rec.get("secondary_store_id", "") or "").strip(),
        })
    return stores


def extract_brik_oven(data: dict) -> list[dict]:
    """Return only Brik Oven products from a search response."""
    products = []
    for widget in data.get("layout", []):
        if widget.get("widgetId") != "PRODUCT_GRID":
            continue
        items = (
            widget.get("data", {})
            .get("resolver", {})
            .get("data", {})
            .get("items") or []
        )
        for item in items:
            pr = item.get("productResponse")
            if not pr:
                continue
            prod = pr.get("product", {}) or {}
            pv   = pr.get("productVariant", {}) or {}
            brand = prod.get("brand", "")
            if BRAND.lower() not in brand.lower():
                continue
            mrp = pv.get("mrp") or pr.get("mrp") or 0
            sp  = pr.get("discountedSellingPrice") or pr.get("sellingPrice") or 0
            products.append({
                "position":    (item.get("position") or 0) + 1,
                "product_name": prod.get("name", ""),
                "pack_size":   pv.get("formattedPacksize", ""),
                "mrp_rs":      round(mrp / 100, 2),
                "sp_rs":       round(sp / 100, 2),
                "discount_pct": pr.get("discountPercent", 0),
                "in_stock":    not pr.get("outOfStock", False),
                "category":    pr.get("primaryCategoryName", ""),
                "variant_id":  pv.get("id", ""),
                "product_id":  prod.get("id", ""),
            })
    return products


async def search_at_store(store: dict, page, target: dict) -> tuple[list[dict], set, int]:
    """Force this store via request headers, then search for Brik Oven.

    lat/lng alone does NOT switch the store for the search API — the store is
    carried in the store_id / storeid / store_ids request headers, which the
    route handler rewrites to this store's id.
    Returns (brik_oven_products, store_ids_seen_in_response, total_products).
    """
    found: list[dict] = []
    seen_store_ids: set = set()
    total = 0
    captured = False

    # Tell the route handler which store to force for the upcoming request
    target["id"] = store["store_id"]
    # Also set lat/lng — keeps the page state consistent with the store
    try:
        await page.evaluate(f"""() => {{
            localStorage.setItem('__probe_lat', '{store["lat"]}');
            localStorage.setItem('__probe_lng', '{store["lng"]}');
        }}""")
    except Exception:
        pass

    async def on_response(response):
        nonlocal captured, total
        if _SEARCH_PATH not in response.url or captured:
            return
        try:
            data = await response.json()
            if not any(w.get("widgetId") == "PRODUCT_GRID" for w in data.get("layout", [])):
                return
            captured = True
            for widget in data.get("layout", []):
                if widget.get("widgetId") != "PRODUCT_GRID":
                    continue
                items = (widget.get("data", {}).get("resolver", {})
                         .get("data", {}).get("items") or [])
                for it in items:
                    pr = it.get("productResponse") or {}
                    if pr.get("storeId"):
                        seen_store_ids.add(pr["storeId"])
                    total += 1
            found.extend(extract_brik_oven(data))
        except Exception:
            pass

    page.on("response", on_response)
    try:
        await page.goto(
            f"{_BASE}/search?query={quote(KEYWORD)}",
            timeout=30000, wait_until="domcontentloaded"
        )
        await page.wait_for_timeout(4000)
    except Exception as e:
        print(f"    [nav error] {e}")
    finally:
        page.remove_listener("response", on_response)

    return found, seen_store_ids, total


async def main():
    ts         = datetime.now().strftime("%Y%m%d_%H%M")
    scraped_at = datetime.now().strftime("%d-%b-%Y %H:%M")

    stores = load_bengaluru_stores()
    print(f"Brik Oven Coverage — Bengaluru — {ts}")
    print(f"Dark stores to check : {len(stores)}")
    print(f"Keyword              : \"{KEYWORD}\"")
    print("=" * 65)

    # Results: one row per store × product
    coverage_rows: list[dict] = []
    store_summary: list[dict] = []

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        ctx = await browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/125.0.0.0 Safari/537.36"
            )
        )
        page = await ctx.new_page()
        await page.add_init_script(_INIT_SCRIPT)

        # ── Force the target dark store on every search request ────────────
        # The search API takes the store from these headers, not from lat/lng.
        target = {"id": None}

        async def route_handler(route):
            req = route.request
            if _SEARCH_PATH in req.url and req.method == "POST" and target["id"]:
                h = dict(req.headers)
                sid = target["id"]
                h["store_id"]   = sid
                h["storeid"]    = sid
                h["store_ids"]  = sid
                h["store_etas"] = json.dumps({sid: -1})
                await route.continue_(headers=h)
            else:
                await route.continue_()

        await page.route("**/user-search-service/**", route_handler)

        print("Warming up browser...")
        await page.goto(_BASE, timeout=30000, wait_until="domcontentloaded")
        await page.wait_for_timeout(3000)
        print("Browser ready.\n")

        for i, store in enumerate(stores, 1):
            print(f"[{i:>3}/{len(stores)}] {store['store_name'][:33]:<33}", end="  ")
            products, seen_ids, total = await search_at_store(store, page, target)

            verified = store["store_id"] in seen_ids
            vflag = "" if verified else " [!store-mismatch]"

            if products:
                best_pos = min(p["position"] for p in products)
                names    = ", ".join(p["product_name"].replace("Brik Oven ", "") for p in products[:2])
                print(f"✓ {len(products):>2} products  best #{best_pos:<3} [{names[:40]}]{vflag}")
            else:
                print(f"✗ not found  (total results: {total}){vflag}")

            store_summary.append({
                "store_name":    store["store_name"],
                "store_id":      store["store_id"],
                "area":          store["area"],
                "lat":           store["lat"],
                "lng":           store["lng"],
                "brik_oven_found": "YES" if products else "NO",
                "product_count":  len(products),
                "best_rank":      min(p["position"] for p in products) if products else "",
                "total_results":  total,
                "store_verified": "YES" if verified else "NO",
                "products_list":  " | ".join(
                    f"#{p['position']} {p['product_name'].replace('Brik Oven ','')}"
                    for p in products
                ) if products else "",
            })

            for prod in products:
                coverage_rows.append({
                    "store_name":   store["store_name"],
                    "store_id":     store["store_id"],
                    "area":         store["area"],
                    "lat":          store["lat"],
                    "lng":          store["lng"],
                    **prod,
                })

            await asyncio.sleep(0.5)

        await browser.close()

    found_stores = [s for s in store_summary if s["brik_oven_found"] == "YES"]
    print(f"\n{'='*65}")
    print(f"Stores checked    : {len(stores)}")
    print(f"Stores with Brik Oven : {len(found_stores)}")
    print(f"Coverage          : {len(found_stores)/len(stores)*100:.1f}%")
    print(f"Total product rows: {len(coverage_rows)}")

    # ── Build Excel ────────────────────────────────────────────────────────
    wb   = openpyxl.Workbook()
    title_fill = PatternFill("solid", fgColor="1A3C5E")
    sub_fill   = ALT_FILL

    def banner(ws, t, s):
        ws.insert_rows(1, amount=2)
        mc = ws.max_column or 1
        ws["A1"] = t
        ws["A2"] = s
        ws["A1"].font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
        ws["A1"].fill = title_fill
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 22
        ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="444444")
        ws["A2"].fill = sub_fill
        ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[2].height = 15
        if mc > 1:
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=mc)
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=mc)

    def col_widths(ws):
        for col in ws.columns:
            cl = get_column_letter(col[0].column)
            ml = max((len(str(c.value)) for c in col if c.value), default=0)
            ws.column_dimensions[cl].width = min(max(ml + 2, 8), 50)

    banner_title = f"Brik Oven Coverage — Bengaluru — {len(found_stores)}/{len(stores)} stores"
    banner_sub   = f"Generated: {scraped_at}  |  Keyword: \"{KEYWORD}\"  |  Coverage: {len(found_stores)/len(stores)*100:.1f}%"

    # Sheet 1: Store Summary
    ws1 = wb.active
    ws1.title = "Store Coverage"
    sum_cols = ["store_name", "store_id", "area", "lat", "lng",
                "brik_oven_found", "product_count", "best_rank",
                "total_results", "store_verified", "products_list"]
    ws1.append(sum_cols)
    for cell in ws1[1]:
        cell.font  = HDR_FONT
        cell.fill  = HDR_FILL
        cell.alignment = HDR_ALIGN
    ws1.row_dimensions[1].height = 18
    ws1.freeze_panes = "A2"

    for r_idx, row in enumerate(
        sorted(store_summary, key=lambda x: (x["brik_oven_found"] == "NO", x["best_rank"] or 999)),
        start=2
    ):
        ws1.append([row.get(c, "") for c in sum_cols])
        fill = YES_FILL if row["brik_oven_found"] == "YES" else NO_FILL
        for c in range(1, len(sum_cols) + 1):
            ws1.cell(row=r_idx, column=c).fill = fill
    ws1.auto_filter.ref = ws1.dimensions
    col_widths(ws1)

    # Sheet 2: Product detail (one row per store × product)
    ws2 = wb.create_sheet("Product Detail")
    det_cols = ["store_name", "area", "position", "product_name", "pack_size",
                "mrp_rs", "sp_rs", "discount_pct", "in_stock", "category",
                "store_id", "product_id", "variant_id"]
    ws2.append(det_cols)
    for cell in ws2[1]:
        cell.font  = HDR_FONT
        cell.fill  = HDR_FILL
        cell.alignment = HDR_ALIGN
    ws2.row_dimensions[1].height = 18
    ws2.freeze_panes = "A2"

    for r_idx, row in enumerate(
        sorted(coverage_rows, key=lambda x: (x["store_name"], x["position"])),
        start=2
    ):
        ws2.append([row.get(c, "") for c in det_cols])
        if r_idx % 2 == 0:
            for c in range(1, len(det_cols) + 1):
                ws2.cell(row=r_idx, column=c).fill = ALT_FILL
    ws2.auto_filter.ref = ws2.dimensions
    col_widths(ws2)

    # Sheet 3: Quick summary stats
    ws3 = wb.create_sheet("Summary")
    ws3.append(["Metric", "Value"])
    for cell in ws3[1]:
        cell.font  = HDR_FONT
        cell.fill  = HDR_FILL
        cell.alignment = HDR_ALIGN
    stats = [
        ("Total Bengaluru dark stores",   len(stores)),
        ("Stores carrying Brik Oven",      len(found_stores)),
        ("Stores NOT carrying Brik Oven",  len(stores) - len(found_stores)),
        ("Coverage %",                     f"{len(found_stores)/len(stores)*100:.1f}%"),
        ("Total Brik Oven product-store rows", len(coverage_rows)),
        ("Unique products found",
            len({r["product_name"] for r in coverage_rows})),
        ("Avg search rank (where found)",
            f"{sum(r['position'] for r in coverage_rows)/len(coverage_rows):.1f}"
            if coverage_rows else "—"),
        ("Best (lowest) rank seen",
            min((r["position"] for r in coverage_rows), default="—")),
        ("Scraped at", scraped_at),
    ]
    for r_idx, (metric, val) in enumerate(stats, start=2):
        ws3.cell(row=r_idx, column=1, value=metric)
        ws3.cell(row=r_idx, column=2, value=val)
        if r_idx % 2 == 0:
            for c in range(1, 3):
                ws3.cell(row=r_idx, column=c).fill = ALT_FILL
    col_widths(ws3)

    for ws in [ws1, ws2, ws3]:
        banner(ws, banner_title, banner_sub)

    out = Path(__file__).parent.parent / f"zepto_brik_oven_coverage_bengaluru_{ts}.xlsx"
    wb.save(out)
    print(f"\nSaved → {out}")


if __name__ == "__main__":
    asyncio.run(main())
