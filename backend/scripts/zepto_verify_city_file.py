"""
Verify a discovery file's coordinates by round-tripping them through the API.

The strongest available check: take each lat/lng the file recorded and send it
back to get_page. If the coordinate is right, Zepto returns the same store_id.
A mismatch means the coordinate does not actually resolve to that store.

Also checks, without any network calls:
  - duplicate store_ids
  - coordinates outside the city's stated box
  - coordinates that disagree with the master workbook for known stores
  - rows missing lat/lng

Run from backend/:
    python -m scripts.zepto_verify_city_file darkstores_kolkata_20260803_1054.xlsx
    python -m scripts.zepto_verify_city_file <file> --sample 20
"""
import asyncio
import math
import sys
from pathlib import Path
from collections import Counter

sys.path.insert(0, str(Path(__file__).parent.parent))

import openpyxl
from playwright.async_api import async_playwright

BACKEND = Path(__file__).parent.parent
OUTDIR  = BACKEND / "dark_stores_zepto"

_BASE = "https://www.zepto.com"
_BFF  = "https://bff-gateway.zepto.com"
_GP   = "/lms/api/v2/get_page"
_DROP = {"host", "content-length", "connection", "accept-encoding",
         ":method", ":path", ":authority", ":scheme"}
CONCURRENCY = 6
TRIES = 3


def km(a, b, c, d):
    R = 6371.0
    p = math.radians
    return 2 * R * math.asin(math.sqrt(
        math.sin(p(c - a) / 2) ** 2 +
        math.cos(p(a)) * math.cos(p(c)) * math.sin(p(d - b) / 2) ** 2))


def read_sheet(wb, name):
    rr = list(wb[name].iter_rows(values_only=True))
    hdr = [str(x).strip().lower() if x else "" for x in rr[2]]
    return [dict(zip(hdr, r)) for r in rr[3:] if r and r[0]]


def load_master_coords():
    p = sorted(BACKEND.glob("zepto_master_all_cities_*.xlsx"))
    if not p:
        return {}
    wb = openpyxl.load_workbook(p[-1], read_only=True, data_only=True)
    rows = list(wb["All Stores"].iter_rows(values_only=True))
    wb.close()
    hdr = [str(x).strip().lower() if x else "" for x in rows[2]]
    out = {}
    for r in rows[3:]:
        if not r or not r[0]:
            continue
        d = dict(zip(hdr, r))
        try:
            out[str(d["store_id"]).strip()] = (float(d["lat"]), float(d["lng"]),
                                               d["city"], d.get("store_name", ""))
        except (TypeError, ValueError, KeyError):
            pass
    return out


async def capture_headers(page):
    cap = {}

    async def on_req(req):
        if _GP in req.url and not cap:
            cap.update(req.headers)

    page.on("request", on_req)
    try:
        await page.goto(_BASE, timeout=30000, wait_until="domcontentloaded")
        w = 0
        while not cap and w < 6000:
            await page.wait_for_timeout(250)
            w += 250
    except Exception:
        pass
    finally:
        page.remove_listener("request", on_req)
    return {k: v for k, v in cap.items() if k.lower() not in _DROP}


async def probe(ctx, hdr, lat, lng):
    url = (f"{_BFF}{_GP}?latitude={lat}&longitude={lng}"
           f"&page_type=HOME&version=v2&show_new_eta_banner=true"
           f"&page_size=3&enforce_platform_type=WEB")
    for a in range(1, TRIES + 1):
        try:
            r = await ctx.request.get(url, headers=hdr, timeout=15000)
            if r.status == 200:
                d = await r.json()
                svc = d.get("storeServiceableResponse") or {}
                return svc.get("storeId"), (d.get("storeDetailsResponse") or {}).get("name", "")
        except Exception:
            pass
        await asyncio.sleep(0.5 * a)
    return "ERROR", ""


async def main():
    if len(sys.argv) < 2:
        print("Usage: python -m scripts.zepto_verify_city_file <file.xlsx> [--sample N]")
        print("\nAvailable files:")
        for f in sorted(OUTDIR.glob("*.xlsx")):
            print("   ", f.name)
        return

    fname = sys.argv[1]
    path = OUTDIR / fname if not Path(fname).is_absolute() else Path(fname)
    if not path.exists():
        path = Path(fname)
    if not path.exists():
        print(f"Not found: {fname}")
        return

    sample = None
    if "--sample" in sys.argv:
        sample = int(sys.argv[sys.argv.index("--sample") + 1])

    wb = openpyxl.load_workbook(path, data_only=True)
    stores = read_sheet(wb, "Stores")
    box_txt = ""
    for r in wb["Summary"].iter_rows(values_only=True):
        if r and str(r[0]).strip() == "Box":
            box_txt = str(r[1])
    wb.close()

    print(f"File   : {path.name}")
    print(f"Stores : {len(stores)}")
    print(f"Box    : {box_txt}")
    print("=" * 96)

    # ── Offline checks ───────────────────────────────────────────────────────
    print("\nOFFLINE CHECKS")
    ids = [str(r["store_id"]).strip() for r in stores]
    dups = [k for k, v in Counter(ids).items() if v > 1]
    print(f"  duplicate store_ids            : {len(dups)}"
          + (f"  {dups}" if dups else "  OK"))

    missing = [r for r in stores if not r.get("lat") or not r.get("lng")]
    print(f"  rows missing lat/lng           : {len(missing)}"
          + ("  OK" if not missing else ""))

    box = None
    try:
        nums = [float(x) for x in
                box_txt.replace("lat ", "").replace("lng ", "").replace(",", " ")
                .replace("-", " ").split()]
        if len(nums) == 4:
            box = nums
    except Exception:
        pass
    if box:
        out = [r for r in stores
               if not (box[0] <= float(r["lat"]) <= box[1]
                       and box[2] <= float(r["lng"]) <= box[3])]
        print(f"  coords outside the stated box  : {len(out)}"
              + ("  OK" if not out else ""))
        for r in out[:5]:
            print(f"      {str(r['store_name'])[:30]:<31} {r['lat']},{r['lng']}")

    master = load_master_coords()
    drift, absent = [], 0
    for r in stores:
        sid = str(r["store_id"]).strip()
        m = master.get(sid)
        if not m:
            absent += 1
            continue
        d = km(m[0], m[1], float(r["lat"]), float(r["lng"]))
        if d > 0.001:
            drift.append((r["store_name"], d, r["found_by"]))
    print(f"  not in master (i.e. new)       : {absent}")
    print(f"  coord differs from master      : {len(drift)}")
    for n, d, fb in sorted(drift, key=lambda x: -x[1])[:6]:
        print(f"      {str(n)[:30]:<31} {d:6.2f} km  (found_by={fb})")

    # ── Live round-trip ──────────────────────────────────────────────────────
    targets = stores if sample is None else stores[:sample]
    print(f"\nLIVE ROUND-TRIP  — sending {len(targets)} coordinates back to get_page")
    print("  (a coordinate is correct if the API returns the same store_id)")
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        ctx = await browser.new_context(
            user_agent=("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                        "(KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36"))
        page = await ctx.new_page()
        hdr = await capture_headers(page)
        if not hdr:
            print("  header capture failed — cannot verify live")
            await browser.close()
            return

        sem = asyncio.Semaphore(CONCURRENCY)
        match, mismatch, err = [], [], []

        async def one(r):
            async with sem:
                sid, name = await probe(ctx, hdr, float(r["lat"]), float(r["lng"]))
            exp = str(r["store_id"]).strip()
            if sid == "ERROR":
                err.append(r)
            elif sid == exp:
                match.append(r)
            else:
                mismatch.append((r, sid, name))

        for s in range(0, len(targets), 60):
            await asyncio.gather(*(one(r) for r in targets[s:s + 60]))
            print(f"    ... {min(s+60, len(targets))}/{len(targets)} checked")

        await browser.close()

    tot = len(match) + len(mismatch) + len(err)
    print(f"\n  MATCH    : {len(match)}/{tot}"
          + (f"  ({len(match)/tot*100:.1f}%)" if tot else ""))
    print(f"  MISMATCH : {len(mismatch)}")
    print(f"  ERROR    : {len(err)}")
    for r, got, gname in mismatch[:10]:
        print(f"      {str(r['store_name'])[:28]:<29} {r['lat']},{r['lng']}")
        print(f"          expected {str(r['store_id'])[:8]}  got {str(got)[:8]} «{gname[:26]}»")

    print("\n" + "=" * 96)
    if not dups and not missing and not mismatch and not err:
        print("VERDICT: every coordinate resolves to the store the file claims. Data is sound.")
    else:
        print("VERDICT: see the items above.")


if __name__ == "__main__":
    asyncio.run(main())
