"""
Scrape Zepto dark stores — pincode-only mode.
For each of the 180 NCR pincodes, takes ONLY the postal-code centroid suggestion
(location_type = APPROXIMATE). One row per pincode. Max 180 rows.

Run from backend/:
    python -m scripts.zepto_pincode_only
"""
import asyncio
import sys
from pathlib import Path
from collections import Counter
from datetime import datetime

sys.path.insert(0, str(Path(__file__).parent.parent))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from playwright.async_api import async_playwright
from scraper.platforms.zepto.dark_store.endpoints import DELHI_NCR_PINCODES

_BFF = "https://bff-gateway.zepto.com"

_INIT_SCRIPT = """
(() => {
    const injectCoords = (url) => {
        if (typeof url !== 'string' || !url.includes('/lms/api/v2/get_page')) return url;
        const lat = localStorage.getItem('__probe_lat');
        const lng = localStorage.getItem('__probe_lng');
        if (!lat || !lng) return url;
        try {
            const u = new URL(url, location.href);
            u.searchParams.set('latitude', lat);
            u.searchParams.set('longitude', lng);
            return u.toString();
        } catch(e) { return url; }
    };
    const origFetch = window.fetch;
    window.fetch = function(url, opts) {
        return origFetch.apply(this, [injectCoords(url), opts]);
    };
    const origOpen = XMLHttpRequest.prototype.open;
    XMLHttpRequest.prototype.open = function(method, url, ...rest) {
        return origOpen.call(this, method, injectCoords(url), ...rest);
    };
})();
"""

_HEADERS = {
    "Accept": "application/json",
    "Origin": "https://www.zepto.com",
    "Referer": "https://www.zepto.com/",
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "app_sub_platform": "WEB", "app_version": "16.20.0",
    "platform": "WEB", "tenant": "ZEPTO",
}


async def get_predictions(query: str, ctx) -> list[dict]:
    for attempt in range(2):
        try:
            r = await ctx.request.get(
                f"{_BFF}/api/v1/maps/place/autocomplete/",
                params={"place_name": query}, headers=_HEADERS, timeout=10000,
            )
            if r.status != 200:
                return []
            data = await r.json()
            return data.get("data", {}).get("predictions") or data.get("predictions") or []
        except Exception as e:
            if attempt == 0:
                await asyncio.sleep(1)
            else:
                print(f"    [autocomplete error] {query}: {e}")
    return []


async def get_details(place_id: str, ctx) -> dict | None:
    for attempt in range(2):
        try:
            r = await ctx.request.get(
                f"{_BFF}/api/v1/maps/place/details/",
                params={"place_id": place_id}, headers=_HEADERS, timeout=10000,
            )
            if r.status != 200:
                return None
            data = await r.json()
            result = data.get("data", {}).get("result") or data.get("result") or {}
            geo = result.get("geometry", {})
            loc = geo.get("location", {})
            lat, lng = loc.get("lat"), loc.get("lng")
            if not lat or not lng:
                return None
            pincode = city = state = None
            for comp in result.get("address_components", []):
                t = comp.get("types", [])
                if "postal_code" in t:
                    pincode = comp.get("long_name")
                if ("locality" in t or "administrative_area_level_2" in t) and not city:
                    city = comp.get("long_name")
                if "administrative_area_level_1" in t and not state:
                    state = comp.get("long_name")
            return {
                "lat": lat, "lng": lng,
                "location_type": geo.get("location_type", ""),
                "pincode": pincode or "",
                "city": city or "",
                "state": state or "",
            }
        except Exception as e:
            if attempt == 0:
                await asyncio.sleep(1)
            else:
                print(f"    [details error] {place_id}: {e}")
    return None


async def probe_store_at(lat: float, lng: float, page) -> dict:
    result = {}
    done = asyncio.Event()

    async def on_response(response):
        if "/lms/api/v2/get_page" in response.url and not done.is_set():
            try:
                data = await response.json()
                svc = data.get("storeServiceableResponse", {})
                sid = svc.get("storeId")
                if sid:
                    result["store_id"] = str(sid)
                    result["secondary_ids"] = [str(s) for s in svc.get("secondaryStoreIds", [])]
                    result["store_name"] = data.get("storeDetailsResponse", {}).get("name", "")
                done.set()
            except Exception:
                pass

    try:
        await page.evaluate(f"""() => {{
            localStorage.setItem('__probe_lat', '{lat}');
            localStorage.setItem('__probe_lng', '{lng}');
        }}""")
    except Exception:
        pass

    page.on("response", on_response)
    try:
        await page.reload(timeout=20000, wait_until="domcontentloaded")
        try:
            await asyncio.wait_for(done.wait(), timeout=8)
        except asyncio.TimeoutError:
            pass
    except Exception:
        pass
    finally:
        page.remove_listener("response", on_response)
    return result


def apply_heading(ws, title: str, generated: str):
    ws.insert_rows(1, amount=2)
    max_col = ws.max_column or 1
    ws["A1"] = title
    ws["A2"] = f"Generated: {generated}  |  Source: zepto.com"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="2E4057")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="444444")
    ws["A2"].fill = PatternFill("solid", fgColor="F0F4F8")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 16
    for cell in ws[3]:
        if cell.value:
            cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="048A81")
            cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 18
    if max_col > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
    for col_idx in range(1, max_col + 1):
        max_len = 0
        for row_idx in range(3, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 40)


async def main():
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    print(f"Zepto NCR Dark Store Scraper — PINCODE ONLY MODE — {ts}")
    print(f"Pincodes: {len(DELHI_NCR_PINCODES)} | one row per pincode | APPROXIMATE centroid only")
    print("=" * 60)

    rows: list[dict] = []
    skipped: list[str] = []

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        ctx = await browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
        )
        page = await ctx.new_page()
        await page.add_init_script(_INIT_SCRIPT)

        print("Warming up browser (WAF)...")
        try:
            await page.goto("https://www.zepto.com", timeout=30000, wait_until="domcontentloaded")
            await page.wait_for_timeout(3000)
            print("Browser ready.\n")
        except Exception as e:
            print(f"Warmup: {e}\n")

        total = len(DELHI_NCR_PINCODES)
        for i, (pincode, label) in enumerate(DELHI_NCR_PINCODES.items(), 1):
            print(f"\n[{i}/{total}] {pincode} — {label}")

            preds = await get_predictions(pincode, ctx)
            if not preds:
                print(f"  no predictions")
                skipped.append(pincode)
                await asyncio.sleep(0.4)
                continue

            # Find the APPROXIMATE (postal code centroid) suggestion
            chosen_pred = None
            chosen_details = None
            fallback_pred = None
            fallback_details = None

            for pred in preds:
                place_id = pred.get("place_id")
                if not place_id:
                    continue
                details = await get_details(place_id, ctx)
                if not details:
                    await asyncio.sleep(0.15)
                    continue

                lat, lng = details["lat"], details["lng"]
                if not (27.8 <= lat <= 29.2 and 76.5 <= lng <= 78.0):
                    await asyncio.sleep(0.15)
                    continue

                if details["location_type"] == "APPROXIMATE":
                    chosen_pred = pred
                    chosen_details = details
                    break  # found pincode centroid — stop here

                # Keep first valid non-APPROXIMATE as fallback
                if fallback_pred is None:
                    fallback_pred = pred
                    fallback_details = details

                await asyncio.sleep(0.15)

            # Use APPROXIMATE if found, else fallback to first valid suggestion
            final_pred = chosen_pred or fallback_pred
            final_details = chosen_details or fallback_details

            if not final_pred or not final_details:
                print(f"  no valid suggestion found")
                skipped.append(pincode)
                await asyncio.sleep(0.3)
                continue

            lat, lng = final_details["lat"], final_details["lng"]
            loc_type = final_details["location_type"]
            name = final_pred.get("description", "")
            print(f"  ✓ {name[:60]} [{loc_type}] ({lat:.5f},{lng:.5f})")

            r = await probe_store_at(lat, lng, page)
            sid = r.get("store_id")

            rows.append({
                "pincode": pincode,
                "area_label": label,
                "suggestion": name,
                "store_id": sid or "",
                "store_name": r.get("store_name", ""),
                "city": final_details["city"],
                "state": final_details["state"],
                "lat": lat,
                "lng": lng,
                "location_type": loc_type,
                "secondary_store_id": ",".join(r.get("secondary_ids", [])) if sid else "",
            })

            if sid:
                print(f"    → {sid} «{r.get('store_name', '')}»")
            else:
                print(f"    no store found")

            await asyncio.sleep(0.5)

        try:
            await browser.close()
        except Exception:
            pass

    print(f"\n{'='*60}")
    print(f"Total pincodes: {total} | Found: {len(rows)} | Skipped: {len(skipped)}")
    if skipped:
        print(f"Skipped pincodes: {', '.join(skipped)}")

    out = Path(__file__).parent.parent / f"zepto_pincode_stores_{ts}.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "pincode_stores"

    hdr = ["pincode", "area_label", "suggestion", "store_id", "store_name",
           "city", "state", "lat", "lng", "location_type", "secondary_store_id"]
    ws.append(hdr)
    for row in sorted(rows, key=lambda x: x["pincode"]):
        ws.append([row[h] for h in hdr])

    ws2 = wb.create_sheet("by_city")
    ws2.append(["city", "pincode_count", "unique_stores"])
    city_pincodes: dict[str, int] = Counter(r["city"] for r in rows)
    city_stores: dict[str, set] = {}
    for r in rows:
        city_stores.setdefault(r["city"], set()).add(r["store_id"])
    for city in sorted(city_pincodes):
        ws2.append([city, city_pincodes[city], len(city_stores.get(city, set()))])
        print(f"  {city}: {city_pincodes[city]} pincodes, {len(city_stores.get(city, set()))} stores")

    generated = datetime.now().strftime("%d-%b-%Y  %H:%M")
    apply_heading(ws, "Zepto Dark Stores NCR — Pincode Only (One Row Per Pincode)", generated)
    apply_heading(ws2, "Zepto NCR — Pincode Count by City", generated)

    wb.save(out)
    print(f"\nSaved → {out}")


if __name__ == "__main__":
    asyncio.run(main())
