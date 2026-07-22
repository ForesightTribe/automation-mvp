"""
City-wide position verification script for Bengaluru.

Scrapes Blinkit from hardcoded lat/lon points across Bengaluru
and records Dobra's Ad + Organic position at each locality.
Saves results to Excel to verify position is consistent city-wide.

Usage:
    cd backend
    python -m scripts.city_position_check --keyword "cotton candy" --brand dobra
"""
import argparse
import asyncio
import sys
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

BENGALURU_ZONES = [
    # Central / South
    {"locality": "MG Road / GPO",           "pincode": "560001", "lat": 12.9757, "lon": 77.6011},
    {"locality": "Shivajinagar",            "pincode": "560001", "lat": 12.9819, "lon": 77.5987},
    {"locality": "Bangalore City",          "pincode": "560002", "lat": 12.9784, "lon": 77.5997},
    {"locality": "Malleshwaram",            "pincode": "560003", "lat": 13.0031, "lon": 77.5649},
    {"locality": "Basavanagudi",            "pincode": "560004", "lat": 12.9432, "lon": 77.5758},
    {"locality": "Rajajinagar",             "pincode": "560010", "lat": 12.9933, "lon": 77.5546},
    {"locality": "Jayanagar (560011)",      "pincode": "560011", "lat": 12.9299, "lon": 77.5827},
    {"locality": "Yeshwanthpur",            "pincode": "560022", "lat": 13.0198, "lon": 77.5400},
    {"locality": "Hebbal",                  "pincode": "560024", "lat": 13.0450, "lon": 77.5930},
    {"locality": "Sarjapur Road",           "pincode": "560035", "lat": 12.9102, "lon": 77.6784},
    {"locality": "Marathahalli",            "pincode": "560037", "lat": 12.9591, "lon": 77.7007},
    {"locality": "Brookefield",             "pincode": "560037", "lat": 12.9517, "lon": 77.7134},
    {"locality": "Indiranagar",             "pincode": "560038", "lat": 12.9784, "lon": 77.6408},
    {"locality": "Vijayanagar",             "pincode": "560040", "lat": 12.9716, "lon": 77.5430},
    {"locality": "Jayanagar (560041)",      "pincode": "560041", "lat": 12.9250, "lon": 77.5933},
    {"locality": "Nagawara",                "pincode": "560045", "lat": 13.0433, "lon": 77.6163},
    {"locality": "Manyata Tech Park",       "pincode": "560045", "lat": 13.0454, "lon": 77.6211},
    {"locality": "Banashankari (560050)",   "pincode": "560050", "lat": 12.9261, "lon": 77.5470},
    {"locality": "Vasanth Nagar",           "pincode": "560052", "lat": 12.9889, "lon": 77.5832},
    {"locality": "Peenya",                  "pincode": "560058", "lat": 13.0290, "lon": 77.5165},
    {"locality": "Kengeri",                 "pincode": "560060", "lat": 12.9124, "lon": 77.4820},
    {"locality": "Yelahanka (560063)",      "pincode": "560063", "lat": 13.0940, "lon": 77.5951},
    {"locality": "Yelahanka (560064)",      "pincode": "560064", "lat": 13.1007, "lon": 77.5963},
    {"locality": "Jakkur",                  "pincode": "560064", "lat": 13.0810, "lon": 77.6027},
    {"locality": "Whitefield (560066)",     "pincode": "560066", "lat": 12.9698, "lon": 77.7500},
    {"locality": "Whitefield (560067)",     "pincode": "560067", "lat": 12.9766, "lon": 77.7479},
    {"locality": "Bommanahalli",            "pincode": "560068", "lat": 12.8987, "lon": 77.6355},
    {"locality": "Banashankari (560070)",   "pincode": "560070", "lat": 12.9020, "lon": 77.5607},
    {"locality": "Domlur",                  "pincode": "560071", "lat": 12.9616, "lon": 77.6391},
    {"locality": "Nagarbhavi",              "pincode": "560072", "lat": 12.9573, "lon": 77.5187},
    {"locality": "BTM Layout",              "pincode": "560076", "lat": 12.9166, "lon": 77.6101},
    {"locality": "JP Nagar",                "pincode": "560078", "lat": 12.9072, "lon": 77.5936},
    {"locality": "Banashankari (560085)",   "pincode": "560085", "lat": 12.8848, "lon": 77.5500},
    {"locality": "Koramangala (560034)",    "pincode": "560034", "lat": 12.9352, "lon": 77.6245},
    {"locality": "Koramangala (560095)",    "pincode": "560095", "lat": 12.9271, "lon": 77.6269},
    {"locality": "CV Raman Nagar",          "pincode": "560093", "lat": 12.9849, "lon": 77.6609},
    {"locality": "RR Nagar",                "pincode": "560098", "lat": 12.9261, "lon": 77.5218},
    {"locality": "Electronic City",         "pincode": "560100", "lat": 12.8399, "lon": 77.6770},
    {"locality": "HSR Layout",              "pincode": "560102", "lat": 12.9116, "lon": 77.6389},
    {"locality": "Bellandur",               "pincode": "560103", "lat": 12.9257, "lon": 77.6761},
    # Outskirts
    {"locality": "Anekal",                  "pincode": "562106", "lat": 12.7124, "lon": 77.6970},
    {"locality": "Attibele",                "pincode": "562107", "lat": 12.7741, "lon": 77.7575},
    {"locality": "Devanahalli",             "pincode": "562110", "lat": 13.2468, "lon": 77.7167},
    {"locality": "Bagalur",                 "pincode": "562149", "lat": 13.1641, "lon": 77.7872},
]


async def check_position(keyword: str, lat: float, lon: float, brand: str) -> dict:
    from ad_campaigns.live_position import get_live_positions

    try:
        results = await get_live_positions(keyword, lat=lat, lon=lon)
    except Exception as e:
        print(f"    ERROR: {e}", flush=True)
        return {"sponsored_position": None, "organic_position": None, "total_results": 0, "error": str(e)}

    sponsored_pos = None
    organic_pos = None
    brand_lower = brand.lower()

    for item in results:
        if brand_lower in item["name"].lower():
            if item.get("is_ad") and sponsored_pos is None:
                sponsored_pos = item["position"]
            elif not item.get("is_ad") and organic_pos is None:
                organic_pos = item["position"]

    return {
        "sponsored_position": sponsored_pos,
        "organic_position": organic_pos,
        "total_results": len(results),
        "error": None,
    }


async def get_zones_from_db(city_filter: str) -> list[dict]:
    """Pull dark store coordinates from marketplace_locations table."""
    try:
        from app.core.database import AsyncSessionLocal
        from app.models.search import MarketplaceLocation
        from sqlmodel import select
        async with AsyncSessionLocal() as session:
            q = (
                select(MarketplaceLocation)
                .where(
                    MarketplaceLocation.mp_slug == "blinkit",
                    MarketplaceLocation.is_active == True,
                    MarketplaceLocation.lat.is_not(None),
                    MarketplaceLocation.lon.is_not(None),
                    MarketplaceLocation.city.ilike(f"%{city_filter}%"),
                )
                .order_by(MarketplaceLocation.city, MarketplaceLocation.location_name)
            )
            rows = (await session.execute(q)).scalars().all()

        seen: set[str] = set()
        zones = []
        for r in rows:
            area = (r.location_name.strip() if r.location_name and r.location_name.strip()
                    else (r.pincode.strip() if r.pincode and r.pincode.strip() else ""))
            key = f"{r.city}|{area}"
            if key in seen:
                continue
            seen.add(key)
            zones.append({
                "locality": area or r.merchant_id,
                "pincode":  r.pincode or "",
                "lat":      r.lat,
                "lon":      r.lon,
            })
        return zones
    except Exception as e:
        print(f"[DB] Could not load zones: {e} — falling back to hardcoded list", flush=True)
        return []


async def run(keyword: str, brand: str, output_dir: Path, city: str = "bengaluru"):
    zones = await get_zones_from_db(city)
    if not zones:
        print("[DB] No zones found — using hardcoded BENGALURU_ZONES", flush=True)
        zones = BENGALURU_ZONES

    print(f"\n{city.title()} City-Wide Position Check", flush=True)
    print(f"  Keyword : {keyword}", flush=True)
    print(f"  Brand   : {brand}", flush=True)
    print(f"  Zones   : {len(zones)}\n", flush=True)

    rows = []
    for i, z in enumerate(zones, 1):
        print(f"[{i}/{len(zones)}] {z['locality']} ({z['pincode']}) lat={z['lat']} lon={z['lon']}", flush=True)
        result = await check_position(keyword, z["lat"], z["lon"], brand)
        rows.append({
            "Locality":          z["locality"],
            "Pincode":           z["pincode"],
            "Lat":               z["lat"],
            "Lon":               z["lon"],
            "Ad Position":       result["sponsored_position"] if result["sponsored_position"] else "Not found",
            "Organic Position":  result["organic_position"] if result["organic_position"] else "Not found",
            "Total Results":     result["total_results"],
            "Error":             result["error"] or "",
        })
        print(
            f"  → Ad pos: {result['sponsored_position'] or 'Not found'} | "
            f"Organic pos: {result['organic_position'] or 'Not found'} | "
            f"Total results: {result['total_results']}",
            flush=True,
        )
        print("", flush=True)

    # ── Excel output ──────────────────────────────────────────────────────────
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()

        # ── Sheet 1: Raw Data ─────────────────────────────────────────────────
        ws = wb.active
        ws.title = "Zone-wise Results"

        headers = list(rows[0].keys())
        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(bold=True, color="FFFFFF")

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        green_font = Font(bold=True, color="00B050")
        red_font   = Font(bold=True, color="FF0000")

        for r, row in enumerate(rows, 2):
            for col, key in enumerate(headers, 1):
                cell = ws.cell(row=r, column=col, value=row[key])
                cell.alignment = Alignment(horizontal="center")
                if key == "Ad Position":
                    if isinstance(row[key], int):
                        cell.font = green_font
                    else:
                        cell.font = red_font
                if key == "Organic Position" and isinstance(row[key], int):
                    cell.font = Font(color="0070C0")

        for col in range(1, len(headers) + 1):
            max_len = max(
                len(str(ws.cell(row=r, column=col).value or ""))
                for r in range(1, len(rows) + 2)
            )
            ws.column_dimensions[get_column_letter(col)].width = min(max_len + 4, 35)

        # ── Sheet 2: Summary ──────────────────────────────────────────────────
        ws2 = wb.create_sheet("Summary")

        ad_positions      = [r["Ad Position"] for r in rows if isinstance(r["Ad Position"], int)]
        organic_positions = [r["Organic Position"] for r in rows if isinstance(r["Organic Position"], int)]
        unique_ad_pos     = sorted(set(ad_positions))
        consistent        = len(unique_ad_pos) == 1

        summary_data = [
            ("City",                    city.title()),
            ("Keyword",                 keyword),
            ("Brand",                   brand),
            ("Total Zones Checked",     len(zones)),
            ("Zones with Ad Found",     len(ad_positions)),
            ("Zones with Organic Found",len(organic_positions)),
            ("Ad Positions Found",      ", ".join(str(p) for p in unique_ad_pos) if unique_ad_pos else "None"),
            ("Organic Positions Found", ", ".join(str(p) for p in sorted(set(organic_positions))) if organic_positions else "None"),
            ("All Same Ad Position?",   "YES ✓ — City-wide auction confirmed" if consistent else f"NO — varies: {unique_ad_pos}"),
            ("Run At",                  datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ]

        ws2["A1"] = f"{city.title()} City-Wide Position Check — Summary"
        ws2["A1"].font = Font(bold=True, size=14, color="1F4E79")
        ws2.merge_cells("A1:B1")

        for i, (label, value) in enumerate(summary_data, 3):
            ws2.cell(row=i, column=1, value=label).font = Font(bold=True)
            cell = ws2.cell(row=i, column=2, value=value)
            if label == "All Same Ad Position?":
                cell.font = Font(bold=True, color="00B050" if consistent else "FF0000")

        ws2.column_dimensions["A"].width = 30
        ws2.column_dimensions["B"].width = 45

        # ── Save ──────────────────────────────────────────────────────────────
        safe_keyword = keyword.replace(" ", "_").replace("/", "-")
        filename = f"{city.lower()}_position_check_{safe_keyword}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        out_path = output_dir / filename
        wb.save(out_path)
        print(f"Saved to: {out_path}", flush=True)

    except ImportError:
        import csv
        safe_keyword = keyword.replace(" ", "_")
        filename = f"bengaluru_position_check_{safe_keyword}.csv"
        out_path = output_dir / filename
        with open(out_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
            writer.writeheader()
            writer.writerows(rows)
        print(f"openpyxl not installed — saved as CSV: {out_path}", flush=True)
        print("Run: pip install openpyxl   to get Excel output", flush=True)

    # ── Console summary ───────────────────────────────────────────────────────
    ad_positions  = [r["Ad Position"] for r in rows if isinstance(r["Ad Position"], int)]
    unique_ad_pos = sorted(set(ad_positions))
    print(f"\n{'='*55}", flush=True)
    print(f"SUMMARY — {city.title()} / '{keyword}'", flush=True)
    print(f"  Zones checked     : {len(zones)}", flush=True)
    print(f"  Zones with Ad     : {len(ad_positions)}", flush=True)
    if ad_positions:
        print(f"  Ad positions      : {unique_ad_pos}", flush=True)
        if len(unique_ad_pos) == 1:
            print(f"  Result            : CONSISTENT — #{unique_ad_pos[0]} across all zones ✓", flush=True)
        else:
            print(f"  Result            : VARIES — {unique_ad_pos}", flush=True)
    print(f"{'='*55}\n", flush=True)


def main():
    parser = argparse.ArgumentParser(description="City-wide Blinkit position checker")
    parser.add_argument("--keyword", required=True, help="Search keyword e.g. 'cotton candy'")
    parser.add_argument("--brand",   default="dobra", help="Brand name (default: dobra)")
    parser.add_argument("--city",    default="bengaluru", help="City to check (default: bengaluru)")
    parser.add_argument("--output",  default=".", help="Output directory (default: current)")
    args = parser.parse_args()

    if hasattr(asyncio, "WindowsProactorEventLoopPolicy"):
        asyncio.set_event_loop_policy(asyncio.WindowsProactorEventLoopPolicy())

    asyncio.run(run(args.keyword, args.brand, Path(args.output), args.city))


if __name__ == "__main__":
    main()
