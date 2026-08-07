"""
Tighten the coordinates of grid-discovered Bengaluru stores.

Problem: a grid-discovered store's only coordinate is the lattice node where the
scan happened to hit it. That node sits somewhere inside the store's ~3 km
catchment, so it can be a couple of km off — and any pincode reverse-geocoded
from it inherits that error.

Fix: for each such store, probe a dense local grid (~330 m) around the known node
and collect EVERY point that store serves. The centroid of that service area is a
much better estimate of where the store actually is than any single node. Then
reverse-geocode the centroid instead.

Only stores whose coord_source is `grid_node` are touched; `place_details` rows
already carry Zepto's own coordinate and are left alone.

Run from backend/:
    python -m scripts.zepto_refine_new_store_coords
Output: zepto_bengaluru_stores_refined_TIMESTAMP.xlsx
"""
import asyncio
import math
import sys
import time
from pathlib import Path
from datetime import datetime

sys.path.insert(0, str(Path(__file__).parent.parent))

import httpx
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from playwright.async_api import async_playwright

BACKEND = Path(__file__).parent.parent

# Local probe box around each grid node, and its resolution.
HALF_BOX = 0.027      # ~3 km each way — a store's catchment fits inside this
LOCAL_STEP = 0.003    # ~330 m
CONCURRENCY = 8
PROBE_TRIES = 2

_BASE = "https://www.zepto.com"
_BFF  = "https://bff-gateway.zepto.com"
_GP   = "/lms/api/v2/get_page"

NOMINATIM = "https://nominatim.openstreetmap.org/reverse"
UA = "zepto-darkstore-research/1.0 (contact: soumi97bag@gmail.com)"
RATE_LIMIT_S = 1.1

_DROP = {"host", "content-length", "connection", "accept-encoding",
         ":method", ":path", ":authority", ":scheme"}

HDR_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
HDR_FILL = PatternFill("solid", fgColor="2E4057")
HDR_ALIGN = Alignment(horizontal="center", vertical="center")
ALT_FILL = PatternFill("solid", fgColor="F0F4F8")
NEW_FILL = PatternFill("solid", fgColor="D9F7E6")
TITLE_FILL = PatternFill("solid", fgColor="1A3C5E")


def km(a, b, c, d):
    R = 6371.0
    p = math.radians
    return 2 * R * math.asin(math.sqrt(
        math.sin(p(c - a) / 2) ** 2 +
        math.cos(p(a)) * math.cos(p(c)) * math.sin(p(d - b) / 2) ** 2))


def latest_enriched() -> Path:
    files = sorted(BACKEND.glob("zepto_bengaluru_stores_with_pincode_*.xlsx"))
    if not files:
        raise SystemExit("Run zepto_enrich_bengaluru_pincodes first.")
    return files[-1]


def read_rows(path: Path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rr = list(wb["Bengaluru Stores"].iter_rows(values_only=True))
    wb.close()
    hdr = [str(h).strip().lower() if h else "" for h in rr[2]]
    return hdr, [dict(zip(hdr, r)) for r in rr[3:] if r and r[0]]


async def capture_headers(page) -> dict:
    cap = {}

    async def on_req(req):
        if _GP in req.url and not cap:
            cap.update(req.headers)

    page.on("request", on_req)
    try:
        await page.goto(_BASE, timeout=30000, wait_until="domcontentloaded")
        await page.wait_for_timeout(4000)
    except Exception:
        pass
    finally:
        page.remove_listener("request", on_req)
    return {k: v for k, v in cap.items() if k.lower() not in _DROP}


def gp_url(lat, lng):
    return (f"{_BFF}{_GP}?latitude={lat}&longitude={lng}"
            f"&page_type=HOME&version=v2&show_new_eta_banner=true"
            f"&page_size=3&enforce_platform_type=WEB")


async def probe(ctx, lat, lng, state):
    for attempt in range(1, PROBE_TRIES + 1):
        try:
            r = await ctx.request.get(gp_url(lat, lng), headers=state["headers"], timeout=15000)
            if r.status == 200:
                d = await r.json()
                return ((d.get("storeServiceableResponse") or {}).get("storeId")) or None
            if r.status in (401, 403, 429):
                state["stale"] = True
                await asyncio.sleep(1.5 * attempt)
                continue
        except Exception:
            pass
        await asyncio.sleep(0.5 * attempt)
    return None


def reverse_geocode(client, lat, lng):
    for attempt in range(1, 4):
        try:
            r = client.get(NOMINATIM,
                           params={"lat": lat, "lon": lng, "format": "jsonv2",
                                   "addressdetails": 1, "zoom": 18},
                           headers={"User-Agent": UA, "Accept-Language": "en"},
                           timeout=20)
            if r.status_code == 200:
                a = (r.json() or {}).get("address") or {}
                area = next((a[k] for k in ("neighbourhood", "suburb", "village", "town",
                                            "city_district", "residential", "quarter", "hamlet")
                             if a.get(k)), "")
                return {"pincode": (a.get("postcode") or "").replace(" ", ""),
                        "area": area,
                        "city": a.get("city") or a.get("town") or a.get("state_district") or "",
                        "state": a.get("state") or ""}
            if r.status_code in (429, 503):
                time.sleep(2.0 * attempt)
                continue
        except Exception:
            pass
        time.sleep(1.0 * attempt)
    return None


def widths(ws):
    for col in ws.columns:
        cl = get_column_letter(col[0].column)
        ml = max((len(str(c.value)) for c in col if c.value), default=0)
        ws.column_dimensions[cl].width = min(max(ml + 2, 9), 52)


async def main():
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    scraped_at = datetime.now().strftime("%d-%b-%Y %H:%M")

    src = latest_enriched()
    hdr, rows = read_rows(src)
    targets = [r for r in rows if str(r.get("coord_source")) == "grid_node"]

    span = int(HALF_BOX / LOCAL_STEP) * 2 + 1
    print("Refine grid-discovered store coordinates")
    print(f"Source     : {src.name}")
    print(f"Total rows : {len(rows)}   to refine: {len(targets)}")
    print(f"Local box  : +/-{HALF_BOX} deg (~{HALF_BOX*111:.1f} km) @ {LOCAL_STEP} deg "
          f"(~{LOCAL_STEP*111:.0f} m)  ->  {span}x{span} = {span*span} points each")
    print(f"Total probes: ~{len(targets)*span*span}")
    print("=" * 78)

    refined: dict[str, dict] = {}

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        ctx = await browser.new_context(
            user_agent=("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                        "(KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36"))
        page = await ctx.new_page()
        state = {"headers": {}, "stale": False}
        print("Capturing session headers...")
        state["headers"] = await capture_headers(page)
        if not state["headers"]:
            print("Header capture failed - aborting.")
            await browser.close()
            return
        print(f"Captured {len(state['headers'])} headers.\n")

        sem = asyncio.Semaphore(CONCURRENCY)

        for i, row in enumerate(targets, 1):
            sid = str(row["store_id"]).strip()
            clat, clng = float(row["lat"]), float(row["lng"])
            pts = []
            la = clat - HALF_BOX
            while la <= clat + HALF_BOX + 1e-9:
                ln = clng - HALF_BOX
                while ln <= clng + HALF_BOX + 1e-9:
                    pts.append((round(la, 5), round(ln, 5)))
                    ln += LOCAL_STEP
                la += LOCAL_STEP

            hits = []

            async def one(lat, lng):
                async with sem:
                    got = await probe(ctx, lat, lng, state)
                if got == sid:
                    hits.append((lat, lng))

            await asyncio.gather(*(one(a, b) for a, b in pts))

            if hits:
                nlat = sum(h[0] for h in hits) / len(hits)
                nlng = sum(h[1] for h in hits) / len(hits)
                moved = km(clat, clng, nlat, nlng)
                refined[sid] = {"lat": round(nlat, 6), "lng": round(nlng, 6),
                                "serves": len(hits), "moved_km": round(moved, 2)}
                print(f"[{i:>2}/{len(targets)}] {str(row['store_name'])[:28]:<28} "
                      f"serves {len(hits):>3} pts  centroid moved {moved:.2f} km")
            else:
                print(f"[{i:>2}/{len(targets)}] {str(row['store_name'])[:28]:<28} "
                      f"no serving points found - keeping grid node")

            if state["stale"]:
                nh = await capture_headers(page)
                if nh:
                    state["headers"] = nh
                state["stale"] = False

        await browser.close()

    # Re-geocode the improved centroids
    print(f"\nRe-geocoding {len(refined)} refined coordinates...")
    with httpx.Client() as client:
        for sid, info in refined.items():
            g = reverse_geocode(client, info["lat"], info["lng"])
            if g and g["pincode"]:
                info.update(pincode=g["pincode"], area=g["area"],
                            city=g["city"], state=g["state"])
            time.sleep(RATE_LIMIT_S)

    changed_pin = 0
    for r in rows:
        sid = str(r["store_id"]).strip()
        if sid in refined:
            info = refined[sid]
            r["grid_lat"], r["grid_lng"] = r["lat"], r["lng"]
            r["lat"], r["lng"] = info["lat"], info["lng"]
            r["coord_source"] = "service_area_centroid"
            r["serves_points"] = info["serves"]
            r["moved_km"] = info["moved_km"]
            if info.get("pincode"):
                if str(r.get("pincode")) != info["pincode"]:
                    changed_pin += 1
                    r["pincode_old"] = r.get("pincode")
                r["pincode"] = info["pincode"]
                r["area"] = info["area"] or r.get("area")
                r["city"] = info["city"] or r.get("city")
                r["state"] = info["state"] or r.get("state")
                r["pincode_source"] = "reverse_geocode(centroid)"

    moves = [i["moved_km"] for i in refined.values()]
    print(f"\nRefined      : {len(refined)}/{len(targets)}")
    if moves:
        print(f"Mean move    : {sum(moves)/len(moves):.2f} km")
        print(f"Max move     : {max(moves):.2f} km")
    print(f"Pincodes changed by refinement: {changed_pin}")

    COLS = ["store_id", "store_name", "pincode", "pincode_old", "area", "city", "state",
            "lat", "lng", "coord_source", "serves_points", "moved_km",
            "grid_lat", "grid_lng", "is_new", "truly_new", "known_city",
            "hit_count", "pincode_source", "secondary"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bengaluru Stores"
    ws.append(COLS)
    for cell in ws[1]:
        cell.font, cell.fill, cell.alignment = HDR_FONT, HDR_FILL, HDR_ALIGN
    ws.freeze_panes = "A2"
    for i, r in enumerate(sorted(rows, key=lambda x: (str(x.get("pincode")),
                                                      str(x.get("store_name")))), start=2):
        ws.append([r.get(c, "") for c in COLS])
        if str(r.get("coord_source")) == "service_area_centroid":
            for c in range(1, len(COLS) + 1):
                ws.cell(row=i, column=c).fill = NEW_FILL
        elif i % 2 == 0:
            for c in range(1, len(COLS) + 1):
                ws.cell(row=i, column=c).fill = ALT_FILL
    ws.auto_filter.ref = ws.dimensions
    widths(ws)

    ws.insert_rows(1, amount=2)
    mc = ws.max_column
    ws["A1"] = f"Bengaluru Dark Stores - refined coordinates - {len(rows)} stores"
    ws["A2"] = (f"Generated: {scraped_at}  |  {len(refined)} coords upgraded from grid node "
                f"to service-area centroid  |  {changed_pin} pincodes corrected")
    ws["A1"].font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="444444")
    ws["A2"].fill = ALT_FILL
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=mc)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=mc)

    out = BACKEND / f"zepto_bengaluru_stores_refined_{ts}.xlsx"
    wb.save(out)
    print(f"\nSaved -> {out}")


if __name__ == "__main__":
    asyncio.run(main())
