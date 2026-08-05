"""
Brik Oven public data across every Bengaluru dark store — lat/lon driven.

Architecture (two hops, both through Zepto's own APIs):

    lat/lon  ──►  get_page API   ──►  storeId  ──►  search API  ──►  products
                  (coords injected            (store forced via
                   into the page's own         store_id request
                   get_page request)           headers)

Every coordinate in the Excel is sent through the API and the serving dark
store is resolved live — the store_id column is used only to cross-check.
Both hops retry with backoff, because Zepto intermittently returns an empty
page under sustained navigation; without retries those look like "not stocked"
when they are really "not measured".

Run from backend/:
    python -m scripts.zepto_brik_oven_latlon_bengaluru
Output: zepto_brik_oven_latlon_bengaluru_TIMESTAMP.xlsx
"""
import asyncio
import json
import sys
from pathlib import Path
from datetime import datetime
from urllib.parse import quote

sys.path.insert(0, str(Path(__file__).parent.parent))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from playwright.async_api import async_playwright

# ── Config ────────────────────────────────────────────────────────────────────
BRAND      = "Brik Oven"
KEYWORD    = "brik oven"
CITY_SHEET = "Bangalore"
COMBINED   = Path(__file__).parent.parent / "zepto_all_stores_combined_20260731_1147.xlsx"

_BASE    = "https://www.zepto.com"
_GETPAGE = "/lms/api/v2/get_page"
_SEARCH  = "/user-search-service/api/v3/search"

GETPAGE_TRIES = 3
SEARCH_TRIES  = 3

# Rewrites the page's OWN get_page request to carry our coordinates.
# A direct fetch() to bff-gateway is blocked by CORS, so we ride the page's call.
_INIT_SCRIPT = """
(() => {
    const inj = (url) => {
        if (typeof url !== 'string' || !url.includes('/lms/api/v2/get_page')) return url;
        const lat = localStorage.getItem('__probe_lat');
        const lng = localStorage.getItem('__probe_lng');
        if (!lat || !lng) return url;
        try { const u = new URL(url, location.href);
              u.searchParams.set('latitude', lat); u.searchParams.set('longitude', lng);
              return u.toString(); } catch(e) { return url; }
    };
    const of = window.fetch;
    window.fetch = function(u, o) { return of.apply(this, [inj(u), o]); };
    const oo = XMLHttpRequest.prototype.open;
    XMLHttpRequest.prototype.open = function(m, u, ...r) { return oo.call(this, m, inj(u), ...r); };
})();
"""

HDR_FONT  = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
HDR_FILL  = PatternFill("solid", fgColor="2E4057")
HDR_ALIGN = Alignment(horizontal="center", vertical="center")
ALT_FILL  = PatternFill("solid", fgColor="F0F4F8")
YES_FILL  = PatternFill("solid", fgColor="D9F7E6")   # stocked
NO_FILL   = PatternFill("solid", fgColor="FDE8E8")   # genuinely not stocked
ERR_FILL  = PatternFill("solid", fgColor="FFF4CE")   # could not measure
TITLE_FILL = PatternFill("solid", fgColor="1A3C5E")


# ── Input ─────────────────────────────────────────────────────────────────────

def load_stores() -> list[dict]:
    wb = openpyxl.load_workbook(COMBINED, read_only=True, data_only=True)
    rows = list(wb[CITY_SHEET].iter_rows(values_only=True))
    wb.close()
    hdr = [str(h).strip().lower() for h in rows[0]]
    out, seen = [], set()
    for r in rows[1:]:
        d = dict(zip(hdr, r))
        sid = str(d.get("store_id", "") or "").strip()
        if not sid or sid in seen:
            continue
        try:
            lat, lng = float(d.get("lat") or 0), float(d.get("lng") or 0)
        except (TypeError, ValueError):
            continue
        if not lat or not lng:
            continue
        seen.add(sid)
        out.append({
            "xl_store_id": sid,
            "store_name":  str(d.get("store_name", "") or "").strip(),
            "area":        str(d.get("area", "") or "").strip(),
            "lat": lat, "lng": lng,
        })
    return out


# ── HOP 1: lat/lon ──► storeId ────────────────────────────────────────────────

async def resolve_store_from_latlon(page, lat: float, lng: float) -> tuple[str | None, list[str]]:
    """Send this coordinate through get_page and read back the serving store."""
    for attempt in range(1, GETPAGE_TRIES + 1):
        found = {"id": None, "sec": []}

        async def on_getpage(resp):
            if _GETPAGE not in resp.url or found["id"]:
                return
            try:
                d = await resp.json()
                svc = d.get("storeServiceableResponse") or {}
                if svc.get("storeId"):
                    found["id"] = svc["storeId"]
                    found["sec"] = [str(x) for x in (svc.get("secondaryStoreIds") or [])]
            except Exception:
                pass

        try:
            await page.evaluate(
                f"() => {{ localStorage.setItem('__probe_lat','{lat}');"
                f" localStorage.setItem('__probe_lng','{lng}'); }}")
        except Exception:
            pass

        page.on("response", on_getpage)
        try:
            await page.goto(_BASE, timeout=30000, wait_until="domcontentloaded")
            await page.wait_for_timeout(3000)
        except Exception:
            pass
        finally:
            page.remove_listener("response", on_getpage)

        if found["id"]:
            return found["id"], found["sec"]
        await asyncio.sleep(1.5 * attempt)
    return None, []


# ── HOP 2: storeId ──► products ───────────────────────────────────────────────

def extract_all(data: dict) -> tuple[list[dict], set, int]:
    """Return (brik_oven_products, store_ids_seen, total_results)."""
    brik, ids, total = [], set(), 0
    for widget in data.get("layout", []):
        if widget.get("widgetId") != "PRODUCT_GRID":
            continue
        items = (widget.get("data", {}).get("resolver", {})
                 .get("data", {}).get("items") or [])
        for item in items:
            pr = item.get("productResponse")
            if not pr:
                continue
            total += 1
            if pr.get("storeId"):
                ids.add(pr["storeId"])
            prod = pr.get("product", {}) or {}
            if BRAND.lower() not in (prod.get("brand") or "").lower():
                continue
            pv  = pr.get("productVariant", {}) or {}
            rat = pv.get("ratingSummary", {}) or {}
            mrp = pv.get("mrp") or pr.get("mrp") or 0
            sp  = pr.get("discountedSellingPrice") or pr.get("sellingPrice") or 0
            brik.append({
                "position":     (item.get("position") or 0) + 1,
                "product_name": prod.get("name", ""),
                "brand":        prod.get("brand", ""),
                "pack_size":    pv.get("formattedPacksize", ""),
                "mrp_rs":       round(mrp / 100, 2),
                "sp_rs":        round(sp / 100, 2),
                "discount_pct": pr.get("discountPercent", 0),
                "in_stock":     not pr.get("outOfStock", False),
                "available_qty": pr.get("availableQuantity", 0),
                "rating":       rat.get("averageRating", ""),
                "rating_count": rat.get("totalRatings", 0),
                "category":     pr.get("primaryCategoryName", ""),
                "match_bucket": (pr.get("meta") or {}).get("query_matching_bucket", ""),
                "product_id":   prod.get("id", ""),
                "variant_id":   pv.get("id", ""),
                "resp_store_id": pr.get("storeId", ""),
            })
    return brik, ids, total


async def search_at_store(page, target: dict, sid: str, all_ids: str):
    """Search KEYWORD with the store forced via headers. Retries on empty page.

    Returns (products, store_ids_seen, total_results, attempts, measured).
    `measured` is False when every attempt came back empty — that is a failed
    measurement, NOT evidence the brand is absent.
    """
    target["id"] = sid
    target["all"] = all_ids
    for attempt in range(1, SEARCH_TRIES + 1):
        box = {"data": None}

        async def on_search(resp):
            if _SEARCH not in resp.url or box["data"] is not None:
                return
            try:
                d = await resp.json()
                if any(w.get("widgetId") == "PRODUCT_GRID" for w in d.get("layout", [])):
                    box["data"] = d
            except Exception:
                pass

        page.on("response", on_search)
        try:
            await page.goto(f"{_BASE}/search?query={quote(KEYWORD)}",
                            timeout=30000, wait_until="domcontentloaded")
            await page.wait_for_timeout(4000)
        except Exception:
            pass
        finally:
            page.remove_listener("response", on_search)

        if box["data"] is not None:
            brik, ids, total = extract_all(box["data"])
            if total > 0:
                return brik, ids, total, attempt, True
        await asyncio.sleep(2.0 * attempt)
    return [], set(), 0, SEARCH_TRIES, False


# ── Excel helpers ─────────────────────────────────────────────────────────────

def banner(ws, title, sub):
    ws.insert_rows(1, amount=2)
    mc = ws.max_column or 1
    ws["A1"], ws["A2"] = title, sub
    ws["A1"].font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="444444")
    ws["A2"].fill = ALT_FILL
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 15
    if mc > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=mc)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=mc)


def widths(ws):
    for col in ws.columns:
        cl = get_column_letter(col[0].column)
        ml = max((len(str(c.value)) for c in col if c.value), default=0)
        ws.column_dimensions[cl].width = min(max(ml + 2, 8), 50)


def head(ws, cols):
    ws.append(cols)
    for cell in ws[1]:
        cell.font, cell.fill, cell.alignment = HDR_FONT, HDR_FILL, HDR_ALIGN
    ws.row_dimensions[1].height = 18
    ws.freeze_panes = "A2"


# ── Main ──────────────────────────────────────────────────────────────────────

async def main():
    ts         = datetime.now().strftime("%Y%m%d_%H%M")
    scraped_at = datetime.now().strftime("%d-%b-%Y %H:%M")

    stores = load_stores()
    print(f"Brik Oven — Bengaluru — lat/lon driven — {ts}")
    print(f"Coordinates to send through API : {len(stores)}")
    print(f"Keyword                         : \"{KEYWORD}\"")
    print("=" * 78)

    detail_rows: list[dict] = []
    summary_rows: list[dict] = []

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        ctx = await browser.new_context(
            user_agent=("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                        "(KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36"))
        page = await ctx.new_page()
        await page.add_init_script(_INIT_SCRIPT)

        target = {"id": None, "all": None}

        async def route_handler(route):
            req = route.request
            if _SEARCH in req.url and req.method == "POST" and target["id"]:
                h = dict(req.headers)
                h["store_id"]   = target["id"]
                h["storeid"]    = target["id"]
                h["store_ids"]  = target["all"]
                h["store_etas"] = json.dumps({s: -1 for s in target["all"].split(",")})
                await route.continue_(headers=h)
            else:
                await route.continue_()

        await page.route("**/user-search-service/**", route_handler)

        print("Warming up browser...")
        await page.goto(_BASE, timeout=30000, wait_until="domcontentloaded")
        await page.wait_for_timeout(2500)
        print("Ready.\n")

        for i, st in enumerate(stores, 1):
            label = st["store_name"][:30]
            print(f"[{i:>3}/{len(stores)}] {label:<30}", end="  ")

            # HOP 1 — coordinate through the API
            api_sid, sec = await resolve_store_from_latlon(page, st["lat"], st["lng"])
            if not api_sid:
                print("get_page failed — could not resolve store from lat/lon")
                summary_rows.append({**st, "api_store_id": "", "id_match": "",
                                     "status": "GETPAGE_FAILED", "brik_found": "",
                                     "product_count": "", "best_rank": "",
                                     "total_results": "", "products_list": ""})
                continue

            match = "SAME" if api_sid == st["xl_store_id"] else "DIFF"

            # HOP 2 — search at the store that coordinate resolved to
            prods, ids, total, tries, measured = await search_at_store(
                page, target, api_sid, ",".join([api_sid] + sec))

            if not measured:
                print(f"[{match}] search empty after {tries} tries — NOT MEASURED")
                status, found = "SEARCH_FAILED", ""
            elif prods:
                best = min(p["position"] for p in prods)
                names = ", ".join(p["product_name"].replace("Brik Oven ", "") for p in prods[:2])
                print(f"[{match}] {len(prods)} products  best #{best:<3} {names[:34]}")
                status, found = "OK", "YES"
            else:
                print(f"[{match}] no Brik Oven  (page had {total} products)")
                status, found = "OK", "NO"

            summary_rows.append({
                **st,
                "api_store_id":  api_sid,
                "id_match":      match,
                "status":        status,
                "brik_found":    found,
                "product_count": len(prods) if measured else "",
                "best_rank":     min((p["position"] for p in prods), default="") if prods else "",
                "total_results": total if measured else "",
                "products_list": " | ".join(
                    f"#{p['position']} {p['product_name'].replace('Brik Oven ','')} ₹{p['sp_rs']}"
                    for p in prods) if prods else "",
            })

            for pr in prods:
                detail_rows.append({
                    "store_name": st["store_name"], "area": st["area"],
                    "lat": st["lat"], "lng": st["lng"],
                    "api_store_id": api_sid, **pr,
                })

            await asyncio.sleep(0.4)

        await browser.close()

    ok       = [r for r in summary_rows if r["status"] == "OK"]
    stocked  = [r for r in ok if r["brik_found"] == "YES"]
    failed   = [r for r in summary_rows if r["status"] != "OK"]
    same_ids = [r for r in summary_rows if r["id_match"] == "SAME"]

    print(f"\n{'='*78}")
    print(f"Coordinates sent through API : {len(stores)}")
    print(f"Successfully measured        : {len(ok)}")
    print(f"Failed to measure            : {len(failed)}")
    print(f"Stores stocking Brik Oven    : {len(stocked)}")
    if ok:
        print(f"Coverage (of measured)       : {len(stocked)/len(ok)*100:.1f}%")
    print(f"lat/lon → same store as Excel: {len(same_ids)}/{len(stores)}")
    print(f"Product rows                 : {len(detail_rows)}")

    # ── Excel ────────────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    cov = f"{len(stocked)}/{len(ok)} measured" if ok else "n/a"
    b_title = f"Brik Oven — Bengaluru dark stores — {cov}"
    b_sub   = (f"Generated: {scraped_at}  |  Keyword: \"{KEYWORD}\"  |  "
               f"lat/lon → get_page → storeId → search  |  "
               f"{len(failed)} store(s) could not be measured")

    # Sheet 1 — per store
    ws1 = wb.active
    ws1.title = "Store Coverage"
    c1 = ["store_name", "area", "lat", "lng", "api_store_id", "xl_store_id",
          "id_match", "status", "brik_found", "product_count", "best_rank",
          "total_results", "products_list"]
    head(ws1, c1)
    order = {"YES": 0, "NO": 1, "": 2}
    for r_idx, row in enumerate(
            sorted(summary_rows, key=lambda x: (order.get(x["brik_found"], 3),
                                                x["best_rank"] if x["best_rank"] != "" else 999)),
            start=2):
        ws1.append([row.get(c, "") for c in c1])
        fill = (ERR_FILL if row["status"] != "OK"
                else YES_FILL if row["brik_found"] == "YES" else NO_FILL)
        for c in range(1, len(c1) + 1):
            ws1.cell(row=r_idx, column=c).fill = fill
    ws1.auto_filter.ref = ws1.dimensions
    widths(ws1)

    # Sheet 2 — per store × product
    ws2 = wb.create_sheet("Product Detail")
    c2 = ["store_name", "area", "lat", "lng", "position", "product_name", "brand",
          "pack_size", "mrp_rs", "sp_rs", "discount_pct", "in_stock", "available_qty",
          "rating", "rating_count", "category", "match_bucket",
          "api_store_id", "product_id", "variant_id"]
    head(ws2, c2)
    for r_idx, row in enumerate(
            sorted(detail_rows, key=lambda x: (x["store_name"], x["position"])), start=2):
        ws2.append([row.get(c, "") for c in c2])
        if r_idx % 2 == 0:
            for c in range(1, len(c2) + 1):
                ws2.cell(row=r_idx, column=c).fill = ALT_FILL
    ws2.auto_filter.ref = ws2.dimensions
    widths(ws2)

    # Sheet 3 — per product rollup (the compact view)
    ws3 = wb.create_sheet("By Product")
    c3 = ["product_name", "pack_size", "stores_stocking", "avg_rank", "best_rank",
          "min_price_rs", "max_price_rs", "avg_price_rs", "mrp_rs", "in_stock_stores"]
    head(ws3, c3)
    by_prod: dict[str, list[dict]] = {}
    for r in detail_rows:
        by_prod.setdefault(r["product_name"], []).append(r)
    for r_idx, (pname, rs) in enumerate(
            sorted(by_prod.items(), key=lambda kv: -len(kv[1])), start=2):
        prices = [r["sp_rs"] for r in rs if r["sp_rs"]]
        ranks  = [r["position"] for r in rs if r["position"]]
        ws3.append([
            pname, rs[0]["pack_size"], len(rs),
            round(sum(ranks) / len(ranks), 1) if ranks else "",
            min(ranks) if ranks else "",
            min(prices) if prices else "", max(prices) if prices else "",
            round(sum(prices) / len(prices), 2) if prices else "",
            rs[0]["mrp_rs"], sum(1 for r in rs if r["in_stock"]),
        ])
        if r_idx % 2 == 0:
            for c in range(1, len(c3) + 1):
                ws3.cell(row=r_idx, column=c).fill = ALT_FILL
    widths(ws3)

    # Sheet 4 — run summary
    ws4 = wb.create_sheet("Summary")
    head(ws4, ["Metric", "Value"])
    stats = [
        ("Coordinates sent through API",  len(stores)),
        ("Successfully measured",         len(ok)),
        ("Failed to measure",             len(failed)),
        ("Stores stocking Brik Oven",     len(stocked)),
        ("Stores NOT stocking",           len(ok) - len(stocked)),
        ("Coverage % (of measured)",      f"{len(stocked)/len(ok)*100:.1f}%" if ok else "—"),
        ("lat/lon resolved to same store as Excel", f"{len(same_ids)}/{len(stores)}"),
        ("Total product rows",            len(detail_rows)),
        ("Unique products",               len(by_prod)),
        ("Avg rank where stocked",
         f"{sum(r['position'] for r in detail_rows)/len(detail_rows):.1f}" if detail_rows else "—"),
        ("Best rank seen",                min((r["position"] for r in detail_rows), default="—")),
        ("Keyword",                       KEYWORD),
        ("Method", "lat/lon → get_page → storeId → search (headers)"),
        ("Scraped at",                    scraped_at),
    ]
    for r_idx, (m, v) in enumerate(stats, start=2):
        ws4.cell(row=r_idx, column=1, value=m)
        ws4.cell(row=r_idx, column=2, value=v)
        if r_idx % 2 == 0:
            for c in range(1, 3):
                ws4.cell(row=r_idx, column=c).fill = ALT_FILL
    widths(ws4)

    for ws in (ws1, ws2, ws3, ws4):
        banner(ws, b_title, b_sub)

    out = Path(__file__).parent.parent / f"zepto_brik_oven_latlon_bengaluru_{ts}.xlsx"
    wb.save(out)
    print(f"\nSaved → {out}")


if __name__ == "__main__":
    asyncio.run(main())
