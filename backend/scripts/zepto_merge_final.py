"""
Merge all 58 city discovery files into one deduplicated master.

This is "layer 3". Each city file lists everything THAT city's scan saw,
including other cities' stores — so Bengaluru's file and Hosur's file both
contain BLR-Nagdevanahalli. Stacking the files would duplicate it.

The merge keys on store_id and keeps ONE row per store, filed under the city
its own name prefix says it belongs to, not whichever file happened to find it.

Rules applied:
  - one row per store_id; the record with a real store_name wins
  - phantom rows (serviceable=False, no storeDetailsResponse) are excluded from
    counts and listed separately — they are not operational stores
  - the SAS/PNK prefix overrides are re-applied, so files scanned before that
    fix (Ambala) get the same attribution as the rest
  - a city with no store of its own is labelled "Covered in <city>", derived
    from what the scans actually found

Run from backend/:
    python -m scripts.zepto_merge_final
Output: zepto_FINAL_master_TIMESTAMP.xlsx
"""
import re
import sys
from pathlib import Path
from datetime import datetime
from collections import Counter, defaultdict

sys.path.insert(0, str(Path(__file__).parent.parent))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BACKEND = Path(__file__).parent.parent
INDIR = BACKEND / "dark_stores_zepto"

PREFIX_OVERRIDE = {"SAS": "Mohali", "PNK": "Panchkula"}

TIERS = {
    **{c: "Tier 1" for c in ["Ahmedabad", "Bengaluru", "Chennai", "Delhi NCR",
                             "Hyderabad", "Kolkata", "Mumbai", "Pune"]},
    **{c: "Tier 2" for c in ["Agra", "Amritsar", "Chandigarh", "Coimbatore", "Dehradun",
                             "Hubballi", "Indore", "Jaipur", "Jalandhar", "Kanpur", "Kochi",
                             "Kota", "Lucknow", "Ludhiana", "Madurai", "Mysuru", "Nagpur",
                             "Nashik", "Patiala", "Prayagraj", "Puducherry", "Rajkot",
                             "Surat", "Udaipur", "Vadodara", "Varanasi", "Vijayawada",
                             "Warangal"]},
    **{c: "Tier 3" for c in ["Ambala", "Bareilly", "Belagavi", "Bhiwadi",
                             "Chhatrapati Sambhaji Nagar", "Davangere", "Gorakhpur",
                             "Guntur", "Hisar", "Hosur", "Karimnagar", "Karnal", "Meerut",
                             "Mehsana", "Mohali", "Palakkad", "Panchkula", "Panipat",
                             "Rewari", "Sonipat", "Tumkuru", "Vellore"]},
}

HDR_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
HDR_FILL = PatternFill("solid", fgColor="2E4057")
HDR_ALIGN = Alignment(horizontal="center", vertical="center")
ALT_FILL = PatternFill("solid", fgColor="F0F4F8")
NEW_FILL = PatternFill("solid", fgColor="D9F7E6")
GHOST_FILL = PatternFill("solid", fgColor="FFF4CE")
T1 = PatternFill("solid", fgColor="1A3C5E")
T2 = PatternFill("solid", fgColor="2E6DA4")
T3 = PatternFill("solid", fgColor="5B9BD5")
COV = PatternFill("solid", fgColor="A9C4E0")
TOT = PatternFill("solid", fgColor="048A81")
TITLE_FILL = PatternFill("solid", fgColor="1A3C5E")


def latest_per_city():
    best = {}
    for f in INDIR.glob("darkstores_*.xlsx"):
        if f.name.startswith("~$"):
            continue
        m = re.match(r"darkstores_(.+)_(\d{8}_\d{4})\.xlsx$", f.name)
        if not m:
            continue
        slug, ts = m.group(1), m.group(2)
        if slug not in best or ts > best[slug][0]:
            best[slug] = (ts, f)
    return {k: v[1] for k, v in best.items()}


def read(f):
    wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
    rr = list(wb["Stores"].iter_rows(values_only=True))
    wb.close()
    hdr = [str(x).strip().lower() if x else "" for x in rr[2]]
    return [dict(zip(hdr, r)) for r in rr[3:] if r and r[0]]


def prefix_of(name):
    m = re.match(r"^\s*([A-Za-z]{2,4})\s*-", str(name or ""))
    return m.group(1).upper() if m else None


def widths(ws, mx=44):
    for col in ws.columns:
        cl = get_column_letter(col[0].column)
        ml = max((len(str(c.value)) for c in col[:500] if c.value), default=0)
        ws.column_dimensions[cl].width = min(max(ml + 2, 9), mx)


def head(ws, cols):
    ws.append(cols)
    for c in ws[1]:
        c.font, c.fill, c.alignment = HDR_FONT, HDR_FILL, HDR_ALIGN
    ws.row_dimensions[1].height = 18
    ws.freeze_panes = "A2"


def banner(ws, t, s):
    ws.insert_rows(1, amount=2)
    mc = ws.max_column or 1
    ws["A1"], ws["A2"] = t, s
    ws["A1"].font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="444444")
    ws["A2"].fill = ALT_FILL
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 15
    if mc > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=mc)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=mc)


def main():
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    at = datetime.now().strftime("%d-%b-%Y %H:%M")
    files = latest_per_city()
    print(f"Merging {len(files)} city files from {INDIR}")

    # prefix -> city, learned from every file, then overridden explicitly
    votes = defaultdict(Counter)
    raw_rows = 0
    all_rows = []
    for slug, f in files.items():
        for r in read(f):
            raw_rows += 1
            all_rows.append(r)
            p = prefix_of(r.get("store_name"))
            if p and r.get("city"):
                votes[p][r["city"]] += 1
    prefix_city = {p: c.most_common(1)[0][0] for p, c in votes.items()
                   if c.most_common(1)[0][1] / sum(c.values()) >= 0.8}
    prefix_city.update(PREFIX_OVERRIDE)

    # one row per store_id; prefer the record that carries a real name
    master = {}
    for r in all_rows:
        sid = str(r.get("store_id") or "").strip()
        if not sid:
            continue
        nm = str(r.get("store_name") or "").strip()
        cur = master.get(sid)
        if cur is None or (not str(cur.get("store_name") or "").strip() and nm):
            master[sid] = dict(r)
        else:
            for k in ("pincode", "area_suggestion", "state", "source_pincode"):
                if not str(cur.get(k) or "").strip() and str(r.get(k) or "").strip():
                    cur[k] = r[k]

    # re-apply attribution so pre-fix files agree with the rest
    moved = []
    for sid, r in master.items():
        p = prefix_of(r.get("store_name"))
        home = prefix_city.get(p) if p else None
        if home and home != r.get("city"):
            moved.append((r.get("store_name"), r.get("city"), home))
            r["city"] = home
        r["tier"] = TIERS.get(r.get("city"), "")
        r["is_phantom"] = str(r.get("store_record")) == "NO STORE RECORD"

    real = {k: v for k, v in master.items() if not v["is_phantom"]}
    ghost = {k: v for k, v in master.items() if v["is_phantom"]}

    print(f"  rows across all files : {raw_rows}")
    print(f"  unique store_ids      : {len(master)}")
    print(f"  duplicates collapsed  : {raw_rows - len(master)}")
    print(f"  real stores           : {len(real)}")
    print(f"  phantom (no record)   : {len(ghost)}")
    print(f"  re-attributed by merge: {len(moved)}")

    per_city = Counter(v["city"] for v in real.values())
    cities = sorted(set(TIERS) | set(per_city))
    zero = [c for c in cities if per_city.get(c, 0) == 0]

    # a zero-store city is "covered in" wherever its old stores ended up
    covered = {}
    for c in zero:
        near = Counter()
        for v in real.values():
            if v.get("already_in") == c:
                near[v["city"]] += 1
        covered[c] = near.most_common(1)[0][0] if near else ""

    # ── Workbook ────────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    T = f"Zepto Dark Stores — FINAL — {len(real)} unique stores across {len([c for c in cities if per_city.get(c,0)])} cities"
    S = (f"Generated: {at}  |  merged from {len(files)} city scans  |  "
         f"{raw_rows} rows collapsed to {len(master)} unique store_ids  |  "
         f"{len(ghost)} phantom rows excluded")

    COLS = ["store_id", "store_name", "city", "tier", "source_pincode",
            "area_suggestion", "state", "lat", "lng", "found_by", "is_new",
            "serviceable", "assigned_by", "secondary"]
    ws = wb.active
    ws.title = "All Stores"
    head(ws, COLS)
    order = {"Tier 1": 0, "Tier 2": 1, "Tier 3": 2, "": 3}
    rows_sorted = sorted(real.values(),
                         key=lambda r: (order.get(r.get("tier"), 3), r.get("city", ""),
                                        str(r.get("source_pincode") or ""),
                                        str(r.get("store_name") or "")))
    for i, r in enumerate(rows_sorted, start=2):
        ws.append([r.get(c, "") for c in COLS])
        if str(r.get("is_new")) == "True":
            for c in range(1, len(COLS) + 1):
                ws.cell(row=i, column=c).fill = NEW_FILL
        elif i % 2 == 0:
            for c in range(1, len(COLS) + 1):
                ws.cell(row=i, column=c).fill = ALT_FILL
    ws.auto_filter.ref = ws.dimensions
    widths(ws)

    # City summary
    ws2 = wb.create_sheet("City Summary")
    head(ws2, ["City", "Tier", "Unique Dark Stores"])
    total = 0
    r_i = 2
    for tier in ("Tier 1", "Tier 2", "Tier 3"):
        for c in sorted(x for x in cities if TIERS.get(x) == tier):
            n = per_city.get(c, 0)
            val = n if n else f"Covered in {covered.get(c)}" if covered.get(c) else 0
            ws2.cell(row=r_i, column=1, value=c)
            ws2.cell(row=r_i, column=2, value=tier)
            ws2.cell(row=r_i, column=3, value=val)
            fill = {"Tier 1": T1, "Tier 2": T2, "Tier 3": T3}[tier] if n else COV
            for cc in (1, 2, 3):
                cell = ws2.cell(row=r_i, column=cc)
                cell.fill = fill
                cell.font = (Font(name="Calibri", size=10, bold=True, color="FFFFFF")
                             if n else Font(name="Calibri", size=10, italic=True,
                                            color="1A3C5E"))
                cell.alignment = Alignment(horizontal="center" if cc > 1 else "left")
            total += n
            r_i += 1
    ws2.cell(row=r_i, column=1, value="TOTAL")
    ws2.cell(row=r_i, column=3, value=total)
    for cc in (1, 2, 3):
        ws2.cell(row=r_i, column=cc).fill = TOT
        ws2.cell(row=r_i, column=cc).font = Font(name="Calibri", size=12, bold=True,
                                                 color="FFFFFF")
    widths(ws2)

    # Re-attributions
    ws3 = wb.create_sheet("Re-attributed")
    head(ws3, ["store_name", "was filed under", "belongs to", "reason"])
    for i, (n, was, now) in enumerate(sorted(moved, key=lambda x: str(x[0])), start=2):
        ws3.append([n, was, now, f"name prefix '{prefix_of(n)}' = {now}"])
        if i % 2 == 0:
            for c in range(1, 5):
                ws3.cell(row=i, column=c).fill = ALT_FILL
    widths(ws3)

    # Phantoms
    ws4 = wb.create_sheet("Excluded (no record)")
    head(ws4, ["store_id", "city", "lat", "lng", "found_by", "why"])
    for i, v in enumerate(ghost.values(), start=2):
        ws4.append([v.get("store_id"), v.get("city"), v.get("lat"), v.get("lng"),
                    v.get("found_by"),
                    "serviceable=False and no storeDetailsResponse — not an operational store"])
        for c in range(1, 7):
            ws4.cell(row=i, column=c).fill = GHOST_FILL
    widths(ws4)

    # New stores
    ws5 = wb.create_sheet("New Stores")
    head(ws5, COLS)
    newr = [r for r in real.values() if str(r.get("is_new")) == "True"]
    for i, r in enumerate(sorted(newr, key=lambda x: (x.get("city", ""),
                                                      str(x.get("store_name")))), start=2):
        ws5.append([r.get(c, "") for c in COLS])
        for c in range(1, len(COLS) + 1):
            ws5.cell(row=i, column=c).fill = NEW_FILL
    widths(ws5)

    for w in (ws, ws2, ws3, ws4, ws5):
        banner(w, T, S)

    out = BACKEND / f"zepto_FINAL_master_{ts}.xlsx"
    wb.save(out)

    print()
    print("=" * 60)
    print(f"FINAL UNIQUE DARK STORES : {len(real)}")
    print(f"  new vs the old 1129    : {len(newr)}")
    print(f"  cities with stores     : {len([c for c in cities if per_city.get(c,0)])}")
    print(f"  cities with none       : {len(zero)}  {zero if zero else ''}")
    print(f"Saved -> {out}")


if __name__ == "__main__":
    main()
