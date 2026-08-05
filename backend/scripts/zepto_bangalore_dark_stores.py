"""
Scrape Zepto dark stores in Bangalore — delivery areas + grid scan.

How it works:
  1. Open zepto.com in Playwright once (WAF token obtained automatically by real browser)
  2. Register init_script: overrides window.fetch + XMLHttpRequest to inject lat/lng
     from localStorage before every get_page call — survives page reloads
  3. For each coordinate (delivery areas + 2 km grid):
     - Write target lat/lng to localStorage
     - Reload → zepto.com fires get_page with our injected coordinates
     - Capture store_id via page.on("response")
  4. Save deduplicated results to Excel (same format as zepto_ncr_dark_stores.xlsx)

No cookies to paste, no tokens to copy.
Run from backend/:
    python -m scripts.zepto_bangalore_dark_stores
"""
import asyncio
import re
import sys
import itertools
from pathlib import Path
from collections import Counter, defaultdict

sys.path.insert(0, str(Path(__file__).parent.parent))

import httpx
import openpyxl
from playwright.async_api import async_playwright

_BFF = "https://bff-gateway.zepto.com"

# ── Bangalore delivery areas from zepto.com/s/del-areas ─────────────────────
BANGALORE_AREAS = [
    "Hebbal", "Koramangala", "Peenya", "Malleswaram", "Jayanagar", "Chickpet",
    "Indiranagar", "Kr Puram", "Ramamurthy Nagar", "Bhadrappa Layout",
    "C V Raman Nagar", "Bannerghatta", "Hennur", "Bellandur", "Richmond Town",
    "Rn Nagar", "Nagarbhavi", "Sarjapur", "Electronic City", "Gunjur", "Jp Nagar",
    "Btm Layout", "Singasandra", "Hsr Layout", "Marathahalli", "Sonnenahalli",
    "Kothnur", "Basavanapura", "Brookefield", "Whitefield", "Kalyan Nagar New",
    "Yeshwantpur", "Mico Layout", "Raghavendra Layout", "Banashankari",
    "Yelahanka New Town", "Jakkur", "Jalahalli", "Vijay Nagar", "Shivaji Nagar",
]

# Bangalore bounding box (covers all delivery areas with buffer)
BANGALORE_BBOX = {
    "lat_min": 12.82, "lat_max": 13.12,
    "lng_min": 77.45, "lng_max": 77.80,
}

# All Bangalore pincodes for autosuggestion sweep
BANGALORE_PINCODES: dict[str, str] = {
    "560001": "MG Road",
    "560002": "Shivajinagar",
    "560003": "Rajajinagar",
    "560004": "Basavanagudi",
    "560005": "Frazer Town",
    "560006": "Malleswaram",
    "560007": "Seshadripuram",
    "560008": "Sadashivanagar",
    "560009": "Rajajinagar West",
    "560010": "Vijayanagar",
    "560011": "Rajajinagar East",
    "560012": "Tumkur Road",
    "560013": "Neelasandra",
    "560014": "Langford Town",
    "560015": "Ashok Nagar",
    "560016": "Malleswaram West",
    "560017": "Indiranagar",
    "560018": "Kammanahalli",
    "560019": "Laggere",
    "560020": "Peenya",
    "560021": "Yeshwantpur",
    "560022": "Rajajinagar Industrial",
    "560023": "Dockyard Road",
    "560024": "Ulsoor",
    "560025": "Benson Town",
    "560026": "RT Nagar",
    "560027": "Hebbal",
    "560029": "Yelahanka",
    "560030": "Kodigehalli",
    "560032": "Nagavara",
    "560033": "Kalyan Nagar",
    "560034": "HMT Layout",
    "560035": "Nagarbhavi",
    "560036": "Mysore Road",
    "560037": "Banashankari",
    "560038": "Kumaraswamy Layout",
    "560039": "JP Nagar",
    "560040": "Uttarahalli",
    "560041": "Jayanagar",
    "560042": "Koramangala",
    "560043": "Ejipura",
    "560045": "BTM Layout",
    "560046": "Domlur",
    "560047": "Richmond Town",
    "560048": "Shanthinagar",
    "560049": "Vasanth Nagar",
    "560050": "Cleveland Town",
    "560051": "Whitefield",
    "560053": "Nandini Layout",
    "560054": "Kengeri",
    "560055": "Bannerghatta Road",
    "560056": "Bommanahalli",
    "560058": "Sarjapur",
    "560059": "KR Puram",
    "560060": "Ramamurthy Nagar",
    "560061": "CV Raman Nagar",
    "560062": "Tin Factory",
    "560063": "Bellandur",
    "560064": "Marathahalli East",
    "560065": "Hennur",
    "560066": "Thanisandra",
    "560067": "Jakkur",
    "560068": "Aero Industries",
    "560069": "Kammanahalli East",
    "560070": "Bannerghatta",
    "560072": "Bommasandra",
    "560073": "Hoskote",
    "560076": "Iblur",
    "560077": "Brookefield",
    "560078": "Gunjur",
    "560079": "Devanahalli Road",
    "560083": "Nagavara North",
    "560084": "Kogilu",
    "560085": "Doddaballapur Road",
    "560086": "Devanahalli",
    "560087": "Manyata Tech Park",
    "560091": "Doddaballapur",
    "560092": "Kothanur",
    "560093": "Bannerghatta North",
    "560094": "JP Nagar Phase 7",
    "560095": "Basavanapura",
    "560097": "Begur",
    "560098": "Kodichikkanahalli",
    "560099": "Attibele",
    "560100": "Hebbagodi",
    "560102": "Kengeri Satellite Town",
    "560103": "Bidadi",
    "560104": "Kuduregere",
    "560108": "Maheshwaram",
    "560300": "Electronic City Phase 2",
    "562106": "Doddaballapur Rural",
    "562107": "Gauribidanur",
    "562114": "Nelamangala",
    "562123": "Doddaballapur North",
    "562125": "Nandagudi",
    "562130": "Magadi",
    "562149": "Kanakapura",
    "562157": "Ramanagara",
    "562162": "Channapatna",
    "562164": "Bidadi Rural",
}

# JS injected before any page script — overrides fetch + XHR to inject lat/lng
_INIT_SCRIPT = """
(() => {
    const injectCoords = (url) => {
        if (typeof url !== 'string' || !url.includes('/lms/api/v2/get_page')) return url;
        const lat = localStorage.getItem('__probe_lat');
        const lng = localStorage.getItem('__probe_lng');
        if (!lat || !lng) return url;
        try {
            const u = new URL(url, location.href);
            u.searchParams.set('latitude', lat);
            u.searchParams.set('longitude', lng);
            return u.toString();
        } catch(e) { return url; }
    };

    // Override fetch
    const origFetch = window.fetch;
    window.fetch = function(url, opts) {
        return origFetch.apply(this, [injectCoords(url), opts]);
    };

    // Override XMLHttpRequest.open (fallback for axios/XHR usage)
    const origOpen = XMLHttpRequest.prototype.open;
    XMLHttpRequest.prototype.open = function(method, url, ...rest) {
        return origOpen.call(this, method, injectCoords(url), ...rest);
    };
})();
"""


def make_grid(lat_min, lat_max, lng_min, lng_max, step_km=2.0):
    step = step_km / 111.0
    lats, lngs = [], []
    lat = lat_min
    while lat <= lat_max:
        lats.append(round(lat, 5))
        lat += step
    lng = lng_min
    while lng <= lng_max:
        lngs.append(round(lng, 5))
        lng += step
    return list(itertools.product(lats, lngs))


async def get_coords(query: str, ctx) -> dict | None:
    """Resolve query to lat/lng via Zepto autocomplete (no WAF on these endpoints)."""
    headers = {
        "Accept": "application/json",
        "Origin": "https://www.zepto.com",
        "Referer": "https://www.zepto.com/",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "app_sub_platform": "WEB",
        "app_version": "16.20.0",
        "platform": "WEB",
        "tenant": "ZEPTO",
    }
    for q in [query, query.split(",")[0].strip(), f"{query} India"]:
        try:
            r = await ctx.request.get(
                f"{_BFF}/api/v1/maps/place/autocomplete/",
                params={"place_name": q},
                headers=headers,
                timeout=10000,
            )
            if r.status != 200:
                continue
            data = await r.json()
            preds = data.get("data", {}).get("predictions") or data.get("predictions") or []
            if not preds:
                continue
            place_id = preds[0]["place_id"]

            r2 = await ctx.request.get(
                f"{_BFF}/api/v1/maps/place/details/",
                params={"place_id": place_id},
                headers=headers,
                timeout=10000,
            )
            if r2.status != 200:
                continue
            d2 = await r2.json()
            result = d2.get("data", {}).get("result") or d2.get("result") or {}
            geo = result.get("geometry", {}).get("location", {})
            lat, lng = geo.get("lat"), geo.get("lng")
            if not lat or not lng:
                continue

            pincode = state = city = None
            for comp in result.get("address_components", []):
                types = comp.get("types", [])
                if "postal_code" in types:
                    pincode = comp.get("long_name")
                if "locality" in types or "administrative_area_level_2" in types:
                    city = city or comp.get("long_name")
                if "administrative_area_level_1" in types:
                    state = state or comp.get("long_name")

            return {
                "lat": lat, "lng": lng,
                "pincode": pincode or "", "city": city or "", "state": state or "Karnataka",
            }
        except Exception as e:
            print(f"    [coord error] {q}: {e}")
            continue
    return None


async def reverse_geocode(lat: float, lng: float, client: httpx.AsyncClient) -> dict:
    """Reverse geocode lat/lng → area name + pincode using OpenStreetMap Nominatim (free)."""
    try:
        r = await client.get(
            "https://nominatim.openstreetmap.org/reverse",
            params={"lat": lat, "lon": lng, "format": "json", "addressdetails": 1},
            headers={"User-Agent": "zepto-store-scraper/1.0"},
            timeout=10,
        )
        if r.status_code != 200:
            return {}
        data = r.json()
        addr = data.get("address", {})
        pincode = addr.get("postcode", "")
        area = (
            addr.get("suburb")
            or addr.get("neighbourhood")
            or addr.get("quarter")
            or addr.get("village")
            or addr.get("town")
            or ""
        )
        return {"pincode": pincode, "area": area}
    except Exception:
        return {}


async def geocode_store_name(store_name: str, city: str, client: httpx.AsyncClient) -> dict:
    """
    Geocode store name using Zepto's own autocomplete API — gives the same
    coordinates the TL gets when manually searching on zepto.com.
    "BLR-Whitefield New" → query "Whitefield Bangalore" → Zepto autocomplete → lat/lng
    """
    if not store_name:
        return {}
    area = re.sub(r'^[A-Z]{2,5}-', '', store_name)
    area = re.sub(r'\s+(New|Old|\d+[A-Za-z]?|Phase\s*\d+|Block\s*\d+)$', '', area, flags=re.IGNORECASE).strip()
    if not area:
        return {}
    headers = {
        "Accept": "application/json",
        "Origin": "https://www.zepto.com",
        "Referer": "https://www.zepto.com/",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "app_sub_platform": "WEB", "app_version": "16.20.0",
        "platform": "WEB", "tenant": "ZEPTO",
    }
    for query in [f"{area} {city}", area]:
        try:
            r = await client.get(
                f"{_BFF}/api/v1/maps/place/autocomplete/",
                params={"place_name": query}, headers=headers, timeout=10,
            )
            if r.status_code != 200:
                continue
            preds = r.json().get("data", {}).get("predictions") or r.json().get("predictions") or []
            if not preds:
                continue
            r2 = await client.get(
                f"{_BFF}/api/v1/maps/place/details/",
                params={"place_id": preds[0]["place_id"]}, headers=headers, timeout=10,
            )
            if r2.status_code != 200:
                continue
            result = r2.json().get("data", {}).get("result") or r2.json().get("result") or {}
            geo = result.get("geometry", {}).get("location", {})
            lat, lng = geo.get("lat"), geo.get("lng")
            if lat and lng:
                return {"estimated_lat": lat, "estimated_lng": lng}
        except Exception:
            continue
    return {}


async def probe_store_at(lat: float, lng: float, page) -> dict:
    """
    Set localStorage coords, reload page.
    The init_script overrides fetch/XHR to inject lat/lng → captures natural get_page response.
    """
    result = {}
    done = asyncio.Event()

    async def on_response(response):
        if "/lms/api/v2/get_page" in response.url and not done.is_set():
            try:
                data = await response.json()
                svc = data.get("storeServiceableResponse", {})
                sid = svc.get("storeId")
                if sid:
                    result["store_id"] = str(sid)
                    result["secondary_ids"] = [str(s) for s in svc.get("secondaryStoreIds", [])]
                    result["store_name"] = data.get("storeDetailsResponse", {}).get("name", "")
                done.set()
            except Exception:
                pass

    # Write target coords to localStorage (persists through reload, read by init_script)
    try:
        await page.evaluate(f"""() => {{
            localStorage.setItem('__probe_lat', '{lat}');
            localStorage.setItem('__probe_lng', '{lng}');
        }}""")
    except Exception:
        pass

    page.on("response", on_response)
    try:
        await page.reload(timeout=20000, wait_until="domcontentloaded")
        try:
            await asyncio.wait_for(done.wait(), timeout=8)
        except asyncio.TimeoutError:
            pass
    except Exception:
        pass
    finally:
        page.remove_listener("response", on_response)

    return result


async def main():
    print("Zepto Bangalore Dark Store Scraper")
    print("=" * 55)

    seen_primary: set[str] = set()
    primary_rows: list[dict] = []
    secondary_rows: list[dict] = []
    # All probe hits per store — used to compute centroid (actual store location estimate)
    probe_hits: dict[str, list] = defaultdict(list)

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        ctx = await browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36"
            ),
        )
        page = await ctx.new_page()

        # Register init_script BEFORE first navigation — runs on every reload
        await page.add_init_script(_INIT_SCRIPT)

        # Warm up: open zepto.com so browser solves WAF challenge
        print("Warming up browser (WAF challenge)...")
        try:
            await page.goto("https://www.zepto.com", timeout=30000, wait_until="domcontentloaded")
            await page.wait_for_timeout(3000)
            print("Browser ready.\n")
        except Exception as e:
            print(f"Warmup: {e}\n")

        # ── Step 1: Resolve delivery area names to coordinates ──────────────────
        print("Step 1: Resolving delivery area coordinates via autocomplete...")
        probes: list[dict] = []

        for area in BANGALORE_AREAS:
            coords = await get_coords(f"{area} Bangalore", ctx)
            if coords:
                probes.append({
                    "area": area, "lat": coords["lat"], "lng": coords["lng"],
                    "pincode": coords["pincode"], "state": coords["state"],
                    "found_by": "autosuggestion",
                })
                print(f"  ✓ {area} → ({coords['lat']:.4f}, {coords['lng']:.4f})")
            else:
                print(f"  ✗ {area} → not resolved")
            await asyncio.sleep(0.4)

        # ── Step 1b: Pincode sweep ───────────────────────────────────────────────
        print(f"\nStep 1b: Pincode sweep ({len(BANGALORE_PINCODES)} pincodes)...")
        for pincode, loc_name in BANGALORE_PINCODES.items():
            coords = await get_coords(f"{pincode} {loc_name} Bangalore", ctx)
            if coords:
                probes.append({
                    "area": loc_name, "lat": coords["lat"], "lng": coords["lng"],
                    "pincode": pincode, "state": coords["state"],
                    "found_by": "pincode",
                })
            await asyncio.sleep(0.3)

        # ── Step 2: Grid scan ────────────────────────────────────────────────────
        grid = make_grid(**BANGALORE_BBOX, step_km=1.0)
        print(f"\nStep 2: Grid scan — {len(grid)} points @ 1 km step")
        for lat, lng in grid:
            probes.append({
                "area": "", "lat": lat, "lng": lng,
                "pincode": "", "state": "Karnataka",
                "found_by": "grid",
            })

        print(f"\nTotal probes: {len(probes)} (areas + pincodes + grid)")
        print("Step 3: Probing each coordinate...\n")

        # ── Step 3: Probe each coordinate ───────────────────────────────────────
        for i, item in enumerate(probes, 1):
            lat, lng = item["lat"], item["lng"]
            r = await probe_store_at(lat, lng, page)
            sid = r.get("store_id")

            label = item["area"] or f"grid ({lat:.4f},{lng:.4f})"

            if sid:
                # Always record every hit — used later to compute centroid
                probe_hits[sid].append((lat, lng))

                if sid not in seen_primary:
                    seen_primary.add(sid)
                    secondary = ",".join(r.get("secondary_ids", []))
                    primary_rows.append({
                        "store_id": sid,
                        "store_name": r.get("store_name", ""),
                        "city": "Bangalore",
                        "state": item.get("state", "Karnataka"),
                        "pincode": item.get("pincode", ""),
                        "area": item.get("area", ""),
                        "probe_lat": lat,
                        "probe_lng": lng,
                        "estimated_lat": "",
                        "estimated_lng": "",
                        "hit_count": 0,
                        "secondary_store_id": secondary,
                        "found_by": item["found_by"],
                    })
                    for ssid in r.get("secondary_ids", []):
                        secondary_rows.append({
                            "primary_store_id": sid, "secondary_store_id": ssid,
                            "probe_lat": lat, "probe_lng": lng,
                        })
                    print(f"  [{i}/{len(probes)}] NEW {sid} ← {label}")
            else:
                if item["found_by"] != "grid":
                    print(f"  [{i}/{len(probes)}] no store ← {label}")

        try:
            await browser.close()
        except Exception:
            pass  # browser may have already crashed — data is safe, proceed to save

    # ── Save to Excel (always runs, even if browser crashed) ─────────────────────
    print(f"\n{'='*55}")
    print(f"Total unique Bangalore dark stores: {len(primary_rows)}")

    if not primary_rows:
        print("No stores found. Try running again — WAF may have blocked initial warm-up.")
        return

    from datetime import datetime
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    # ── Geocode store names + fill missing pincode/area ─────────────────────────
    print(f"\nGeocoding {len(primary_rows)} stores (store name → estimated lat/lng, plus missing pincode/area)...")
    async with httpx.AsyncClient(follow_redirects=True) as client:
        for i, row in enumerate(primary_rows, 1):
            sid = row["store_id"]
            hits = probe_hits[sid]
            row["hit_count"] = len(hits)

            # ── Centroid: average of all probe hits (best estimate of store location) ──
            if len(hits) >= 3:
                row["estimated_lat"] = round(sum(h[0] for h in hits) / len(hits), 6)
                row["estimated_lng"] = round(sum(h[1] for h in hits) / len(hits), 6)
                print(f"  [{i}/{len(primary_rows)}] {row.get('store_name') or sid[:8]} "
                      f"→ centroid from {len(hits)} hits "
                      f"({row['estimated_lat']}, {row['estimated_lng']})")
            else:
                # Too few hits for reliable centroid — fall back to store name geocoding
                if row.get("store_name"):
                    geo = await geocode_store_name(row["store_name"], "Bangalore", client)
                    row["estimated_lat"] = geo.get("estimated_lat", "")
                    row["estimated_lng"] = geo.get("estimated_lng", "")
                    print(f"  [{i}/{len(primary_rows)}] {row['store_name']} "
                          f"→ geocoded ({row['estimated_lat']}, {row['estimated_lng']}) [only {len(hits)} hit]")
                    await asyncio.sleep(1.1)

            # Reverse geocode the CENTROID (not probe point) for correct pincode/area
            est_lat = row.get("estimated_lat") or row["probe_lat"]
            est_lng = row.get("estimated_lng") or row["probe_lng"]
            if not row["pincode"] or not row["area"]:
                geo = await reverse_geocode(est_lat, est_lng, client)
                if not row["pincode"] and geo.get("pincode"):
                    row["pincode"] = geo["pincode"]
                if not row["area"] and geo.get("area"):
                    row["area"] = geo["area"]
                await asyncio.sleep(1.1)
    print("Geocoding done.\n")

    out = Path(__file__).parent.parent / f"zepto_bangalore_dark_stores_{ts}.xlsx"
    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "primary_stores"
    hdr1 = ["store_id", "store_name", "city", "state", "pincode", "area",
            "probe_lat", "probe_lng", "estimated_lat", "estimated_lng",
            "hit_count", "secondary_store_id", "found_by"]
    ws1.append(hdr1)
    for row in sorted(primary_rows, key=lambda x: x["pincode"] or "999999"):
        ws1.append([row[h] for h in hdr1])

    if secondary_rows:
        ws2 = wb.create_sheet("secondary_stores")
        ws2.append(["primary_store_id", "secondary_store_id", "probe_lat", "probe_lng"])
        for row in secondary_rows:
            ws2.append([row["primary_store_id"], row["secondary_store_id"],
                        row["probe_lat"], row["probe_lng"]])

    ws3 = wb.create_sheet("summary")
    ws3.append(["found_by", "count"])
    for method, count in sorted(Counter(r["found_by"] for r in primary_rows).items()):
        ws3.append([method, count])
        print(f"  {method}: {count} stores")

    wb.save(out)
    print(f"\nSaved → {out}")


if __name__ == "__main__":
    asyncio.run(main())
