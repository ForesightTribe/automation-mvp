"""
Full dark store discovery per city — pincode method FIRST, then grid.

PHASE 1 — pincode / autosuggest (the original method)
    pincode  --> /maps/place/autocomplete  --> area suggestions
             --> /maps/place/details       --> lat, lng, city, state
             --> /lms/api/v2/get_page      --> storeId
    Records the pincode, the autosuggested area text, and the coordinates that
    produced each store.  found_by = "pincode"

PHASE 2 — grid (the exhaustive sweep)
    every lat/lng node at ~670 m --> get_page --> storeId
    A store already found in phase 1 is NOT written again; it only has its
    grid hit count recorded.  A store that only the grid finds is added with
    found_by = "grid" — that is the value the grid adds over pincodes alone.

Deduplication is global: a store already present anywhere in the master
workbook is written with is_new = FALSE and already_in = <the city that owns
it>, so re-discovering it can never inflate the total.

Run from backend/:
    python -m scripts.zepto_discover_city_full --list
    python -m scripts.zepto_discover_city_full --only Ambala,Kota
    python -m scripts.zepto_discover_city_full            # every pending city
Output: dark_stores_zepto/
"""
import asyncio
import math
import re
import statistics
import sys
from pathlib import Path
from datetime import datetime
from collections import Counter, defaultdict

sys.path.insert(0, str(Path(__file__).parent.parent))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from playwright.async_api import async_playwright

import scraper.platforms.zepto.dark_store.endpoints as EP

BACKEND = Path(__file__).parent.parent
OUTDIR  = BACKEND / "dark_stores_zepto"

STEP        = 0.006
MIN_HALF    = 0.13
MARGIN      = 0.12
OUTLIER_KM  = 45
CONCURRENCY = 8
TRIES       = 3

# No city is skipped. The six already grid-scanned (Bengaluru, Kolkata, Chennai,
# Pune, Hyderabad, Ahmedabad) never had the Phase-1 pincode pass, so their old
# files lack the Pincodes sheet, area_suggestion and found_by. Re-running them
# gives all 58 cities one consistent format.
DONE = set()

_BASE = "https://www.zepto.com"
_BFF  = "https://bff-gateway.zepto.com"
_GP   = "/lms/api/v2/get_page"
_AC   = "/api/v1/maps/place/autocomplete/"
_PD   = "/api/v1/maps/place/details/"
_DROP = {"host", "content-length", "connection", "accept-encoding",
         ":method", ":path", ":authority", ":scheme"}

_API_HDRS = {
    "Accept": "application/json",
    "Origin": "https://www.zepto.com",
    "Referer": "https://www.zepto.com/",
    "User-Agent": ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                   "(KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36"),
    "app_sub_platform": "WEB", "app_version": "16.20.0",
    "platform": "WEB", "tenant": "ZEPTO",
}

# Prefixes Zepto uses that the automatic map cannot resolve, because the existing
# data has them split across several cities (the Chandigarh/Mohali/Panchkula
# tri-city sits 9-15 km apart, so the original pincode scrapes cross-contaminated
# each other). Without these, the code falls back to nearest-centroid, which sends
# four SAS- stores to Panchkula even though SAS Nagar IS Mohali.
# Attribution only — changes which city a store is filed under, never which
# stores are found, so no re-scraping is needed.
PREFIX_OVERRIDE = {
    "SAS": "Mohali",       # SAS Nagar is Mohali's official name
    "PNK": "Panchkula",
}

# Pincode-dict name in endpoints.py -> city name in the master workbook.
PIN_CITY = {
    "BANGALORE": "Bengaluru", "DELHI_NCR": "Delhi NCR", "CSN": "Chhatrapati Sambhaji Nagar",
    "HARIDWAR": "Dehradun", "SAHARANPUR": "Dehradun", "KURUKSHETRA": "Ambala",
    "VALSAD": "Surat", "HAPUR": "Delhi NCR",
}

HDR_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
HDR_FILL = PatternFill("solid", fgColor="2E4057")
HDR_ALIGN = Alignment(horizontal="center", vertical="center")
ALT_FILL = PatternFill("solid", fgColor="F0F4F8")
NEW_FILL = PatternFill("solid", fgColor="D9F7E6")
GRID_FILL = PatternFill("solid", fgColor="FFF2CC")
TITLE_FILL = PatternFill("solid", fgColor="1A3C5E")


def km(a, b, c, d):
    R = 6371.0
    p = math.radians
    return 2 * R * math.asin(math.sqrt(
        math.sin(p(c - a) / 2) ** 2 +
        math.cos(p(a)) * math.cos(p(c)) * math.sin(p(d - b) / 2) ** 2))


def pincodes_for(city: str) -> dict:
    """Merge every pincode dict that maps to this city."""
    out = {}
    for k, v in vars(EP).items():
        if not (k.endswith("_PINCODES") and isinstance(v, dict)):
            continue
        base = k[:-9]
        mapped = PIN_CITY.get(base, base.replace("_", " ").title())
        if mapped == city:
            out.update(v)
    return out


# ── Master workbook ───────────────────────────────────────────────────────────

def load_master():
    p = sorted(BACKEND.glob("zepto_master_all_cities_*.xlsx"))
    if not p:
        raise SystemExit("No master workbook found.")
    wb = openpyxl.load_workbook(p[-1], read_only=True, data_only=True)
    rows = list(wb["All Stores"].iter_rows(values_only=True))
    wb.close()
    hdr = [str(h).strip().lower() if h else "" for h in rows[2]]
    return p[-1], [dict(zip(hdr, r)) for r in rows[3:] if r and r[0]]


def build_context(recs):
    city = defaultdict(list)
    known = {}
    for r in recs:
        sid = str(r.get("store_id") or "").strip()
        if not sid:
            continue
        try:
            la, ln = float(r["lat"]), float(r["lng"])
        except (TypeError, ValueError):
            continue
        rec = {"store_id": sid, "store_name": str(r.get("store_name") or ""),
               "city": r["city"], "lat": la, "lng": ln}
        city[r["city"]].append(rec)
        known[sid] = rec
    centroid = {c: (statistics.median(p["lat"] for p in v),
                    statistics.median(p["lng"] for p in v)) for c, v in city.items()}
    pref = defaultdict(Counter)
    for r in known.values():
        m = re.match(r"^\s*([A-Za-z]{2,4})\s*-", r["store_name"])
        if m:
            pref[m.group(1).upper()][r["city"]] += 1
    prefix_city = {p: c.most_common(1)[0][0] for p, c in pref.items()
                   if c.most_common(1)[0][1] / sum(c.values()) >= 0.8}
    prefix_city.update(PREFIX_OVERRIDE)      # explicit wins over inferred
    return city, centroid, prefix_city, known


def core(c, city_stores):
    pts = city_stores[c]
    if len(pts) < 3:
        return pts
    mla = statistics.median(p["lat"] for p in pts)
    mln = statistics.median(p["lng"] for p in pts)
    return [p for p in pts if km(mla, mln, p["lat"], p["lng"]) <= OUTLIER_KM] or pts


def box_for(c, city_stores):
    pts = core(c, city_stores)
    la = [p["lat"] for p in pts]
    ln = [p["lng"] for p in pts]
    cla, cln = (min(la) + max(la)) / 2, (min(ln) + max(ln)) / 2
    hla = max((max(la) - min(la)) / 2 + MARGIN, MIN_HALF)
    hln = max((max(ln) - min(ln)) / 2 + MARGIN, MIN_HALF)
    return (round(cla - hla, 4), round(cla + hla, 4),
            round(cln - hln, 4), round(cln + hln, 4))


def build_grid(box):
    lo_la, hi_la, lo_ln, hi_ln = box
    pts, la = [], lo_la
    while la <= hi_la + 1e-9:
        ln = lo_ln
        while ln <= hi_ln + 1e-9:
            pts.append((round(la, 5), round(ln, 5)))
            ln += STEP
        la += STEP
    return pts


# ── API calls ─────────────────────────────────────────────────────────────────

async def capture_headers(page):
    cap = {}

    async def on_req(req):
        if _GP in req.url and not cap:
            cap.update(req.headers)

    page.on("request", on_req)
    try:
        await page.goto(_BASE, timeout=30000, wait_until="domcontentloaded")
        w = 0
        while not cap and w < 6000:
            await page.wait_for_timeout(250)
            w += 250
    except Exception:
        pass
    finally:
        page.remove_listener("request", on_req)
    return {k: v for k, v in cap.items() if k.lower() not in _DROP}


async def autocomplete(ctx, query):
    for a in range(1, 3):
        try:
            r = await ctx.request.get(f"{_BFF}{_AC}", params={"place_name": query},
                                      headers=_API_HDRS, timeout=12000)
            if r.status == 200:
                d = await r.json()
                return d.get("data", {}).get("predictions") or d.get("predictions") or []
        except Exception:
            pass
        await asyncio.sleep(0.4 * a)
    return []


async def place_details(ctx, place_id):
    for a in range(1, 3):
        try:
            r = await ctx.request.get(f"{_BFF}{_PD}", params={"place_id": place_id},
                                      headers=_API_HDRS, timeout=12000)
            if r.status == 200:
                d = await r.json()
                res = d.get("data", {}).get("result") or d.get("result") or {}
                loc = (res.get("geometry") or {}).get("location") or {}
                if not loc.get("lat"):
                    return None
                cty = st = ""
                for comp in res.get("address_components", []):
                    t = comp.get("types", [])
                    if not cty and ("locality" in t or "administrative_area_level_2" in t):
                        cty = comp.get("long_name", "")
                    if not st and "administrative_area_level_1" in t:
                        st = comp.get("long_name", "")
                return {"lat": loc["lat"], "lng": loc["lng"], "city": cty, "state": st,
                        "location_type": (res.get("geometry") or {}).get("location_type", "")}
        except Exception:
            pass
        await asyncio.sleep(0.4 * a)
    return None


def gp_url(lat, lng):
    return (f"{_BFF}{_GP}?latitude={lat}&longitude={lng}"
            f"&page_type=HOME&version=v2&show_new_eta_banner=true"
            f"&page_size=3&enforce_platform_type=WEB")


async def get_store(ctx, state, lat, lng):
    for a in range(1, TRIES + 1):
        try:
            r = await ctx.request.get(gp_url(lat, lng), headers=state["hdr"], timeout=15000)
            if r.status == 200:
                d = await r.json()
                svc = d.get("storeServiceableResponse") or {}
                if not svc.get("storeId"):
                    return None
                det = d.get("storeDetailsResponse")
                # Some coordinates return a storeId with serviceable=False and NO
                # storeDetailsResponse at all — no name, no city, nothing. Those are
                # not operational dark stores, so they are flagged rather than
                # silently counted as discoveries.
                return {"store_id": svc["storeId"],
                        "secondary": [str(x) for x in (svc.get("secondaryStoreIds") or [])],
                        "store_name": (det or {}).get("name", ""),
                        "serviceable": bool(svc.get("serviceable")),
                        "has_details": det is not None}
            if r.status in (401, 403, 429):
                state["stale"] = True
                await asyncio.sleep(1.2 * a)
                continue
        except Exception:
            pass
        await asyncio.sleep(0.5 * a)
    return None


def assign_city(name, lat, lng, city, prefix_city, centroid):
    m = re.match(r"^\s*([A-Za-z]{2,4})\s*-", name or "")
    if m:
        c = prefix_city.get(m.group(1).upper())
        if c:
            return c, "prefix"
    best, bd = city, 1e9
    for c, (cla, cln) in centroid.items():
        d = km(lat, lng, cla, cln)
        if d < bd:
            best, bd = c, d
    return best, "nearest"


# ── Excel ─────────────────────────────────────────────────────────────────────

COLS = ["store_id", "store_name", "city", "state", "area", "source_pincode",
        "area_suggestion", "lat", "lng", "found_by", "is_new", "already_in",
        "note", "serviceable", "store_record", "assigned_by",
        "pincode_hits", "grid_hits", "secondary"]

PIN_COLS = ["pincode", "pincode_label", "suggestions", "coords_resolved",
            "in_box", "stores_found", "store_names"]


def widths(ws, mx=46):
    for col in ws.columns:
        cl = get_column_letter(col[0].column)
        ml = max((len(str(c.value)) for c in col[:400] if c.value), default=0)
        ws.column_dimensions[cl].width = min(max(ml + 2, 9), mx)


def head(ws, cols):
    ws.append(cols)
    for c in ws[1]:
        c.font, c.fill, c.alignment = HDR_FONT, HDR_FILL, HDR_ALIGN
    ws.row_dimensions[1].height = 18
    ws.freeze_panes = "A2"


def banner(ws, t, s):
    ws.insert_rows(1, amount=2)
    mc = ws.max_column or 1
    ws["A1"], ws["A2"] = t, s
    ws["A1"].font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="444444")
    ws["A2"].fill = ALT_FILL
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 15
    if mc > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=mc)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=mc)


def write_city(city, found, pin_rows, box, stats):
    OUTDIR.mkdir(exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    slug = re.sub(r"[^a-z0-9]+", "_", city.lower())[:40]
    real = [v for v in found.values() if v.get("store_record") != "NO STORE RECORD"]
    ghost = [v for v in found.values() if v.get("store_record") == "NO STORE RECORD"]
    new = [v for v in real if v["is_new"]]
    new_ghost = [v for v in ghost if v["is_new"]]
    by_pin = [v for v in found.values() if v["found_by"] == "pincode"]
    by_grid = [v for v in found.values() if v["found_by"] == "grid"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stores"
    head(ws, COLS)
    order = {"pincode": 0, "grid": 1}
    for i, r in enumerate(sorted(found.values(),
                                 key=lambda x: (not x["is_new"], order.get(x["found_by"], 2),
                                                str(x["source_pincode"]), x["store_name"])),
                          start=2):
        ws.append([r.get(c, "") for c in COLS])
        fill = NEW_FILL if r["is_new"] else (GRID_FILL if r["found_by"] == "grid" else
                                             (ALT_FILL if i % 2 == 0 else None))
        if fill:
            for c in range(1, len(COLS) + 1):
                ws.cell(row=i, column=c).fill = fill
    ws.auto_filter.ref = ws.dimensions
    widths(ws)

    ws2 = wb.create_sheet("Pincodes")
    head(ws2, PIN_COLS)
    for i, r in enumerate(sorted(pin_rows, key=lambda x: str(x["pincode"])), start=2):
        ws2.append([r.get(c, "") for c in PIN_COLS])
        if not r["stores_found"]:
            for c in range(1, len(PIN_COLS) + 1):
                ws2.cell(row=i, column=c).fill = PatternFill("solid", fgColor="FDE8E8")
        elif i % 2 == 0:
            for c in range(1, len(PIN_COLS) + 1):
                ws2.cell(row=i, column=c).fill = ALT_FILL
    ws2.auto_filter.ref = ws2.dimensions
    widths(ws2)

    ws3 = wb.create_sheet("Summary")
    head(ws3, ["Metric", "Value"])
    rows = [("City", city),
            ("Box", f"lat {box[0]}-{box[1]}, lng {box[2]}-{box[3]}"),
            ("", ""),
            ("PHASE 1 - pincode method", ""),
            ("  pincodes tried", stats["pins"]),
            ("  autosuggest predictions", stats["preds"]),
            ("  coordinates resolved", stats["coords"]),
            ("  coordinates inside box", stats["inbox"]),
            ("  stores found", len(by_pin)),
            ("", ""),
            ("PHASE 2 - grid", ""),
            ("  grid points", stats["gpts"]),
            ("  points with no store", stats["gno"]),
            ("  stores found ONLY by grid", len(by_grid)),
            ("", ""),
            ("TOTAL unique storeIds hit", len(found)),
            ("  real stores (have a store record)", len(real)),
            ("  no store record (serviceable=False)", len(ghost)),
            ("", ""),
            ("NEW real stores (count these)", len(new)),
            ("NEW with no store record (ignore)", len(new_ghost)),
            ("Already known - this city", stats["mine"]),
            ("Already known - ANOTHER city", stats["other"])]
    for c, n in stats["other_by"].most_common():
        rows.append((f"    already counted in {c}", n))
    rows += [("", ""), ("Scraped at", datetime.now().strftime("%d-%b-%Y %H:%M"))]
    for i, (m, v) in enumerate(rows, start=2):
        ws3.cell(row=i, column=1, value=m)
        # section headers carry no value; write blank, not the string "None"
        ws3.cell(row=i, column=2, value="" if v in (None, "") else v)
        if str(m).startswith("PHASE") or str(m).startswith("TOTAL"):
            for c in (1, 2):
                ws3.cell(row=i, column=c).font = Font(name="Calibri", size=10, bold=True)
    widths(ws3)

    T = f"{city} — dark store discovery — {len(new)} new"
    S = (f"Phase 1 pincode: {len(by_pin)} stores from {stats['pins']} pincodes  |  "
         f"Phase 2 grid: +{len(by_grid)} more  |  green = new, yellow = grid-only")
    for w in (ws, ws2, ws3):
        banner(w, T, S)
    out = OUTDIR / f"darkstores_{slug}_{ts}.xlsx"
    wb.save(out)
    return out, new, by_pin, by_grid


# ── Per-city run ──────────────────────────────────────────────────────────────

async def run_city(ctx, state, city, city_stores, centroid, prefix_city, known):
    box = box_for(city, city_stores)
    pins = pincodes_for(city)
    found, pin_rows = {}, []
    stats = {"pins": len(pins), "preds": 0, "coords": 0, "inbox": 0,
             "gpts": 0, "gno": 0, "mine": 0, "other": 0, "other_by": Counter()}

    def record(sid, name, lat, lng, how, pin="", sugg="", det=None, sec="",
               serviceable=True, has_details=True):
        prev = known.get(sid)
        c, by = assign_city(name, lat, lng, city, prefix_city, centroid)
        found[sid] = {
            "serviceable": "YES" if serviceable else "NO",
            "store_record": "YES" if has_details else "NO STORE RECORD",
            "store_id": sid, "store_name": name, "city": c,
            "state": (det or {}).get("state", ""), "area": (det or {}).get("area", ""),
            "source_pincode": pin, "area_suggestion": sugg,
            "lat": lat, "lng": lng, "found_by": how,
            "is_new": prev is None,
            "already_in": "" if prev is None else prev["city"],
            "note": "" if prev is None else f"already counted in {prev['city']}",
            "assigned_by": by, "pincode_hits": 1 if how == "pincode" else 0,
            "grid_hits": 0 if how == "pincode" else 1, "secondary": sec,
        }

    # ── PHASE 1 ──────────────────────────────────────────────────────────────
    print(f"    PHASE 1 pincode method — {len(pins)} pincodes")
    sem = asyncio.Semaphore(CONCURRENCY)
    lock = asyncio.Lock()

    async def do_pin(pin, label):
        async with sem:
            preds = await autocomplete(ctx, pin)
            if label:
                preds += await autocomplete(ctx, f"{label} {pin}")
        seen_pid, names, coords, inbox, hits = set(), [], 0, 0, []
        for p in preds:
            pid = p.get("place_id")
            if not pid or pid in seen_pid:
                continue
            seen_pid.add(pid)
            desc = p.get("description") or ""
            names.append(desc)
            async with sem:
                det = await place_details(ctx, pid)
            if not det:
                continue
            coords += 1
            la, ln = det["lat"], det["lng"]
            if not (box[0] <= la <= box[1] and box[2] <= ln <= box[3]):
                continue
            inbox += 1
            async with sem:
                r = await get_store(ctx, state, la, ln)
            if not r:
                continue
            async with lock:
                sid = r["store_id"]
                if sid in found:
                    found[sid]["pincode_hits"] += 1
                else:
                    record(sid, r["store_name"], la, ln, "pincode", pin, desc,
                           {"state": det["state"], "area": desc},
                           ",".join(r["secondary"]),
                           r.get("serviceable", True), r.get("has_details", True))
                    hits.append(r["store_name"])
        async with lock:
            stats["preds"] += len(seen_pid)
            stats["coords"] += coords
            stats["inbox"] += inbox
            pin_rows.append({"pincode": pin, "pincode_label": label,
                             "suggestions": len(seen_pid), "coords_resolved": coords,
                             "in_box": inbox, "stores_found": len(hits),
                             "store_names": " | ".join(hits)})

    items = list(pins.items())
    for s in range(0, len(items), 40):
        await asyncio.gather(*(do_pin(p, l) for p, l in items[s:s + 40]))
        if state["stale"]:
            nh = await capture_headers(state["page"])
            if nh:
                state["hdr"] = nh
            state["stale"] = False
    n_pin = len(found)
    print(f"      -> {n_pin} stores from pincodes "
          f"({stats['preds']} suggestions, {stats['inbox']} coords in box)")

    # ── PHASE 2 ──────────────────────────────────────────────────────────────
    grid = build_grid(box)
    stats["gpts"] = len(grid)
    print(f"    PHASE 2 grid — {len(grid)} points @ {STEP} deg")
    done = {"n": 0}

    async def do_pt(la, ln):
        async with sem:
            r = await get_store(ctx, state, la, ln)
        async with lock:
            done["n"] += 1
            if not r:
                stats["gno"] += 1
                return
            sid = r["store_id"]
            if sid in found:
                found[sid]["grid_hits"] += 1     # already had it from pincodes
            else:
                record(sid, r["store_name"] or known.get(sid, {}).get("store_name", ""),
                       la, ln, "grid", sec=",".join(r["secondary"]),
                       serviceable=r.get("serviceable", True),
                       has_details=r.get("has_details", True))
                tag = "NEW " if found[sid]["is_new"] else "known"
                print(f"      grid-only {tag} {sid[:8]}  "
                      f"{found[sid]['store_name'][:28]:<29} -> {found[sid]['city']}")

    for s in range(0, len(grid), 300):
        await asyncio.gather(*(do_pt(a, b) for a, b in grid[s:s + 300]))
        if state["stale"]:
            nh = await capture_headers(state["page"])
            if nh:
                state["hdr"] = nh
            state["stale"] = False

    for v in found.values():
        if v["already_in"]:
            if v["already_in"] == city:
                stats["mine"] += 1
            else:
                stats["other"] += 1
                stats["other_by"][v["already_in"]] += 1

    return found, pin_rows, box, stats


async def main():
    args = sys.argv[1:]
    only = ({x.strip() for x in args[args.index("--only") + 1].split(",")}
            if "--only" in args else None)
    list_only = "--list" in args

    path, recs = load_master()
    city_stores, centroid, prefix_city, known = build_context(recs)
    targets = sorted(only) if only else sorted(c for c in city_stores if c not in DONE)

    plan = []
    for c in targets:
        if c not in city_stores:
            print(f"  [skip] {c} not in master")
            continue
        plan.append((c, len(pincodes_for(c)), len(build_grid(box_for(c, city_stores)))))
    plan.sort(key=lambda x: -(x[2]))

    print(f"Master   : {path.name}")
    print(f"Output   : {OUTDIR}")
    print(f"Dedup    : {len(known)} known store ids across {len(city_stores)} cities")
    print(f"Cities   : {len(plan)}")
    print("=" * 84)
    print(f"{'city':<32}{'pincodes':>10}{'grid pts':>10}{'est min':>10}")
    print("-" * 84)
    tp = tg = 0.0
    for c, np_, ng in plan:
        est = (np_ * 9 * 0.33 + ng * 0.33) / CONCURRENCY / 60
        tp += np_
        tg += ng
        print(f"{c[:31]:<32}{np_:>10}{ng:>10}{est:>10.1f}")
    print("-" * 84)
    tot = (tp * 9 * 0.33 + tg * 0.33) / CONCURRENCY / 60
    print(f"{'TOTAL':<32}{int(tp):>10}{int(tg):>10}{tot:>10.0f}")
    if list_only:
        print("\n--list given: nothing scanned.")
        return
    print()

    grand = []
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        ctx = await browser.new_context(user_agent=_API_HDRS["User-Agent"])
        page = await ctx.new_page()
        state = {"hdr": {}, "stale": False, "page": page}
        print("Capturing session headers...")
        state["hdr"] = await capture_headers(page)
        if not state["hdr"]:
            print("Header capture failed — aborting.")
            await browser.close()
            return
        print(f"Captured {len(state['hdr'])} headers.\n")

        for i, (city, _, _) in enumerate(plan, 1):
            print(f"[{i}/{len(plan)}] {city}")
            found, pin_rows, box, stats = await run_city(
                ctx, state, city, city_stores, centroid, prefix_city, known)
            out, new, bp, bg = write_city(city, found, pin_rows, box, stats)
            grand.extend(new)
            print(f"    -> {len(found)} stores | pincode {len(bp)} | grid-only {len(bg)} "
                  f"| NEW {len(new)}")
            print(f"    saved {out.name}\n")

        await browser.close()

    print("=" * 84)
    print(f"Cities scanned   : {len(plan)}")
    print(f"TOTAL NEW stores : {len(grand)}")
    for c, n in Counter(v["city"] for v in grand).most_common():
        print(f"   {c:<30} +{n}")
    print(f"\nFiles in: {OUTDIR}")


if __name__ == "__main__":
    asyncio.run(main())
