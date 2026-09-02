"""Zepto supply-chain report: fill-rate analysis plus the raw PO/ASN/GRN rows.

Answers "how much of what Zepto ordered actually got received, and where does
the loss happen" — a question `zepto_po` alone cannot answer, because Zepto
never populates its `total_asn_qty` / `total_grn_qty` columns (NULL on all 379
rows as of 2026-08-31). The real figures come from `zepto_grn`, which carries
`po_qty` and `grn_qty` side by side, and from `zepto_asn`, which carries the
shipped quantity that sits between them.

Analysis sheets first, then the raw tables behind them so the working is visible.

Run from backend/:
    python -m scripts.zepto_supply_report -t <tenant_id> --from 2026-04-01 --to 2026-08-31
"""
import argparse
import asyncio
import uuid
from datetime import date, datetime
from pathlib import Path

import openpyxl
from dotenv import load_dotenv
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import text

load_dotenv(Path(__file__).parent.parent / ".env")

from app.core.database import AsyncSessionLocal  # noqa: E402

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(color="FFFFFF", bold=True)

# Plumbing columns, dropped from every sheet — same set the other exports strip.
_STRIP = {"id", "upsert_key", "tenant_id", "scrape_job_id", "platform"}

# Each query is scoped by :t and the :start/:end window, applied on the table's
# OWN date column — so a GRN counts in the month it was received, not the month
# its PO was raised. Those differ for 94% of deliveries.
_ANALYSIS: list[tuple[str, str]] = [
    # Ordered/shipped/received come from ONE matched set (ASN joined to GRN), not
    # from three separately-scoped queries — those select slightly different
    # deliveries, which made "shipped" read higher than "ordered".
    ("Summary", """
        select 'POs raised' metric, count(*)::text value from zepto_po
          where tenant_id=:t and po_date between :start and :end
        union all select 'Matched deliveries (ASN+GRN)', count(*)::text
          from zepto_asn a join zepto_grn g on g.po_id=a.po_id and g.tenant_id=a.tenant_id
          where a.tenant_id=:t and g.grn_date between :start and :end
        union all select 'Units ordered', sum(a.po_qty)::text
          from zepto_asn a join zepto_grn g on g.po_id=a.po_id and g.tenant_id=a.tenant_id
          where a.tenant_id=:t and g.grn_date between :start and :end
        union all select 'Units shipped (ASN)', sum(a.asn_qty)::text
          from zepto_asn a join zepto_grn g on g.po_id=a.po_id and g.tenant_id=a.tenant_id
          where a.tenant_id=:t and g.grn_date between :start and :end
        union all select 'Units received (GRN)', sum(g.grn_qty)::text
          from zepto_asn a join zepto_grn g on g.po_id=a.po_id and g.tenant_id=a.tenant_id
          where a.tenant_id=:t and g.grn_date between :start and :end
        union all select 'Shipped as % of ordered',
          round(100.0*sum(a.asn_qty)/nullif(sum(a.po_qty),0),1)::text
          from zepto_asn a join zepto_grn g on g.po_id=a.po_id and g.tenant_id=a.tenant_id
          where a.tenant_id=:t and g.grn_date between :start and :end
        union all select 'Accepted as % of shipped',
          round(100.0*sum(g.grn_qty)/nullif(sum(a.asn_qty),0),1)::text
          from zepto_asn a join zepto_grn g on g.po_id=a.po_id and g.tenant_id=a.tenant_id
          where a.tenant_id=:t and g.grn_date between :start and :end
        union all select 'Units short', (sum(a.asn_qty)-sum(g.grn_qty))::text
          from zepto_asn a join zepto_grn g on g.po_id=a.po_id and g.tenant_id=a.tenant_id
          where a.tenant_id=:t and g.grn_date between :start and :end
        union all select 'Expired POs', count(*)::text from zepto_po
          where tenant_id=:t and status='EXPIRED' and po_date between :start and :end
        union all select 'Expired PO value', round(sum(total_value)::numeric)::text from zepto_po
          where tenant_id=:t and status='EXPIRED' and po_date between :start and :end
    """),
    ("Fill by month", """
        select to_char(grn_date,'YYYY-MM') grn_month, count(*) receipts,
               sum(po_qty) ordered, sum(grn_qty) received,
               sum(po_qty)-sum(grn_qty) short,
               round(100.0*sum(grn_qty)/nullif(sum(po_qty),0),1) fill_pct
        from zepto_grn where tenant_id=:t and grn_date between :start and :end
        group by 1 order by 1
    """),
    ("Fill by dock time", """
        select (g.grn_date - a.asn_date) days_on_dock, count(*) deliveries,
               sum(a.asn_qty) shipped, sum(g.grn_qty) received,
               round(100.0*sum(g.grn_qty)/nullif(sum(a.asn_qty),0),1) accepted_pct
        from zepto_asn a join zepto_grn g
          on g.po_id = a.po_id and g.tenant_id = a.tenant_id
        where a.tenant_id=:t and g.grn_date between :start and :end
        group by 1 order by 1
    """),
    ("Fill by warehouse", """
        select location, count(*) receipts, sum(po_qty) ordered, sum(grn_qty) received,
               sum(po_qty)-sum(grn_qty) short,
               round(100.0*sum(grn_qty)/nullif(sum(po_qty),0),1) fill_pct
        from zepto_grn where tenant_id=:t and grn_date between :start and :end
        group by 1 order by short desc
    """),
    ("Fill by SKU", """
        select i.sku_name, sum(i.po_qty) ordered, sum(i.grn_qty) received,
               sum(i.po_qty)-sum(coalesce(i.grn_qty,0)) short,
               round(100.0*sum(coalesce(i.grn_qty,0))/nullif(sum(i.po_qty),0),1) fill_pct
        from zepto_po_items i join zepto_po p
          on p.po_id = i.po_id and p.tenant_id = i.tenant_id
        where i.tenant_id=:t and i.grn_qty is not null
          and p.po_date between :start and :end
        group by 1 order by short desc
    """),
    ("Expired POs", """
        select po_id, po_date, expiry_date, city, location, items_count, total_qty,
               round(total_value::numeric) value,
               (select count(*) from zepto_asn a
                 where a.po_id = p.po_id and a.tenant_id = p.tenant_id) asn_sent
        from zepto_po p
        where tenant_id=:t and status='EXPIRED' and po_date between :start and :end
        order by po_date
    """),
    ("Open POs", """
        select po_id, po_date, expiry_date, status, city, location, total_qty,
               round(total_value::numeric) value
        from zepto_po
        where tenant_id=:t and status not in ('COMPLETED','EXPIRED')
          and po_date between :start and :end
        order by po_date
    """),
]

_RAW: list[tuple[str, str, str | None]] = [
    ("Raw POs", "zepto_po", "po_date"),
    ("Raw PO Items", "zepto_po_items", None),
    ("Raw ASNs", "zepto_asn", "asn_date"),
    ("Raw GRNs", "zepto_grn", "grn_date"),
]


def _to_header(key: str) -> str:
    return key.replace("_", " ").title()


def _clean(v):
    if isinstance(v, datetime):
        return v.replace(tzinfo=None)
    if isinstance(v, uuid.UUID):
        return str(v)
    return v


def _write(ws, rows: list[dict]) -> None:
    if not rows:
        ws.append(["(no data for this window)"])
        return
    keys = [k for k in rows[0] if k not in _STRIP]
    for c, k in enumerate(keys, 1):
        cell = ws.cell(row=1, column=c, value=_to_header(k))
        cell.fill, cell.font = _HEADER_FILL, _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    for r, row in enumerate(rows, 2):
        for c, k in enumerate(keys, 1):
            ws.cell(row=r, column=c, value=_clean(row.get(k)))
    for c, k in enumerate(keys, 1):
        width = max(
            len(_to_header(k)),
            max((len(str(row.get(k) or "")) for row in rows), default=0),
        )
        ws.column_dimensions[get_column_letter(c)].width = min(width + 2, 55)
    ws.freeze_panes = "A2"


async def build(tenant_id: str, start: date, end: date, out: str) -> None:
    params = {"t": uuid.UUID(tenant_id), "start": start, "end": end}
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    async with AsyncSessionLocal() as db:
        for name, sql in _ANALYSIS:
            rows = [dict(r) for r in (await db.execute(text(sql), params)).mappings()]
            _write(wb.create_sheet(name), rows)
            print(f"  {name:24} {len(rows):>5} rows")

        for name, table, date_col in _RAW:
            if date_col is None:
                # No date of its own — scope through the parent PO.
                sql = (
                    "select i.* from zepto_po_items i join zepto_po p "
                    "on p.po_id=i.po_id and p.tenant_id=i.tenant_id "
                    "where i.tenant_id=:t and p.po_date between :start and :end "
                    "order by p.po_date, i.sku_name"
                )
            else:
                sql = (
                    f"select * from {table} where tenant_id=:t "
                    f"and {date_col} between :start and :end order by {date_col}"
                )
            rows = [dict(r) for r in (await db.execute(text(sql), params)).mappings()]
            _write(wb.create_sheet(name), rows)
            print(f"  {name:24} {len(rows):>5} rows")

    wb.save(out)
    print(f"\nSaved {out}")


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("-t", "--tenant", required=True, help="Tenant ID")
    ap.add_argument("--from", dest="start", required=True, help="Start date YYYY-MM-DD")
    ap.add_argument("--to", dest="end", required=True, help="End date YYYY-MM-DD")
    ap.add_argument("-o", "--output", help="Output .xlsx")
    a = ap.parse_args()
    out = a.output or f"zepto_supply_{a.start}_to_{a.end}.xlsx"
    print(f"Building Zepto supply-chain report {a.start}..{a.end}")
    asyncio.run(
        build(a.tenant, date.fromisoformat(a.start), date.fromisoformat(a.end), out)
    )


if __name__ == "__main__":
    main()
