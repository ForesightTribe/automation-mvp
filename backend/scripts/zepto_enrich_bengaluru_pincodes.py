"""
Add pincode + area name to the Bengaluru grid scan output.

Two sources, in priority order:
  1. The combined workbook — for stores it already knows, `source_pincode` and
     `area` are authoritative (that pincode is literally what the scraper searched
     to find the store). No network call needed.
  2. Reverse geocoding — for grid-discovered stores there is no source pincode,
     only a coordinate, so the coordinate is resolved via OpenStreetMap Nominatim.
     Nominatim's usage policy requires a real User-Agent and max 1 request/second;
     both are honoured below.

Anything that still fails to resolve is written as "UNRESOLVED" rather than
guessed, so the output never implies precision it does not have.

Run from backend/:
    python -m scripts.zepto_enrich_bengaluru_pincodes
Output: zepto_bengaluru_stores_with_pincode_TIMESTAMP.xlsx
"""
import sys
import time
from pathlib import Path
from datetime import datetime

sys.path.insert(0, str(Path(__file__).parent.parent))

import httpx
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BACKEND   = Path(__file__).parent.parent
COMBINED  = BACKEND / "zepto_all_stores_combined_20260731_1147.xlsx"
CITY_SHEET = "Bangalore"

NOMINATIM = "https://nominatim.openstreetmap.org/reverse"
# Nominatim requires an identifying User-Agent; anonymous requests get blocked.
UA = "zepto-darkstore-research/1.0 (contact: soumi97bag@gmail.com)"
RATE_LIMIT_S = 1.1          # policy is max 1 req/sec — stay just under
GEO_TRIES = 3

HDR_FONT  = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
HDR_FILL  = PatternFill("solid", fgColor="2E4057")
HDR_ALIGN = Alignment(horizontal="center", vertical="center")
ALT_FILL  = PatternFill("solid", fgColor="F0F4F8")
NEW_FILL  = PatternFill("solid", fgColor="D9F7E6")
SPILL_FILL = PatternFill("solid", fgColor="FFF4CE")
BAD_FILL  = PatternFill("solid", fgColor="FDE8E8")
TITLE_FILL = PatternFill("solid", fgColor="1A3C5E")


def latest_grid_file() -> Path:
    files = sorted(BACKEND.glob("zepto_bengaluru_grid_scan_*.xlsx"))
    if not files:
        raise SystemExit("No zepto_bengaluru_grid_scan_*.xlsx found — run the grid scan first.")
    return files[-1]


def read_grid(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows = list(wb["Grid Stores"].iter_rows(values_only=True))
    wb.close()
    hdr = [str(h).strip().lower() if h else "" for h in rows[2]]   # rows 0-1 banner
    return [dict(zip(hdr, r)) for r in rows[3:] if r and r[0]]


def read_known() -> dict[str, dict]:
    """store_id -> {source_pincode, area, city, state} from EVERY city sheet."""
    wb = openpyxl.load_workbook(COMBINED, read_only=True, data_only=True)
    known: dict[str, dict] = {}
    for sheet in wb.sheetnames:
        if sheet in ("All Stores", "Summary"):
            continue
        rows = list(wb[sheet].iter_rows(values_only=True))
        if not rows:
            continue
        hdr = [str(h).strip().lower() if h else "" for h in rows[0]]
        if "store_id" not in hdr:
            continue
        for r in rows[1:]:
            if not r or not r[0]:
                continue
            d = dict(zip(hdr, r))
            sid = str(d.get("store_id") or "").strip()
            if sid and sid not in known:
                try:
                    la, ln = float(d.get("lat") or 0), float(d.get("lng") or 0)
                except (TypeError, ValueError):
                    la = ln = 0
                known[sid] = {
                    "pincode": str(d.get("source_pincode") or "").strip(),
                    "area":    str(d.get("area") or "").strip(),
                    "city":    str(d.get("city") or "").strip(),
                    "state":   str(d.get("state") or "").strip(),
                    "lat": la, "lng": ln,
                    "sheet":   sheet,
                }
    wb.close()
    return known


def reverse_geocode(client: httpx.Client, lat: float, lng: float) -> dict | None:
    """Coordinate -> {pincode, area, city, state} via Nominatim."""
    for attempt in range(1, GEO_TRIES + 1):
        try:
            r = client.get(
                NOMINATIM,
                params={"lat": lat, "lon": lng, "format": "jsonv2",
                        "addressdetails": 1, "zoom": 18},
                headers={"User-Agent": UA, "Accept-Language": "en"},
                timeout=20,
            )
            if r.status_code == 200:
                a = (r.json() or {}).get("address") or {}
                # Nominatim varies which key holds the locality — take the first present.
                area = next((a[k] for k in
                             ("neighbourhood", "suburb", "village", "town",
                              "city_district", "residential", "quarter", "hamlet")
                             if a.get(k)), "")
                return {
                    "pincode": (a.get("postcode") or "").replace(" ", ""),
                    "area":    area,
                    "city":    a.get("city") or a.get("town") or a.get("state_district") or "",
                    "state":   a.get("state") or "",
                }
            if r.status_code in (429, 503):
                time.sleep(2.0 * attempt)
                continue
        except Exception:
            pass
        time.sleep(1.0 * attempt)
    return None


def widths(ws):
    for col in ws.columns:
        cl = get_column_letter(col[0].column)
        ml = max((len(str(c.value)) for c in col if c.value), default=0)
        ws.column_dimensions[cl].width = min(max(ml + 2, 9), 52)


def main():
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    scraped_at = datetime.now().strftime("%d-%b-%Y %H:%M")

    grid_file = latest_grid_file()
    stores = read_grid(grid_file)
    known = read_known()

    print(f"Bengaluru store enrichment — pincode + area")
    print(f"Source grid file : {grid_file.name}")
    print(f"Stores           : {len(stores)}")
    print("=" * 74)

    # The grid file's lat/lng is the LATTICE NODE where the store was first hit —
    # an arbitrary point in its ~3 km catchment, not the store's location. The
    # combined file's coordinate came from place/details for a real named place,
    # so it is the better value and is what the rest of the dataset already uses.
    # Prefer it whenever it exists; keep the grid node only as a fallback.
    from_known, need_geo = [], []
    for s in stores:
        sid = str(s.get("store_id") or "").strip()
        s["grid_lat"], s["grid_lng"] = s.get("lat"), s.get("lng")
        k = known.get(sid)
        if k and k["pincode"]:
            s.update(pincode=k["pincode"], area=k["area"],
                     city=k["city"], state=k["state"], pincode_source="scraper")
            if k["lat"] and k["lng"]:
                s["lat"], s["lng"] = k["lat"], k["lng"]
                s["coord_source"] = "place_details"
            else:
                s["coord_source"] = "grid_node"
            from_known.append(s)
        else:
            s["coord_source"] = "grid_node"
            need_geo.append(s)

    print(f"Pincode from combined file : {len(from_known)}")
    print(f"Need reverse geocoding     : {len(need_geo)}")
    print(f"Estimated geocode time     : ~{len(need_geo) * RATE_LIMIT_S / 60:.1f} min\n")

    resolved = failed = 0
    with httpx.Client() as client:
        for i, s in enumerate(need_geo, 1):
            lat, lng = float(s["lat"]), float(s["lng"])
            g = reverse_geocode(client, lat, lng)
            name = str(s.get("store_name") or "")[:30]
            if g and g["pincode"]:
                s.update(pincode=g["pincode"], area=g["area"] or str(s.get("store_name") or ""),
                         city=g["city"], state=g["state"], pincode_source="reverse_geocode")
                resolved += 1
                print(f"  [{i:>2}/{len(need_geo)}] {name:<30} -> {g['pincode']}  {g['area'][:26]}")
            else:
                s.update(pincode="UNRESOLVED", area=g["area"] if g else "",
                         city="", state="", pincode_source="failed")
                failed += 1
                print(f"  [{i:>2}/{len(need_geo)}] {name:<30} -> UNRESOLVED")
            time.sleep(RATE_LIMIT_S)

    print(f"\nGeocoded ok : {resolved}   failed : {failed}")

    # ── Excel ────────────────────────────────────────────────────────────────
    COLS = ["store_id", "store_name", "pincode", "area", "city", "state",
            "lat", "lng", "coord_source", "grid_lat", "grid_lng",
            "is_new", "truly_new", "known_city",
            "hit_count", "pincode_source", "secondary"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bengaluru Stores"
    ws.append(COLS)
    for cell in ws[1]:
        cell.font, cell.fill, cell.alignment = HDR_FONT, HDR_FILL, HDR_ALIGN
    ws.row_dimensions[1].height = 18
    ws.freeze_panes = "A2"

    def sort_key(s):
        p = str(s.get("pincode") or "")
        return (p == "UNRESOLVED", p, str(s.get("store_name") or ""))

    for i, s in enumerate(sorted(stores, key=sort_key), start=2):
        ws.append([s.get(c, "") for c in COLS])
        if s.get("pincode") == "UNRESOLVED":
            fill = BAD_FILL
        elif str(s.get("known_city") or ""):
            fill = SPILL_FILL
        elif str(s.get("truly_new") or "").upper() in ("TRUE", "YES", "1"):
            fill = NEW_FILL
        else:
            fill = ALT_FILL if i % 2 == 0 else None
        if fill:
            for c in range(1, len(COLS) + 1):
                ws.cell(row=i, column=c).fill = fill
    ws.auto_filter.ref = ws.dimensions
    widths(ws)

    # By-pincode rollup
    ws2 = wb.create_sheet("By Pincode")
    ws2.append(["pincode", "area", "store_count", "store_names"])
    for cell in ws2[1]:
        cell.font, cell.fill, cell.alignment = HDR_FONT, HDR_FILL, HDR_ALIGN
    ws2.freeze_panes = "A2"
    bypin: dict[str, list[dict]] = {}
    for s in stores:
        bypin.setdefault(str(s.get("pincode") or "UNRESOLVED"), []).append(s)
    for i, (pin, group) in enumerate(sorted(bypin.items()), start=2):
        ws2.append([pin,
                    group[0].get("area", ""),
                    len(group),
                    " | ".join(str(g.get("store_name") or "") for g in group)])
        if i % 2 == 0:
            for c in range(1, 5):
                ws2.cell(row=i, column=c).fill = ALT_FILL
    widths(ws2)

    # Banner
    for w in (ws, ws2):
        w.insert_rows(1, amount=2)
        mc = w.max_column or 1
        w["A1"] = f"Bengaluru Dark Stores with Pincode + Area — {len(stores)} stores"
        w["A2"] = (f"Generated: {scraped_at}  |  source: {grid_file.name}  |  "
                   f"{len(from_known)} pincodes from scraper data, {resolved} reverse-geocoded, "
                   f"{failed} unresolved")
        w["A1"].font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
        w["A1"].fill = TITLE_FILL
        w["A1"].alignment = Alignment(horizontal="left", vertical="center")
        w.row_dimensions[1].height = 22
        w["A2"].font = Font(name="Calibri", size=10, italic=True, color="444444")
        w["A2"].fill = ALT_FILL
        w["A2"].alignment = Alignment(horizontal="left", vertical="center")
        w.row_dimensions[2].height = 15
        if mc > 1:
            w.merge_cells(start_row=1, start_column=1, end_row=1, end_column=mc)
            w.merge_cells(start_row=2, start_column=1, end_row=2, end_column=mc)

    out = BACKEND / f"zepto_bengaluru_stores_with_pincode_{ts}.xlsx"
    wb.save(out)
    print(f"\nUnique pincodes : {len([p for p in bypin if p != 'UNRESOLVED'])}")
    print(f"Saved -> {out}")


if __name__ == "__main__":
    main()
