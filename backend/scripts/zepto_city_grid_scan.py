"""
Generic Zepto dark store GRID SCAN — exhaustive store discovery for any city.

Why a grid: the pincode method is only as good as Zepto's autocomplete. Where
suggestions are sparse, stores go undiscovered. A grid asks get_page "which store
serves this exact point?" at every node, so nothing can hide between samples.
On Bengaluru this lifted 134 -> 169 confirmed stores (+26%).

Speed: probing via page.reload() costs ~3 s/point. Instead we capture the session
headers from the page's OWN get_page request once, then replay them through
ctx.request — ~0.33 s/point, verified correct against known store coordinates.
Headers are re-captured automatically if the session goes stale (401/403/429).

Dedup is against EVERY city sheet, not just the target city's, so a store
belonging to a neighbouring city (border spill) is never miscounted as new.

Run from backend/:
    python -m scripts.zepto_city_grid_scan kolkata
    python -m scripts.zepto_city_grid_scan bengaluru
    python -m scripts.zepto_city_grid_scan            # lists available cities
Output: zepto_<city>_grid_scan_TIMESTAMP.xlsx
"""
import asyncio
import sys
from pathlib import Path
from datetime import datetime

sys.path.insert(0, str(Path(__file__).parent.parent))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from playwright.async_api import async_playwright

# ── City configs ──────────────────────────────────────────────────────────────
# `sheet` must match the tab name in the combined workbook.
# Boxes are set generously BEYOND the known store span, because the pincode
# method may have been cut off on any edge — the run reports edge proximity so
# you can tell whether the box needs widening.
CITIES = {
    "bengaluru": {
        "sheet": "Bangalore",
        "label": "Bengaluru",
        # known span: lat 12.801-13.117, lng 77.490-77.783
        "lat": (12.58, 13.45),
        "lng": (77.25, 77.95),
    },
    "kolkata": {
        "sheet": "Kolkata",
        "label": "Kolkata",
        # known span: lat 22.413-22.767, lng 88.241-88.505
        # widened north (Naihati/Kalyani), south (Budge Budge/Falta),
        # west (Howrah/Uluberia), east (Rajarhat/New Town fringe)
        "lat": (22.25, 23.00),
        "lng": (88.05, 88.65),
    },
    "chennai": {
        "sheet": "Chennai",
        "label": "Chennai",
        # known span: lat 12.790-13.200, lng 80.062-80.293
        # south widened to 12.60: the old 12.75 left only 4.4 km below the
        # southernmost store (Kelambakkam) and the OMR corridor runs on past it
        # (Thiruporur / Mahabalipuram). West widened for Sriperumbudur /
        # Poonamallee. East stays at 80.35 — beyond that is the Bay of Bengal.
        "lat": (12.60, 13.40),
        "lng": (79.85, 80.35),
    },
    "hyderabad": {
        "sheet": "Hyderabad",
        "label": "Hyderabad",
        # known span: lat 17.302-17.572, lng 78.284-78.643
        # No edge was strictly cut off, but north/east/west margins were only
        # 6-9 km and Hyderabad sprawls past the ORR. Widened north for Medchal /
        # Shamirpet, south for Shamshabad airport / Adibatla, east for Ghatkesar /
        # Pocharam, west for Patancheru. No other city within 150 km.
        "lat": (17.18, 17.72),
        "lng": (78.15, 78.80),
    },
    "mumbai": {
        "sheet": "Mumbai",
        "label": "Mumbai",
        "lat": (18.85, 19.55),
        "lng": (72.75, 73.15),
    },
    "pune": {
        "sheet": "Pune",
        "label": "Pune",
        # known span: lat 18.448-18.724, lng 73.682-73.994
        # The old box was cut off on TWO edges — north 2.9 km and west 3.4 km,
        # both at PUN-Talegaon in the NW corner. Widened north for Chakan /
        # Rajgurunagar, west for Kamshet / Lonavala, south for Saswad, east for
        # the Ranjangaon-Shirur corridor. Stops short of Mumbai's nearest store
        # (MUM-Manjarli at 19.167, 73.239) so no border spill.
        "lat": (18.35, 18.90),
        "lng": (73.40, 74.15),
    },
    "delhi_ncr": {
        "sheet": "Delhi NCR",
        "label": "Delhi NCR",
        "lat": (28.30, 28.95),
        "lng": (76.85, 77.55),
    },
    "ahmedabad": {
        "sheet": "Ahmedabad",
        "label": "Ahmedabad",
        # known span: lat 22.982-23.193, lng 72.469-72.669
        # Old box was badly cut off: north only 0.8 km (AHM-Gandhinagar sat almost
        # on the edge, and Gandhinagar city runs well past it) and west 1.9 km
        # (AHM-South Bopal, with Bopal/Shela/Sanand beyond). Widened north through
        # Gandhinagar toward Adalaj/Kalol, west for Sanand, south for Narol/Aslali,
        # east for Naroda/Kathwada. Stops short of Mehsana's store (23.588, 72.369).
        "lat": (22.85, 23.40),
        "lng": (72.30, 72.80),
    },
}

STEP        = 0.006     # ~670 m — finer than 1 km, which provably still hid stores
CONCURRENCY = 8
PROBE_TRIES = 3

COMBINED = Path(__file__).parent.parent / "zepto_all_stores_combined_20260731_1147.xlsx"

_BASE = "https://www.zepto.com"
_BFF  = "https://bff-gateway.zepto.com"
_GP   = "/lms/api/v2/get_page"

HDR_FONT  = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
HDR_FILL  = PatternFill("solid", fgColor="2E4057")
HDR_ALIGN = Alignment(horizontal="center", vertical="center")
ALT_FILL  = PatternFill("solid", fgColor="F0F4F8")
NEW_FILL  = PatternFill("solid", fgColor="D9F7E6")
SEC_FILL  = PatternFill("solid", fgColor="FFF4CE")
TITLE_FILL = PatternFill("solid", fgColor="1A3C5E")

_DROP_HEADERS = {"host", "content-length", "connection", "accept-encoding",
                 ":method", ":path", ":authority", ":scheme"}


def load_known(city_sheet: str):
    """(city_primary, city_secondary, {id: name}, {id: owning_city}) across ALL sheets."""
    wb = openpyxl.load_workbook(COMBINED, read_only=True, data_only=True)
    primary, secondary, names, owner = set(), set(), {}, {}
    coords = []
    for sheet in wb.sheetnames:
        if sheet in ("All Stores", "Summary"):
            continue
        rows = list(wb[sheet].iter_rows(values_only=True))
        if not rows:
            continue
        hdr = [str(h).strip().lower() if h else "" for h in rows[0]]
        if "store_id" not in hdr:
            continue
        for r in rows[1:]:
            if not r or not r[0]:
                continue
            d = dict(zip(hdr, r))
            sid = str(d.get("store_id") or "").strip()
            if not sid:
                continue
            owner.setdefault(sid, sheet)
            nm = str(d.get("store_name") or "").strip()
            if nm and sid not in names:
                names[sid] = nm
            if sheet == city_sheet:
                primary.add(sid)
                for x in str(d.get("secondary_store_id") or "").split(","):
                    x = x.strip()
                    if x:
                        secondary.add(x)
                try:
                    la, ln = float(d.get("lat") or 0), float(d.get("lng") or 0)
                    if la and ln:
                        coords.append((la, ln))
                except (TypeError, ValueError):
                    pass
    wb.close()
    return primary, secondary, names, owner, coords


def build_grid(lat_rng, lng_rng):
    pts = []
    lat = lat_rng[0]
    while lat <= lat_rng[1] + 1e-9:
        lng = lng_rng[0]
        while lng <= lng_rng[1] + 1e-9:
            pts.append((round(lat, 5), round(lng, 5)))
            lng += STEP
        lat += STEP
    return pts


async def capture_headers(page) -> dict:
    captured: dict[str, str] = {}

    async def on_req(req):
        if _GP in req.url and not captured:
            captured.update(req.headers)

    page.on("request", on_req)
    try:
        await page.goto(_BASE, timeout=30000, wait_until="domcontentloaded")
        await page.wait_for_timeout(4000)
    except Exception:
        pass
    finally:
        page.remove_listener("request", on_req)
    return {k: v for k, v in captured.items() if k.lower() not in _DROP_HEADERS}


def gp_url(lat, lng):
    return (f"{_BFF}{_GP}?latitude={lat}&longitude={lng}"
            f"&page_type=HOME&version=v2&show_new_eta_banner=true"
            f"&page_size=3&enforce_platform_type=WEB")


async def probe(ctx, lat, lng, state):
    for attempt in range(1, PROBE_TRIES + 1):
        try:
            r = await ctx.request.get(gp_url(lat, lng), headers=state["headers"], timeout=15000)
            if r.status == 200:
                d = await r.json()
                svc = d.get("storeServiceableResponse") or {}
                sid = svc.get("storeId")
                if not sid:
                    return None
                return {
                    "store_id": sid,
                    "secondary": [str(x) for x in (svc.get("secondaryStoreIds") or [])],
                    "store_name": (d.get("storeDetailsResponse") or {}).get("name", ""),
                }
            if r.status in (401, 403, 429):
                state["stale"] = True
                await asyncio.sleep(1.5 * attempt)
                continue
        except Exception:
            pass
        await asyncio.sleep(0.6 * attempt)
    state["failed"] += 1
    return None


def widths(ws):
    for col in ws.columns:
        cl = get_column_letter(col[0].column)
        ml = max((len(str(c.value)) for c in col if c.value), default=0)
        ws.column_dimensions[cl].width = min(max(ml + 2, 8), 46)


def head(ws, cols):
    ws.append(cols)
    for cell in ws[1]:
        cell.font, cell.fill, cell.alignment = HDR_FONT, HDR_FILL, HDR_ALIGN
    ws.row_dimensions[1].height = 18
    ws.freeze_panes = "A2"


def banner(ws, title, sub):
    ws.insert_rows(1, amount=2)
    mc = ws.max_column or 1
    ws["A1"], ws["A2"] = title, sub
    ws["A1"].font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="444444")
    ws["A2"].fill = ALT_FILL
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 15
    if mc > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=mc)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=mc)


def edge_report(found, lat_rng, lng_rng):
    """Warn if stores sit near a box edge — that edge probably cut off more."""
    if not found:
        return []
    lats = [v["lat"] for v in found.values()]
    lngs = [v["lng"] for v in found.values()]
    out = []
    for name, gap_km in [
        ("NORTH", (lat_rng[1] - max(lats)) * 111),
        ("SOUTH", (min(lats) - lat_rng[0]) * 111),
        ("EAST",  (lng_rng[1] - max(lngs)) * 103),
        ("WEST",  (min(lngs) - lng_rng[0]) * 103),
    ]:
        flag = "  <-- CUT OFF, widen this edge" if gap_km < 5 else ""
        out.append(f"  {name:<6} nearest store {gap_km:5.1f} km from edge{flag}")
    return out


async def main():
    if len(sys.argv) < 2 or sys.argv[1].lower() not in CITIES:
        print("Usage: python -m scripts.zepto_city_grid_scan <city>")
        print("Available:", ", ".join(sorted(CITIES)))
        return

    key  = sys.argv[1].lower()
    cfg  = CITIES[key]
    ts   = datetime.now().strftime("%Y%m%d_%H%M")
    scraped_at = datetime.now().strftime("%d-%b-%Y %H:%M")

    known_primary, known_secondary, known_names, owner, known_coords = load_known(cfg["sheet"])
    grid = build_grid(cfg["lat"], cfg["lng"])

    print(f"{cfg['label']} Dark Store GRID SCAN - {ts}")
    print(f"Box     : lat {cfg['lat'][0]}-{cfg['lat'][1]}, lng {cfg['lng'][0]}-{cfg['lng'][1]}")
    print(f"Step    : {STEP} deg (~{STEP*111:.0f} m)")
    print(f"Points  : {len(grid)}   (concurrency {CONCURRENCY})")
    print(f"Known   : {len(known_primary)} {cfg['label']} primary, {len(known_secondary)} secondary")
    if known_coords:
        la = [c[0] for c in known_coords]; ln = [c[1] for c in known_coords]
        print(f"          known span lat {min(la):.3f}-{max(la):.3f}, lng {min(ln):.3f}-{max(ln):.3f}")
    print(f"          {len(owner)} store ids known across ALL cities (dedup base)")
    print("=" * 80)

    found: dict[str, dict] = {}
    sec_seen: dict[str, set] = {}
    state = {"headers": {}, "stale": False, "failed": 0}
    done = {"n": 0, "no_store": 0}

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        ctx = await browser.new_context(
            user_agent=("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                        "(KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36"))
        page = await ctx.new_page()

        print("Capturing session headers...")
        state["headers"] = await capture_headers(page)
        if not state["headers"]:
            print("FAILED to capture session headers - aborting.")
            await browser.close()
            return
        print(f"Captured {len(state['headers'])} headers. Scanning...\n")

        sem = asyncio.Semaphore(CONCURRENCY)
        lock = asyncio.Lock()

        async def worker(lat, lng):
            async with sem:
                r = await probe(ctx, lat, lng, state)
            async with lock:
                done["n"] += 1
                n = done["n"]
                if not r:
                    done["no_store"] += 1
                else:
                    sid = r["store_id"]
                    for s in r["secondary"]:
                        sec_seen.setdefault(s, set()).add((lat, lng))
                    if sid not in found:
                        is_new    = sid not in known_primary
                        truly_new = sid not in owner
                        found[sid] = {
                            "store_id": sid,
                            "store_name": r["store_name"] or known_names.get(sid, ""),
                            "lat": lat, "lng": lng,
                            "secondary": ",".join(r["secondary"]),
                            "is_new": is_new, "truly_new": truly_new,
                            "known_city": "" if truly_new else owner.get(sid, ""),
                            "hit_count": 1,
                        }
                        n_new = sum(1 for v in found.values() if v["truly_new"])
                        tag = ("NEW  " if truly_new
                               else f"dup:{owner.get(sid,'')[:6]}" if is_new else "known")
                        print(f"[{n:>5}/{len(grid)}] ({lat},{lng}) {tag:<10} {sid[:8]} "
                              f"{(r['store_name'] or '')[:28]:<28} | uniq {len(found)} new {n_new}")
                    else:
                        found[sid]["hit_count"] += 1
                if n % 500 == 0:
                    n_new = sum(1 for v in found.values() if v["truly_new"])
                    print(f"    ... {n}/{len(grid)} probed | unique {len(found)} "
                          f"| new {n_new} | no-store {done['no_store']}")

        CHUNK = 250
        for start in range(0, len(grid), CHUNK):
            batch = grid[start:start + CHUNK]
            await asyncio.gather(*(worker(la, ln) for la, ln in batch))
            if state["stale"]:
                print("    [session stale - re-capturing headers]")
                nh = await capture_headers(page)
                if nh:
                    state["headers"] = nh
                state["stale"] = False
                await asyncio.sleep(2)

        await browser.close()

    sec_only    = {s: c for s, c in sec_seen.items() if s not in found}
    new_primary = [v for v in found.values() if v["is_new"]]
    brand_new   = [v for v in found.values() if v["truly_new"]]
    cross_city  = [v for v in found.values() if v["is_new"] and not v["truly_new"]]
    all_known   = known_primary | known_secondary
    new_sec     = [s for s in sec_only if s not in all_known and s not in owner]
    total_after = len(known_primary | known_secondary | set(found) | set(sec_only))
    missed      = [s for s in known_primary if s not in found]
    confirmed   = len(found) - len(cross_city) + len(missed)

    print(f"\n{'='*80}")
    print(f"Grid points probed          : {done['n']}")
    print(f"  no store at point         : {done['no_store']}")
    print(f"  probe failures            : {state['failed']}")
    print(f"Unique stores hit           : {len(found)}")
    print(f"  already known in {cfg['label']:<10}: {len(found) - len(new_primary)}")
    print(f"  new vs known {len(known_primary):<3}          : {len(new_primary)}")
    print(f"    BRAND NEW to dataset    : {len(brand_new)}")
    print(f"    another city's (spill)  : {len(cross_city)}")
    for v in cross_city:
        print(f"        {v['store_id'][:8]}  {v['store_name'][:30]:<30} -> {v['known_city']}")
    print(f"Secondary-only stores       : {len(sec_only)}  ({len(new_sec)} brand new)")
    print(f"Known stores not hit by grid: {len(missed)}")
    print()
    print("Edge check (is the box big enough?):")
    for line in edge_report(found, cfg["lat"], cfg["lng"]):
        print(line)
    print()
    print(f"{cfg['label']} CONFIRMED dark stores : {confirmed}   (was {len(known_primary)})")
    print(f"TOTAL unique store IDs after this scan: {total_after}")

    # ── Excel ────────────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    b_title = (f"{cfg['label']} Grid Scan - {confirmed} confirmed stores "
               f"({len(brand_new)} brand new, was {len(known_primary)})")
    b_sub   = (f"Generated: {scraped_at}  |  {len(grid)} points @ {STEP} deg "
               f"(~{STEP*111:.0f} m)  |  box lat {cfg['lat'][0]}-{cfg['lat'][1]}, "
               f"lng {cfg['lng'][0]}-{cfg['lng'][1]}")

    c1 = ["store_id", "store_name", "lat", "lng", "is_new", "truly_new",
          "known_city", "hit_count", "secondary"]

    ws1 = wb.active
    ws1.title = "Grid Stores"
    head(ws1, c1)
    for i, row in enumerate(sorted(found.values(),
                                   key=lambda x: (not x["truly_new"], x["store_name"])), start=2):
        ws1.append([row.get(c, "") for c in c1])
        fill = NEW_FILL if row["truly_new"] else (SEC_FILL if row["is_new"] else
                                                  (ALT_FILL if i % 2 == 0 else None))
        if fill:
            for c in range(1, len(c1) + 1):
                ws1.cell(row=i, column=c).fill = fill
    ws1.auto_filter.ref = ws1.dimensions
    widths(ws1)

    ws2 = wb.create_sheet("NEW Stores")
    head(ws2, c1)
    for i, row in enumerate(sorted(brand_new, key=lambda x: x["store_name"]), start=2):
        ws2.append([row.get(c, "") for c in c1])
        for c in range(1, len(c1) + 1):
            ws2.cell(row=i, column=c).fill = NEW_FILL
    widths(ws2)

    ws3 = wb.create_sheet("Border Spill")
    head(ws3, c1)
    for i, row in enumerate(sorted(cross_city, key=lambda x: x["store_name"]), start=2):
        ws3.append([row.get(c, "") for c in c1])
        for c in range(1, len(c1) + 1):
            ws3.cell(row=i, column=c).fill = SEC_FILL
    widths(ws3)

    ws4 = wb.create_sheet("Secondary Only")
    c4 = ["store_id", "times_referenced", "sample_lat", "sample_lng", "brand_new"]
    head(ws4, c4)
    for i, (s, coords) in enumerate(sorted(sec_only.items(), key=lambda kv: -len(kv[1])), start=2):
        c0 = sorted(coords)[0]
        ws4.append([s, len(coords), c0[0], c0[1], "YES" if s not in owner else "no"])
        for c in range(1, len(c4) + 1):
            ws4.cell(row=i, column=c).fill = SEC_FILL
    widths(ws4)

    ws5 = wb.create_sheet("Summary")
    head(ws5, ["Metric", "Value"])
    stats = [
        ("City", cfg["label"]),
        ("Grid box", f"lat {cfg['lat'][0]}-{cfg['lat'][1]}, lng {cfg['lng'][0]}-{cfg['lng'][1]}"),
        ("Grid step", f"{STEP} deg (~{STEP*111:.0f} m)"),
        ("Grid points probed", done["n"]),
        ("Points with no store", done["no_store"]),
        ("Probe failures", state["failed"]),
        ("Unique stores hit by grid", len(found)),
        ("Previously known", len(known_primary)),
        ("New vs previously known", len(new_primary)),
        ("  BRAND NEW to dataset", len(brand_new)),
        ("  border spill (other city)", len(cross_city)),
        ("Secondary-only stores", len(sec_only)),
        ("Known stores not hit by grid", len(missed)),
        (f"{cfg['label']} CONFIRMED stores", confirmed),
        ("TOTAL unique store IDs now", total_after),
        ("Scraped at", scraped_at),
    ]
    for i, (m, v) in enumerate(stats, start=2):
        ws5.cell(row=i, column=1, value=m)
        ws5.cell(row=i, column=2, value=v)
        if i % 2 == 0:
            for c in range(1, 3):
                ws5.cell(row=i, column=c).fill = ALT_FILL
    for line in edge_report(found, cfg["lat"], cfg["lng"]):
        r = ws5.max_row + 1
        ws5.cell(row=r, column=1, value="Edge check")
        ws5.cell(row=r, column=2, value=line.strip())
    widths(ws5)

    for ws in (ws1, ws2, ws3, ws4, ws5):
        banner(ws, b_title, b_sub)

    out = Path(__file__).parent.parent / f"zepto_{key}_grid_scan_{ts}.xlsx"
    wb.save(out)
    print(f"\nSaved -> {out}")


if __name__ == "__main__":
    asyncio.run(main())
