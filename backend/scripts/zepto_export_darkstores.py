"""
Export Zepto dark stores in the same shape as blinkit_darkstores_export.xlsx.

Only columns that genuinely exist on both platforms are kept. Blinkit columns
with no Zepto equivalent are dropped rather than filled with blanks or invented
values:

    type                - Blinkit tiers stores express/longtail/super_longtail.
                          Zepto has no such tiering, only a primary store plus
                          secondaryStoreIds, so the column would be a constant.
    super_longtail_ids  - no equivalent
    unicorn_ids         - no equivalent

Kept, with the Blinkit name on the left:

    merchant_id   -> store_id      (UUID, not an integer like Blinkit's)
    city          -> city
    state         -> state
    region        -> derived from state
    pincode       -> source_pincode
    lat           -> lat
    lon           -> lng
    active        -> serviceable
    address       -> area_suggestion (full address on pincode-found rows)
    location_name -> locality extracted from the address, or the store name

store_name is also kept: it is Zepto's own label for the store and carries the
city prefix used throughout this dataset.

Run from backend/:
    python -m scripts.zepto_export_darkstores
Output: zepto_darkstores_export.xlsx
"""
import re
import sys
from pathlib import Path
from datetime import datetime
from collections import Counter

sys.path.insert(0, str(Path(__file__).parent.parent))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BACKEND = Path(__file__).parent.parent

# Column order mirrors blinkit_darkstores_export.xlsx for the shared fields.
COLS = ["store_id", "store_name", "city", "state", "region", "pincode",
        "lat", "lon", "active", "address", "location_name"]

REGION = {
    "North India": ["Delhi", "Haryana", "Punjab", "Uttar Pradesh", "Uttarakhand",
                    "Himachal Pradesh", "Jammu and Kashmir", "Chandigarh",
                    "Rajasthan", "Ladakh"],
    "South India": ["Karnataka", "Tamil Nadu", "Kerala", "Andhra Pradesh",
                    "Telangana", "Puducherry", "Pondicherry"],
    "West India": ["Maharashtra", "Gujarat", "Goa", "Dadra and Nagar Haveli",
                   "Daman and Diu"],
    "East India": ["West Bengal", "Bihar", "Odisha", "Jharkhand", "Assam",
                   "Sikkim", "Tripura", "Meghalaya", "Manipur", "Nagaland",
                   "Mizoram", "Arunachal Pradesh"],
    "Central India": ["Madhya Pradesh", "Chhattisgarh"],
}
STATE_REGION = {s: r for r, ss in REGION.items() for s in ss}

# Every city -> its state. Needed because the details API sometimes returns a
# blank state, and sometimes returns it in Devanagari ('महाराष्ट्र' for
# Maharashtra), neither of which the region lookup can use. The city is always
# known, so it is the reliable source.
CITY_STATE = {
    "Delhi NCR": "Delhi", "Bengaluru": "Karnataka", "Mumbai": "Maharashtra",
    "Chennai": "Tamil Nadu", "Kolkata": "West Bengal", "Hyderabad": "Telangana",
    "Pune": "Maharashtra", "Ahmedabad": "Gujarat",
    "Agra": "Uttar Pradesh", "Amritsar": "Punjab", "Chandigarh": "Chandigarh",
    "Coimbatore": "Tamil Nadu", "Dehradun": "Uttarakhand", "Hubballi": "Karnataka",
    "Indore": "Madhya Pradesh", "Jaipur": "Rajasthan", "Jalandhar": "Punjab",
    "Kanpur": "Uttar Pradesh", "Kochi": "Kerala", "Kota": "Rajasthan",
    "Lucknow": "Uttar Pradesh", "Ludhiana": "Punjab", "Madurai": "Tamil Nadu",
    "Mysuru": "Karnataka", "Nagpur": "Maharashtra", "Nashik": "Maharashtra",
    "Patiala": "Punjab", "Prayagraj": "Uttar Pradesh", "Puducherry": "Puducherry",
    "Rajkot": "Gujarat", "Surat": "Gujarat", "Udaipur": "Rajasthan",
    "Vadodara": "Gujarat", "Varanasi": "Uttar Pradesh",
    "Vijayawada": "Andhra Pradesh", "Warangal": "Telangana",
    "Ambala": "Haryana", "Bareilly": "Uttar Pradesh", "Belagavi": "Karnataka",
    "Bhiwadi": "Rajasthan", "Chhatrapati Sambhaji Nagar": "Maharashtra",
    "Davangere": "Karnataka", "Gorakhpur": "Uttar Pradesh",
    "Guntur": "Andhra Pradesh", "Hisar": "Haryana", "Hosur": "Tamil Nadu",
    "Karimnagar": "Telangana", "Karnal": "Haryana", "Meerut": "Uttar Pradesh",
    "Mehsana": "Gujarat", "Mohali": "Punjab", "Palakkad": "Kerala",
    "Panchkula": "Haryana", "Panipat": "Haryana", "Rewari": "Haryana",
    "Sonipat": "Haryana", "Tumkuru": "Karnataka", "Vellore": "Tamil Nadu",
}

DROP_PARTS = re.compile(r"^\d{6}-?$|^india$|^[A-Za-z ]+ \d{6}$", re.I)

HDR_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
HDR_FILL = PatternFill("solid", fgColor="2E4057")
HDR_ALIGN = Alignment(horizontal="center", vertical="center")
ALT_FILL = PatternFill("solid", fgColor="F0F4F8")


def blank(v):
    return not str(v or "").strip() or str(v).strip().lower() == "none"


def locality(address, store_name, state, city, raw_state=""):
    """Short area name — Blinkit's location_name equivalent.

    Pincode-found rows carry a full address ("560035, Kodathi Village Road,
    Kodathi, ..."); geocoded rows already carry a bare locality. Falls back to
    the store name minus its city prefix.
    """
    if not blank(address):
        parts = [p.strip() for p in str(address).split(",") if p.strip()]
        drop = {str(state).lower(), str(raw_state).lower(),
                str(city).lower(), "india"}
        parts = [p for p in parts
                 if not DROP_PARTS.match(p) and p.lower() not in drop]
        if parts:
            return parts[-1]
    s = re.sub(r"^\s*[A-Za-z]{2,4}\s*-\s*", "", str(store_name or "")).strip()
    return s or ""


def main():
    src = sorted(BACKEND.glob("zepto_FINAL_master_filled_*.xlsx"))
    if not src:
        src = sorted(BACKEND.glob("zepto_FINAL_master_*.xlsx"))
    if not src:
        raise SystemExit("No FINAL master found — run the merge first.")
    src = src[-1]

    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    rr = list(wb["All Stores"].iter_rows(values_only=True))
    wb.close()
    hdr = [str(x).strip().lower() if x else "" for x in rr[2]]
    rows = [dict(zip(hdr, r)) for r in rr[3:] if r and r[0]]

    out = []
    no_state = no_region = 0
    for r in rows:
        city = str(r.get("city") or "").strip()
        state = str(r.get("state") or "").strip()
        # Fall back to the city whenever the API gave nothing, or gave a
        # localized name the region lookup cannot match.
        if blank(state) or not state.isascii() or state not in STATE_REGION:
            fixed = CITY_STATE.get(city, "")
            if fixed:
                if fixed != state:
                    no_state += 1
                state = fixed
        region = STATE_REGION.get(state, "")
        if not region:
            no_region += 1
        addr = "" if blank(r.get("area_suggestion")) else str(r["area_suggestion"]).strip()
        pin = str(r.get("source_pincode") or "").strip()
        out.append({
            "store_id": str(r.get("store_id") or "").strip(),
            "store_name": str(r.get("store_name") or "").strip(),
            "city": city,
            "state": state,
            "region": region,
            "pincode": "" if pin in ("", "UNRESOLVED") else pin,
            "lat": r.get("lat"),
            "lon": r.get("lng"),
            # every row in the master is a real store; the unserviceable ones
            # were already excluded into their own sheet by the merge
            "active": "no" if str(r.get("serviceable")).upper() == "NO" else "yes",
            "address": addr,
            "location_name": locality(addr, r.get("store_name"), state, city,
                                      str(r.get("state") or "")),
        })

    out.sort(key=lambda x: (x["region"], x["state"], x["city"],
                            str(x["pincode"]), x["store_name"]))

    wb2 = openpyxl.Workbook()
    ws = wb2.active
    ws.title = "zepto_darkstores_export"
    ws.append(COLS)
    for c in ws[1]:
        c.font, c.fill, c.alignment = HDR_FONT, HDR_FILL, HDR_ALIGN
    ws.row_dimensions[1].height = 18
    ws.freeze_panes = "A2"
    for i, r in enumerate(out, start=2):
        ws.append([r.get(c, "") for c in COLS])
        if i % 2 == 0:
            for c in range(1, len(COLS) + 1):
                ws.cell(row=i, column=c).fill = ALT_FILL
    ws.auto_filter.ref = ws.dimensions
    for col in ws.columns:
        cl = get_column_letter(col[0].column)
        ml = max((len(str(c.value)) for c in col[:400] if c.value), default=0)
        ws.column_dimensions[cl].width = min(max(ml + 2, 9), 46)

    dest = BACKEND / "zepto_darkstores_export.xlsx"
    wb2.save(dest)

    print(f"Source : {src.name}")
    print(f"Rows   : {len(out)}")
    print(f"Columns: {', '.join(COLS)}")
    print()
    print(f"  state filled from city fallback : {no_state}")
    print(f"  rows with no region             : {no_region}")
    print(f"  blank pincode                   : {sum(1 for r in out if not r['pincode'])}")
    print(f"  blank location_name             : {sum(1 for r in out if not r['location_name'])}")
    print()
    print("region  :", dict(Counter(r["region"] for r in out)))
    print("active  :", dict(Counter(r["active"] for r in out)))
    print()
    print(f"Saved -> {dest}")


if __name__ == "__main__":
    main()
