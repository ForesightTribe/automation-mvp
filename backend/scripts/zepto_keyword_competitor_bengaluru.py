"""
Keyword-level competitive scan across every Bengaluru dark store.

For each (dark store x keyword) it captures the FULL result page — every brand,
not just the client's — up to RESULT_CAP products, preserving search position.
That is what makes share-of-voice and rank-vs-competitor measurable.

Chain, both hops through Zepto's own APIs:

    lat/lng --> get_page --> storeId --> search(keyword, page N) --> products

get_page runs once per store; the resolved storeId is then reused for every
keyword, so the coordinate still drives the store but is not re-sent 6 times.

Client brand rows are colour-coded green throughout the workbook.

Run from backend/:
    python -m scripts.zepto_keyword_competitor_bengaluru
Output: zepto_keyword_competitors_bengaluru_TIMESTAMP.xlsx
"""
import asyncio
import json
import sys
from pathlib import Path
from datetime import datetime

sys.path.insert(0, str(Path(__file__).parent.parent))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from playwright.async_api import async_playwright

# ── Config ────────────────────────────────────────────────────────────────────
CLIENT_BRAND = "Brik Oven"

KEYWORDS = [
    "sourdough",
    "sourdough bread",
    "sourdough bread loaf",
    "rosemary sourdough",
    "sour cream",
    "mozzarella",
    "ricotta",
    "cheesecake",
    "buratta",
]

RESULT_CAP = 30          # products to keep per (store, keyword)
CITY_SHEET = "Bengaluru"
MAX_PAGES  = 3           # pagination ceiling while filling up to RESULT_CAP

BACKEND = Path(__file__).parent.parent
_BASE = "https://www.zepto.com"
_BFF  = "https://bff-gateway.zepto.com"
_GP   = "/lms/api/v2/get_page"
_SEARCH = "/user-search-service/api/v3/search"

TRIES = 3
# Zepto rate-limits at roughly 35-40 searches, and it is enforced at the IP
# level: two separate fixes were measured and both failed at ~22% loss —
# rotating the captured session headers, and tearing down the whole browser
# context for a new device_id/cookie jar. Failures resumed on the very first
# store after each rotation in both cases.
#
# What the data does show is that the window recovers by itself in ~60-90 s.
# So the only thing that works is pacing: run at a sustainable rate, and when a
# block is detected, stop and wait it out rather than burning retries into it
# (retries during a block keep the limit pegged and lengthen it).
# Measured: at a 1.4 s gap (~33 searches/min) the limit still trips and ~3% of
# searches are lost. At 3.0 s (~15.7/min) a 60-search run had ZERO failures —
# comfortably under the ceiling. The extra ~13 min buys the last 3% of the data
# and makes the run steady instead of burst-then-stall.
# 3.0 s measured 0% failures on a cold IP. It collapsed from store 8 when run
# straight after the 58-city discovery sweep — hours of get_page traffic leaves
# the IP in a penalty state, so the threshold that held before no longer does.
# 6.0 s halves the request rate again to buy margin.
SEARCH_GAP_S   = 6.0     # buffer after every single search
# 75 s was not enough once the IP was already penalised: the cooldown fired,
# waited, and the very next search failed again. 300 s gives the window real
# time to decay instead of poking it every 75 s.
COOLDOWN_S     = 300     # wait this long once a block is detected
COOLDOWN_AFTER = 2       # consecutive failed searches that mean "blocked"
REFRESH_EVERY  = 40      # identity refresh is near-useless; keep it rare
_DROP = {"host", "content-length", "connection", "accept-encoding",
         ":method", ":path", ":authority", ":scheme"}

# ── Styling ───────────────────────────────────────────────────────────────────
HDR_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
HDR_FILL = PatternFill("solid", fgColor="2E4057")
HDR_ALIGN = Alignment(horizontal="center", vertical="center")
ALT_FILL = PatternFill("solid", fgColor="F0F4F8")
CLIENT_FILL = PatternFill("solid", fgColor="C6EFCE")   # green - client brand
CLIENT_FONT = Font(name="Calibri", size=10, bold=True, color="0B6B3A")
TOP3_FILL = PatternFill("solid", fgColor="FFF2CC")     # amber - top 3 rank
OOS_FONT  = Font(name="Calibri", size=10, color="C00000")
TITLE_FILL = PatternFill("solid", fgColor="1A3C5E")
TOT_FILL = PatternFill("solid", fgColor="048A81")
TOT_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")


def latest(pat):
    f = sorted(BACKEND.glob(pat))
    return f[-1] if f else None


def load_stores() -> list[dict]:
    """Bengaluru stores from the FINAL merged master (169 after the 58-city sweep).

    That file has one 'All Stores' sheet with a city column, not per-city sheets,
    and a 2-row banner — hence the different read compared with the old workbook.
    """
    path = latest("zepto_FINAL_master_filled_*.xlsx") or latest("zepto_FINAL_master_*.xlsx")
    if not path:
        raise SystemExit("No FINAL master found. Run zepto_merge_final first.")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rr = list(wb["All Stores"].iter_rows(values_only=True))
    wb.close()
    hdr = [str(h).strip().lower() if h else "" for h in rr[2]]
    rows = [None] + [r for r in rr[3:]
                     if r and r[0] and str(dict(zip(hdr, r)).get("city")) == "Bengaluru"]
    out, seen = [], set()
    for r in rows[1:]:
        if not r or not r[0]:
            continue
        d = dict(zip(hdr, r))
        sid = str(d.get("store_id") or "").strip()
        if not sid or sid in seen:
            continue
        try:
            la, ln = float(d.get("lat") or 0), float(d.get("lng") or 0)
        except (TypeError, ValueError):
            continue
        if not la or not ln:
            continue
        seen.add(sid)
        out.append({"store_id": sid,
                    "store_name": str(d.get("store_name") or "").strip(),
                    "pincode": str(d.get("source_pincode") or d.get("pincode") or "").strip(),
                    "area": str(d.get("area_suggestion") or d.get("area") or "").strip(),
                    "lat": la, "lng": ln})
    print(f"Store source: {path.name}")
    return out


async def capture(page, part, do_nav, settle_ms=6000):
    cap = {"h": None, "body": None}

    async def on_req(req):
        if part in req.url and cap["h"] is None:
            cap["h"] = dict(req.headers)
            try:
                cap["body"] = req.post_data
            except Exception:
                pass

    page.on("request", on_req)
    try:
        await do_nav()
        waited = 0
        while cap["h"] is None and waited < settle_ms:
            await page.wait_for_timeout(250)
            waited += 250
    except Exception:
        pass
    finally:
        page.remove_listener("request", on_req)
    if not cap["h"]:
        return {}, None
    return {k: v for k, v in cap["h"].items() if k.lower() not in _DROP}, cap["body"]


async def new_identity(browser, state):
    """Tear down the browser context and build a fresh one.

    Rotating only the captured headers does NOT reset Zepto's counter: the run
    that tried it failed on the very first store after every rotation. The limit
    tracks the browser identity (device_id + cookies), which survives a header
    swap. Closing the context and opening a new one issues a new device_id and a
    clean cookie jar, which is what actually resets it.
    """
    old = state.get("ctx")
    if old is not None:
        try:
            await old.close()
        except Exception:
            pass

    ctx = await browser.new_context(
        user_agent=("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                    "(KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36"))
    page = await ctx.new_page()
    state["ctx"], state["page"] = ctx, page

    gp, _ = await capture(page, _GP,
                          lambda: page.goto(_BASE, timeout=30000,
                                            wait_until="domcontentloaded"))
    if gp:
        state["gp"] = gp
    await page.wait_for_timeout(1200)
    se, raw = await capture(page, _SEARCH,
                            lambda: page.goto(f"{_BASE}/search?query=bread",
                                              timeout=30000,
                                              wait_until="domcontentloaded"),
                            settle_ms=9000)
    if se:
        state["se"] = se
        if raw:
            try:
                state["body"] = json.loads(raw)
            except Exception:
                pass
    return bool(gp and se)


def gp_url(lat, lng):
    return (f"{_BFF}{_GP}?latitude={lat}&longitude={lng}"
            f"&page_type=HOME&version=v2&show_new_eta_banner=true"
            f"&page_size=3&enforce_platform_type=WEB")


async def resolve_store(state, lat, lng):
    for a in range(1, TRIES + 1):
        try:
            r = await state["ctx"].request.get(gp_url(lat, lng),
                                              headers=state["gp"], timeout=15000)
            if r.status == 200:
                svc = (await r.json()).get("storeServiceableResponse") or {}
                return svc.get("storeId"), [str(x) for x in (svc.get("secondaryStoreIds") or [])]
            if r.status in (401, 403, 429):
                state["stale"] = True
                await asyncio.sleep(1.2 * a)
                continue
        except Exception:
            pass
        await asyncio.sleep(0.5 * a)
    return None, []


def _rs(p):
    return round((p or 0) / 100, 2)


def extract_all(data: dict, page_no: int) -> tuple[list[dict], int, bool]:
    """EVERY product on the page — all brands. Returns (rows, total, reached_end)."""
    rows, total = [], 0
    for w in data.get("layout", []):
        if w.get("widgetId") != "PRODUCT_GRID":
            continue
        for it in (w.get("data", {}).get("resolver", {}).get("data", {}).get("items") or []):
            pr = it.get("productResponse")
            if not pr:
                continue
            total += 1
            prod = pr.get("product", {}) or {}
            pv   = pr.get("productVariant", {}) or {}
            rat  = pv.get("ratingSummary", {}) or {}
            meta = pr.get("meta", {}) or {}
            brand = (prod.get("brand") or "").strip()
            mrp = pv.get("mrp") or pr.get("mrp") or 0
            sp  = pr.get("discountedSellingPrice") or pr.get("sellingPrice") or 0
            rows.append({
                "position":      (it.get("position") or 0) + 1,
                "page":          page_no,
                "brand":         brand or "(unbranded)",
                "is_client":     CLIENT_BRAND.lower() in brand.lower(),
                "product_name":  prod.get("name", ""),
                "pack_size":     pv.get("formattedPacksize", ""),
                "mrp_rs":        _rs(mrp),
                "selling_price_rs": _rs(sp),
                "discount_pct":  pr.get("discountPercent", 0),
                "discount_rs":   _rs(pr.get("discountAmount")),
                "in_stock":      not pr.get("outOfStock", False),
                "available_qty": pr.get("availableQuantity", 0),
                "rating":        rat.get("averageRating", ""),
                "rating_count":  rat.get("totalRatings", 0),
                "category":      pr.get("primaryCategoryName", ""),
                "match_bucket":  meta.get("query_matching_bucket", ""),
                "weight_g":      pv.get("weightInGms", ""),
                "product_id":    prod.get("id", ""),
                "variant_id":    pv.get("id", ""),
                "resp_store_id": pr.get("storeId", ""),
            })
    return rows, total, bool(data.get("hasReachedEnd", True))


async def search_keyword(state, sid, all_ids, keyword):
    """Search one keyword at one store, paging until RESULT_CAP or end."""
    collected, seen = [], set()
    reached_end = False
    for page_no in range(MAX_PAGES):
        if len(collected) >= RESULT_CAP or reached_end:
            break
        body = dict(state["body"])
        body["query"] = keyword
        body["pageNumber"] = page_no
        body["mode"] = "SHOW_ALL_RESULTS"
        got_page = False
        for a in range(1, TRIES + 1):
            h = dict(state["se"])
            h["store_id"] = sid
            h["storeid"] = sid
            h["store_ids"] = all_ids
            h["store_etas"] = json.dumps({s: -1 for s in all_ids.split(",")})
            try:
                r = await state["ctx"].request.post(f"{_BFF}{_SEARCH}", headers=h,
                                                   data=json.dumps(body), timeout=20000)
                # 299 is Zepto's own throttle signal — non-standard, so it was
                # previously mistaken for an empty result and retried 3x, which
                # kept hammering an active block and prolonged it. Bail out at
                # once and let the cooldown do its job.
                if r.status == 299:
                    state["search_stale"] = True
                    return collected[:RESULT_CAP], False
                if r.status == 200:
                    rows, total, end = extract_all(await r.json(), page_no)
                    if total > 0 or page_no > 0:
                        for x in rows:
                            k = x["variant_id"] or (x["product_name"], x["position"])
                            if k in seen:
                                continue
                            seen.add(k)
                            collected.append(x)
                        reached_end = end
                        got_page = True
                        break
                    state["search_stale"] = True
                elif r.status in (401, 403, 429, 299):
                    state["search_stale"] = True
            except Exception:
                pass
            await asyncio.sleep(0.7 * a)
        if not got_page:
            return collected[:RESULT_CAP], False
    collected.sort(key=lambda x: x["position"])
    return collected[:RESULT_CAP], True


def widths(ws, mx=44):
    for col in ws.columns:
        cl = get_column_letter(col[0].column)
        ml = max((len(str(c.value)) for c in col[:400] if c.value), default=0)
        ws.column_dimensions[cl].width = min(max(ml + 2, 9), mx)


def head(ws, cols):
    ws.append(cols)
    for c in ws[1]:
        c.font, c.fill, c.alignment = HDR_FONT, HDR_FILL, HDR_ALIGN
    ws.row_dimensions[1].height = 18
    ws.freeze_panes = "A2"


def banner(ws, t, s):
    ws.insert_rows(1, amount=2)
    mc = ws.max_column or 1
    ws["A1"], ws["A2"] = t, s
    ws["A1"].font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="444444")
    ws["A2"].fill = ALT_FILL
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 15
    if mc > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=mc)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=mc)


async def main():
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    at = datetime.now().strftime("%d-%b-%Y %H:%M")

    stores = load_stores()
    print(f"\nKeyword competitive scan — Bengaluru — {ts}")
    print(f"Client brand : {CLIENT_BRAND}")
    print(f"Dark stores  : {len(stores)}")
    print(f"Keywords     : {len(KEYWORDS)}  -> {', '.join(KEYWORDS)}")
    print(f"Cap          : {RESULT_CAP} products per store x keyword")
    print(f"Searches     : {len(stores) * len(KEYWORDS)}")
    print("=" * 112)

    detail, store_kw = [], []

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        state = {"gp": {}, "se": {}, "body": {}, "stale": False,
                 "search_stale": False, "ctx": None, "page": None,
                 "identities": 1, "consec_fail": 0, "cooldowns": 0}

        print("Creating browser identity...")
        if not await new_identity(browser, state):
            print("Header capture failed — aborting (re-run; cold-session WAF challenge).")
            await browser.close()
            return
        print(f"  get_page {len(state['gp'])} headers | search {len(state['se'])} headers\n")

        for i, st in enumerate(stores, 1):
            sid, sec = await resolve_store(state, st["lat"], st["lng"])
            if not sid:
                print(f"[{i:>3}/{len(stores)}] {st['store_name'][:28]:<28} get_page failed")
                for kw in KEYWORDS:
                    store_kw.append({**st, "api_store_id": "", "keyword": kw,
                                     "status": "GETPAGE_FAILED", "results": "",
                                     "client_products": "", "client_best_rank": "",
                                     "client_sov_pct": ""})
                continue

            match = "SAME" if sid == st["store_id"] else "DIFF"
            print(f"[{i:>3}/{len(stores)}] {st['store_name'][:28]:<28} {st['pincode']:<7} "
                  f"{sid[:8]} [{match}]")

            all_ids = ",".join([sid] + sec)
            for kw in KEYWORDS:
                rows, ok = await search_keyword(state, sid, all_ids, kw)
                client = [r for r in rows if r["is_client"]]
                sov = round(len(client) / len(rows) * 100, 1) if rows else 0
                best = min((r["position"] for r in client), default="")

                if not ok and not rows:
                    print(f"        {kw:<20} search failed")
                    status = "SEARCH_FAILED"
                    state["consec_fail"] += 1
                    if state["consec_fail"] >= COOLDOWN_AFTER:
                        print(f"        [rate limited — cooling down {COOLDOWN_S}s "
                              f"so the window can reset]")
                        await asyncio.sleep(COOLDOWN_S)
                        state["consec_fail"] = 0
                        state["search_stale"] = False
                        state["cooldowns"] += 1
                else:
                    status = "OK"
                    state["consec_fail"] = 0
                    tag = (f"{CLIENT_BRAND} #{best} ({len(client)} SKU, {sov}% SoV)"
                           if client else f"no {CLIENT_BRAND}")
                    print(f"        {kw:<20} {len(rows):>2} products | {tag}")
                    for r in sorted(client, key=lambda x: x["position"])[:3]:
                        stock = "in stock" if r["in_stock"] else "OUT OF STOCK"
                        print(f"             #{r['position']:<3} {r['product_name'][:42]:<42} "
                              f"MRP Rs{r['mrp_rs']:<7.2f} SP Rs{r['selling_price_rs']:<7.2f} "
                              f"-{r['discount_pct']:<3}% {stock}")

                store_kw.append({**st, "api_store_id": sid, "keyword": kw,
                                 "status": status, "results": len(rows),
                                 "client_products": len(client),
                                 "client_best_rank": best,
                                 "client_sov_pct": sov})
                for r in rows:
                    detail.append({"store_name": st["store_name"],
                                   "pincode": st["pincode"], "area": st["area"],
                                   "lat": st["lat"], "lng": st["lng"],
                                   "api_store_id": sid, "keyword": kw, **r})
                await asyncio.sleep(SEARCH_GAP_S)

            await asyncio.sleep(4)   # buffer between stores, on top of SEARCH_GAP_S

            if state["search_stale"] or state["stale"] or i % REFRESH_EVERY == 0:
                why = ("search stale" if state["search_stale"]
                       else "get_page stale" if state["stale"] else "routine")
                state["identities"] += 1
                print(f"        [new browser identity #{state['identities']} — {why}]")
                await asyncio.sleep(5 if (state["search_stale"] or state["stale"]) else 2)
                await new_identity(browser, state)
                state["stale"] = state["search_stale"] = False

        await browser.close()

    ok_rows = [s for s in store_kw if s["status"] == "OK"]
    fail_rows = [s for s in store_kw if s["status"] != "OK"]
    client_rows = [d for d in detail if d["is_client"]]
    brands = {}
    for d in detail:
        b = brands.setdefault(d["brand"], {"n": 0, "ranks": [], "prices": [], "client": d["is_client"]})
        b["n"] += 1
        if d["position"]:
            b["ranks"].append(d["position"])
        if d["selling_price_rs"]:
            b["prices"].append(d["selling_price_rs"])

    print(f"\n{'='*112}")
    fail_pct = len(fail_rows) / len(store_kw) * 100 if store_kw else 0
    print(f"Searches run        : {len(store_kw)}   ok: {len(ok_rows)}   "
          f"failed: {len(fail_rows)} ({fail_pct:.1f}%)")
    print(f"Rate-limit cooldowns: {state['cooldowns']}")
    if fail_rows:
        bad = {}
        for r in fail_rows:
            bad[r["store_name"]] = bad.get(r["store_name"], 0) + 1
        print(f"  stores with any failed keyword: {len(bad)}")
        print("  NOTE: failed != not stocked. Those rows are yellow in Store x Keyword.")
    print(f"Product rows        : {len(detail)}")
    print(f"{CLIENT_BRAND} rows      : {len(client_rows)}")
    print(f"Distinct brands     : {len(brands)}")
    if detail:
        print(f"{CLIENT_BRAND} share of voice: "
              f"{len(client_rows)/len(detail)*100:.1f}%")
    print("\nTop competitor brands by appearances:")
    for b, v in sorted(brands.items(), key=lambda kv: -kv[1]["n"])[:12]:
        mark = "  <-- CLIENT" if v["client"] else ""
        avg = sum(v["ranks"]) / len(v["ranks"]) if v["ranks"] else 0
        ap = sum(v["prices"]) / len(v["prices"]) if v["prices"] else 0
        print(f"   {b[:30]:<30} {v['n']:>5} rows   avg rank {avg:5.1f}   "
              f"avg Rs{ap:7.2f}{mark}")

    # ── Excel ────────────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    T = f"Keyword Competitive Scan — Bengaluru — {CLIENT_BRAND}"
    S = (f"Generated: {at}  |  {len(stores)} dark stores x {len(KEYWORDS)} keywords  |  "
         f"cap {RESULT_CAP}/search  |  {len(detail)} product rows  |  "
         f"green = {CLIENT_BRAND}")

    DET = ["store_name", "pincode", "area", "keyword", "position", "brand",
           "product_name", "pack_size", "mrp_rs", "selling_price_rs",
           "discount_pct", "discount_rs", "in_stock", "available_qty",
           "rating", "rating_count", "category", "match_bucket", "weight_g",
           "lat", "lng", "api_store_id", "product_id", "variant_id"]
    ws1 = wb.active
    ws1.title = "All Products"
    head(ws1, DET)
    for i, r in enumerate(sorted(detail, key=lambda x: (x["store_name"], x["keyword"],
                                                        x["position"])), start=2):
        ws1.append([r.get(c, "") for c in DET])
        if r["is_client"]:
            for c in range(1, len(DET) + 1):
                cell = ws1.cell(row=i, column=c)
                cell.fill = CLIENT_FILL
                cell.font = CLIENT_FONT
        elif r["position"] <= 3:
            ws1.cell(row=i, column=5).fill = TOP3_FILL
        if not r["in_stock"]:
            ws1.cell(row=i, column=13).font = OOS_FONT
    ws1.auto_filter.ref = ws1.dimensions
    widths(ws1)

    ws2 = wb.create_sheet(f"{CLIENT_BRAND} Only")
    head(ws2, DET)
    for i, r in enumerate(sorted(client_rows, key=lambda x: (x["keyword"], x["position"])),
                          start=2):
        ws2.append([r.get(c, "") for c in DET])
        for c in range(1, len(DET) + 1):
            ws2.cell(row=i, column=c).fill = CLIENT_FILL
    ws2.auto_filter.ref = ws2.dimensions
    widths(ws2)

    ws3 = wb.create_sheet("By Keyword")
    K = ["keyword", "stores_measured", "avg_results", f"{CLIENT_BRAND} SKUs",
         f"{CLIENT_BRAND} best rank", f"{CLIENT_BRAND} avg rank",
         f"{CLIENT_BRAND} SoV %", "stores stocking client", "top competitor"]
    head(ws3, K)
    for i, kw in enumerate(KEYWORDS, start=2):
        rs = [d for d in detail if d["keyword"] == kw]
        cl = [d for d in rs if d["is_client"]]
        sk = [s for s in store_kw if s["keyword"] == kw and s["status"] == "OK"]
        stocking = len([s for s in sk if s["client_products"]])
        cb = {}
        for d in rs:
            if not d["is_client"]:
                cb[d["brand"]] = cb.get(d["brand"], 0) + 1
        top = max(cb.items(), key=lambda kv: kv[1])[0] if cb else ""
        ws3.append([kw, len(sk),
                    round(sum(s["results"] for s in sk) / len(sk), 1) if sk else 0,
                    len(cl),
                    min((d["position"] for d in cl), default=""),
                    round(sum(d["position"] for d in cl) / len(cl), 1) if cl else "",
                    round(len(cl) / len(rs) * 100, 1) if rs else 0,
                    stocking, top])
        if i % 2 == 0:
            for c in range(1, len(K) + 1):
                ws3.cell(row=i, column=c).fill = ALT_FILL
    widths(ws3)

    ws4 = wb.create_sheet("Brand Ranking")
    B = ["brand", "is_client", "appearances", "share_of_voice_pct", "avg_rank",
         "best_rank", "avg_price_rs", "min_price_rs", "max_price_rs", "in_stock_pct"]
    head(ws4, B)
    tot = len(detail) or 1
    for i, (b, v) in enumerate(sorted(brands.items(), key=lambda kv: -kv[1]["n"]), start=2):
        rs = [d for d in detail if d["brand"] == b]
        ws4.append([b, "YES" if v["client"] else "", v["n"],
                    round(v["n"] / tot * 100, 2),
                    round(sum(v["ranks"]) / len(v["ranks"]), 1) if v["ranks"] else "",
                    min(v["ranks"]) if v["ranks"] else "",
                    round(sum(v["prices"]) / len(v["prices"]), 2) if v["prices"] else "",
                    min(v["prices"]) if v["prices"] else "",
                    max(v["prices"]) if v["prices"] else "",
                    round(sum(1 for d in rs if d["in_stock"]) / len(rs) * 100, 1)])
        if v["client"]:
            for c in range(1, len(B) + 1):
                ws4.cell(row=i, column=c).fill = CLIENT_FILL
                ws4.cell(row=i, column=c).font = CLIENT_FONT
        elif i % 2 == 0:
            for c in range(1, len(B) + 1):
                ws4.cell(row=i, column=c).fill = ALT_FILL
    widths(ws4)

    ws5 = wb.create_sheet("Store x Keyword")
    SK = ["store_name", "pincode", "area", "keyword", "status", "results",
          "client_products", "client_best_rank", "client_sov_pct",
          "lat", "lng", "api_store_id"]
    head(ws5, SK)
    for i, r in enumerate(sorted(store_kw, key=lambda x: (x["store_name"], x["keyword"])),
                          start=2):
        ws5.append([r.get(c, "") for c in SK])
        if r["status"] != "OK":
            for c in range(1, len(SK) + 1):
                ws5.cell(row=i, column=c).fill = PatternFill("solid", fgColor="FFF4CE")
        elif r["client_products"]:
            for c in range(1, len(SK) + 1):
                ws5.cell(row=i, column=c).fill = CLIENT_FILL
    ws5.auto_filter.ref = ws5.dimensions
    widths(ws5)

    ws6 = wb.create_sheet("Summary")
    head(ws6, ["Metric", "Value"])
    prices = [d["selling_price_rs"] for d in client_rows if d["selling_price_rs"]]
    stats = [
        ("Client brand", CLIENT_BRAND), ("City", "Bengaluru"),
        ("Dark stores", len(stores)), ("Keywords", len(KEYWORDS)),
        ("Searches run", len(store_kw)),
        ("Searches OK", len(ok_rows)),
        ("Searches failed", len(store_kw) - len(ok_rows)),
        ("Total product rows", len(detail)),
        (f"{CLIENT_BRAND} rows", len(client_rows)),
        (f"{CLIENT_BRAND} share of voice",
         f"{len(client_rows)/len(detail)*100:.1f}%" if detail else "—"),
        ("Distinct brands seen", len(brands)),
        (f"{CLIENT_BRAND} best rank",
         min((d["position"] for d in client_rows), default="—")),
        (f"{CLIENT_BRAND} avg rank",
         f"{sum(d['position'] for d in client_rows)/len(client_rows):.1f}"
         if client_rows else "—"),
        (f"{CLIENT_BRAND} avg price",
         f"Rs{sum(prices)/len(prices):.2f}" if prices else "—"),
        ("Result cap per search", RESULT_CAP),
        ("Scraped at", at),
    ]
    for i, (m, v) in enumerate(stats, start=2):
        ws6.cell(row=i, column=1, value=m)
        ws6.cell(row=i, column=2, value=v)
        if i % 2 == 0:
            for c in (1, 2):
                ws6.cell(row=i, column=c).fill = ALT_FILL
    widths(ws6)

    for w in (ws1, ws2, ws3, ws4, ws5, ws6):
        banner(w, T, S)

    out = BACKEND / f"zepto_keyword_competitors_bengaluru_{ts}.xlsx"
    wb.save(out)
    print(f"\nSaved -> {out}")


if __name__ == "__main__":
    asyncio.run(main())
