"""
Build the master all-India dark store workbook.

Merges:
  1. zepto_all_stores_combined_*.xlsx  — the original 1064 stores (pincode, area,
     city, state all present)
  2. zepto_<city>_grid_scan_*.xlsx     — the grid scans, contributing brand-new
     stores the pincode method never found

Grid-discovered stores arrive with only a coordinate, so their pincode/area are
filled by reverse geocoding (OpenStreetMap Nominatim, 1 req/sec per its policy).
Border-spill stores are skipped — they already belong to another city's sheet and
counting them again would double-count.

Run from backend/:
    python -m scripts.zepto_build_master_all_cities
Output: zepto_master_all_cities_TIMESTAMP.xlsx
"""
import re
import sys
import time
from pathlib import Path
from datetime import datetime

sys.path.insert(0, str(Path(__file__).parent.parent))

import httpx
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BACKEND = Path(__file__).parent.parent

# grid-scan file prefix -> the city sheet it belongs to
GRID_CITY = {
    "bengaluru":  "Bangalore",
    "hyderabad":  "Hyderabad",
    "chennai":    "Chennai",
    "pune":       "Pune",
    "kolkata":    "Kolkata",
    "ahmedabad":  "Ahmedabad",
    "mumbai":     "Mumbai",
    "delhi_ncr":  "Delhi NCR",
}

# Combined-file sheet name -> display name used in the summary
DISPLAY = {"Bangalore": "Bengaluru"}

TIERS = {
    **{c: "Tier 1" for c in ["Ahmedabad", "Bengaluru", "Chennai", "Delhi NCR",
                             "Hyderabad", "Kolkata", "Mumbai", "Pune"]},
    **{c: "Tier 2" for c in ["Agra", "Amritsar", "Chandigarh", "Coimbatore", "Dehradun",
                             "Hubballi", "Indore", "Jaipur", "Jalandhar", "Kanpur", "Kochi",
                             "Kota", "Lucknow", "Ludhiana", "Madurai", "Mysuru", "Nagpur",
                             "Nashik", "Patiala", "Prayagraj", "Puducherry", "Rajkot",
                             "Surat", "Udaipur", "Vadodara", "Varanasi", "Vijayawada",
                             "Warangal"]},
    **{c: "Tier 3" for c in ["Ambala", "Bareilly", "Belagavi", "Bhiwadi",
                             "Chhatrapati Sambhaji Nagar", "Davangere", "Gorakhpur",
                             "Guntur", "Hisar", "Hosur", "Karimnagar", "Karnal", "Meerut",
                             "Mehsana", "Mohali", "Palakkad", "Panchkula", "Panipat",
                             "Rewari", "Sonipat", "Tumkuru", "Vellore"]},
}

NOMINATIM = "https://nominatim.openstreetmap.org/reverse"
UA = "zepto-darkstore-research/1.0 (contact: soumi97bag@gmail.com)"
RATE_LIMIT_S = 1.1

COLS = ["store_id", "store_name", "city", "tier", "pincode", "area", "state",
        "lat", "lng", "discovery", "hit_count", "secondary_store_id"]

HDR_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
HDR_FILL = PatternFill("solid", fgColor="2E4057")
HDR_ALIGN = Alignment(horizontal="center", vertical="center")
ALT_FILL = PatternFill("solid", fgColor="F0F4F8")
NEW_FILL = PatternFill("solid", fgColor="D9F7E6")
T1 = PatternFill("solid", fgColor="1A3C5E")
T2 = PatternFill("solid", fgColor="2E6DA4")
T3 = PatternFill("solid", fgColor="5B9BD5")
TOT = PatternFill("solid", fgColor="048A81")
TITLE_FILL = PatternFill("solid", fgColor="1A3C5E")


def latest(pattern: str) -> Path | None:
    f = sorted(BACKEND.glob(pattern))
    return f[-1] if f else None


def read_combined() -> tuple[dict, dict]:
    """(store_id -> record, sheet -> [store_ids])"""
    path = latest("zepto_all_stores_combined_*.xlsx")
    if not path:
        raise SystemExit("No combined workbook found.")
    print(f"Base file : {path.name}")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    stores, by_city = {}, {}
    for sheet in wb.sheetnames:
        if sheet in ("All Stores", "Summary"):
            continue
        rows = list(wb[sheet].iter_rows(values_only=True))
        if not rows:
            continue
        hdr = [str(h).strip().lower() if h else "" for h in rows[0]]
        if "store_id" not in hdr:
            continue
        ids = []
        for r in rows[1:]:
            if not r or not r[0]:
                continue
            d = dict(zip(hdr, r))
            sid = str(d.get("store_id") or "").strip()
            if not sid:
                continue
            ids.append(sid)
            if sid in stores:
                continue
            city = DISPLAY.get(sheet, sheet)
            stores[sid] = {
                "store_id": sid,
                "store_name": str(d.get("store_name") or "").strip(),
                "city": city,
                "tier": TIERS.get(city, ""),
                "pincode": str(d.get("source_pincode") or "").strip(),
                "area": str(d.get("area") or "").strip(),
                "state": str(d.get("state") or "").strip(),
                "lat": d.get("lat"), "lng": d.get("lng"),
                "discovery": "pincode_scan",
                "hit_count": d.get("hit_count") or "",
                "secondary_store_id": str(d.get("secondary_store_id") or "").strip(),
            }
        by_city[DISPLAY.get(sheet, sheet)] = ids
    wb.close()
    return stores, by_city


def read_grid_new(key: str, sheet_city: str) -> list[dict]:
    """Brand-new stores from a grid scan (border spill excluded)."""
    path = latest(f"zepto_{key}_grid_scan_*.xlsx")
    if not path:
        return []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "NEW Stores" not in wb.sheetnames:
        wb.close()
        return []
    rows = list(wb["NEW Stores"].iter_rows(values_only=True))
    wb.close()
    if len(rows) < 4:
        return []
    hdr = [str(h).strip().lower() if h else "" for h in rows[2]]
    out = []
    city = DISPLAY.get(sheet_city, sheet_city)
    for r in rows[3:]:
        if not r or not r[0]:
            continue
        d = dict(zip(hdr, r))
        out.append({
            "store_id": str(d.get("store_id") or "").strip(),
            "store_name": str(d.get("store_name") or "").strip(),
            "city": city,
            "tier": TIERS.get(city, ""),
            "pincode": "", "area": "", "state": "",
            "lat": d.get("lat"), "lng": d.get("lng"),
            "discovery": "grid_scan",
            "hit_count": d.get("hit_count") or "",
            "secondary_store_id": str(d.get("secondary") or "").strip(),
        })
    print(f"  {city:<12} +{len(out)} new from grid  ({path.name})")
    return out


def read_bengaluru_pincodes() -> dict:
    """Re-use pincodes already geocoded for Bengaluru, to avoid repeat lookups."""
    path = latest("zepto_bengaluru_stores_with_pincode_*.xlsx")
    if not path:
        return {}
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows = list(wb["Bengaluru Stores"].iter_rows(values_only=True))
    wb.close()
    hdr = [str(h).strip().lower() if h else "" for h in rows[2]]
    out = {}
    for r in rows[3:]:
        if not r or not r[0]:
            continue
        d = dict(zip(hdr, r))
        sid = str(d.get("store_id") or "").strip()
        pin = str(d.get("pincode") or "").strip()
        if sid and pin and pin != "UNRESOLVED":
            out[sid] = {"pincode": pin,
                        "area": str(d.get("area") or "").strip(),
                        "state": str(d.get("state") or "").strip()}
    return out


def reverse_geocode(client, lat, lng):
    for attempt in range(1, 4):
        try:
            r = client.get(NOMINATIM,
                           params={"lat": lat, "lon": lng, "format": "jsonv2",
                                   "addressdetails": 1, "zoom": 18},
                           headers={"User-Agent": UA, "Accept-Language": "en"},
                           timeout=20)
            if r.status_code == 200:
                a = (r.json() or {}).get("address") or {}
                area = next((a[k] for k in ("neighbourhood", "suburb", "village", "town",
                                            "city_district", "residential", "quarter",
                                            "hamlet") if a.get(k)), "")
                return {"pincode": (a.get("postcode") or "").replace(" ", ""),
                        "area": area, "state": a.get("state") or ""}
            if r.status_code in (429, 503):
                time.sleep(2.0 * attempt)
                continue
        except Exception:
            pass
        time.sleep(1.0 * attempt)
    return None


def widths(ws, mx=48):
    for col in ws.columns:
        cl = get_column_letter(col[0].column)
        ml = max((len(str(c.value)) for c in col if c.value), default=0)
        ws.column_dimensions[cl].width = min(max(ml + 2, 9), mx)


def head(ws, cols):
    ws.append(cols)
    for cell in ws[1]:
        cell.font, cell.fill, cell.alignment = HDR_FONT, HDR_FILL, HDR_ALIGN
    ws.row_dimensions[1].height = 18
    ws.freeze_panes = "A2"


def banner(ws, title, sub):
    ws.insert_rows(1, amount=2)
    mc = ws.max_column or 1
    ws["A1"], ws["A2"] = title, sub
    ws["A1"].font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="444444")
    ws["A2"].fill = ALT_FILL
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 15
    if mc > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=mc)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=mc)


def main():
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    scraped_at = datetime.now().strftime("%d-%b-%Y %H:%M")

    print("Building master all-India dark store workbook")
    print("=" * 74)
    stores, by_city = read_combined()
    base_n = len(stores)
    print(f"Base unique stores: {base_n}\n")

    print("Grid scan contributions:")
    new_rows = []
    for key, sheet_city in GRID_CITY.items():
        for rec in read_grid_new(key, sheet_city):
            if rec["store_id"] and rec["store_id"] not in stores:
                new_rows.append(rec)
    print(f"Total brand new: {len(new_rows)}\n")

    # Fill pincode/area — reuse Bengaluru's existing work, geocode the rest.
    cached = read_bengaluru_pincodes()
    todo = []
    for rec in new_rows:
        c = cached.get(rec["store_id"])
        if c:
            rec.update(pincode=c["pincode"], area=c["area"], state=c["state"])
        else:
            todo.append(rec)
    print(f"Pincodes reused from Bengaluru file : {len(new_rows) - len(todo)}")
    print(f"Need reverse geocoding              : {len(todo)}"
          f"  (~{len(todo)*RATE_LIMIT_S/60:.1f} min)\n")

    if todo:
        with httpx.Client() as client:
            for i, rec in enumerate(todo, 1):
                g = reverse_geocode(client, float(rec["lat"]), float(rec["lng"]))
                if g and g["pincode"]:
                    rec.update(pincode=g["pincode"],
                               area=g["area"] or rec["store_name"],
                               state=g["state"])
                    print(f"  [{i:>2}/{len(todo)}] {rec['store_name'][:30]:<30} -> "
                          f"{g['pincode']}  {g['area'][:24]}")
                else:
                    rec["pincode"] = "UNRESOLVED"
                    print(f"  [{i:>2}/{len(todo)}] {rec['store_name'][:30]:<30} -> UNRESOLVED")
                time.sleep(RATE_LIMIT_S)

    for rec in new_rows:
        stores[rec["store_id"]] = rec
        by_city.setdefault(rec["city"], []).append(rec["store_id"])

    total = len(stores)
    print(f"\nTOTAL unique stores: {base_n} -> {total}  (+{total - base_n})")

    # ── Workbook ─────────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    b_title = f"Zepto Dark Stores - All India Master - {total} unique stores"
    b_sub = (f"Generated: {scraped_at}  |  base {base_n} from pincode scans "
             f"+ {len(new_rows)} from 670 m grid scans  |  each store counted once")

    # Sheet 1 — every store
    ws = wb.active
    ws.title = "All Stores"
    head(ws, COLS)
    ordered = sorted(stores.values(),
                     key=lambda r: (r["tier"] or "Tier 9", r["city"],
                                    str(r["pincode"]), r["store_name"]))
    for i, r in enumerate(ordered, start=2):
        ws.append([r.get(c, "") for c in COLS])
        if r["discovery"] == "grid_scan":
            for c in range(1, len(COLS) + 1):
                ws.cell(row=i, column=c).fill = NEW_FILL
        elif i % 2 == 0:
            for c in range(1, len(COLS) + 1):
                ws.cell(row=i, column=c).fill = ALT_FILL
    ws.auto_filter.ref = ws.dimensions
    widths(ws)

    # Sheet 2 — city summary
    ws2 = wb.create_sheet("City Summary")
    head(ws2, ["City", "Tier", "Unique Dark Stores", "From Grid Scan", "Previously"])
    counts = {}
    for r in stores.values():
        counts.setdefault(r["city"], {"n": 0, "grid": 0})
        counts[r["city"]]["n"] += 1
        if r["discovery"] == "grid_scan":
            counts[r["city"]]["grid"] += 1
    order = {"Tier 1": 0, "Tier 2": 1, "Tier 3": 2, "": 3}
    rows_sorted = sorted(counts.items(),
                         key=lambda kv: (order.get(TIERS.get(kv[0], ""), 3), kv[0]))
    for i, (city, d) in enumerate(rows_sorted, start=2):
        tier = TIERS.get(city, "")
        ws2.append([city, tier, d["n"], d["grid"] or "", d["n"] - d["grid"]])
        fill = {"Tier 1": T1, "Tier 2": T2, "Tier 3": T3}.get(tier)
        if fill:
            for c in (1, 2):
                ws2.cell(row=i, column=c).fill = fill
                ws2.cell(row=i, column=c).font = Font(name="Calibri", size=10,
                                                      bold=True, color="FFFFFF")
        if d["grid"]:
            ws2.cell(row=i, column=4).fill = NEW_FILL
    tr = len(rows_sorted) + 2
    ws2.cell(row=tr, column=1, value="TOTAL")
    ws2.cell(row=tr, column=3, value=total)
    ws2.cell(row=tr, column=4, value=len(new_rows))
    ws2.cell(row=tr, column=5, value=base_n)
    for c in range(1, 6):
        ws2.cell(row=tr, column=c).fill = TOT
        ws2.cell(row=tr, column=c).font = Font(name="Calibri", size=11,
                                               bold=True, color="FFFFFF")
    widths(ws2)

    # Sheet 3 — by pincode
    ws3 = wb.create_sheet("By Pincode")
    head(ws3, ["City", "Pincode", "Stores", "Store Names"])
    bp = {}
    for r in stores.values():
        bp.setdefault((r["city"], str(r["pincode"] or "")), []).append(r["store_name"])
    for i, ((city, pin), names) in enumerate(sorted(bp.items()), start=2):
        ws3.append([city, pin, len(names), " | ".join(names)])
        if i % 2 == 0:
            for c in range(1, 5):
                ws3.cell(row=i, column=c).fill = ALT_FILL
    widths(ws3)

    # One sheet per city
    for city, _ in rows_sorted:
        safe = re.sub(r"[\\/*?:\[\]]", "-", city)[:28]
        wsx = wb.create_sheet(safe)
        head(wsx, COLS)
        rs = sorted((r for r in stores.values() if r["city"] == city),
                    key=lambda r: (str(r["pincode"]), r["store_name"]))
        for i, r in enumerate(rs, start=2):
            wsx.append([r.get(c, "") for c in COLS])
            if r["discovery"] == "grid_scan":
                for c in range(1, len(COLS) + 1):
                    wsx.cell(row=i, column=c).fill = NEW_FILL
            elif i % 2 == 0:
                for c in range(1, len(COLS) + 1):
                    wsx.cell(row=i, column=c).fill = ALT_FILL
        wsx.auto_filter.ref = wsx.dimensions
        widths(wsx)

    for w in (ws, ws2, ws3):
        banner(w, b_title, b_sub)

    out = BACKEND / f"zepto_master_all_cities_{ts}.xlsx"
    wb.save(out)
    print(f"Cities: {len(counts)}   Sheets: {len(wb.sheetnames)}")
    print(f"Saved -> {out}")


if __name__ == "__main__":
    main()
