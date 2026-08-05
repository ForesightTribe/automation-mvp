"""
Scrape Zepto dark stores for all Lucknow pincodes.
For each pincode, fetches ALL autocomplete suggestions, calls place/details for
each, and takes the lat/lng as-is. No grid scan. No reverse geocoding.

Run from backend/:
    python -m scripts.zepto_lucknow_areas
"""
import asyncio
import sys
from pathlib import Path
from collections import Counter, defaultdict
from datetime import datetime

sys.path.insert(0, str(Path(__file__).parent.parent))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from playwright.async_api import async_playwright
from scraper.platforms.zepto.dark_store.endpoints import LUCKNOW_PINCODES

_BFF = "https://bff-gateway.zepto.com"

_INIT_SCRIPT = """
(() => {
    const injectCoords = (url) => {
        if (typeof url !== 'string' || !url.includes('/lms/api/v2/get_page')) return url;
        const lat = localStorage.getItem('__probe_lat');
        const lng = localStorage.getItem('__probe_lng');
        if (!lat || !lng) return url;
        try {
            const u = new URL(url, location.href);
            u.searchParams.set('latitude', lat);
            u.searchParams.set('longitude', lng);
            return u.toString();
        } catch(e) { return url; }
    };
    const origFetch = window.fetch;
    window.fetch = function(url, opts) {
        return origFetch.apply(this, [injectCoords(url), opts]);
    };
    const origOpen = XMLHttpRequest.prototype.open;
    XMLHttpRequest.prototype.open = function(method, url, ...rest) {
        return origOpen.call(this, method, injectCoords(url), ...rest);
    };
})();
"""

_HEADERS = {
    "Accept": "application/json",
    "Origin": "https://www.zepto.com",
    "Referer": "https://www.zepto.com/",
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "app_sub_platform": "WEB", "app_version": "16.20.0",
    "platform": "WEB", "tenant": "ZEPTO",
}

# Lucknow bounding box — rejects coordinates outside Lucknow region
_LAT_MIN, _LAT_MAX = 26.6, 27.2
_LNG_MIN, _LNG_MAX = 80.5, 81.3


async def get_predictions(query: str, ctx) -> list[dict]:
    for attempt in range(2):
        try:
            r = await ctx.request.get(
                f"{_BFF}/api/v1/maps/place/autocomplete/",
                params={"place_name": query}, headers=_HEADERS, timeout=10000,
            )
            if r.status != 200:
                return []
            data = await r.json()
            return data.get("data", {}).get("predictions") or data.get("predictions") or []
        except Exception as e:
            if attempt == 0:
                await asyncio.sleep(1)
            else:
                print(f"    [autocomplete error] {query}: {e}")
    return []


async def get_details(place_id: str, ctx) -> dict | None:
    for attempt in range(2):
        try:
            r = await ctx.request.get(
                f"{_BFF}/api/v1/maps/place/details/",
                params={"place_id": place_id}, headers=_HEADERS, timeout=10000,
            )
            if r.status != 200:
                return None
            data = await r.json()
            result = data.get("data", {}).get("result") or data.get("result") or {}
            geo = result.get("geometry", {})
            loc = geo.get("location", {})
            lat, lng = loc.get("lat"), loc.get("lng")
            if not lat or not lng:
                return None
            city = state = None
            for comp in result.get("address_components", []):
                t = comp.get("types", [])
                if ("locality" in t or "administrative_area_level_2" in t) and not city:
                    city = comp.get("long_name")
                if "administrative_area_level_1" in t and not state:
                    state = comp.get("long_name")
            return {
                "lat": lat, "lng": lng,
                "location_type": geo.get("location_type", ""),
                "city": city or "",
                "state": state or "",
            }
        except Exception as e:
            if attempt == 0:
                await asyncio.sleep(1)
            else:
                print(f"    [details error] {place_id}: {e}")
    return None


async def probe_store_at(lat: float, lng: float, page) -> dict:
    result = {}
    done = asyncio.Event()

    async def on_response(response):
        if "/lms/api/v2/get_page" in response.url and not done.is_set():
            try:
                data = await response.json()
                svc = data.get("storeServiceableResponse", {})
                sid = svc.get("storeId")
                if sid:
                    result["store_id"] = str(sid)
                    result["secondary_ids"] = [str(s) for s in svc.get("secondaryStoreIds", [])]
                    result["store_name"] = data.get("storeDetailsResponse", {}).get("name", "")
                done.set()
            except Exception:
                pass

    try:
        await page.evaluate(f"""() => {{
            localStorage.setItem('__probe_lat', '{lat}');
            localStorage.setItem('__probe_lng', '{lng}');
        }}""")
    except Exception:
        pass

    page.on("response", on_response)
    try:
        await page.reload(timeout=20000, wait_until="domcontentloaded")
        try:
            await asyncio.wait_for(done.wait(), timeout=8)
        except asyncio.TimeoutError:
            pass
    except Exception:
        pass
    finally:
        page.remove_listener("response", on_response)
    return result


def apply_heading(ws, title: str, generated: str):
    ws.insert_rows(1, amount=2)
    max_col = ws.max_column or 1
    ws["A1"] = title
    ws["A2"] = f"Generated: {generated}  |  Source: zepto.com"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="2E4057")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="444444")
    ws["A2"].fill = PatternFill("solid", fgColor="F0F4F8")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 16
    for cell in ws[3]:
        if cell.value:
            cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="048A81")
            cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 18
    if max_col > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
    for col_idx in range(1, max_col + 1):
        max_len = 0
        for row_idx in range(3, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 40)


async def main():
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    print(f"Zepto Lucknow Dark Store Scraper — {ts}")
    print(f"Pincodes: {len(LUCKNOW_PINCODES)} | all suggestions | exact details-API coords")
    print("=" * 60)

    seen_stores: set[str] = set()
    seen_coords: set[tuple] = set()
    rows: list[dict] = []
    hit_counts: dict[str, int] = defaultdict(int)

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        ctx = await browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
        )
        page = await ctx.new_page()
        await page.add_init_script(_INIT_SCRIPT)

        print("Warming up browser (WAF)...")
        try:
            await page.goto("https://www.zepto.com", timeout=30000, wait_until="domcontentloaded")
            await page.wait_for_timeout(3000)
            print("Browser ready.\n")
        except Exception as e:
            print(f"Warmup: {e}\n")

        total = len(LUCKNOW_PINCODES)
        for i, (pincode, label) in enumerate(LUCKNOW_PINCODES.items(), 1):
            print(f"\n[{i}/{total}] {pincode} — {label}")

            preds_by_pin = await get_predictions(pincode, ctx)
            await asyncio.sleep(0.2)
            preds_by_name = await get_predictions(f"{label} {pincode}", ctx)

            seen_ids: set[str] = set()
            preds = []
            for p in preds_by_pin:
                pid = p.get("place_id")
                if pid and pid not in seen_ids:
                    seen_ids.add(pid)
                    p["_search_source"] = "pincode"
                    preds.append(p)
            for p in preds_by_name:
                pid = p.get("place_id")
                if pid and pid not in seen_ids:
                    seen_ids.add(pid)
                    p["_search_source"] = "area_name"
                    preds.append(p)

            if not preds:
                print(f"  no predictions")
                await asyncio.sleep(0.4)
                continue

            for pred in preds:
                place_id = pred.get("place_id")
                name = pred.get("description") or ""
                search_source = pred.get("_search_source", "pincode")
                if not place_id:
                    continue

                details = await get_details(place_id, ctx)
                if not details:
                    print(f"  ✗ {name[:55]} → no coords")
                    await asyncio.sleep(0.15)
                    continue

                lat, lng = details["lat"], details["lng"]
                if not (_LAT_MIN <= lat <= _LAT_MAX and _LNG_MIN <= lng <= _LNG_MAX):
                    print(f"  ✗ {name[:55]} → outside Lucknow ({lat:.3f},{lng:.3f}), skipped")
                    await asyncio.sleep(0.15)
                    continue

                coord_key = (round(lat, 4), round(lng, 4))
                if coord_key in seen_coords:
                    print(f"  ~ {name[:55]} → dup coords")
                    await asyncio.sleep(0.15)
                    continue
                seen_coords.add(coord_key)

                loc_type = details["location_type"]
                print(f"  ✓ {name[:55]} [{loc_type}] ({lat:.5f},{lng:.5f})")

                r = await probe_store_at(lat, lng, page)
                sid = r.get("store_id")
                if sid:
                    hit_counts[sid] += 1
                    if sid not in seen_stores:
                        seen_stores.add(sid)
                        rows.append({
                            "store_id": sid,
                            "store_name": r.get("store_name", ""),
                            "area": name,
                            "city": details["city"],
                            "state": details["state"],
                            "lat": lat,
                            "lng": lng,
                            "location_type": loc_type,
                            "hit_count": 0,
                            "secondary_store_id": ",".join(r.get("secondary_ids", [])),
                            "source_pincode": pincode,
                            "search_source": search_source,
                        })
                        print(f"    NEW {sid} «{r.get('store_name', '')}»")
                    else:
                        print(f"    dup {sid[:8]}")
                else:
                    print(f"    no store")
                await asyncio.sleep(0.3)

            await asyncio.sleep(0.2)

        try:
            await browser.close()
        except Exception:
            pass

    for row in rows:
        row["hit_count"] = hit_counts[row["store_id"]]

    print(f"\n{'='*60}")
    print(f"Unique stores found: {len(rows)}")

    out = Path(__file__).parent.parent / f"zepto_lucknow_stores_{ts}.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "stores"

    hdr = ["store_id", "store_name", "city", "state", "area",
           "lat", "lng", "location_type", "hit_count", "secondary_store_id", "source_pincode", "search_source"]
    ws.append(hdr)
    for row in sorted(rows, key=lambda x: (x["city"], x["source_pincode"])):
        ws.append([row[h] for h in hdr])

    ws2 = wb.create_sheet("by_city")
    ws2.append(["city", "store_count"])
    city_counts = Counter(r["city"] for r in rows)
    for city, count in sorted(city_counts.items()):
        ws2.append([city, count])
        print(f"  {city}: {count} stores")

    generated = datetime.now().strftime("%d-%b-%Y  %H:%M")
    apply_heading(ws, "Zepto Dark Stores Lucknow — All Pincode Suggestions", generated)
    apply_heading(ws2, "Zepto Lucknow — Store Count by City", generated)

    wb.save(out)
    print(f"\nSaved → {out}")


if __name__ == "__main__":
    asyncio.run(main())
