import asyncio
from datetime import date as _date, timedelta
from typing import Optional
import typer
from rich.console import Console
from rich.markup import escape
from rich.table import Table
from app.core.database import AsyncSessionLocal
from app.utils.logger import logger
from platform_auth import service as auth_service
from platform_auth.errors import AuthError
from scraper.platforms.zepto.dashboard_data.seller.session_health import (
    ensure_healthy_session,
    SessionUnhealthy,
)
from scraper.utils.jobs import create_scrape_job, complete_scrape_job, fail_scrape_job
from scraper.platforms.zepto.dashboard_data.seller.scraper import (
    validate as zepto_seller_validate,
    discover_ids as zepto_discover_ids,
    fetch_sales_overview as zepto_fetch_sales_overview,
    fetch_product_performance as zepto_fetch_product_performance,
    fetch_product_performance_by_city as zepto_fetch_product_perf_by_city,
    NoDataYet as ZeptoNoDataYet,
    fetch_pos as zepto_fetch_pos,
    fetch_grns as zepto_fetch_grns,
    fetch_asns as zepto_fetch_asns,
    fetch_po_items as zepto_fetch_po_items,
    capture_ads_headers as zepto_capture_ads_headers,
    fetch_ad_campaigns as zepto_fetch_ad_campaigns,
    fetch_ads_tabular as zepto_fetch_ads_tabular,
)
from scraper.platforms.zepto.dashboard_data.seller.parser import (
    parse_sales_daily as parse_zepto_sales_daily,
    parse_product_perf as parse_zepto_product_perf,
    parse_product_city as parse_zepto_product_city,
    parse_pos as parse_zepto_pos,
    parse_grns as parse_zepto_grns,
    parse_asns as parse_zepto_asns,
    parse_po_items as parse_zepto_po_items,
    parse_ad_campaigns as parse_zepto_ad_campaigns,
    parse_ad_tabular_campaigns as parse_zepto_ad_tabular_campaigns,
    parse_ad_keywords as parse_zepto_ad_keywords,
    parse_ad_products as parse_zepto_ad_products,
    parse_ad_breakdown as parse_zepto_ad_breakdown,
)
from scraper.platforms.zepto.dashboard_data.seller.storage import (
    save_sales_results as zepto_save_sales_results,
    save_po_results as zepto_save_po_results,
    save_ad_results as zepto_save_ad_results,
)
from scraper.platforms.blinkit.dashboard_data.marketing.scraper import scrape
from scraper.platforms.blinkit.dashboard_data.marketing.parser import (
    parse_campaign,
    parse_campaign_daily,
    parse_campaign_detail,
    parse_campaign_keywords,
    parse_sponsored_sov,
    parse_brand_collection,
    parse_visibility_plan,
)
from scraper.platforms.blinkit.dashboard_data.marketing.storage import save_scrape_results
from scraper.platforms.blinkit.dashboard_data.seller import scraper as seller_scraper
from scraper.platforms.blinkit.dashboard_data.seller.parser import (
    parse_sale_row,
    parse_sales_summary,
    parse_po_row,
    parse_po_item,
    parse_po_summary,
    parse_soh_row,
    parse_scorecard_weekly,
    parse_scorecard_facility,
    parse_scorecard_key_sku,
)
from scraper.platforms.blinkit.dashboard_data.seller.storage import (
    save_scrape_results as seller_save_results,
    save_po_results,
    save_soh_results,
    save_scorecard_results,
)

app = typer.Typer(help="Run scrapers and view results.")
console = Console()

# Gap between per-day Zepto product calls. A daily production run makes one
# call and never waits; this only paces multi-day backfills, where a burst of
# back-to-back requests is the kind of pattern this site's WAF reacts to.
_ZEPTO_DAY_GAP_S = 1.5


@app.command("blinkit")
def scrape_blinkit(
    tenant_id: str = typer.Option(..., "--tenant", "-t", help="Tenant ID"),
    date_from: str = typer.Option(None, "--from", help="Start date YYYY-MM-DD (default: 7 days ago)"),
    date_to: str = typer.Option(None, "--to", help="End date YYYY-MM-DD (default: today)"),
    limit: int = typer.Option(None, "--limit", help="Test mode: only the N most-active campaigns"),
    save: bool = typer.Option(True, "--save/--no-save", help="Save results to PostgreSQL"),
):
    """Scrape the Blinkit marketing dashboard for a date window.

    One pass fetches the campaign list, each campaign's daily metric series + its
    keyword/recommendation breakdown, plus SOV / collections / plans. Use --from
    to backfill (e.g. --from 30 days ago); the daily run defaults to the last week
    so late metric revisions are picked up. Use --limit to smoke-test a few
    campaigns without the full-volume run.
    """
    asyncio.run(_scrape_blinkit(tenant_id, date_from, date_to, limit, save))


async def _scrape_blinkit(
    tenant_id: str, date_from: str | None, date_to: str | None, limit: int | None, save: bool
) -> None:
    start = _date.fromisoformat(date_from) if date_from else _date.today() - timedelta(days=7)
    end = _date.fromisoformat(date_to) if date_to else _date.today()
    snapshot_date = end.isoformat()

    async with AsyncSessionLocal() as db:
        job_id = None
        try:
            # ensure() = load → probe → refresh → re-login, doing the least work
            # that yields a session known to work. Replaces a bare load, which
            # happily returned a session that had died days earlier and let the
            # scrape fail deep inside Playwright instead.
            storage_state = (await auth_service.ensure(db, tenant_id, "blinkit")).storage_state

            job_id = await create_scrape_job(db, tenant_id, "blinkit_marketing")

            with console.status(f"[cyan]Scraping Blinkit marketing {start}→{end}...[/cyan]"):
                raw = await scrape(storage_state, start, end, limit=limit)

            campaign_detail = raw.get("campaign_detail") or {}
            cities = raw.get("cities") or {}
            campaigns = [
                parse_campaign(c, tenant_id, job_id,
                               detail=campaign_detail.get(c["id"]), cities=cities)
                for c in raw["campaigns"]
            ]
            type_by_id = {c["id"]: c.get("campaign_type") for c in raw["campaigns"]}

            keyword_bids = [
                row
                for cid, attrs in (raw.get("keyword_attributes") or {}).items()
                for row in parse_campaign_keywords(
                    attrs, campaign_detail.get(cid) or {}, cid, tenant_id, job_id
                )
            ]

            daily = [
                parse_campaign_daily(row, cid, type_by_id.get(cid), tenant_id, job_id)
                for cid, rows in raw["daily"].items()
                for row in rows
            ]
            detail = [
                d
                for cid, report in raw["detail"].items()
                for d in parse_campaign_detail(
                    report, cid, type_by_id.get(cid), snapshot_date, tenant_id, job_id
                )
            ]
            sov = [
                parse_sponsored_sov(s, tenant_id, job_id, snapshot_date)
                for s in raw["sponsored_sov"]
            ]
            collections = [parse_brand_collection(c, tenant_id, job_id) for c in raw["brand_collections"]]
            plans = [parse_visibility_plan(p, tenant_id, job_id) for p in raw["visibility_plans"]]

            if save:
                await save_scrape_results(db, campaigns, daily, detail, sov, collections,
                                          plans, keywords=keyword_bids)
                await complete_scrape_job(db, job_id)

            console.print(
                f"\n[bold cyan]Blinkit marketing[/bold cyan] ({start}→{end})\n"
                f"  campaigns: {len(campaigns)}\n"
                f"  daily rows: {len(daily)}\n"
                f"  detail rows: {len(detail)}\n"
                f"  keyword bid ranges: {len(keyword_bids)}\n"
                f"  sov: {len(sov)}  collections: {len(collections)}  plans: {len(plans)}"
            )
            _print_campaigns(campaigns)
            _print_targeting(campaigns)
            _print_keyword_bids(keyword_bids)
            _print_daily(daily)
            _print_detail(detail)
            _print_sov(sov)
            _print_collections(collections)
            _print_plans(plans)

        except typer.Exit:
            raise
        except AuthError:
            # Must escape the generic handler below. cli/main.py turns AuthError
            # into exit code 3, which the job runner records as `auth_expired` —
            # collapsing it into typer.Exit(1) here would bury every auth failure
            # among anonymous exit_1s, which is exactly how the seller breakage
            # went unnoticed for weeks.
            if job_id:
                await fail_scrape_job(db, job_id, "auth_expired")
            raise
        except Exception as e:
            # For DB errors, e.orig is the short asyncpg message; str(e) would dump
            # the entire (huge) statement + params, flooding the terminal.
            err = getattr(e, "orig", None) or e
            if job_id:
                await fail_scrape_job(db, job_id, str(err))
            console.print(f"[red]Scrape failed: {escape(str(err))}[/red]")
            raise typer.Exit(1)


def _print_campaigns(campaigns: list) -> None:
    if not campaigns:
        console.print("\n[dim]No campaigns found.[/dim]")
        return
    preview = campaigns[:5]
    console.print(f"\n[bold cyan]Campaigns[/bold cyan] (showing {len(preview)} of {len(campaigns)})")
    table = Table(show_header=True, header_style="bold")
    table.add_column("Name")
    table.add_column("Type", style="dim")
    table.add_column("Status")
    for c in preview:
        status_style = {"ACTIVE": "green", "PAUSED": "yellow", "STOPPED": "red"}.get(c["status"], "")
        status_text = f"[{status_style}]{c['status']}[/{status_style}]" if status_style else c["status"]
        table.add_row(c["name"], c["type"], status_text)
    console.print(table)


def _print_targeting(campaigns: list) -> None:
    """City targeting + the budget/pacing facts, from the per-campaign detail call (V7).
    Shows the CITY-targeted campaigns first — they are the ones the bid-rule form
    auto-fills from, and the rare case worth eyeballing after a scrape."""
    scoped = [c for c in campaigns if c.get("region_type")]
    if not scoped:
        console.print("\n[dim]No campaign targeting captured (detail calls returned nothing).[/dim]")
        return
    targeted = [c for c in scoped if c.get("cities")]
    preview = (targeted + [c for c in scoped if not c.get("cities")])[:10]
    console.print(
        f"\n[bold cyan]Targeting & budget floor[/bold cyan] "
        f"({len(targeted)} city-targeted of {len(scoped)}; showing {len(preview)})"
    )
    table = Table(show_header=True, header_style="bold")
    table.add_column("Campaign")
    table.add_column("Region")
    table.add_column("Cities")
    table.add_column("Pacing", style="dim")
    table.add_column("Budget", justify="right")
    table.add_column("Spent", justify="right")
    table.add_column("min_cpm", justify="right")
    for c in preview:
        cities = ", ".join(x["name"] for x in (c.get("cities") or [])) or "—"
        budget = c.get("daily_budget")
        table.add_row(
            (c.get("name") or "")[:34],
            c.get("region_type") or "—",
            cities[:40],
            c.get("pacing_type") or "—",
            "—" if budget is None else f"{budget:,.0f}",
            f"{c.get('billed_amount') or 0:,.0f}",
            "—" if c.get("min_cpm") is None else f"{c['min_cpm']:,}",
        )
    console.print(table)


def _print_keyword_bids(rows: list) -> None:
    """Blinkit's published bid range per keyword (V7). The interesting column is
    `Min` — the floor a bid rule may not go under — and whether the live bid is under it."""
    if not rows:
        console.print("\n[dim]No keyword bid ranges captured.[/dim]")
        return
    exact = [r for r in rows if r["match_type"] == "EXACT"]
    below = [
        r for r in exact
        if r.get("current_cpm") is not None and r.get("min_bid") is not None
        and r["current_cpm"] < r["min_bid"]
    ]
    preview = (below + [r for r in exact if r not in below])[:15]
    console.print(
        f"\n[bold cyan]Keyword bid ranges[/bold cyan] "
        f"(showing {len(preview)} of {len(exact)} EXACT, {len(rows)} rows all match types)"
    )
    table = Table(show_header=True, header_style="bold")
    table.add_column("Campaign", style="dim")
    table.add_column("Keyword")
    table.add_column("Bid", justify="right")
    table.add_column("Min", justify="right")
    table.add_column("Max", justify="right")
    table.add_column("Suggested", justify="right")
    table.add_column("Searches", justify="right")
    for r in preview:
        bid = r.get("current_cpm")
        under = bid is not None and r.get("min_bid") is not None and bid < r["min_bid"]
        sugg = (
            f"{r['suggested_min']:,}–{r['suggested_max']:,}"
            if r.get("suggested_min") and r.get("suggested_max") else "—"
        )
        table.add_row(
            str(r["campaign_id"]),
            r["keyword"][:28],
            "—" if bid is None else (f"[red]{bid:,}[/red]" if under else f"{bid:,}"),
            f"{r['min_bid']:,}" if r.get("min_bid") is not None else "—",
            f"{r['max_bid']:,}" if r.get("max_bid") is not None else "—",
            sugg,
            f"{r['keyword_searches']:,}" if r.get("keyword_searches") else "—",
        )
    console.print(table)
    if below:
        console.print(f"[yellow]{len(below)} live bid(s) sit below Blinkit's published minimum.[/yellow]")


def _print_daily(rows: list) -> None:
    if not rows:
        console.print("\n[dim]No daily rows captured.[/dim]")
        return
    preview = rows[:15]
    console.print(f"\n[bold cyan]Campaign Daily[/bold cyan] (showing {len(preview)} of {len(rows)} rows)")
    table = Table(show_header=True, header_style="bold")
    table.add_column("Campaign", style="dim")
    table.add_column("Date")
    table.add_column("Budget", justify="right")
    table.add_column("Impr", justify="right")
    table.add_column("ATC", justify="right")
    table.add_column("Qty", justify="right")
    table.add_column("Ad Sales", justify="right")
    table.add_column("RoAS", justify="right")
    for r in preview:
        table.add_row(
            str(r["campaign_id"]),
            str(r["date"]),
            f"₹{r['budget_consumed']:,.0f}",
            f"{r['impressions']:,}",
            f"{r['atc']:,}",
            f"{r['quantities_sold']:,}",
            f"₹{r['ad_sales']:,.0f}",
            f"{r['roas']}",
        )
    console.print(table)


def _print_detail(rows: list) -> None:
    if not rows:
        console.print("\n[dim]No detail rows captured.[/dim]")
        return
    preview = rows[:15]
    console.print(f"\n[bold cyan]Campaign Detail[/bold cyan] (showing {len(preview)} of {len(rows)} rows)")
    table = Table(show_header=True, header_style="bold")
    table.add_column("Campaign", style="dim")
    table.add_column("Type", style="dim")
    table.add_column("Target")
    table.add_column("Impr", justify="right")
    table.add_column("Budget", justify="right")
    table.add_column("Direct RoAS", justify="right")
    table.add_column("Total RoAS", justify="right")
    for r in preview:
        table.add_row(
            str(r["campaign_id"]),
            r["target_type"],
            r["target"],
            f"{r['impressions']:,}",
            f"₹{r['budget_consumed']:,.0f}",
            f"{r['direct_roas']}",
            f"{r['total_roas']}",
        )
    console.print(table)


def _print_sov(sov: list) -> None:
    if not sov:
        return
    preview = sov[:5]
    console.print(f"\n[bold cyan]Sponsored Share of Voice[/bold cyan] (showing {len(preview)} of {len(sov)} keywords)")
    table = Table(show_header=True, header_style="bold")
    table.add_column("Keyword")
    table.add_column("Monthly Searches", justify="right")
    table.add_column("SOV %", justify="right")
    for s in preview:
        table.add_row(s["keyword"], f"{s['monthly_searches']:,}", f"{s['sov']}%")
    console.print(table)


def _print_collections(collections: list) -> None:
    if not collections:
        return
    preview = collections[:5]
    console.print(f"\n[bold cyan]Brand Collections[/bold cyan] (showing {len(preview)} of {len(collections)})")
    table = Table(show_header=True, header_style="bold")
    table.add_column("Name")
    table.add_column("Type", style="dim")
    table.add_column("Products", justify="right")
    table.add_column("Created On")
    for c in preview:
        collection_type = "DYNAMIC" if c["is_dynamic"] else "STATIC"
        table.add_row(c["name"], collection_type, str(c["number_of_products"]), c.get("created_on", ""))
    console.print(table)


def _print_plans(plans: list) -> None:
    if not plans:
        return
    preview = plans[:5]
    console.print(f"\n[bold cyan]Visibility Plans[/bold cyan] (showing {len(preview)} of {len(plans)})")
    table = Table(show_header=True, header_style="bold")
    table.add_column("Name")
    table.add_column("Type", style="dim")
    table.add_column("Budget", justify="right")
    table.add_column("Start Date")
    table.add_column("End Date")
    table.add_column("Status")
    for p in preview:
        table.add_row(
            p["name"],
            p["type"],
            f"₹{p['budget']:,.0f}",
            (p.get("start_date") or "")[:10],
            (p.get("end_date") or "")[:10],
            p["status"],
        )
    console.print(table)


# ── Blinkit Seller ─────────────────────────────────────────────────────────────

@app.command("blinkit-seller")
def scrape_blinkit_seller(
    tenant_id: str = typer.Option(..., "--tenant", "-t", help="Tenant ID"),
    date_from: str = typer.Option(None, "--from", help="Start date YYYY-MM-DD (default: yesterday)"),
    date_to: str = typer.Option(None, "--to", help="End date YYYY-MM-DD (default: --from)"),
    sales: bool = typer.Option(False, "--sales", help="Scrape sales data"),
    po: bool = typer.Option(False, "--po", help="Scrape PO data"),
    soh: bool = typer.Option(False, "--soh", help="Scrape stock on hand"),
    po_days_back: int = typer.Option(90, "--po-days-back", help="Rolling window for PO fetch"),
    refetch_po_items: bool = typer.Option(
        False, "--refetch-po-items",
        help="Re-fetch line items for every PO in the window (one-time backfill of stale received qty)",
    ),
    save: bool = typer.Option(True, "--save/--no-save", help="Save results to MongoDB"),
):
    """Scrape Blinkit seller data. Pass --sales, --po, --soh, or none to run all three."""
    asyncio.run(_scrape_blinkit_seller(tenant_id, date_from, date_to, sales, po, soh, po_days_back, refetch_po_items, save))


def _date_range(date_from: str | None, date_to: str | None) -> list[str]:
    yesterday = (_date.today() - timedelta(days=1)).isoformat()
    start = _date.fromisoformat(date_from or yesterday)
    end = _date.fromisoformat(date_to or yesterday)
    days = (end - start).days + 1
    return [(start + timedelta(days=i)).isoformat() for i in range(days)]


async def _scrape_blinkit_seller(
    tenant_id: str,
    date_from: str | None,
    date_to: str | None,
    sales_flag: bool,
    po_flag: bool,
    soh_flag: bool,
    po_days_back: int,
    refetch_po_items: bool,
    save: bool,
) -> None:
    run_all = not sales_flag and not po_flag and not soh_flag
    run_sales = sales_flag or run_all
    run_po = po_flag or run_all
    run_soh = soh_flag or run_all

    async with AsyncSessionLocal() as db:
        storage_state = (
            await auth_service.ensure(db, tenant_id, "blinkit_seller")
        ).storage_state

        # ── Sales (loops per day) ──────────────────────────────────────────────
        if run_sales:
            days = _date_range(date_from, date_to)
            for day in days:
                job_id = None
                try:
                    job_id = await create_scrape_job(db, tenant_id, "blinkit_seller_sales")

                    with console.status(f"[cyan]Scraping sales {day}...[/cyan]"):
                        raw = await seller_scraper.scrape(storage_state, day)

                    parsed_sales = [parse_sale_row(r, tenant_id, job_id, raw["date"]) for r in raw["sales"]]
                    summary = parse_sales_summary(raw, tenant_id, job_id, raw["date"])

                    if save:
                        await seller_save_results(db, parsed_sales, summary)
                        await complete_scrape_job(db, job_id)

                    _print_seller_summary(summary, len(parsed_sales))
                    _print_seller_sales(parsed_sales)

                except Exception as e:
                    if job_id:
                        await fail_scrape_job(db, job_id, str(e))
                    console.print(f"[red]{day} failed: {escape(str(e))}[/red]")
                    if len(days) == 1:
                        raise typer.Exit(1)

        # ── PO (runs once) ────────────────────────────────────────────────────
        if run_po:
            po_job_id = None
            try:
                # po_number -> (po_state, total_grn_quantity) for the targeted
                # item refetch: a changed GRN or non-terminal state means the
                # line items may have moved and must be re-pulled.
                known_pos: dict[str, tuple[str | None, int | None]] = {}
                if save:
                    import uuid as _uuid
                    from sqlmodel import select as _select
                    from app.models.blinkit_seller import BlinkitPO as _BlinkitPO
                    rows = await db.execute(
                        _select(
                            _BlinkitPO.po_number,
                            _BlinkitPO.po_state,
                            _BlinkitPO.total_grn_quantity,
                        ).where(_BlinkitPO.tenant_id == _uuid.UUID(tenant_id))
                    )
                    known_pos = {pn: (state, grn) for pn, state, grn in rows.all()}
                    detail = (
                        "re-fetching ALL their line items"
                        if refetch_po_items
                        else "re-fetching only changed/in-flight ones"
                    )
                    console.print(
                        f"\n[dim]{len(known_pos)} existing POs in DB — {detail}[/dim]"
                    )

                po_job_id = await create_scrape_job(db, tenant_id, "blinkit_seller_po")

                with console.status("[cyan]Scraping PO data...[/cyan]"):
                    raw_po = await seller_scraper.scrape_po(
                        storage_state,
                        po_days_back=po_days_back,
                        known_pos=known_pos,
                        refetch_all_items=refetch_po_items,
                    )

                pos = [parse_po_row(r, tenant_id, po_job_id) for r in raw_po["pos"]]
                po_items = [
                    parse_po_item(it, r["po_number"], tenant_id, po_job_id)
                    for r in raw_po["pos"]
                    for it in r.get("items", [])
                ]
                snapshot = parse_po_summary(
                    raw_po["po_summary"], tenant_id, po_job_id, raw_po["po_window_start"]
                )

                if save:
                    await save_po_results(db, pos, po_items, snapshot)
                    await complete_scrape_job(db, po_job_id)

                _print_po_snapshot(snapshot, len(pos))
                _print_po_list(pos)

            except Exception as e:
                if po_job_id:
                    await fail_scrape_job(db, po_job_id, str(e))
                console.print(f"[red]PO scrape failed: {escape(str(e))}[/red]")
                raise typer.Exit(1)

        # ── SOH (runs once) ───────────────────────────────────────────────────
        if run_soh:
            soh_job_id = None
            try:
                soh_job_id = await create_scrape_job(db, tenant_id, "blinkit_seller_soh")

                with console.status("[cyan]Scraping stock on hand...[/cyan]"):
                    raw_soh = await seller_scraper.scrape_soh(storage_state)

                rows = [parse_soh_row(r, tenant_id, soh_job_id, raw_soh["date"]) for r in raw_soh["rows"]]

                if save:
                    await save_soh_results(db, rows)
                    await complete_scrape_job(db, soh_job_id)

                _print_soh(rows, raw_soh["date"])

            except Exception as e:
                if soh_job_id:
                    await fail_scrape_job(db, soh_job_id, str(e))
                console.print(f"[red]SOH scrape failed: {escape(str(e))}[/red]")
                raise typer.Exit(1)


# ── Print helpers ─────────────────────────────────────────────────────────────

def _print_seller_summary(summary: dict, total_rows: int) -> None:
    console.print(f"\n[bold cyan]Sales Summary[/bold cyan] ({summary['date']})")
    table = Table(show_header=False, box=None, padding=(0, 2))
    table.add_column(style="dim")
    table.add_column(style="bold")
    table.add_row("Total Rows", str(total_rows))
    table.add_row("Distinct SKUs", str(summary["distinct_skus"]))
    table.add_row("Distinct Categories", str(summary["distinct_categories"]))
    table.add_row("Max Selling Item", summary["max_sell_item"])
    console.print(table)


def _print_seller_sales(sales: list) -> None:
    if not sales:
        console.print("\n[dim]No sales data found.[/dim]")
        return
    preview = sales[:5]
    console.print(f"\n[bold cyan]Sales[/bold cyan] (showing {len(preview)} of {len(sales)} rows)")
    table = Table(show_header=True, header_style="bold")
    table.add_column("Item")
    table.add_column("Category", style="dim")
    table.add_column("City")
    table.add_column("Qty Sold", justify="right")
    table.add_column("MRP Value", justify="right")
    for r in preview:
        table.add_row(
            r["item_name"],
            r["category"],
            r["city_name"],
            str(r["qty_sold"]),
            f"₹{r['mrp_value']:,.0f}",
        )
    console.print(table)


def _print_po_snapshot(snapshot: dict, total_rows: int) -> None:
    console.print(f"\n[bold cyan]PO Summary[/bold cyan] (window: {snapshot['window_start']} → today)")
    table = Table(show_header=False, box=None, padding=(0, 2))
    table.add_column(style="dim")
    table.add_column(style="bold")
    table.add_row("Total POs fetched", str(total_rows))
    table.add_row("Total Raised", str(snapshot.get("total_raised", 0)))
    table.add_row("Scheduled", str(snapshot.get("scheduled", 0)))
    table.add_row("Created (unscheduled)", str(snapshot.get("created", 0)))
    table.add_row("Cancelled", str(snapshot.get("cancelled", 0)))
    table.add_row("Expired (unfulfilled)", str(snapshot.get("expired_unfulfilled", 0)))
    table.add_row("Expired (partial GRN)", str(snapshot.get("expired_partial", 0)))
    table.add_row("PO Amount", f"₹{snapshot.get('po_amount', 0):,.0f}")
    table.add_row("Items Delivered", f"{snapshot.get('items_delivered', 0):,}")
    console.print(table)


def _print_po_list(pos: list) -> None:
    if not pos:
        console.print("\n[dim]No PO data found.[/dim]")
        return
    preview = pos[:5]
    console.print(f"\n[bold cyan]POs[/bold cyan] (showing {len(preview)} of {len(pos)})")
    table = Table(show_header=True, header_style="bold")
    table.add_column("PO Number")
    table.add_column("State")
    table.add_column("City")
    table.add_column("Facility", style="dim")
    table.add_column("Units", justify="right")
    table.add_column("Amount", justify="right")
    table.add_column("SKUs", justify="right")
    for po in preview:
        state = po.get("po_state", "")
        state_colour = {
            "Fulfilled": "green", "Delivered": "green",
            "Scheduled": "cyan", "Rescheduled": "cyan",
            "Unscheduled": "yellow", "Created": "yellow",
            "Expired": "red", "Cancelled": "red",
            "Cancelled post Creation": "red",
        }.get(state, "")
        state_str = f"[{state_colour}]{state}[/{state_colour}]" if state_colour else state
        table.add_row(
            po.get("po_number", ""),
            state_str,
            po.get("city_name", ""),
            po.get("facility_name", ""),
            str(po.get("total_units_ordered", 0)),
            f"₹{po.get('total_po_amount', 0):,.0f}",
            str(po.get("item_count") or 0),
        )
    console.print(table)


# ── Blinkit Scorecard ──────────────────────────────────────────────────────────

@app.command("blinkit-scorecard")
def scrape_blinkit_scorecard(
    tenant_id: str = typer.Option(..., "--tenant", "-t", help="Tenant ID"),
    week: str = typer.Option(None, "--week", help="Week start date YYYY-MM-DD (must be a Monday, default: last Monday)"),
    save: bool = typer.Option(True, "--save/--no-save", help="Save results to MongoDB"),
):
    """Scrape Blinkit scorecard (fill rates). Data refreshes every Monday."""
    asyncio.run(_scrape_blinkit_scorecard(tenant_id, week, save))


async def _scrape_blinkit_scorecard(tenant_id: str, week: str | None, save: bool) -> None:
    async with AsyncSessionLocal() as db:
        job_id = None
        try:
            storage_state = (
                await auth_service.ensure(db, tenant_id, "blinkit_seller")
            ).storage_state

            job_id = await create_scrape_job(db, tenant_id, "blinkit_seller_scorecard")

            with console.status("[cyan]Scraping scorecard...[/cyan]"):
                raw = await seller_scraper.scrape_scorecard(storage_state, week=week)

            manufacturer_id = raw["manufacturer_id"]
            from_date = raw["from_date_ist"]

            weekly = parse_scorecard_weekly(raw, tenant_id, job_id)
            facilities = [
                parse_scorecard_facility(f, tenant_id, job_id, manufacturer_id, from_date)
                for f in raw["facilities"]
            ]
            key_skus = [
                parse_scorecard_key_sku(s, tenant_id, job_id, manufacturer_id, from_date)
                for s in raw["key_skus"]
            ]

            if save:
                await save_scorecard_results(db, weekly, facilities, key_skus)
                await complete_scrape_job(db, job_id)

            _print_scorecard_summary(weekly)
            _print_scorecard_facilities(facilities)
            _print_scorecard_key_skus(key_skus)

        except typer.Exit:
            raise
        except AuthError:
            # See the note in _scrape_blinkit — must not collapse into exit 1.
            if job_id:
                await fail_scrape_job(db, job_id, "auth_expired")
            raise
        except Exception as e:
            if job_id:
                await fail_scrape_job(db, job_id, str(e))
            console.print(f"[red]Scrape failed: {escape(str(e))}[/red]")
            raise typer.Exit(1)


@app.command("zepto-sales")
def scrape_zepto_sales(
    tenant_id: str = typer.Option(..., "--tenant", "-t", help="Tenant ID"),
    date_from: str = typer.Option(None, "--from", help="Start date YYYY-MM-DD (default: 7 days ago)"),
    date_to: str = typer.Option(None, "--to", help="End date YYYY-MM-DD (default: yesterday)"),
    save_xlsx: str = typer.Option(None, "--save-xlsx", help="Also write the results to this .xlsx path"),
    all_cities: bool = typer.Option(
        False, "--all-cities",
        help=(
            "Sweep every city for the per-city split instead of only those "
            "already known to sell. 138 calls rather than a handful — run it "
            "occasionally to pick up a new city, not daily."
        ),
    ),
    save: bool = typer.Option(True, "--save/--no-save", help="Save results to PostgreSQL"),
):
    """Fetch Zepto Sales Analytics (GMV/Units + per-SKU breakdown).

    Browser-free end to end: a pre-flight session health check, fresh
    brand/city/category ID discovery, then direct API calls — with an auth-only
    browser fallback if a call comes back 401/403. Tenant-general, no hardcoded
    IDs. Requires a session saved by `cli auth zepto-seller`.
    """
    asyncio.run(_scrape_zepto_sales(tenant_id, date_from, date_to, save_xlsx, all_cities, save))


def _write_zepto_sales_xlsx(
    path: str, data: dict, products: list[dict], date_from: str, date_to: str, ids: dict
) -> None:
    from openpyxl import Workbook

    gmv_daily = data["metrics"]["gmv"]["data"]
    units_daily = {row["key"]: next(v for k, v in row.items() if k != "key") for row in data["metrics"]["units"]["data"]}

    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Sales"
    ws.append(["Date", "GMV (Rs)", "Units"])
    for row in gmv_daily:
        day = row["key"]
        gmv_val = next(v for k, v in row.items() if k != "key")
        ws.append([day, gmv_val, units_daily.get(day)])

    ws3 = wb.create_sheet("Top Products")
    ws3.append([
        "Product", "Pack Size", "Category", "Subcategory", "GMV (Rs)", "Units Sold",
        "Sales Contribution %", "Available Stores %", "WoW Growth %", "MoM Growth %",
        "Stock On Hand",
    ])
    for p in products:
        ws3.append([
            p.get("productName"), p.get("packSize"), p.get("categoryName"), p.get("subcategoryName"),
            p.get("gmv"), p.get("qtySold"), p.get("salesContribution"), p.get("availableStores"),
            p.get("weekOnWeekGrowth"), p.get("monthOnMonthGrowth"), p.get("stockOnHand"),
        ])
    products_gmv_sum = sum(p.get("gmv") or 0 for p in products)

    ws2 = wb.create_sheet("Summary")
    ws2.append(["Brand", ids["brand_name"]])
    ws2.append(["Date range", f"{date_from} to {date_to}"])
    ws2.append(["Total GMV (Sales Overview)", data["headers"]["gmv"]["value"]])
    ws2.append(["Total Units", data["headers"]["units"]["value"]])
    ws2.append(["Sum of Top-Products GMV", products_gmv_sum])
    ws2.append(["Note", "Top-Products GMV may not fully reconcile with the overview total — see scraper.py"])
    ws2.append(["Cities", len(ids["city_ids"])])
    ws2.append(["Subcategories", ", ".join(ids["subcategory_names"])])

    wb.save(path)


async def _scrape_zepto_sales(
    tenant_id: str,
    date_from: str | None,
    date_to: str | None,
    save_xlsx: str | None,
    all_cities: bool,
    save: bool,
) -> None:
    yesterday = (_date.today() - timedelta(days=1)).isoformat()
    date_from = date_from or (_date.today() - timedelta(days=8)).isoformat()
    date_to = date_to or yesterday

    async with AsyncSessionLocal() as db:
        job_id = None
        try:
            # `ensure()` = load -> probe -> refresh -> re-login, doing the least
            # work that yields a session known to work — the same call Blinkit
            # makes. It replaces the old `zepto_seller` path, which stopped the
            # run and asked a human to type the OTP. The Zepto seller account is
            # shared, so it gets evicted mid-run routinely; self-healing matters
            # more here than anywhere else.
            #
            # The JWT lives in `raw`, not in cookies: platform_auth leaves
            # `storage_state` empty because the token travels in a header.
            with console.status("[cyan]Pre-flight: Zepto session…[/cyan]"):
                storage_state = {
                    "jwt": (await auth_service.ensure(db, tenant_id, "zepto")).raw["jwt"]
                }
            console.print("[green]Session healthy.[/green]")

            job_id = await create_scrape_job(db, tenant_id, "zepto_seller_sales", platform="zepto")

            with console.status("[cyan]Discovering current brand/city/category IDs...[/cyan]"):
                ids = await zepto_discover_ids(storage_state)
            console.print(f"[green]IDs discovered:[/green] {ids['brand_name']} — {len(ids['city_ids'])} cities, {len(ids['subcategory_ids'])} subcategories")

            with console.status(f"[cyan]Fetching Sales Analytics {date_from}..{date_to}...[/cyan]"):
                data = await zepto_fetch_sales_overview(storage_state, date_from, date_to, ids)

            # Per-SKU data is fetched one day at a time, like the Blinkit seller
            # scrape, so the rows land at day grain instead of one aggregate per
            # window — that is what makes a per-day SKU/category trend possible.
            # Only this endpoint needs the loop: sales-overview already returns
            # the whole window broken down by day in a single call.
            days = _date_range(date_from, date_to)
            product_rows: list[dict] = []
            failed_days: list[str] = []
            with console.status("[cyan]Fetching product-level breakdown...[/cyan]") as status:
                for i, day in enumerate(days, 1):
                    status.update(f"[cyan]Fetching product breakdown {day} ({i}/{len(days)})...[/cyan]")
                    try:
                        day_products = await zepto_fetch_product_performance(
                            storage_state, day, day, ids
                        )
                    except Exception as e:
                        # One bad day shouldn't discard the rest of the run; the
                        # count is reported below so a partial result is visible
                        # rather than silently short.
                        logger.warning(f"Zepto product-performance failed for {day}: {e}")
                        failed_days.append(day)
                        continue
                    product_rows.extend(
                        parse_zepto_product_perf(day_products, tenant_id, job_id, day, day)
                    )
                    if i < len(days):
                        await asyncio.sleep(_ZEPTO_DAY_GAP_S)

            daily_rows = parse_zepto_sales_daily(data, ids, tenant_id, job_id, date_from, date_to)

            # Which cities to ask for. Zepto has no city breakdown in a single
            # response, so a split means one call per city — but only for cities
            # already known to sell, which on this account is two of 138.
            # `--all-cities` re-sweeps everything to catch a new one; worth
            # running occasionally, not daily (Hosur went unnoticed for weeks).
            city_names = {c["cityID"]: c["cityName"] for c in ids.get("city_list", [])}
            targets = None if all_cities else await _zepto_known_cities(db, tenant_id)

            # SKU x city x day — the only source that carries city AND category
            # on one row, which is what the Analytics category-x-city heatmap
            # needs. One call per city per day, so it reuses the same short
            # `targets` list as the city split above rather than sweeping all
            # 138 every run.
            product_city_rows: list[dict] = []
            pc_targets = targets if targets is not None else ids["city_ids"]
            if pc_targets:
                with console.status("[cyan]Fetching product breakdown by city...[/cyan]") as status:
                    for i, day in enumerate(days, 1):
                        status.update(
                            f"[cyan]Product-by-city {day} ({i}/{len(days)}), "
                            f"{len(pc_targets)} cities...[/cyan]"
                        )
                        try:
                            by_city_products = await zepto_fetch_product_perf_by_city(
                                storage_state, day, day, ids, pc_targets
                            )
                        except Exception as e:
                            logger.warning(
                                f"Zepto product-performance by city failed for {day}: {e}"
                            )
                            continue
                        product_city_rows.extend(
                            parse_zepto_product_city(
                                by_city_products, city_names, tenant_id, job_id, day
                            )
                        )
                        if i < len(days):
                            await asyncio.sleep(_ZEPTO_DAY_GAP_S)

            written = 0
            if save:
                written = await zepto_save_sales_results(
                    db, daily_rows, product_rows, product_city_rows
                )
                await complete_scrape_job(db, job_id, written)
            else:
                await complete_scrape_job(db, job_id)

            # City summary now comes from the per-SKU city rows, the only
            # place the split is stored since sales_city_daily was dropped.
            if product_city_rows:
                by_name: dict[str, float] = {}
                for r in product_city_rows:
                    key = r["city_name"] or r["city_id"]
                    by_name[key] = by_name.get(key, 0) + r["gmv"]
                top = sorted(by_name.items(), key=lambda kv: -kv[1])[:3]
                console.print(
                    "  Cities: "
                    + ", ".join(f"{n} ₹{v:,.0f}" for n, v in top)
                    + (f" (+{len(by_name) - 3} more)" if len(by_name) > 3 else "")
                )

            gmv = data["headers"]["gmv"]["value"]
            units = data["headers"]["units"]["value"]
            console.print(f"\n[bold cyan]Zepto Sales Overview[/bold cyan] ({date_from} to {date_to})")
            console.print(f"  GMV: [bold]{gmv}[/bold]   Units: [bold]{units}[/bold]   ({len(daily_rows)} days)")
            console.print(f"  Product rows: [bold]{len(product_rows)}[/bold] over {len(days) - len(failed_days)}/{len(days)} days")
            if failed_days:
                console.print(
                    f"  [yellow]{len(failed_days)} day(s) failed and were skipped: "
                    f"{', '.join(failed_days[:5])}{' …' if len(failed_days) > 5 else ''}[/yellow]"
                )
            if save:
                console.print(f"  [green]Saved to DB:[/green] {written} rows")
            else:
                console.print("  [yellow]--no-save: nothing written to the database[/yellow]")

            if save_xlsx:
                _write_zepto_sales_xlsx(save_xlsx, data, products, date_from, date_to, ids)
                console.print(f"[green]Saved:[/green] {save_xlsx}")

        except SessionUnhealthy as e:
            console.print(f"[red]Session not usable: {escape(str(e))}[/red]")
            console.print("[yellow]Run `cli auth zepto-seller --tenant <id>` to re-login.[/yellow]")
            raise typer.Exit(1)
        except typer.Exit:
            raise
        except ZeptoNoDataYet as e:
            # Not a failure — Zepto simply has not computed that day yet. The
            # job is completed with zero records rather than marked failed, so
            # `cli status` does not report a broken scrape for something that is
            # only a matter of timing.
            if job_id:
                await complete_scrape_job(db, job_id, {})
            console.print(f"[yellow]Nothing to scrape yet: {escape(str(e))}[/yellow]")
            raise typer.Exit(0)
        except Exception as e:
            if job_id:
                await fail_scrape_job(db, job_id, str(e))
            console.print(f"[red]Scrape failed: {escape(str(e))}[/red]")
            raise typer.Exit(1)


def _print_scorecard_summary(weekly: dict) -> None:
    overall = weekly.get("overall", {})
    best_cat = weekly.get("best_category") or {}
    console.print(f"\n[bold cyan]Scorecard Summary[/bold cyan] (week of {weekly['from_date_ist']})")
    table = Table(show_header=False, box=None, padding=(0, 2))
    table.add_column(style="dim")
    table.add_column(style="bold")
    table.add_row("Fill Rate", f"{overall.get('fill_rate', 0):.2f}%")
    table.add_row("Weighted Fill Rate", f"{overall.get('weighted_fill_rate_percent', 0):.2f}%")
    table.add_row("PO Quantity", f"{overall.get('total_po_quantity', 0):,}")
    table.add_row("GRN Quantity", f"{overall.get('total_grn_quantity', 0):,}")
    table.add_row("Potential Loss", f"₹{overall.get('potential_loss', 0):,}")
    table.add_row("Total GMV", f"₹{overall.get('total_gmv', 0):,.0f}")
    table.add_row("Overall Rank", str(overall.get("manufacturer_rank", "—")))
    if best_cat:
        table.add_row(
            f"Best Category ({best_cat.get('proxy_category', '')})",
            f"Rank #{best_cat.get('manufacturer_rank', '—')} | {best_cat.get('fill_rate', 0):.2f}%",
        )
    console.print(table)


def _print_scorecard_facilities(facilities: list) -> None:
    if not facilities:
        return
    console.print(f"\n[bold cyan]Facilities[/bold cyan] ({len(facilities)})")
    table = Table(show_header=True, header_style="bold")
    table.add_column("Facility")
    table.add_column("City", style="dim")
    table.add_column("Fill Rate", justify="right")
    table.add_column("PO Qty", justify="right")
    table.add_column("GRN Qty", justify="right")
    table.add_column("Potential Loss", justify="right")
    table.add_column("Rank", justify="right")
    for f in sorted(facilities, key=lambda x: x.get("fill_rate", 0)):
        fill = f.get("fill_rate", 0)
        colour = "green" if fill >= 90 else ("yellow" if fill >= 70 else "red")
        table.add_row(
            f["facility_name"],
            f.get("city_name", ""),
            f"[{colour}]{fill:.0f}%[/{colour}]",
            f"{f.get('total_po_quantity', 0):,}",
            f"{f.get('total_grn_quantity', 0):,}",
            f"₹{f.get('potential_loss', 0):,}",
            str(f.get("manufacturer_rank", "—")),
        )
    console.print(table)


def _print_scorecard_key_skus(key_skus: list) -> None:
    if not key_skus:
        return
    console.print(f"\n[bold cyan]Key SKUs by Potential Loss[/bold cyan] ({len(key_skus)})")
    table = Table(show_header=True, header_style="bold")
    table.add_column("Item")
    table.add_column("Variant", style="dim")
    table.add_column("Category", style="dim")
    table.add_column("Potential Loss", justify="right")
    for s in sorted(key_skus, key=lambda x: x.get("potential_loss", 0), reverse=True):
        table.add_row(
            s.get("item_name", ""),
            s.get("variant_description", ""),
            s.get("proxy_category", ""),
            f"₹{s.get('potential_loss', 0):,.0f}",
        )
    console.print(table)


# ── Public search scraping ────────────────────────────────────────────────────

@app.command("public")
def scrape_public(
    keyword: str = typer.Option(..., "--keyword", "-k", help="Search keyword (e.g. 'cola', 'sunflower oil')"),
    brand: str = typer.Option(..., "--brand", "-b", help="Brand slug for classification (e.g. 'dobra')"),
    city: str = typer.Option("bengaluru", "--city", "-c", help="City slug (see scraper/utils/cities.py)"),
    platform: str = typer.Option("all", "--platform", "-p", help="Platform: blinkit | zepto | instamart | all"),
    all_zones: bool = typer.Option(False, "--all-zones", help="Scrape all zones defined for the city"),
    aliases: Optional[str] = typer.Option(None, "--aliases", help="Comma-separated brand name aliases (e.g. 'dobra,dobra cola')"),
    tenant_id: str = typer.Option(None, "--tenant", "-t", help="Tenant (client) UUID — required to --save (per-tenant storage)"),
    save: bool = typer.Option(False, "--save/--no-save", help="Save results to PostgreSQL (requires --tenant)"),
):
    """Scrape public product search results — no login required.

    Without --save it just scrapes and prints (no tenant needed). With --save it
    writes per-tenant header+detail rows (search_snapshots + search_listings) and
    opens a scrape_job, so --tenant is required.
    """
    alias_list = [a.strip() for a in aliases.split(",")] if aliases else None
    asyncio.run(_scrape_public(keyword, brand, city, platform, all_zones, alias_list, tenant_id, save))


async def _scrape_public(
    keyword: str,
    brand_slug: str,
    city_slug: str,
    platform: str,
    all_zones: bool,
    aliases: list[str] | None,
    tenant_id: str | None,
    save: bool,
) -> None:
    from scraper.utils.cities import CITIES, PLATFORM_CITIES
    from scraper.platforms.blinkit.public_data import scraper as bl_scraper, parser as bl_parser, storage as bl_storage
    from scraper.platforms.instamart.public_data import scraper as im_scraper, parser as im_parser, storage as im_storage
    from scraper.platforms.zepto.public_data import scraper as ze_scraper, parser as ze_parser, storage as ze_storage

    city = CITIES.get(city_slug)
    if not city:
        console.print(f"[red]Unknown city slug '{city_slug}'. Check scraper/utils/cities.py for valid slugs.[/red]")
        raise typer.Exit(1)

    platforms_to_run = (
        [p for p in ["blinkit", "zepto", "instamart"] if p in city["platforms"]]
        if platform == "all"
        else [platform]
    )
    invalid = [p for p in platforms_to_run if p not in {"blinkit", "zepto", "instamart"}]
    if invalid:
        console.print(f"[red]Unknown platform(s): {', '.join(invalid)}[/red]")
        raise typer.Exit(1)

    scrapers = {
        "blinkit":   (bl_scraper, bl_parser, bl_storage),
        "instamart": (im_scraper, im_parser, im_storage),
        "zepto":     (ze_scraper, ze_parser, ze_storage),
    }

    if save and not tenant_id:
        console.print("[red]--tenant is required with --save (storage is per-tenant).[/red]")
        raise typer.Exit(1)

    async with AsyncSessionLocal() as db:
        job_id = None
        rows_written = 0
        if save:
            job_id = await create_scrape_job(db, tenant_id, "public_search", "blinkit")

        try:
            for plat in platforms_to_run:
                if plat not in city["platforms"]:
                    console.print(f"  [dim]{plat} not available in {city['name']}[/dim]")
                    continue

                plat_zones = city["platforms"][plat]["zones"] if all_zones else [
                    {"zone": "", "pincode": city["pincode"], "lat": city["lat"], "lon": city["lon"]}
                ]

                scraper_mod, parser_mod, storage_mod = scrapers[plat]

                for zone_def in plat_zones:
                    zone_label = zone_def.get("zone", "")
                    zone_display = f" [{zone_label}]" if zone_label else ""
                    console.print(f"\n[bold cyan]{city['name']}{zone_display}[/bold cyan]  [dim]{plat}[/dim]")

                    with console.status(f"  [cyan]Scraping {plat}…[/cyan]"):
                        raw = await scraper_mod.scrape(
                            keyword=keyword,
                            brand_slug=brand_slug,
                            city_slug=city_slug,
                            zone=zone_label,
                            pincode=zone_def.get("pincode", city["pincode"]),
                            lat=zone_def.get("lat"),
                            lon=zone_def.get("lon"),
                            aliases=aliases,
                        )
                    result = parser_mod.parse(raw)
                    _print_public_result(plat, result)

                    if save:
                        rows_written += await storage_mod.save(db, result, tenant_id, job_id)

            if save:
                await complete_scrape_job(db, job_id, rows_written)
                console.print(f"\n[green]Saved {rows_written} rows[/green] (job {job_id})")
        except Exception as e:
            if save and job_id:
                await fail_scrape_job(db, job_id, str(e))
            raise


def _validate_marketplace(mp: str) -> str:
    """Normalise + check a --marketplace value against the wired providers.

    Fails fast rather than falling back: a typo'd marketplace silently scraping
    Blinkit would write real rows under the wrong platform.
    """
    from scraper.public import providers

    slug = (mp or "").strip().lower()
    try:
        providers.get_provider(slug)
    except ValueError as e:
        console.print(f"[red]{escape(str(e))}[/red]")
        raise typer.Exit(1)
    return slug


@app.command("public-run")
def public_run(
    tenant_id: str = typer.Option(None, "--tenant", "-t", help="Tenant (client) UUID — omit with --all"),
    all_tenants: bool = typer.Option(False, "--all", help="Run every active tenant"),
    marketplace: str = typer.Option("blinkit", "--marketplace", "-m", help="Marketplace to scrape (blinkit)"),
    cap: int = typer.Option(None, "--cap", help="Max products per search (default: tenant keyword_cap, else the platform floor)"),
    keyword: str = typer.Option(None, "--keyword", "-k", help="Only this keyword (subset of the watchlist)"),
    city: str = typer.Option(None, "--city", "-c", help="Only locations in this city slug"),
    resume: bool = typer.Option(False, "--resume", help="Continue this tenant's last incomplete run on this marketplace (skip already-scraped stores)"),
    workers: int = typer.Option(5, "--workers", "-w", help="Concurrent browser workers (pool size). ~5–6 is a good balance."),
    no_load: bool = typer.Option(False, "--no-load", help="Stage only; don't push to the database afterwards"),
):
    """Orchestrate a tenant's full watchlist (keywords × locations), all sourced
    from the DB (watchlist + tenant_locations). Writes per-tenant snapshot+listing
    rows under one scrape_job per tenant. --marketplace selects the platform (its
    locations, its engine). --keyword/--city narrow a run to a single keyword or
    city. --resume picks up an interrupted run. --workers sets the concurrent
    browser pool size.
    """
    if not tenant_id and not all_tenants:
        console.print("[red]Provide --tenant <id> or --all.[/red]")
        raise typer.Exit(1)
    if resume and all_tenants:
        console.print("[red]--resume works with a single --tenant, not --all.[/red]")
        raise typer.Exit(1)
    mp = _validate_marketplace(marketplace)
    asyncio.run(_public_run(tenant_id, all_tenants, cap, keyword, city, resume, workers, no_load, mp))


async def _auto_load(summary: dict, no_load: bool) -> None:
    """Push one finished scrape's staging file into Postgres, inline.

    Called right after each tenant's scrape (including inside a --all sweep, via the
    orchestrator's on_tenant_done hook) so a later tenant failing can never strand an
    earlier tenant's data. Deliberately NON-FATAL: the rows are already safe on disk,
    so a load failure leaves the file pending and prints the recovery command rather
    than failing the run.

    Only clean runs auto-load. A crashed/partial run still holds real data, but
    sweeping it in unnoticed is the accident worth avoiding — same rule
    `scrape load --all` follows. See docs/staging.md.
    """
    if no_load:
        return
    name = summary.get("staging_file")
    if not name:
        return
    if summary.get("status") != "success":
        console.print(
            f"[yellow]Not auto-loading {name} — the run didn't finish cleanly.[/yellow] "
            f"Review with [bold]cli scrape staged[/bold], then load or discard it."
        )
        return

    from scraper.public import loader, staging

    try:
        path = staging.resolve(name)
    except (FileNotFoundError, ValueError) as e:
        console.print(f"[red]Auto-load skipped:[/red] {escape(str(e))}")
        return

    console.print(f"[dim]Loading {name} into the database…[/dim]")
    try:
        res = await loader.load_with_retry(path)
    except Exception as e:
        err = getattr(e, "orig", None) or e
        console.print(f"[red]Auto-load FAILED:[/red] {escape(str(err))[:300]}")
        console.print(
            f"[dim]Nothing was written; the scrape is safe on disk. Retry with "
            f"[/dim][bold]python -m cli scrape load --file {staging.ref(path)}[/bold]"
        )
        return
    console.print(
        f"[green]Loaded[/green] {res['total']:,} rows "
        f"({res['snapshots']:,} snapshots, {res['listings']:,} listings, "
        f"{res['skus']:,} sku rows)"
    )


async def _public_run(
    tenant_id: str | None, all_tenants: bool, cap: int | None,
    keyword: str | None, city: str | None, resume: bool, workers: int,
    no_load: bool = False, mp_slug: str = "blinkit",
) -> None:
    from scraper.public import orchestrator

    async def _after(summary: dict) -> None:
        await _auto_load(summary, no_load)

    async with AsyncSessionLocal() as db:
        if all_tenants:
            # Load each tenant the moment its scrape finishes, not at the end of the
            # sweep — on a weekly scheduled run, tenant 7 failing must not strand the
            # six already scraped.
            summaries = await orchestrator.run_all(
                db, cap, keyword, city, workers, on_tenant_done=_after, mp_slug=mp_slug
            )
        else:
            summaries = [await orchestrator.run_tenant(
                db, tenant_id, cap, keyword, city, resume, workers, mp_slug=mp_slug
            )]
            await _after(summaries[0])

    if not summaries:
        console.print("[yellow]No active tenants to run.[/yellow]")
        return
    table = Table(show_header=True, header_style="bold")
    table.add_column("Tenant")
    table.add_column("MP")
    table.add_column("Keywords", justify="right")
    table.add_column("Locations", justify="right")
    table.add_column("Snapshots", justify="right")
    table.add_column("Rows", justify="right")
    table.add_column("Skipped", justify="right")
    table.add_column("Errors", justify="right")
    for s in summaries:
        table.add_row(
            s["tenant_id"][:8], s.get("mp_slug", mp_slug),
            str(s["keywords"]), str(s["locations"]),
            str(s["snapshots"]), str(s["rows"]), str(s.get("skipped", 0)),
            f"[red]{s['errors']}[/red]" if s["errors"] else "0",
        )
    console.print(table)


@app.command("public-skus")
def public_skus(
    tenant_id: str = typer.Option(None, "--tenant", "-t", help="Tenant (client) UUID — omit with --all"),
    all_tenants: bool = typer.Option(False, "--all", help="Run every active tenant"),
    marketplace: str = typer.Option("blinkit", "--marketplace", "-m", help="Marketplace to scrape (blinkit)"),
    cap: int = typer.Option(None, "--brand-cap", help="Override brand_cap for this run (default: per-tenant, else the platform floor)"),
    city: str = typer.Option(None, "--city", "-c", help="Only locations in this city slug"),
    resume: bool = typer.Option(False, "--resume", help="Continue this tenant's last incomplete run on this marketplace (skip scraped stores)"),
    workers: int = typer.Option(5, "--workers", "-w", help="Concurrent browser workers (pool size)."),
    no_load: bool = typer.Option(False, "--no-load", help="Stage only; don't push to the database afterwards"),
):
    """Targeted own-SKU scrape: search each tenant's brand name, paginate its whole
    catalog, and write per-product rows to sku_snapshots (price/mrp/discount/stock/
    inventory/rating), keyed on product_id. Complements `public-run` (which covers
    SoV/rank + competitors). --marketplace selects the platform. --resume picks up
    an interrupted run.
    """
    if not tenant_id and not all_tenants:
        console.print("[red]Provide --tenant <id> or --all.[/red]")
        raise typer.Exit(1)
    if resume and all_tenants:
        console.print("[red]--resume works with a single --tenant, not --all.[/red]")
        raise typer.Exit(1)
    mp = _validate_marketplace(marketplace)
    asyncio.run(_public_skus(tenant_id, all_tenants, cap, city, resume, workers, no_load, mp))


async def _public_skus(
    tenant_id: str | None, all_tenants: bool, cap: int | None,
    city: str | None, resume: bool, workers: int, no_load: bool = False,
    mp_slug: str = "blinkit",
) -> None:
    from scraper.public import targeted

    async def _after(summary: dict) -> None:
        await _auto_load(summary, no_load)

    async with AsyncSessionLocal() as db:
        if all_tenants:
            summaries = await targeted.run_all_targeted(
                db, cap, city, workers, on_tenant_done=_after, mp_slug=mp_slug
            )
        else:
            summaries = [await targeted.run_targeted(
                db, tenant_id, cap, city, resume, workers, mp_slug=mp_slug
            )]
            await _after(summaries[0])

    if not summaries:
        console.print("[yellow]No active tenants to run.[/yellow]")
        return
    table = Table(show_header=True, header_style="bold")
    table.add_column("Tenant")
    table.add_column("MP")
    table.add_column("Brands", justify="right")
    table.add_column("Locations", justify="right")
    table.add_column("SKU Rows", justify="right")
    table.add_column("Skipped", justify="right")
    table.add_column("Errors", justify="right")
    for s in summaries:
        table.add_row(
            s["tenant_id"][:8], s.get("mp_slug", mp_slug),
            str(s["brands"]), str(s["locations"]),
            str(s["rows"]), str(s.get("skipped", 0)),
            f"[red]{s['errors']}[/red]" if s["errors"] else "0",
        )
    console.print(table)


def _print_public_result(platform: str, result: dict) -> None:
    sov = result.get("brand_sov_pct", 0)
    rank = result.get("brand_rank")
    total = result.get("total_results", 0)
    brand_count = result.get("brand_product_count", 0)

    rank_str = f"#{rank}" if rank else "not ranked"
    colour = "green" if sov >= 20 else ("yellow" if sov >= 5 else "red")

    console.print(
        f"  [bold]{platform}[/bold]  "
        f"total={total}  brand={brand_count}  "
        f"rank={rank_str}  sov=[{colour}]{sov}%[/{colour}]"
    )

    brand_prods = result.get("brand_products", [])
    if brand_prods:
        table = Table(show_header=True, header_style="bold", box=None, padding=(0, 2))
        table.add_column("#", style="dim", width=4)
        table.add_column("Product")
        table.add_column("Price", justify="right")
        for p in brand_prods:
            price_str = f"₹{p['price']:.0f}" if p.get("price") else "—"
            table.add_row(str(p.get("position", "")), p.get("name", ""), price_str)
        console.print(table)

    comps = result.get("competitors", [])[:3]
    if comps:
        comp_names = ", ".join(f"{c['name']} ({c['count_in_results']})" for c in comps)
        console.print(f"  [dim]Top competitors: {comp_names}[/dim]")


def _print_soh(rows: list, date: str) -> None:
    console.print(f"\n[bold cyan]Stock on Hand[/bold cyan] ({date}) — {len(rows)} rows")
    if not rows:
        console.print("\n[dim]No SOH data found.[/dim]")
        return

    preview = rows[:10]
    table = Table(show_header=True, header_style="bold")
    table.add_column("Item")
    table.add_column("Facility", style="dim")
    table.add_column("Backend Qty", justify="right")
    table.add_column("Frontend Qty", justify="right")
    for r in preview:
        backend = r.get("backend_inv_qty", 0)
        frontend = r.get("frontend_inv_qty", 0)
        backend_str = f"[red]{backend}[/red]" if backend == 0 else str(backend)
        frontend_str = f"[red]{frontend}[/red]" if frontend == 0 else str(frontend)
        table.add_row(
            r.get("item_name", ""),
            r.get("backend_facility_name", ""),
            backend_str,
            frontend_str,
        )
    console.print(table)
    if len(rows) > 10:
        console.print(f"[dim]... and {len(rows) - 10} more rows[/dim]")


# ── Staging: local SQLite → Postgres ─────────────────────────────────────────

@app.command("staged")
def staged_list(
    pending_only: bool = typer.Option(False, "--pending", help="Only runs not yet loaded"),
    marketplace: str = typer.Option(None, "--marketplace", "-m", help="Only runs from this marketplace"),
):
    """List local staging files from public scrapes (what `scrape load` would push).

    Public scrapes write to a local SQLite file, not straight to Postgres — a long
    run no longer dies with the database. See docs/staging.md.
    """
    from scraper.public import staging

    mp = (marketplace or "").strip().lower() or None
    runs = staging.pending(mp_slug=mp) if pending_only else staging.list_runs(mp_slug=mp)
    if not runs:
        console.print("[dim]No staging files.[/dim]")
        return

    table = Table(show_header=True, header_style="bold",
                  title=f"Staging files — {len(runs)}")
    table.add_column("Date")
    table.add_column("Time")
    table.add_column("MP")
    table.add_column("Kind")
    table.add_column("Stores", justify="right")
    table.add_column("Rows", justify="right")
    table.add_column("Err", justify="right")
    table.add_column("State")
    table.add_column("Ref", style="dim")
    for r in runs:
        started, loaded = r["started_at"] or "", r["loaded_at"]
        date, _, tm = started.partition("T")
        done, want = r.get("stores_done"), r.get("stores_total")
        stores = f"{done:,}/{want:,}" if done is not None and want else (
            f"{done:,}" if done is not None else "[dim]—[/dim]")
        # A partial run is not automatically bad — 500 of 2059 stores is still 500
        # stores of real data. Flag it, let the human decide.
        if done is not None and want and done < want:
            stores = f"[yellow]{stores}[/yellow]"
        errs = r.get("errors")
        err_txt = "[dim]—[/dim]" if errs is None else (
            f"[red]{errs:,}[/red]" if errs else "0")
        status = {"success": "[green]ok[/green]", "failed": "[red]failed[/red]"} \
            .get(r["status"], f"[yellow]{r['status']}[/yellow]")
        where = "[green]loaded[/green]" if loaded else "[yellow]pending[/yellow]"
        table.add_row(
            date, tm[:5], r["mp_slug"],
            r["kind"].replace("public_", ""),
            stores, f"{r['rows']:,}", err_txt,
            f"{status} · {where}",
            staging.ref(r["path"]),
        )
    console.print(table)
    n_pending = sum(1 for r in runs if not r["loaded_at"])
    n_bad = sum(1 for r in runs if not r["loaded_at"] and r["status"] != "success")
    if n_pending:
        console.print(f"[yellow]{n_pending} file(s) not yet in the database.[/yellow] "
                      f"Push with [bold]python -m cli scrape load[/bold]")
    if n_bad:
        console.print(
            f"[red]{n_bad} pending file(s) did not finish cleanly.[/red] Review before "
            f"loading — drop one with [bold]python -m cli scrape discard --file <name>[/bold]"
        )


@app.command("load")
def load_staged(
    file: str = typer.Option(None, "--file", "-f", help="Which file — the Ref from `scrape staged` (default: all pending, oldest first)"),
    all_pending: bool = typer.Option(False, "--all", help="Load every pending file"),
    dry_run: bool = typer.Option(False, "--dry-run", help="Show what would be loaded, write nothing"),
    keep: int = typer.Option(None, "--keep", help=f"Loaded files to retain per tenant+kind+marketplace (default {5})"),
):
    """Push staged public-scrape results into Postgres.

    Each file is loaded in ONE all-or-nothing transaction: if the connection drops
    the whole thing rolls back and nothing is written, so re-running is always safe
    (public data is append-only with no upsert — a partial load would duplicate).

    With no --file, loads every pending file oldest-first.
    """
    from scraper.public import staging

    if file:
        try:
            targets = [{"path": staging.resolve(file)}]
        except (FileNotFoundError, ValueError) as e:
            console.print(f"[red]{escape(str(e))}[/red]")
            console.print("[dim]List them with `python -m cli scrape staged`.[/dim]")
            raise typer.Exit(1)
    else:
        targets = list(reversed(staging.pending()))   # oldest first
        if not targets:
            console.print("[dim]Nothing pending — all staging files are loaded.[/dim]")
            return
        if len(targets) > 1 and not all_pending and not dry_run:
            console.print(
                f"[yellow]{len(targets)} files pending.[/yellow] "
                f"Pass --all to load them all, or --file <name> for one:"
            )
            for t in targets:
                console.print(f"   {t['path'].name}  ({t['rows']:,} rows)")
            raise typer.Exit(1)
        # A run that crashed may still hold data worth keeping, so this is never an
        # automatic skip — but sweeping one into the DB unnoticed via --all is the
        # exact accident worth blocking. Force a deliberate choice.
        unclean = [t for t in targets if t.get("status") != "success"]
        if unclean and not dry_run:
            console.print(
                f"[red]{len(unclean)} pending file(s) did not finish cleanly:[/red]"
            )
            for t in unclean:
                done, want = t.get("stores_done"), t.get("stores_total")
                cover = f", {done}/{want} stores" if done is not None and want else ""
                console.print(
                    f"   [red]{t['status']}[/red]  {t['path'].name}  "
                    f"({t['rows']:,} rows{cover}, {t.get('errors') or 0} errors)"
                )
            console.print(
                "\nPartial data is often still worth loading — but decide per file:\n"
                "  [bold]python -m cli scrape load --file <name>[/bold]     load it anyway\n"
                "  [bold]python -m cli scrape discard --file <name>[/bold]  throw it away"
            )
            raise typer.Exit(1)

    if dry_run:
        table = Table(show_header=True, header_style="bold", title="DRY RUN — nothing written")
        table.add_column("File")
        table.add_column("MP")
        table.add_column("Kind")
        table.add_column("Snapshots", justify="right")
        table.add_column("Listings", justify="right")
        table.add_column("SKU rows", justify="right")
        for t in targets:
            c = t.get("counts") or {}
            table.add_row(t["path"].name, t.get("mp_slug", "?"), t.get("kind", "?"),
                          f"{c.get('search_snapshots', 0):,}",
                          f"{c.get('search_listings', 0):,}",
                          f"{c.get('sku_snapshots', 0):,}")
        console.print(table)
        return

    asyncio.run(_load_staged([t["path"] for t in targets], keep))


async def _load_staged(paths, keep) -> None:
    from scraper.public import loader, staging

    if keep is not None:
        staging.KEEP_PER_KIND = keep

    from sqlalchemy.exc import DBAPIError

    results, failed = [], []
    for p in paths:
        console.print(f"[dim]Loading {p.name} …[/dim]")
        # Retry once on a DB-level failure. Safe by construction: the load is one
        # all-or-nothing transaction, so a failed attempt wrote nothing. Each attempt
        # gets a FRESH session — a dropped connection can't be reused, and the old
        # session's pool entry is poisoned.
        for attempt in (1, 2):
            try:
                async with AsyncSessionLocal() as db:
                    results.append(await loader.load_file(db, p))
                break
            except DBAPIError as e:
                err = getattr(e, "orig", None) or e
                if attempt == 1:
                    console.print(
                        f"[yellow]Connection failed — retrying once:[/yellow] "
                        f"{escape(str(err))[:120]}"
                    )
                    continue
                failed.append((p.name, str(err)))
                console.print(f"[red]FAILED[/red] {p.name}: {escape(str(err))[:300]}")
                console.print("[dim]Nothing was written — rerun the same command to retry.[/dim]")
            except Exception as e:
                failed.append((p.name, str(e)))
                console.print(f"[red]FAILED[/red] {p.name}: {escape(str(e))[:300]}")
                console.print("[dim]Nothing was written — rerun the same command to retry.[/dim]")
                break

    if results:
        table = Table(show_header=True, header_style="bold", title="Loaded")
        table.add_column("File")
        table.add_column("MP")
        table.add_column("Kind")
        table.add_column("Snapshots", justify="right")
        table.add_column("Listings", justify="right")
        table.add_column("SKU rows", justify="right")
        table.add_column("Total", justify="right")
        for r in results:
            table.add_row(r["file"], r["mp_slug"], r["kind"], f"{r['snapshots']:,}",
                          f"{r['listings']:,}", f"{r['skus']:,}", f"{r['total']:,}")
        console.print(table)
        pruned = sum(r["pruned"] for r in results)
        if pruned:
            console.print(f"[dim]Pruned {pruned} old staging file(s).[/dim]")
    if failed:
        console.print(f"[red]{len(failed)} file(s) failed to load.[/red]")
        raise typer.Exit(1)


@app.command("discard")
def discard_staged(
    file: str = typer.Option(..., "--file", "-f", help="Which file — the Ref from `scrape staged`, or a filename/path"),
    force: bool = typer.Option(False, "--force", help="Skip the confirmation prompt"),
):
    """Delete a staging file without loading it.

    For a file that was never loaded this destroys scraped data held nowhere else —
    hence the prompt. Use it to drop a bad run (wrong city, wrong cap, a crash that
    produced nothing useful) so `scrape load --all` can't sweep it into the database.
    """
    from scraper.public import staging

    try:
        path = staging.resolve(file)
    except (FileNotFoundError, ValueError) as e:
        console.print(f"[red]{escape(str(e))}[/red]")
        console.print("[dim]List them with `python -m cli scrape staged`.[/dim]")
        raise typer.Exit(1)

    match = next((r for r in staging.list_runs() if r["path"].resolve() == path.resolve()), None)
    if match is None:
        console.print(f"[red]{path.name} is not a readable staging file.[/red]")
        raise typer.Exit(1)

    loaded = match["loaded_at"]
    console.print(
        f"\n  [bold]{path.name}[/bold]\n"
        f"  kind    : {match['kind']}\n"
        f"  scraped : {(match['started_at'] or '').replace('T', ' ')[:16]}\n"
        f"  status  : {match['status']}   errors: {match.get('errors') or 0}\n"
        f"  rows    : {match['rows']:,}\n"
        f"  loaded  : {'yes — ' + loaded[:16].replace('T', ' ') if loaded else 'NO'}\n"
    )
    if loaded:
        console.print("[dim]Already in the database — deleting only removes the local backup.[/dim]")
    else:
        console.print(
            f"[red]NOT loaded.[/red] These {match['rows']:,} rows exist nowhere else — "
            f"deleting is irreversible."
        )
    if not force and not typer.confirm("Delete this file?"):
        console.print("[dim]Cancelled.[/dim]")
        raise typer.Exit(0)

    staging.discard(path)
    console.print(f"[green]Deleted[/green] {path.name}")



async def _zepto_known_cities(db, tenant_id: str) -> list[str]:
    """City ids that have ever recorded sales for this tenant.

    Keeps the daily scrape to one call per selling city instead of 138. Returns
    an empty list on the very first run, which the caller treats as "sweep
    everything" via --all-cities, or simply skips.
    """
    from sqlalchemy import text

    rows = (
        await db.execute(
            text(
                "SELECT DISTINCT city_id FROM zepto_seller_product_city_daily "
                "WHERE tenant_id = :t"
            ),
            {"t": tenant_id},
        )
    ).scalars().all()
    return list(rows)


@app.command("zepto-ads")
def scrape_zepto_ads(
    tenant_id: str = typer.Option(..., "--tenant", "-t", help="Tenant ID"),
    date_from: str = typer.Option(None, "--from", help="Start date YYYY-MM-DD (default: 7 days ago)"),
    date_to: str = typer.Option(None, "--to", help="End date YYYY-MM-DD (default: yesterday)"),
    category: str = typer.Option(
        "all", "--category",
        help=(
            "sponsored_products | sponsored_display | sponsored_brands | all. "
            "Applies to the Analytics tables only, which DO partition by it "
            "(the tabs return disjoint campaigns). The campaign list itself "
            "ignores the filter, so it is fetched once regardless. Default "
            "'all' — anything narrower silently drops the other tabs' spend."
        ),
    ),
    save: bool = typer.Option(True, "--save/--no-save", help="Save results to PostgreSQL"),
):
    """Fetch Zepto ads per day: campaigns (spend, RoAS, clicks, SOV) and keywords.

    Two endpoints, because neither is complete on its own. Campaign Management
    (`/campaigns`) has the operational fields — budgets, base bid, targeting,
    status, start/end dates — but no revenue, add-to-carts or keywords. The
    Analytics tables (`/metrics/tabular`) have those, but none of the
    operational ones. Campaign rows merge both; keyword rows come only from
    the second.

    Unlike `zepto-sales`, this needs a browser: the ads-bff service rejects the
    saved session's WAF token with a 202 challenge, so headers are harvested
    from one short page load and reused for the HTTP calls.
    """
    asyncio.run(_scrape_zepto_ads(tenant_id, date_from, date_to, category, save))


async def _scrape_zepto_ads(
    tenant_id: str, date_from: str | None, date_to: str | None, category: str, save: bool
) -> None:
    from scraper.platforms.zepto.dashboard_data.seller import endpoints as zep

    categories = list(zep.ADS_CATEGORIES) if category == "all" else [category]
    days = _date_range(date_from, date_to)

    async with AsyncSessionLocal() as db:
        job_id = None
        try:
            with console.status("[cyan]Pre-flight: checking Zepto seller session health...[/cyan]"):
                storage_state = await ensure_healthy_session(db, tenant_id, "zepto_seller", zepto_seller_validate)
            console.print("[green]Session healthy.[/green]")

            with console.status("[cyan]Discovering brand...[/cyan]"):
                ids = await zepto_discover_ids(storage_state)
            brand_id = ids["brand_id"]
            console.print(f"[green]Brand:[/green] {ids['brand_name']}")

            # ads-bff will not take the stored WAF token; harvest a live one.
            with console.status("[cyan]Capturing ads headers (browser)...[/cyan]"):
                headers = await zepto_capture_ads_headers(storage_state)
            console.print("[green]Ads headers captured.[/green]")

            job_id = await create_scrape_job(db, tenant_id, "zepto_ads", platform="zepto")

            rows: list[dict] = []
            kw_rows: list[dict] = []
            prod_rows: list[dict] = []
            bd_rows: list[dict] = []
            failed: list[str] = []
            not_ready: list[str] = []
            # Per day: one campaign list (the filter is ignored there, so
            # fetching it per category would just repeat the same call) plus
            # four tabular views per category — campaign, keyword, product and
            # category. Those ARE per-category: the tabs return disjoint data.
            total_calls = len(days) * (1 + 6 * len(categories))

            def _has_any(day_rows: list[dict]) -> bool:
                return any(r["spend"] or r["impressions"] or r["clicks"] for r in day_rows)

            # A dead session fails every remaining call, so keep going and the
            # run burns ~150 requests to save nothing — which has happened three
            # times. One 401 can be a blip; several in a row is the session
            # being gone (expired, or displaced by someone else logging in).
            # Bail out and keep whatever was fetched before it died.
            class _SessionGone(Exception):
                pass

            auth_fails = 0

            def _note(exc: Exception) -> None:
                nonlocal auth_fails
                if "401" in str(exc) or "403" in str(exc):
                    auth_fails += 1
                    if auth_fails >= 3:
                        raise _SessionGone(
                            "3 consecutive auth failures — the Zepto session is gone. "
                            "Stopping so the rows already fetched can be saved; "
                            "re-run `cli auth zepto-seller` and scrape the missing days."
                        )
                else:
                    auth_fails = 0

            session_died = False
            try:
                with console.status("[cyan]Fetching campaigns...[/cyan]") as status:
                    n = 0
                    for day in days:
                        n += 1
                        status.update(f"[cyan]Campaigns {day} ({n}/{total_calls})...[/cyan]")
                        try:
                            camps = await zepto_fetch_ad_campaigns(
                                headers, brand_id, day, day, categories[0]
                            )
                            day_rows = parse_zepto_ad_campaigns(
                                camps, tenant_id, job_id, day, categories[0]
                            )

                            # ads-bff intermittently returns the campaign list with
                            # every metric as "-", then real figures for the same
                            # window seconds later. Retry once; if it is still bare,
                            # skip the day rather than upserting zeros over data that
                            # a previous run got right.
                            if day_rows and not _has_any(day_rows):
                                await asyncio.sleep(6)
                                camps = await zepto_fetch_ad_campaigns(
                                    headers, brand_id, day, day, categories[0]
                                )
                                day_rows = parse_zepto_ad_campaigns(
                                    camps, tenant_id, job_id, day, categories[0]
                                )
                                if not _has_any(day_rows):
                                    not_ready.append(day)
                                    continue
                        except Exception as e:
                            logger.warning(f"Zepto ads failed for {day}: {e}")
                            failed.append(day)
                            _note(e)
                            continue
                        await asyncio.sleep(_ZEPTO_DAY_GAP_S)

                        by_id = {r["campaign_id"]: r for r in day_rows}
                        for cat in categories:
                            n += 1
                            status.update(f"[cyan]Analytics {cat} {day} ({n}/{total_calls})...[/cyan]")
                            try:
                                tab = await zepto_fetch_ads_tabular(
                                    headers, brand_id, day, day, zep.ADS_VIEW_CAMPAIGN, cat
                                )
                            except Exception as e:
                                logger.warning(f"Zepto campaign_table failed for {cat} {day}: {e}")
                                failed.append(f"{cat}/{day} analytics")
                                _note(e)
                            else:
                                for cid, patch in parse_zepto_ad_tabular_campaigns(tab).items():
                                    row = by_id.get(cid)
                                    if row is None:
                                        # A campaign the Analytics view knows about
                                        # but the campaign list did not return. Not
                                        # observed, but it would silently lose that
                                        # campaign's revenue if it ever happened.
                                        logger.warning(
                                            f"Zepto campaign {cid} is in the {cat} analytics table "
                                            f"but not in the campaign list for {day} — metrics dropped"
                                        )
                                        continue
                                    row.update(patch)
                                    # The tabular tabs partition properly, unlike the
                                    # campaign list, so this is the campaign's real
                                    # category rather than the tab that was asked for.
                                    row["campaign_category"] = cat
                            await asyncio.sleep(_ZEPTO_DAY_GAP_S)

                            n += 1
                            status.update(f"[cyan]Keywords {cat} {day} ({n}/{total_calls})...[/cyan]")
                            try:
                                kws = await zepto_fetch_ads_tabular(
                                    headers, brand_id, day, day, zep.ADS_VIEW_KEYWORD, cat
                                )
                            except Exception as e:
                                logger.warning(f"Zepto keyword_table failed for {cat} {day}: {e}")
                                failed.append(f"{cat}/{day} keywords")
                                _note(e)
                            else:
                                kw_rows.extend(
                                    parse_zepto_ad_keywords(
                                        kws, tenant_id, job_id, day, cat, brand_id
                                    )
                                )
                            await asyncio.sleep(_ZEPTO_DAY_GAP_S)

                            # Product performance, then the three breakdown views
                            # (category / city / page) which share one shape and one
                            # parser. All partition by campaign_category the way the
                            # keyword view does.
                            n += 1
                            status.update(f"[cyan]Products {cat} {day} ({n}/{total_calls})...[/cyan]")
                            try:
                                tab_rows = await zepto_fetch_ads_tabular(
                                    headers, brand_id, day, day, zep.ADS_VIEW_PRODUCT, cat
                                )
                            except Exception as e:
                                logger.warning(f"Zepto product_table failed for {cat} {day}: {e}")
                                failed.append(f"{cat}/{day} products")
                                _note(e)
                            else:
                                prod_rows.extend(
                                    parse_zepto_ad_products(
                                        tab_rows, tenant_id, job_id, day, cat, brand_id
                                    )
                                )
                            await asyncio.sleep(_ZEPTO_DAY_GAP_S)

                            for view, dim in (
                                (zep.ADS_VIEW_CATEGORY, "category"),
                                (zep.ADS_VIEW_CITY, "city"),
                                (zep.ADS_VIEW_PAGE, "page"),
                            ):
                                n += 1
                                status.update(
                                    f"[cyan]{dim.title()} {cat} {day} ({n}/{total_calls})...[/cyan]"
                                )
                                try:
                                    tab_rows = await zepto_fetch_ads_tabular(
                                        headers, brand_id, day, day, view, cat
                                    )
                                except Exception as e:
                                    logger.warning(f"Zepto {view} failed for {cat} {day}: {e}")
                                    failed.append(f"{cat}/{day} {dim}")
                                    _note(e)
                                else:
                                    bd_rows.extend(
                                        parse_zepto_ad_breakdown(
                                            tab_rows, tenant_id, job_id, day, cat, brand_id, dim
                                        )
                                    )
                                await asyncio.sleep(_ZEPTO_DAY_GAP_S)

                        rows.extend(day_rows)
            except _SessionGone as e:
                session_died = True
                console.print(f"[yellow]{escape(str(e))}[/yellow]")

            written: dict[str, int] = {}
            if save:
                written = await zepto_save_ad_results(
                    db, rows, kw_rows, prod_rows, bd_rows
                )
                await complete_scrape_job(db, job_id, sum(written.values()))
            else:
                await complete_scrape_job(db, job_id)

            # Report on de-duplicated rows, matching what storage actually
            # writes. The campaign list is now fetched once per day rather than
            # once per category, so this should be a no-op — but it was not
            # always, and a run that double-counted spend read as plausible.
            unique = {r["upsert_key"]: r for r in rows}.values()
            spend = sum(r["spend"] for r in unique)
            clicks = sum(r["clicks"] for r in unique)
            revenue = sum(r.get("revenue") or 0 for r in unique)
            rows_reported = len(unique)
            kw_unique = {r["upsert_key"]: r for r in kw_rows}.values()
            prod_unique = {r["upsert_key"]: r for r in prod_rows}.values()
            bd_unique = {r["upsert_key"]: r for r in bd_rows}.values()
            console.print(f"\n[bold cyan]Zepto Ads[/bold cyan] ({days[0]} to {days[-1]}) — {', '.join(categories)}")
            console.print(f"  Rows: [bold]{rows_reported}[/bold]   Spend: [bold]₹{spend:,.0f}[/bold]   Clicks: [bold]{clicks:,}[/bold]   Revenue: [bold]₹{revenue:,.0f}[/bold]")
            console.print(f"  Keywords: [bold]{len(kw_unique)}[/bold] row(s), spend ₹{sum(k['spend'] for k in kw_unique):,.0f}")
            console.print(f"  Products: [bold]{len(prod_unique)}[/bold] row(s), spend ₹{sum(k['spend'] for k in prod_unique):,.0f}")
            console.print(f"  Breakdown (category/city/page): [bold]{len(bd_unique)}[/bold] row(s)")
            if failed:
                console.print(f"  [yellow]{len(failed)} fetch(es) failed: {', '.join(failed[:5])}{' …' if len(failed) > 5 else ''}[/yellow]")
            if not_ready:
                console.print(
                    f"  [yellow]{len(not_ready)} day(s) had no metrics yet and were skipped "
                    f"(not written as zero): {', '.join(not_ready[:5])}{' …' if len(not_ready) > 5 else ''}[/yellow]"
                )
            if save:
                console.print(
                    "  [green]Saved to DB:[/green] "
                    + ", ".join(f"{v} {k}" for k, v in written.items())
                )
            else:
                console.print("  [yellow]--no-save: nothing written[/yellow]")

        except SessionUnhealthy as e:
            console.print(f"[red]Session not usable: {escape(str(e))}[/red]")
            console.print("[yellow]Run `cli auth zepto-seller --tenant <id>` to re-login.[/yellow]")
            raise typer.Exit(1)
        except typer.Exit:
            raise
        except Exception as e:
            if job_id:
                await fail_scrape_job(db, job_id, str(e))
            console.print(f"[red]Scrape failed: {escape(str(e))}[/red]")
            raise typer.Exit(1)


# ── Zepto PO Management ────────────────────────────────────────────────────────

@app.command("zepto-po")
def scrape_zepto_po(
    tenant_id: str = typer.Option(..., "--tenant", "-t", help="Tenant ID"),
    date_from: str = typer.Option(None, "--from", help="Start date YYYY-MM-DD (default: 30 days ago)"),
    date_to: str = typer.Option(None, "--to", help="End date YYYY-MM-DD (default: today)"),
    save: bool = typer.Option(True, "--save/--no-save", help="Save results to PostgreSQL"),
):
    """Fetch Zepto purchase orders, goods receipts and shipping notices.

    Three endpoints on the `/vendor` app (same host and session as the sales
    scrape, no browser needed): what Zepto ordered, what shipped, what arrived.
    `grn_qty / po_qty` is the fill rate — the figure Blinkit's seller scorecard
    reports and the only route to one for Zepto.

    Windows here are inclusive of TODAY by default, unlike the sales scrape:
    POs are forward-looking (an order issued today expires in three weeks), so
    stopping at yesterday would miss the ones that most need acting on.
    """
    asyncio.run(_scrape_zepto_po(tenant_id, date_from, date_to, save))


async def _scrape_zepto_po(
    tenant_id: str, date_from: str | None, date_to: str | None, save: bool
) -> None:
    date_to = date_to or _date.today().isoformat()
    date_from = date_from or (_date.today() - timedelta(days=30)).isoformat()

    async with AsyncSessionLocal() as db:
        job_id = None
        try:
            # See the note on the sales scrape: same self-healing `ensure()`.
            # `/api/v1/po/*` needs only the JWT — measured 2026-09-01 returning
            # 200 with no WAF token.
            with console.status("[cyan]Pre-flight: Zepto session…[/cyan]"):
                storage_state = {
                    "jwt": (await auth_service.ensure(db, tenant_id, "zepto")).raw["jwt"]
                }
            console.print("[green]Session healthy.[/green]")

            job_id = await create_scrape_job(db, tenant_id, "zepto_po", platform="zepto")

            # Each endpoint is fetched independently: Zepto returned a 500 on
            # asn/filter on 2026-08-27 and aborting the run discarded 74 POs and
            # 72 GRNs that had already come back. One flaky endpoint should cost
            # its own data, not the whole scrape.
            async def _try(label, coro):
                try:
                    return await coro
                except Exception as e:
                    logger.warning(f"Zepto {label} failed, continuing without it: {e}")
                    console.print(f"[yellow]{label} failed — continuing[/yellow]")
                    return []

            with console.status(f"[cyan]Fetching POs {date_from}..{date_to}...[/cyan]"):
                raw_pos = await _try("po/filter", zepto_fetch_pos(storage_state, date_from, date_to))
            with console.status(f"[cyan]Fetching GRNs {date_from}..{date_to}...[/cyan]"):
                raw_grns = await _try("grn/filter", zepto_fetch_grns(storage_state, date_from, date_to))
            with console.status(f"[cyan]Fetching ASNs {date_from}..{date_to}...[/cyan]"):
                raw_asns = await _try("asn/filter", zepto_fetch_asns(storage_state, date_from, date_to))

            pos = parse_zepto_pos(raw_pos, tenant_id, job_id)
            grns = parse_zepto_grns(raw_grns, tenant_id, job_id)
            asns = parse_zepto_asns(raw_asns, tenant_id, job_id)

            # Line items: one GET per PO. Carries unit_price (cost) and mrp,
            # which appear on no other Zepto endpoint, plus per-SKU fill rate.
            po_ids = [p["po_id"] for p in pos]
            with console.status(f"[cyan]Fetching line items for {len(po_ids)} POs...[/cyan]"):
                raw_items = await zepto_fetch_po_items(storage_state, po_ids)
            po_items = parse_zepto_po_items(raw_items, tenant_id, job_id)

            written = 0
            if save:
                counts = await zepto_save_po_results(db, pos, grns, asns, po_items)
                written = sum(counts.values())
                await complete_scrape_job(db, job_id, written)
            else:
                await complete_scrape_job(db, job_id)

            ordered = sum(p["total_qty"] or 0 for p in pos)
            received = sum(g["grn_qty"] or 0 for g in grns)
            po_value = sum(p["total_value"] or 0.0 for p in pos)

            console.print("")
            console.print(f"[bold]Zepto PO Management ({date_from} to {date_to})[/bold]")
            console.print(f"  POs: {len(pos)}   units ordered: {ordered:,}   value: Rs {po_value:,.0f}")
            console.print(f"  GRNs: {len(grns)}   units received: {received:,}")
            console.print(f"  ASNs: {len(asns)}")
            console.print(f"  PO line items: {len(po_items)}")

            # Fill rate from GRN rows, where ordered and received sit together.
            po_q = sum(g["po_qty"] or 0 for g in grns)
            grn_q = sum(g["grn_qty"] or 0 for g in grns)
            if po_q:
                console.print(f"  [bold]Fill rate: {grn_q:,}/{po_q:,} = {100 * grn_q / po_q:.1f}%[/bold]")
            if save:
                console.print(f"  Saved to DB: {written} rows")
        except Exception as e:
            if job_id:
                await fail_scrape_job(db, job_id, str(e))
            console.print(f"[red]Scrape failed: {escape(str(e))}[/red]")
            raise typer.Exit(1)
