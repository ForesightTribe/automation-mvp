"""Config sync: reconcile the DB to a source-of-truth workbook.

One command reads `config.xlsx` (sheets: locations / brands / coverage / city_map) and makes
the DB match it — upserting by default, deleting only with --prune. The workbook
is the source of truth; scrapers never read it. Updates are infrequent, so this
replaces the per-row add/remove CLI verbs.

Sheets
  locations : the per-marketplace store catalog (keyed on mp + merchant_id)
              cols: mp, merchant_id, city, state, region, pincode, lat, lon,
                    active, location_name, address
  brands    : per-tenant keywords/aliases (the watchlist)
              cols: tenant, brand, relationship, keywords, aliases,
                    keyword_cap, brand_cap   (caps optional; own rows only)
  coverage  : which catalog locations each tenant scrapes
              cols: mp, tenant, city   (one row per marketplace+city covered)
  city_map  : how a marketplace's own city names map onto the catalog (V7.3)
              cols: source, alias, city, pincode_prefixes
              Two row kinds: an ALIAS (source+alias+city) for a name that IS
              one city, or a PREFIX row (city+pincode_prefixes) saying which
              pincodes belong to it — that is how a grouped catalog name like
              `hr-ncr` splits across several real cities.
              Only EXCEPTIONS need a row — a marketplace city whose name already
              equals a catalog city resolves without one (228 of Blinkit's 242).

`mp` is optional and defaults to `blinkit`, so a pre-multi-marketplace workbook
syncs unchanged. **Each marketplace has its own catalog** — a store's coordinate is
that platform's catchment and is never borrowed from another (see docs/zepto.md).

⚠️ `--prune` only ever deletes within the marketplaces the FILE mentions. A
workbook with no Zepto rows can never prune Zepto — but a workbook whose `mp`
column is blank on Zepto rows reads them as Blinkit and will propose deleting the
real Blinkit catalog. Always `--dry-run` first and treat a non-zero `locations`
deletion as a stop-and-check.
"""
import asyncio

import typer
from openpyxl import Workbook, load_workbook
from rich.console import Console
from rich.table import Table
from sqlalchemy import text
from sqlmodel import select

from app.core.database import AsyncSessionLocal
from app.models.search import City, CityAlias, MarketplaceLocation, TenantLocation
from app.models.tenant import Tenant, TenantWatchlist
from scraper.utils.storage import _MP_NAMES, ensure_refs

console = Console()
DEFAULT_MP = "blinkit"


def _mp(v) -> str:
    """The row's marketplace slug; blank → blinkit (back-compat for older files)."""
    return _str(v).lower() or DEFAULT_MP


# ── cell coercion ────────────────────────────────────────────────────────────

def _str(v) -> str:
    return "" if v is None else str(v).strip()


def _pincode(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _float(v):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _int(v):
    if v is None or v == "":
        return None
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return None


def _bool(v, default=True) -> bool:
    if v is None or v == "":
        return default
    return str(v).strip().lower() in {"1", "yes", "y", "true", "t", "active"}


def _csv(v) -> list[str]:
    return [x.strip() for x in str(v).split(",") if x.strip()] if v else []


def _read_sheet(wb, name: str) -> list[dict]:
    if name not in wb.sheetnames:
        console.print(f"[yellow]Sheet '{name}' not found — skipping.[/yellow]")
        return []
    rows = list(wb[name].iter_rows(values_only=True))
    if not rows:
        return []
    header = [(_str(c).lower()) for c in rows[0]]
    out = []
    for r in rows[1:]:
        if all(c is None for c in r):
            continue
        out.append({header[i]: r[i] for i in range(len(header)) if header[i]})
    return out


def _read_config(path: str) -> dict:
    wb = load_workbook(path, data_only=True)
    return {
        "locations": _read_sheet(wb, "locations"),
        "brands": _read_sheet(wb, "brands"),
        "coverage": _read_sheet(wb, "coverage"),
        # Optional — an older workbook without it syncs as before (_read_sheet warns
        # and returns []), and an absent sheet must never be read as "delete the map".
        "city_map": _read_sheet(wb, "city_map"),
    }


# ── reconciliation ───────────────────────────────────────────────────────────

_LOC_FIELDS = ("city", "state", "region", "pincode", "location_name", "address",
               "lat", "lon", "is_active")


async def _sync_locations(db, rows, prune) -> dict:
    desired = {}
    for r in rows:
        mid = _str(r.get("merchant_id"))
        if not mid:
            continue
        desired[(_mp(r.get("mp")), mid)] = {
            "city": _str(r.get("city")).lower(),
            "state": _str(r.get("state")),
            "region": _str(r.get("region")),
            "pincode": _pincode(r.get("pincode")),
            "location_name": _str(r.get("location_name")),
            "address": _str(r.get("address")),
            "lat": _float(r.get("lat")),
            "lon": _float(r.get("lon")),
            "is_active": _bool(r.get("active")),
        }
    # Only load (and therefore only ever prune) the marketplaces this file declares.
    # A workbook that says nothing about a marketplace must not be able to delete it.
    mps = {mp for mp, _ in desired}
    existing = {
        (l.mp_slug, l.merchant_id): l
        for l in (await db.execute(
            select(MarketplaceLocation).where(MarketplaceLocation.mp_slug.in_(mps or [None]))
        )).scalars().all()
    }
    added = updated = deleted = 0
    for (mp, mid), vals in desired.items():
        cur = existing.get((mp, mid))
        if cur is None:
            db.add(MarketplaceLocation(mp_slug=mp, merchant_id=mid, **vals))
            added += 1
        elif any(getattr(cur, f) != vals[f] for f in _LOC_FIELDS):
            for f in _LOC_FIELDS:
                setattr(cur, f, vals[f])
            updated += 1
    if prune:
        for key, cur in existing.items():
            if key not in desired:
                await db.execute(
                    TenantLocation.__table__.delete().where(TenantLocation.location_id == cur.id)
                )
                await db.delete(cur)
                deleted += 1
    await db.flush()
    return {"added": added, "updated": updated, "deleted": deleted}


async def _sync_brands(db, rows, tenants, prune, tenant_mps: dict | None = None) -> dict:
    desired = {}
    for r in rows:
        tname = _str(r.get("tenant"))
        brand = _str(r.get("brand")).lower()
        if not tname or not brand:
            continue
        desired[(tenants[tname], brand)] = {
            "relationship": _str(r.get("relationship")).lower() or "own",
            "keywords": _csv(r.get("keywords")),
            "aliases": _csv(r.get("aliases")),
            "keyword_cap": _int(r.get("keyword_cap")),
            "brand_cap": _int(r.get("brand_cap")),
        }
    ref_tenants = {tid for tid, _ in desired}
    existing = {
        (w.tenant_id, w.brand_slug): w
        for w in (await db.execute(
            select(TenantWatchlist).where(TenantWatchlist.tenant_id.in_(ref_tenants or [None]))
        )).scalars().all()
    }
    tenant_mps = tenant_mps or {}
    added = updated = deleted = 0
    for (tid, brand), vals in desired.items():
        mps = tenant_mps.get(tid) or [DEFAULT_MP]
        for mp in mps:
            await ensure_refs(db, brand, mp)
        cur = existing.get((tid, brand))
        if cur is None:
            # `marketplaces` is descriptive metadata (the scrape path selects the
            # platform via --marketplace, not from here), so it is seeded on INSERT
            # only — rewriting it on every sync would churn rows for no effect.
            db.add(TenantWatchlist(
                tenant_id=tid, brand_slug=brand, cities=[], marketplaces=mps, **vals
            ))
            added += 1
        elif (cur.relationship, cur.keywords, cur.aliases, cur.keyword_cap, cur.brand_cap) != (
            vals["relationship"], vals["keywords"], vals["aliases"],
            vals["keyword_cap"], vals["brand_cap"],
        ):
            (cur.relationship, cur.keywords, cur.aliases, cur.keyword_cap, cur.brand_cap) = (
                vals["relationship"], vals["keywords"], vals["aliases"],
                vals["keyword_cap"], vals["brand_cap"],
            )
            updated += 1
    if prune:
        for key, cur in existing.items():
            if key not in desired:
                await db.delete(cur)
                deleted += 1
    await db.flush()
    return {"added": added, "updated": updated, "deleted": deleted}


async def _sync_coverage(db, rows, tenants, prune) -> dict:
    # Resolve each (marketplace, tenant, city) to catalog location ids.
    desired: dict[tuple, str] = {}
    covered_tenants = set()
    mps = set()
    for r in rows:
        tname = _str(r.get("tenant"))
        city = _str(r.get("city")).lower()
        if not tname or not city:
            continue
        mp = _mp(r.get("mp"))
        tid = tenants[tname]
        covered_tenants.add(tid)
        mps.add(mp)
        q = select(MarketplaceLocation.id).where(
            MarketplaceLocation.mp_slug == mp, MarketplaceLocation.city == city
        )
        for lid in (await db.execute(q)).scalars().all():
            desired[(tid, lid)] = mp

    # Scoped to the marketplaces the file declares, for the same reason as locations:
    # silence about a marketplace must never be read as "delete it".
    existing = {
        (t.tenant_id, t.location_id): t
        for t in (await db.execute(
            select(TenantLocation).where(
                TenantLocation.tenant_id.in_(covered_tenants or [None]),
                TenantLocation.mp_slug.in_(mps or [None]),
            )
        )).scalars().all()
    }
    added = deleted = 0
    for (tid, lid), mp in desired.items():
        if (tid, lid) not in existing:
            db.add(TenantLocation(tenant_id=tid, mp_slug=mp, location_id=lid))
            added += 1
    if prune:
        for key, cur in existing.items():
            if key not in desired:
                await db.delete(cur)
                deleted += 1
    await db.flush()
    return {"added": added, "deleted": deleted}


async def _sync_city_map(db, rows, prune) -> dict:
    """Reconcile the `city_map` sheet into `city_aliases` and cities' pincode prefixes (V7.3).

    The sheet carries TWO kinds of row, and the difference matters:

      - **Alias** (`source` + `alias` + `city`) — a name that IS one city, on one surface.
        `blinkit:catalog | aurangabad (maharashtra) | Aurangabad`.
      - **Prefix** (`city` + `pincode_prefixes`, no source/alias) — which pincodes belong
        to a city. This is how a GROUPED catalog name is handled: `hr-ncr` gets no alias at
        all, because it is not one city. Its 77 stores each resolve by their own pincode,
        which is what lets one bucket split into several cities (`up-ncr` → Noida AND
        Ghaziabad). An alias could never express that.

    Only exceptions belong here: a name that already equals a canonical city resolves
    without a row (228 of Blinkit's 242). `scripts/blinkit_city_match.py` generates the
    sheet.
    """
    cities = {c.slug: c for c in (await db.execute(select(City))).scalars().all()}
    by_name = {(c.name or "").strip().lower(): c for c in cities.values()}

    def _city(ref: str):
        """A sheet's `city` cell — canonical slug or display name."""
        key = ref.strip().lower()
        return cities.get(key.replace(" ", "-")) or by_name.get(key) or cities.get(key)

    alias_rows: dict[tuple, int] = {}
    prefix_rows: dict[str, list] = {}
    unknown: list[str] = []
    for r in rows:
        ref = _str(r.get("city"))
        if not ref:
            continue
        city = _city(ref)
        if city is None:
            unknown.append(ref)
            continue
        # Excel turns "1220" into the number 1220 — read prefixes back as strings or a
        # prefix match against a text pincode silently never fires.
        prefixes = [str(p).strip() for p in _csv(r.get("pincode_prefixes"))]
        if prefixes:
            prefix_rows.setdefault(city.slug, []).extend(prefixes)
        source, alias = _str(r.get("source")).lower(), _str(r.get("alias")).lower()
        if source and alias:
            alias_rows[(source, alias)] = city.id

    if unknown:
        raise ValueError(
            f"city_map references unknown cities: {sorted(set(unknown))}. "
            f"Run `cli cities seed` first, or fix the spelling."
        )

    # Prefixes from the sheet are AUTHORITATIVE over the derived ones — they exist
    # precisely for the cities `cli cities seed` cannot derive, because no catalog name
    # matches them.
    updated_cities = 0
    for slug, prefixes in prefix_rows.items():
        want = sorted(set(prefixes))
        if cities[slug].pincode_prefixes != want:
            cities[slug].pincode_prefixes = want
            updated_cities += 1

    existing = {
        (a.source, a.alias): a
        for a in (await db.execute(select(CityAlias))).scalars().all()
    }
    added = updated = deleted = 0
    for (source, alias), city_id in alias_rows.items():
        cur = existing.get((source, alias))
        if cur is None:
            db.add(CityAlias(source=source, alias=alias, city_id=city_id, is_active=True))
            added += 1
        elif cur.city_id != city_id or not cur.is_active:
            cur.city_id, cur.is_active = city_id, True
            updated += 1
    if prune:
        for key, cur in existing.items():
            if key not in alias_rows:
                await db.delete(cur)
                deleted += 1
    await db.flush()
    return {"added": added, "updated": updated + updated_cities, "deleted": deleted}


async def _tag_store_cities(db, mps: set) -> dict:
    """Give every store its canonical `city_id` (V7.3).

    Resolution order, and the order is the whole design:

      1. an ALIAS for this marketplace's catalog — an explicit human decision, so it wins;
      2. an exact canonical NAME match — true for 228 of Blinkit's 242 cities;
      3. the store's PINCODE. Last because it is derived, but it is the only thing that can
         resolve a grouped catalog name (`hr-ncr`, `delhi ncr`), since those match no city
         by name at all and cover several cities at once.

    A store that matches nothing keeps `city_id = NULL` — the picker then simply does not
    offer it, which is visible and recoverable. Guessing would silently measure a campaign's
    ad position in the wrong city.
    """
    cities = (await db.execute(select(City))).scalars().all()
    by_name = {(c.name or "").strip().lower(): c.id for c in cities}
    by_prefix: dict[str, int] = {}
    for c in cities:
        for p in c.pincode_prefixes or []:
            by_prefix[str(p)] = c.id
    aliases = {
        (a.source, a.alias): a.city_id
        for a in (await db.execute(select(CityAlias))).scalars().all() if a.is_active
    }

    stores = (await db.execute(
        select(MarketplaceLocation).where(MarketplaceLocation.mp_slug.in_(mps or [None]))
    )).scalars().all()

    tagged = 0
    for s in stores:
        name = (s.city or "").strip().lower()
        city_id = aliases.get((f"{s.mp_slug}:catalog", name)) or by_name.get(name)
        if city_id is None and s.pincode:
            # Longest prefix first, so a 4-digit rule beats the 3-digit one it sits inside
            # (2013 → Noida must win over any 201 → …).
            for length in (4, 3):
                city_id = by_prefix.get(s.pincode[:length])
                if city_id:
                    break
        if city_id and s.city_id != city_id:
            s.city_id = city_id
            tagged += 1
    await db.flush()
    return {"updated": tagged}


def _tenant_marketplaces(rows, tenants) -> dict:
    """tenant_id -> the marketplaces it covers, from the coverage sheet. Seeds a NEW
    watchlist row's `marketplaces`; existing rows are never rewritten from here."""
    out: dict = {}
    for r in rows:
        tname = _str(r.get("tenant"))
        if not tname or tname not in tenants:
            continue
        out.setdefault(tenants[tname], set()).add(_mp(r.get("mp")))
    return {tid: sorted(mps) for tid, mps in out.items()}


async def _do_sync(data, dry_run, prune) -> dict:
    async with AsyncSessionLocal() as db:
        tenants = {t.name: t.id for t in (await db.execute(select(Tenant))).scalars().all()}
        referenced = {_str(r.get("tenant")) for r in data["brands"] + data["coverage"] if _str(r.get("tenant"))}
        unknown = sorted(referenced - set(tenants))
        if unknown:
            raise ValueError(
                f"Unknown tenant(s) in file: {unknown}. Create them with `cli tenant create` first."
            )

        tenant_mps = _tenant_marketplaces(data["coverage"], tenants)

        # `marketplace_locations.mp_slug` is an FK to `marketplaces.slug`. That row
        # is created by ensure_refs() — but only on the scrape path and inside
        # _sync_brands, which runs AFTER locations. So a marketplace that has never
        # been scraped could not have its catalog synced: the locations insert died
        # on the foreign key. Blinkit never hit it because a scrape had seeded its
        # row long ago. Seed every marketplace the file declares, first.
        for mp in sorted({_mp(r.get("mp")) for r in data["locations"]}):
            await db.execute(
                text("INSERT INTO marketplaces (slug, name) VALUES (:slug, :name) "
                     "ON CONFLICT (slug) DO NOTHING"),
                {"slug": mp, "name": _MP_NAMES.get(mp, mp.title())},
            )

        loc = await _sync_locations(db, data["locations"], prune)
        brands = await _sync_brands(db, data["brands"], tenants, prune, tenant_mps)
        cov = await _sync_coverage(db, data["coverage"], tenants, prune)
        cmap = await _sync_city_map(db, data.get("city_map") or [], prune)
        # After aliases + prefixes exist, give every store its canonical city.
        tagged = await _tag_store_cities(db, {_mp(r.get("mp")) for r in data["locations"]})

        if dry_run:
            await db.rollback()
        else:
            await db.commit()
    return {"locations": loc, "brands": brands, "coverage": cov, "city_map": cmap,
            "store cities": tagged}


# ── template ─────────────────────────────────────────────────────────────────

def _write_template(path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "locations"
    ws.append(["mp", "merchant_id", "city", "state", "region", "pincode", "lat", "lon",
               "active", "location_name", "address"])
    ws.append(["blinkit", "48967", "delhi", "Delhi", "North India", "110001", 28.6315,
               77.2167, "yes", "Connaught Place",
               "Block A, Connaught Place, New Delhi, Delhi 110001, India"])

    b = wb.create_sheet("brands")
    b.append(["tenant", "brand", "relationship", "keywords", "aliases", "keyword_cap", "brand_cap"])
    b.append(["Dobra", "dobra", "own", "dobra, soda, goli soda", "dobra", 12, 60])
    b.append(["Dobra", "bisleri", "competitor", "", "bisleri", "", ""])

    c = wb.create_sheet("coverage")
    c.append(["mp", "tenant", "city"])
    c.append(["blinkit", "Dobra", "delhi"])

    # Only EXCEPTIONS need a row — a marketplace city whose name already matches a
    # catalog city resolves without one. `scripts/blinkit_city_match.py` generates it.
    m = wb.create_sheet("city_map")
    m.append(["source", "alias", "city", "pincode_prefixes"])
    # An ALIAS row: a name that IS one city, on one surface.
    m.append(["blinkit:catalog", "aurangabad (maharashtra)", "Aurangabad", ""])
    m.append(["blinkit:seller", "Gurgaon", "Gurugram", ""])
    # PREFIX rows: which pincodes belong to a city. A grouped catalog name like `hr-ncr`
    # gets NO alias — it is not one city; its stores resolve individually by pincode.
    m.append(["", "", "Gurugram", "122"])
    m.append(["", "", "Noida", "2013"])
    m.append(["", "", "Ghaziabad", "2010,2011"])

    wb.save(path)


# ── command ──────────────────────────────────────────────────────────────────

def run_sync(
    file: str = typer.Option("config.xlsx", "--file", "-f", help="Path to the config workbook"),
    dry_run: bool = typer.Option(False, "--dry-run", help="Preview changes without writing"),
    prune: bool = typer.Option(False, "--prune", help="Also delete DB rows missing from the file"),
    template: bool = typer.Option(False, "--template", help="Write a starter workbook to --file and exit"),
):
    """Reconcile the DB to the config workbook (locations + brands + coverage + city_map)."""
    if template:
        _write_template(file)
        console.print(f"[green]Template written to[/green] {file}")
        return

    data = _read_config(file)
    console.print(
        f"[dim]Read {len(data['locations'])} locations, {len(data['brands'])} brands, "
        f"{len(data['coverage'])} coverage, {len(data['city_map'])} city_map rows from {file}[/dim]"
    )
    result = asyncio.run(_do_sync(data, dry_run, prune))

    table = Table(show_header=True, header_style="bold",
                  title="DRY RUN — no changes written" if dry_run else "Synced")
    table.add_column("Section")
    table.add_column("Added", justify="right")
    table.add_column("Updated", justify="right")
    table.add_column("Deleted", justify="right")
    for section in ("locations", "brands", "coverage", "city_map", "store cities"):
        s = result[section]
        table.add_row(section, str(s.get("added", 0)), str(s.get("updated", 0)), str(s.get("deleted", 0)))
    console.print(table)
    if not prune:
        console.print("[dim]Deletions disabled (no --prune): rows missing from the file were kept.[/dim]")
