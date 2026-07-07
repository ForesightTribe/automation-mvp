"""CLI for the private↔public SKU map (`sku_map`).

  build : auto-match own private items to public products by normalized name,
          upsert the map, and export it to a workbook for review.
  apply : read a reviewed workbook and apply the manual `platform_product_id`
          corrections (marks them method='manual', preserved on future builds).

Typical flow: `sku-map build` → open the .xlsx, fill in `platform_product_id`
for the unmatched rows → `sku-map apply --file <same.xlsx>`.
"""
import asyncio
import uuid

import typer
from openpyxl import Workbook, load_workbook
from rich.console import Console
from rich.table import Table

from app.core.database import AsyncSessionLocal
from app.services import sku_map_service

app = typer.Typer(help="Map private seller item_id ↔ public platform_product_id.")
console = Console()

_COLS = ["item_id", "item_name", "platform_product_id", "product_name", "match_method"]


def _str(v) -> str:
    return "" if v is None else str(v).strip()


async def _build(tenant_id: str, file: str) -> None:
    tid = uuid.UUID(tenant_id)
    async with AsyncSessionLocal() as db:
        report = await sku_map_service.build_map(db, tid)
        rows = await sku_map_service.list_map(db, tid)

    table = Table(show_header=True, header_style="bold", title="sku-map build")
    for k in ("private_own_items", "matched", "unmatched", "preserved_manual"):
        table.add_column(k)
    table.add_row(*[str(report[k]) for k in
                    ("private_own_items", "matched", "unmatched", "preserved_manual")])
    console.print(table)

    unmatched = [r for r in rows if not r.platform_product_id]
    if unmatched:
        console.print(f"\n[yellow]{len(unmatched)} unmatched — fill platform_product_id in the workbook:[/yellow]")
        for r in unmatched:
            console.print(f"  [dim]{r.item_id}[/dim]  {r.item_name}")

    wb = Workbook()
    ws = wb.active
    ws.title = "sku_map"
    ws.append(_COLS)
    for r in rows:
        ws.append([r.item_id, r.item_name, r.platform_product_id or "",
                   r.product_name, r.match_method])
    wb.save(file)
    console.print(f"\n[green]Wrote {len(rows)} rows to[/green] {file} "
                  f"[dim](edit platform_product_id, then `sku-map apply --file {file}`)[/dim]")


async def _apply(tenant_id: str, file: str) -> None:
    tid = uuid.UUID(tenant_id)
    wb = load_workbook(file, data_only=True)
    ws = wb["sku_map"] if "sku_map" in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    header = [_str(c).lower() for c in rows[0]]
    i_item = header.index("item_id")
    i_pid = header.index("platform_product_id")

    pairs = []
    for r in rows[1:]:
        item_id = _str(r[i_item])
        pid = _str(r[i_pid])
        if item_id and pid:
            pairs.append((item_id, pid))

    async with AsyncSessionLocal() as db:
        report = await sku_map_service.apply_corrections(db, tid, pairs)
    console.print(f"[green]Applied {report['applied']} mapping(s) from[/green] {file}")


@app.command("build")
def build(
    tenant_id: str = typer.Option(..., "--tenant", "-t", help="Tenant (client) UUID"),
    file: str = typer.Option("sku_map.xlsx", "--file", "-f", help="Workbook to write for review"),
):
    """Auto-match by name, upsert sku_map, and export a review workbook."""
    asyncio.run(_build(tenant_id, file))


@app.command("apply")
def apply(
    tenant_id: str = typer.Option(..., "--tenant", "-t", help="Tenant (client) UUID"),
    file: str = typer.Option("sku_map.xlsx", "--file", "-f", help="Reviewed workbook to import"),
):
    """Apply manual platform_product_id corrections from a reviewed workbook."""
    asyncio.run(_apply(tenant_id, file))
