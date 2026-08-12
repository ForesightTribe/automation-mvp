"""The shared Excel renderer — the only module that styles a cell.

`write_workbook(report, path)` turns a typed `Report` into an .xlsx. It formats;
it never aggregates. Sections declare *what a column is* (`count`, `pct`,
`money`, …) and this module decides width, format, alignment and emphasis, so
every sheet of every export lines up without anyone thinking about it.

Sheet skeleton, identical everywhere:

    A1  Sheet Title                      13pt semibold
    A2  One plain-English line.          10pt muted
    A3  window · filter · N observed     9pt muted
    A4  (spacer)
    A5  HEADER ROW  — frozen, autofiltered
    A6+ data
"""
from datetime import date, datetime

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.worksheet.worksheet import Worksheet

from exports import glossary
from exports import theme as t
from app.schemas.exports import Column, Report, Section

_ILLEGAL = set(r"[]:*?/\\")
_LINK_COL = 6          # "← Contents" parks at F1 — clear of the title, on screen
_CONTENTS = "Contents"
_GLOSSARY = "How to read this"


# ── Sheet names ───────────────────────────────────────────────────────────────

def _sheet_name(title: str, used: set[str]) -> str:
    """Excel rejects >31 chars and []:*?/\\ — and silently dies at save() rather
    than when the offending section was written. Sanitize up front."""
    clean = "".join(c for c in title if c not in _ILLEGAL).strip() or "Sheet"
    clean = clean[:31]
    name, n = clean, 2
    while name.lower() in used:
        suffix = f" ({n})"
        name = clean[: 31 - len(suffix)] + suffix
        n += 1
    used.add(name.lower())
    return name


# ── Value rendering ───────────────────────────────────────────────────────────

def _display_len(value, col_type: str) -> int:
    """Approximate the *rendered* width of a value — ₹1,234 is 6 characters even
    though the underlying float is 1234.0."""
    if value is None or value == "":
        return 1
    if col_type in ("count",):
        return len(f"{value:,.0f}") if isinstance(value, (int, float)) else len(str(value))
    if col_type == "pct":
        return len(f"{value:,.1f}") + 1 if isinstance(value, (int, float)) else len(str(value))
    if col_type == "money":
        return len(f"{value:,.0f}") + 1 if isinstance(value, (int, float)) else len(str(value))
    if col_type == "money_fine":
        return len(f"{value:,.2f}") + 1 if isinstance(value, (int, float)) else len(str(value))
    if col_type == "rating":
        return 4
    if col_type == "date":
        return 11
    return len(str(value))


def _width(col: Column, rows: list[dict]) -> int:
    """Measured from the 90th percentile, not the maximum: one 90-character
    product name must not blow the column to its cap and push every other column
    off the screen."""
    if col.width:
        return col.width
    spec = t.COLUMN_TYPES[col.type]
    if col.wrap:
        return spec["max"]
    lens = sorted(
        _display_len(r.get(col.key), col.type)
        for r in rows[: t.WIDTH_SAMPLE]
        if r.get(col.key) not in (None, "")
    )
    p90 = lens[min(int(len(lens) * 0.9), len(lens) - 1)] if lens else 0
    return max(spec["min"], min(spec["max"], max(len(col.header), p90) + 3))


def _header_height(cols: list[Column], widths: list[int]) -> float:
    """Tall enough for the longest header to wrap without clipping.

    A fixed height silently truncates: "Stores with none left" in a 16-wide
    column renders as "Stores with none l" and reads as a different metric.
    """
    lines = 1
    for col, width in zip(cols, widths):
        lines = max(lines, min(t.WRAP_MAX_LINES,
                               -(-len(col.header) // max(width * t.PROSE_DENSITY, 1))))
    # Generous padding: Excel measures wrapped text against the font's full line
    # box, so a height that looks arithmetically sufficient still clips.
    return max(t.HEIGHT_HEADER, t.HEIGHT_HEADER + t.LINE_HEIGHT * (lines - 1) + 6)


def _wrap_height(cols: list[Column], rows: list[dict]) -> float | None:
    """Rows needed by the widest wrapping cell, as a height. None when nothing
    actually wraps."""
    lines = 1
    for col in (c for c in cols if c.wrap):
        width = col.width or t.COLUMN_TYPES[col.type]["max"]
        # An Excel width unit is the width of "0" in the default font, and
        # lowercase prose is narrower than a digit — so a 40-wide column holds
        # noticeably more than 40 characters of a sentence. Without this factor
        # every prose column claims it needs a second line.
        capacity = max(width * t.PROSE_DENSITY, 1)
        longest = max((len(str(r.get(col.key) or "")) for r in rows[: t.WIDTH_SAMPLE]), default=0)
        lines = max(lines, min(t.WRAP_MAX_LINES, -(-longest // capacity)))
    return t.LINE_HEIGHT * lines + 4 if lines > 1 else None


def _put(ws: Worksheet, row: int, col: int, value, column: Column, *,
         no_rule: bool = False, blank_empty: bool = False):
    """Write one data cell. Empty becomes a muted em dash — a blank cell reads as
    zero, which is a different claim entirely. `blank_empty` is for the totals
    row, where a missing product id means "not applicable", not "no data";
    `no_rule` suppresses the row hairline where another border takes over."""
    cell = ws.cell(row=row, column=col)
    spec = t.COLUMN_TYPES[column.type]
    if value is None or value == "":
        if not blank_empty:
            cell.value = t.DASH
            cell.font = t.F_MUTED
            cell.alignment = t.AL_CENTER
    else:
        cell.value = value
        cell.font = t.F_BODY
        cell.alignment = t.AL_WRAP if column.wrap else spec["align"]
        # datetime must be included: without it openpyxl falls back to its own
        # "yyyy-mm-dd h:mm:ss" and the date column shows a midnight timestamp.
        if spec["fmt"] and isinstance(value, (int, float, date, datetime)):
            cell.number_format = spec["fmt"]
    if not no_rule:
        cell.border = t.BORDER_ROW
    return cell


def _chip(cell, column: Column, value) -> bool:
    """Soft status fill for a declared value ("Out of stock" → bad). Returns True
    when a chip was applied, so row banding doesn't overwrite it."""
    tone = column.chips.get(value) if isinstance(value, str) else None
    if not tone:
        return False
    fill, font = t.CHIP[tone]
    cell.fill = fill
    cell.font = font
    cell.alignment = t.AL_CENTER
    return True


# ── Blocks ────────────────────────────────────────────────────────────────────

def _head_block(ws: Worksheet, section: Section, *, link_back: bool) -> None:
    ws.cell(t.ROW_TITLE, 1, section.title).font = t.F_TITLE
    ws.cell(t.ROW_DESC, 1, section.description).font = t.F_SUBTITLE
    ws.cell(t.ROW_CONTEXT, 1, section.context).font = t.F_CONTEXT
    for row, height in t.HEIGHT.items():
        ws.row_dimensions[row].height = height
    if link_back:
        cell = ws.cell(t.ROW_TITLE, _LINK_COL, f"← {_CONTENTS}")
        cell.font = t.F_LINK
        cell.hyperlink = Hyperlink(ref=cell.coordinate, location=f"'{_CONTENTS}'!A1")


def _kpi_block(ws: Worksheet, section: Section, start: int) -> int:
    """Headline numbers as a vertical list: label · value · the counts behind it.
    Vertical beats a row of cards here — cards are a web idiom that reads badly in
    a grid, and the detail column is where clarity rule #2 gets satisfied."""
    row = start
    for kpi in section.kpis:
        ws.cell(row, 1, kpi.label).font = t.F_KPI_LABEL
        cell = ws.cell(row, 2, kpi.value if kpi.value is not None else t.DASH)
        cell.font = t.F_KPI_VALUE if kpi.value is not None else t.F_MUTED
        cell.alignment = t.AL_LEFT
        fmt = t.COLUMN_TYPES[kpi.type]["fmt"]
        if fmt and isinstance(kpi.value, (int, float)):
            cell.number_format = fmt
        if kpi.detail:
            ws.cell(row, 3, kpi.detail).font = t.F_MUTED
        if kpi.help:
            ws.cell(row, 1).comment = Comment(kpi.help, "Foresight", width=320, height=90)
        ws.row_dimensions[row].height = 24
        row += 1
    return row + 1      # one blank row after the block


def _emphasis(ws: Worksheet, col: int, kind: str, first: int, last: int) -> None:
    """Data bars for magnitudes, two-colour scales for good/bad. The harsh
    three-colour ramp is `heat`, and it belongs only on a grid whose entire job is
    finding weak cells."""
    if last < first:
        return
    letter = get_column_letter(col)
    rng = f"{letter}{first}:{letter}{last}"
    if kind == "bar":
        ws.conditional_formatting.add(rng, DataBarRule(
            start_type="num", start_value=0, end_type="max",
            color=t.BAR_COLOR, showValue=True,
        ))
    elif kind in ("good_high", "good_low"):
        end = t.SCALE_GOOD if kind == "good_high" else t.SCALE_BAD
        ws.conditional_formatting.add(rng, ColorScaleRule(
            start_type="min", start_color=t.PAPER,
            end_type="max", end_color=end,
        ))

def _heat_block(ws: Worksheet, first_col: int, last_col: int, first: int, last: int) -> None:
    """One colour scale across the WHOLE grid, not one per column.

    Per-column scales rescale to each column's own min and max, so an average
    position of 5 shows green in a city that ranges 1–35 and red in one that
    ranges 1–5. On a "where am I weak" map that inverts the message. A single
    rule over the block makes every cell comparable to every other.
    """
    if last < first or last_col < first_col:
        return
    rng = (f"{get_column_letter(first_col)}{first}:"
           f"{get_column_letter(last_col)}{last}")
    ws.conditional_formatting.add(rng, ColorScaleRule(
        start_type="min", start_color=t.SCALE_GOOD,
        mid_type="percentile", mid_value=50, mid_color=t.SCALE_MID,
        end_type="max", end_color=t.SCALE_BAD,
    ))


def _table(ws: Worksheet, section: Section, start: int) -> int:
    dense = section.dense
    cols = section.columns
    header_row = start
    # Widths first: the header row's height depends on how many lines each
    # header needs, which depends on how wide its column ended up.
    widths = [_width(col, section.rows) for col in cols]

    for j, col in enumerate(cols, 1):
        cell = ws.cell(header_row, j, col.header)
        cell.font = t.F_HEADER
        cell.fill = t.FILL_HEADER
        cell.border = t.BORDER_HEADER
        # Headers align with their data — a right-aligned number column under a
        # centred header is what makes a table look wonky.
        cell.alignment = Alignment(
            horizontal=t.COLUMN_TYPES[col.type]["align"].horizontal,
            vertical="center", wrap_text=True,
        )
        if col.help:
            cell.comment = Comment(col.help, "Foresight", width=320, height=110)
    ws.row_dimensions[header_row].height = _header_height(cols, widths)

    # Excel only auto-fits a wrapped row when it recalculates the height itself,
    # which it does not do for rows openpyxl wrote — so a `wrap` column would
    # silently clip. Set an explicit height, but only when the text genuinely
    # overflows: padding every row of a sheet whose text happens to fit just
    # spreads it out for nothing.
    wrap_height = _wrap_height(cols, section.rows)

    first_data = header_row + 1
    for i, row in enumerate(section.rows):
        r = first_data + i
        if wrap_height:
            ws.row_dimensions[r].height = wrap_height
        own = bool(section.highlight_key and row.get(section.highlight_key))
        for j, col in enumerate(cols, 1):
            value = row.get(col.key)
            cell = _put(ws, r, j, value, col, no_rule=dense)
            if dense:
                continue
            if _chip(cell, col, value):
                continue
            if own:
                cell.fill = t.FILL_OWN
            elif i % 2:
                cell.fill = t.FILL_BAND
    last_data = first_data + len(section.rows) - 1

    if section.total_row is not None:
        r = last_data + 1
        for j, col in enumerate(cols, 1):
            cell = _put(ws, r, j, section.total_row.get(col.key), col,
                        no_rule=True, blank_empty=True)
            cell.font = t.F_TOTAL
            cell.border = t.BORDER_TOTAL
        last_row = r
    else:
        last_row = last_data

    heat = [j for j, col in enumerate(cols, 1) if col.emphasis == "heat"]
    for j, (col, width) in enumerate(zip(cols, widths), 1):
        ws.column_dimensions[get_column_letter(j)].width = width
        if not dense and col.emphasis not in ("none", "heat"):
            _emphasis(ws, j, col.emphasis, first_data, last_data)
    if heat and not dense:
        _heat_block(ws, min(heat), max(heat), first_data, last_data)

    if section.rows:
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(cols))}{last_data}"
    anchor = "B" if section.freeze_label_col and len(cols) > 1 else "A"
    ws.freeze_panes = f"{anchor}{first_data}"
    ws.print_title_rows = f"{header_row}:{header_row}"
    return last_row


def _notes(ws: Worksheet, section: Section, start: int) -> None:
    for i, note in enumerate(section.notes):
        cell = ws.cell(start + i, 1, note)
        cell.font = t.F_CONTEXT
        cell.alignment = t.AL_LEFT


def _finish(ws: Worksheet, group: str) -> None:
    """Gridlines off is the single biggest lever on whether a sheet looks
    designed or dumped — the hairlines carry the structure instead."""
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = t.TAB_COLOR.get(group, t.ACCENT)
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.oddFooter.center.text = "&A  ·  Page &P of &N"
    ws.page_margins.left = ws.page_margins.right = 0.4
    ws.page_margins.top = ws.page_margins.bottom = 0.5


# ── Sheets ────────────────────────────────────────────────────────────────────

def _section_sheet(wb: Workbook, section: Section, name: str) -> None:
    ws = wb.create_sheet(name)
    _head_block(ws, section, link_back=True)

    row = t.ROW_SPACER + 1
    if section.kpis:
        row = _kpi_block(ws, section, row)
        if section.kpis and not section.columns:
            ws.column_dimensions["A"].width = max(
                24, min(40, max(len(k.label) for k in section.kpis) + 3)
            )
            ws.column_dimensions["B"].width = 16
            ws.column_dimensions["C"].width = 40

    if section.columns:
        last = _table(ws, section, row)
        if section.kpis:
            # The KPI labels sit in column A above the table, so column A must
            # clear the longer of the two.
            letter = ws.column_dimensions["A"]
            letter.width = max(letter.width or 0, min(40, max(len(k.label) for k in section.kpis) + 3))
        if section.notes:
            _notes(ws, section, last + 2)
    elif section.notes:
        _notes(ws, section, row)

    _finish(ws, section.group)


def _cover_sheet(wb: Workbook, report: Report, names: list[tuple[str, str]]) -> None:
    ws = wb.create_sheet(_CONTENTS, 0)
    ws.cell(1, 1, report.title).font = t.F_COVER_TITLE
    ws.cell(2, 1, report.subtitle).font = t.F_SUBTITLE
    ws.cell(3, 1, f"Generated {report.generated_at.strftime('%d %b %Y, %H:%M')}").font = t.F_CONTEXT
    ws.row_dimensions[1].height = 30

    row = 5
    for item in report.meta:
        ws.cell(row, 1, item.label).font = t.F_LABEL
        cell = ws.cell(row, 2, item.value if item.value is not None else t.DASH)
        cell.font = t.F_BODY
        # Left-align regardless of type: Excel right-aligns numbers by default,
        # which leaves "5,090" floating away from the text values above it.
        cell.alignment = t.AL_LEFT
        if isinstance(item.value, (int, float)) and not isinstance(item.value, bool):
            cell.number_format = t.FMT_COUNT
        if item.note:
            ws.cell(row, 3, item.note).font = t.F_MUTED
        row += 1

    row += 1
    ws.cell(row, 1, "What's in this workbook").font = t.F_TITLE
    row += 1
    for header, col in (("Sheet", 1), ("What it shows", 2)):
        cell = ws.cell(row, col, header)
        cell.font = t.F_HEADER
        cell.fill = t.FILL_HEADER
        cell.border = t.BORDER_HEADER
        cell.alignment = t.AL_LEFT
    ws.cell(row, 3).fill = t.FILL_HEADER
    ws.cell(row, 3).border = t.BORDER_HEADER
    row += 1

    for name, description in names:
        cell = ws.cell(row, 1, name)
        cell.font = t.F_LINK
        cell.hyperlink = Hyperlink(ref=cell.coordinate, location=f"'{name}'!A1")
        cell.border = t.BORDER_ROW
        desc = ws.cell(row, 2, description or t.DASH)
        desc.font = t.F_BODY
        desc.alignment = t.AL_LEFT
        desc.border = t.BORDER_ROW
        ws.cell(row, 3).border = t.BORDER_ROW
        row += 1

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 54
    ws.column_dimensions["C"].width = 40
    _finish(ws, "cover")


def _glossary_sheet(wb: Workbook, report: Report) -> None:
    ws = wb.create_sheet(_GLOSSARY, 0)
    section = Section(
        key="glossary",
        title=_GLOSSARY,
        description="Every term used in this workbook, in plain English.",
        context="Definitions match the metrics on your Foresight dashboard.",
        group="cover",
    )
    _head_block(ws, section, link_back=True)

    headers = ["Term", "What it means", "How it's calculated", "Watch out for"]
    row = t.ROW_SPACER + 1
    for j, header in enumerate(headers, 1):
        cell = ws.cell(row, j, header)
        cell.font = t.F_HEADER
        cell.fill = t.FILL_HEADER
        cell.border = t.BORDER_HEADER
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = t.HEIGHT_HEADER

    for i, term in enumerate(report.glossary):
        r = row + 1 + i
        for j, value in enumerate((term.term, term.meaning, term.formula or t.DASH,
                                   term.caveat or t.DASH), 1):
            cell = ws.cell(r, j, value)
            cell.font = t.F_LABEL if j == 1 else t.F_BODY
            cell.alignment = t.AL_WRAP
            cell.border = t.BORDER_ROW
            if i % 2:
                cell.fill = t.FILL_BAND
        ws.row_dimensions[r].height = 30

    for letter, width in (("A", 26), ("B", 62), ("C", 40), ("D", 54)):
        ws.column_dimensions[letter].width = width
    ws.freeze_panes = f"A{row + 1}"
    ws.print_title_rows = f"{row}:{row}"
    _finish(ws, "cover")


# ── Entry point ───────────────────────────────────────────────────────────────

def write_workbook(report: Report, path: str) -> str:
    """Render `report` to an .xlsx at `path`. Returns `path`."""
    # Every consumer gets the vocabulary check, not just the client report —
    # Explorer's old sheets said "SoV %" and "Reach %", which is exactly the
    # drift a shared writer is supposed to make impossible.
    for section in report.sections:
        glossary.check_wording(section)

    wb = Workbook()
    wb.remove(wb.active)

    used: set[str] = {_CONTENTS.lower(), _GLOSSARY.lower()}
    named = [(s, _sheet_name(s.title, used)) for s in report.sections]

    for section, name in named:
        _section_sheet(wb, section, name)

    # Both are prepended, so build the contents list first and insert in reverse:
    # glossary at 0, then the cover at 0 pushes it to 1. Reading order ends up
    # Contents → How to read this → sections.
    contents = [(name, section.description) for section, name in named]
    if report.glossary:
        _glossary_sheet(wb, report)
        contents.insert(0, (_GLOSSARY, "Every term in this workbook, in plain English."))
    _cover_sheet(wb, report, contents)

    wb.active = 0
    wb.save(path)
    return path
