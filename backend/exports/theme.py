"""The export design system — every visual decision, in one place.

`workbook.py` reads this and nothing else styles a cell. The old analysis scripts
each styled their own sheets, which is exactly why they drifted apart; here a
column declares *what it is* and the table below decides how it looks.

Palette rule: no fill is more saturated than ~10%. It reads as paper with
structure rather than a dashboard, and it survives black-and-white printing.
"""
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ── Palette ───────────────────────────────────────────────────────────────────
INK = "1F2933"          # primary text
MUTED = "6B7280"        # descriptions, context, the "no data" dash
HEADER_BG = "EEF2F6"    # header row — dark text on light, NOT white-on-navy
RULE = "CBD5E1"         # the single medium border under a header
HAIRLINE = "E5E7EB"     # horizontal row separators (never vertical)
BAND = "FAFBFC"         # alternating row stripe
ACCENT = "2F5D8C"       # titles, insight tab colour, data bars
OWN_TINT = "F2F7FC"     # rows that are the client's own product
PAPER = "FFFFFF"

GOOD_BG, GOOD_INK = "E7F3EA", "2F6B45"
WARN_BG, WARN_INK = "FDF3E0", "8A5A18"
BAD_BG, BAD_INK = "FBEBEA", "9B3E38"

# Conditional-formatting end points. Pale on purpose — the scale should hint,
# not shout. The loud three-colour ramp is `heat`, used on one sheet only.
SCALE_GOOD = "C9E2CF"
SCALE_BAD = "F2CFCB"
SCALE_MID = "FBF3D5"
BAR_COLOR = "BFD4E8"

TAB_COLOR = {"cover": INK, "insight": ACCENT}

# ── Type ──────────────────────────────────────────────────────────────────────
FONT = "Calibri"        # present on every Windows/Mac Excel; no fallback surprises

F_TITLE = Font(name=FONT, size=13, bold=True, color=INK)
F_SUBTITLE = Font(name=FONT, size=10, color=MUTED)
F_CONTEXT = Font(name=FONT, size=9, color=MUTED)
F_HEADER = Font(name=FONT, size=10, bold=True, color=INK)
F_BODY = Font(name=FONT, size=10, color=INK)
F_MUTED = Font(name=FONT, size=10, color=MUTED)
F_TOTAL = Font(name=FONT, size=10, bold=True, color=INK)
F_LABEL = Font(name=FONT, size=10, bold=True, color=INK)
F_LINK = Font(name=FONT, size=10, color=ACCENT, underline="single")
F_KPI_VALUE = Font(name=FONT, size=16, bold=True, color=INK)
F_KPI_LABEL = Font(name=FONT, size=9, color=MUTED)
F_COVER_TITLE = Font(name=FONT, size=20, bold=True, color=INK)

# ── Fills & borders ───────────────────────────────────────────────────────────
FILL_HEADER = PatternFill("solid", fgColor=HEADER_BG)
FILL_BAND = PatternFill("solid", fgColor=BAND)
FILL_OWN = PatternFill("solid", fgColor=OWN_TINT)
FILL_GOOD = PatternFill("solid", fgColor=GOOD_BG)
FILL_WARN = PatternFill("solid", fgColor=WARN_BG)
FILL_BAD = PatternFill("solid", fgColor=BAD_BG)

# Status chips — a soft fill plus matching ink, never a saturated block.
CHIP = {
    "good": (FILL_GOOD, Font(name=FONT, size=10, bold=True, color=GOOD_INK)),
    "warn": (FILL_WARN, Font(name=FONT, size=10, bold=True, color=WARN_INK)),
    "bad": (FILL_BAD, Font(name=FONT, size=10, bold=True, color=BAD_INK)),
}

BORDER_HEADER = Border(bottom=Side(style="medium", color=RULE))
BORDER_ROW = Border(bottom=Side(style="thin", color=HAIRLINE))
BORDER_TOTAL = Border(top=Side(style="medium", color=RULE))

# ── Alignment ─────────────────────────────────────────────────────────────────
AL_LEFT = Alignment(horizontal="left", vertical="center")
AL_RIGHT = Alignment(horizontal="right", vertical="center")
AL_CENTER = Alignment(horizontal="center", vertical="center")
AL_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)

# ── Number formats ────────────────────────────────────────────────────────────
# Percentages arrive already scaled 0-100 (that is what every read service
# returns), so the format appends the sign rather than Excel's `0.0%`, which
# would silently multiply by 100 and report 8,470%.
FMT_COUNT = "#,##0"
FMT_PCT = '0.0"%"'
FMT_MONEY = "₹#,##0"
FMT_MONEY_FINE = "₹#,##0.00"
FMT_DEC = "0.0"
FMT_DATE = "dd mmm yyyy"

DASH = "—"         # empty renders as an em dash: "no data" is not zero

# ── Column types ──────────────────────────────────────────────────────────────
# min/max clamp the *computed* width. Widths are measured from the 90th
# percentile of the values, not the maximum, so one 90-character product name
# can't blow a column out to the cap and shove every other column off-screen.
COLUMN_TYPES: dict[str, dict] = {
    "text":       {"fmt": None,            "align": AL_LEFT,   "min": 18, "max": 42},
    "id":         {"fmt": None,            "align": AL_LEFT,   "min": 12, "max": 16},
    "count":      {"fmt": FMT_COUNT,       "align": AL_RIGHT,  "min": 10, "max": 16},
    "pct":        {"fmt": FMT_PCT,         "align": AL_RIGHT,  "min": 10, "max": 15},
    "money":      {"fmt": FMT_MONEY,       "align": AL_RIGHT,  "min": 11, "max": 16},
    "money_fine": {"fmt": FMT_MONEY_FINE,  "align": AL_RIGHT,  "min": 11, "max": 16},
    "rating":     {"fmt": FMT_DEC,         "align": AL_RIGHT,  "min": 9,  "max": 12},
    "date":       {"fmt": FMT_DATE,        "align": AL_CENTER, "min": 12, "max": 12},
}

# ── Sheet skeleton ────────────────────────────────────────────────────────────
ROW_TITLE, ROW_DESC, ROW_CONTEXT, ROW_SPACER = 1, 2, 3, 4
HEIGHT = {ROW_TITLE: 22, ROW_DESC: 16, ROW_CONTEXT: 14, ROW_SPACER: 6}

# Headers wrap, and the row height is explicit — so it must clear TWO lines of
# 10pt text or Excel silently clips the second one and a header reads as
# truncated ("Stores with st"). 30 was not enough.
HEIGHT_HEADER = 34
LINE_HEIGHT = 13.5        # one line of 10pt Calibri
WRAP_MAX_LINES = 3        # a wrapping cell never grows a row past three lines
PROSE_DENSITY = 1.15      # characters of prose per Excel width unit (see _wrap_height)

# Measuring every cell of a long sheet to size columns costs more than writing
# it. The first slice is representative enough.
WIDTH_SAMPLE = 2000
