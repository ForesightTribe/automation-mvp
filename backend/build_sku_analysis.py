"""
Client-facing analysis of the targeted own-SKU scrape (`sku_snapshots`).

Unlike the keyword scrape, this searches Dobra's brand name and paginates the
whole catalog, so every own SKU is captured per store regardless of whether a
category keyword surfaces it. It carries the true own price / MRP / discount /
stock-depth per (SKU x store).

Distribution is measured two ways:
  • by SKU  — every distinct platform_product_id (29 today, incl. packs/combos)
  • by FLAVOR (14) — single + multipack + promo variants rolled up to the core
    product, since the catalog has largely migrated singles -> multipacks.

Usage:  python build_sku_analysis.py [--output PATH]
"""

import argparse
import asyncio
from collections import Counter, defaultdict
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv
from sqlalchemy import text

load_dotenv(Path(__file__).parent / ".env")
from app.core.database import AsyncSessionLocal  # noqa: E402

TENANT = "a870fd8d-7373-47ec-ad69-5dd08ce35542"  # Dobra

# ── SKU catalog map:  pid -> (flavor, category, format) ───────────────────────
# 14 core flavors; multipack/promo variants roll up to the same flavor; the 4
# combos are multi-flavor bundles kept out of the flavor set.
CAT_SODA, CAT_CC, CAT_CHIP = "Goli Soda", "Cotton Candy", "Tapioca Chips"
CATALOG = {
    # singles (the 14 core flavors)
    "554784": ("Nimbu Masala Soda", CAT_SODA, "Single"),
    "554785": ("Kokum Jeera Soda", CAT_SODA, "Single"),
    "620124": ("Blueberry Soda", CAT_SODA, "Single"),
    "620055": ("Mango Soda", CAT_SODA, "Single"),
    "618158": ("Grape Soda", CAT_SODA, "Single"),
    "554786": ("Apple Mojito Soda", CAT_SODA, "Single"),
    "693070": ("Rose Apple Soda", CAT_SODA, "Single"),
    "554767": ("Strawberry Cotton Candy", CAT_CC, "Single"),
    "554766": ("Bubblegum Cotton Candy", CAT_CC, "Single"),
    "618146": ("Smoky BBQ Chips", CAT_CHIP, "Single"),
    "554783": ("Spicy Kari Chips", CAT_CHIP, "Single"),
    "554779": ("Tangy Tomato Chips", CAT_CHIP, "Single"),
    "618144": ("Garlic Pickle Chips", CAT_CHIP, "Single"),
    "618143": ("Plain Salted Chips", CAT_CHIP, "Single"),
    # multipacks / promos -> same core flavor
    "771095": ("Nimbu Masala Soda", CAT_SODA, "Pack of 2"),
    "770977": ("Grape Soda", CAT_SODA, "Pack of 2"),
    "771103": ("Kokum Jeera Soda", CAT_SODA, "Pack of 2"),
    "771114": ("Mango Soda", CAT_SODA, "Pack of 2"),
    "771117": ("Blueberry Soda", CAT_SODA, "Pack of 2"),
    "771109": ("Apple Mojito Soda", CAT_SODA, "Pack of 2"),
    "771112": ("Apple Mojito Soda", CAT_SODA, "Pack of 3"),
    "771107": ("Kokum Jeera Soda", CAT_SODA, "Pack of 3"),
    "771115": ("Mango Soda", CAT_SODA, "Pack of 3"),
    "782627": ("Rose Apple Soda", CAT_SODA, "Pack of 3"),
    "787339": ("Nimbu Masala Soda", CAT_SODA, "Promo (B2G1)"),
    # combos (multi-flavor bundles, not a core flavor)
    "775149": ("Nimbu+Kokum+Apple Combo", "Combo", "Combo"),
    "775150": ("Grape+Mango+Blueberry Combo", "Combo", "Combo"),
    "783966": ("Spicy Kari+Garlic Combo", "Combo", "Combo"),
    "783965": ("Plain+Smoky Combo", "Combo", "Combo"),
}
# 14 core flavors in display order
FLAVORS = [
    "Nimbu Masala Soda", "Kokum Jeera Soda", "Blueberry Soda", "Mango Soda",
    "Grape Soda", "Apple Mojito Soda", "Rose Apple Soda",
    "Strawberry Cotton Candy", "Bubblegum Cotton Candy",
    "Smoky BBQ Chips", "Spicy Kari Chips", "Tangy Tomato Chips",
    "Garlic Pickle Chips", "Plain Salted Chips",
]
FLAVOR_CAT = {f: CATALOG[[p for p, v in CATALOG.items() if v[0] == f][0]][1] for f in FLAVORS}
N_FLA = len(FLAVORS)

# ── styling ───────────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(color="1F4E79", bold=True, size=18)
SUB_FONT = Font(color="404040", bold=True, size=12)
LABEL_FONT = Font(bold=True, color="1F4E79")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)


def style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL; cell.font = HEADER_FONT
        cell.alignment = CENTER; cell.border = BORDER


def autofit(ws, headers, rows, cap=48, minw=9):
    for i, h in enumerate(headers, 1):
        w = max(len(str(h)), max((len(str(r[i - 1])) for r in rows), default=0)) + 2
        ws.column_dimensions[get_column_letter(i)].width = max(minw, min(w, cap))


def write_table(ws, headers, rows, num_formats=None):
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    style_header(ws, len(headers))
    for r, row in enumerate(rows, 2):
        for i, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=i, value=val)
            cell.border = BORDER
            if num_formats and (i - 1) in num_formats:
                cell.number_format = num_formats[i - 1]; cell.alignment = CENTER
    ws.freeze_panes = "A2"
    if rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows)+1}"


def f(v):
    return float(v) if v is not None else None


async def main(output):
    async with AsyncSessionLocal() as s:
        async def q(sql):
            return (await s.execute(text(sql), {"t": TENANT})).fetchall()

        meta = (await q("""select count(*), count(distinct merchant_id),
            count(distinct platform_product_id), count(distinct city),
            min(scraped_at), max(scraped_at) from sku_snapshots where tenant_id=:t"""))[0]

        # per (store, sku) — dedupe the scrape's repeat rows
        pair = await q("""
            select merchant_id, platform_product_id,
                   bool_or(in_stock) instk, max(inventory) inv,
                   min(price) price, max(mrp) mrp,
                   avg(discount_pct) disc, avg(rating) rating
            from sku_snapshots where tenant_id=:t
            group by merchant_id, platform_product_id""")

        # store dimension (+ region/state from catalog where matched)
        store_rows = await q("""
            select sn.merchant_id, max(sn.city) city,
                   max(ml.state) state, max(ml.region) region,
                   max(sn.lat) lat, max(sn.lon) lon
            from sku_snapshots sn
            left join marketplace_locations ml
              on ml.merchant_id=sn.merchant_id and ml.mp_slug='blinkit'
            where sn.tenant_id=:t group by sn.merchant_id""")

        prod_name = {r[0]: r[1] for r in await q(
            "select distinct platform_product_id, product_name from sku_snapshots where tenant_id=:t")}

    store_dim = {r[0]: {"city": r[1] or "", "state": r[2] or "", "region": r[3] or "",
                        "lat": f(r[4]), "lon": f(r[5])} for r in store_rows}
    stores = sorted(store_dim)
    n_store = len(stores)

    # ── aggregate per SKU and per flavor ─────────────────────────────────────
    sku_stores = defaultdict(set)          # pid -> {store listed}
    sku_instk = defaultdict(set)           # pid -> {store in-stock}
    sku_price, sku_mrp, sku_disc, sku_inv, sku_rating = (defaultdict(list) for _ in range(5))

    fla_listed = defaultdict(set)          # flavor -> {store listed}
    fla_instk = defaultdict(set)           # flavor -> {store in-stock}
    store_listed = defaultdict(set)        # store -> {flavor listed}
    store_instk = defaultdict(set)         # store -> {flavor in-stock}
    store_skus = defaultdict(set)          # store -> {pid}
    store_oos = defaultdict(set)           # store -> {flavor listed but no variant in stock}
    store_inv = defaultdict(list)
    sf_depth = defaultdict(int)            # (store, flavor) -> best buyable depth across variants

    for mid, pid, instk, inv, price, mrp, disc, rating in pair:
        if pid not in CATALOG:
            continue
        flavor, cat, fmt = CATALOG[pid]
        sku_stores[pid].add(mid)
        if cat != "Combo" and inv is not None:
            sf_depth[(mid, flavor)] = max(sf_depth[(mid, flavor)], inv)
        store_skus[mid].add(pid)
        if price is not None: sku_price[pid].append(f(price))
        if mrp is not None: sku_mrp[pid].append(f(mrp))
        if disc is not None: sku_disc[pid].append(f(disc))
        if inv is not None: sku_inv[pid].append(inv); store_inv[mid].append(inv)
        if rating is not None: sku_rating[pid].append(f(rating))
        if instk:
            sku_instk[pid].add(mid)
        if cat != "Combo":                 # only core flavors feed the 14-flavor view
            fla_listed[flavor].add(mid)
            store_listed[mid].add(flavor)
            if instk:
                fla_instk[flavor].add(mid)
                store_instk[mid].add(flavor)
    for mid in stores:
        store_oos[mid] = store_listed[mid] - store_instk[mid]

    # ══════════════════════════════════════════════════════════════════════════
    wb = openpyxl.Workbook(); wb.remove(wb.active)

    # ── 1. OVERVIEW ───────────────────────────────────────────────────────────
    ws = wb.create_sheet("Overview")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Dobra — Catalog Availability, Stock & Pricing (Blinkit)"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "Targeted own-SKU scrape — full catalog per dark store"
    ws["A2"].font = Font(italic=True, color="808080", size=11)

    avg_listed = sum(len(store_listed[m]) for m in stores) / n_store
    avg_instk = sum(len(store_instk[m]) for m in stores) / n_store
    full_range = sum(1 for m in stores if len(store_listed[m]) == N_FLA)
    gap3 = sum(1 for m in stores if N_FLA - len(store_listed[m]) >= 3)
    oos_pairs = sum(1 for mid, pid, instk, *_ in pair if pid in CATALOG and not instk)
    avg_disc = sum(v for lst in sku_disc.values() for v in lst) / max(
        1, sum(len(v) for v in sku_disc.values()))
    avg_inv = sum(store_inv[m][i] for m in stores for i in range(len(store_inv[m]))) / max(
        1, sum(len(store_inv[m]) for m in stores))
    all_ratings = [v for lst in sku_rating.values() for v in lst]
    avg_rating = sum(all_ratings) / len(all_ratings) if all_ratings else 0

    kpis = [
        ("Scrape date", f"{meta[4]:%d %b %Y}"),
        ("Dark stores covered", f"{n_store:,}"),
        ("Cities", f"{meta[3]:,}"),
        ("Distinct SKUs found", f"{meta[2]} (incl. packs & combos)"),
        ("Core flavors tracked", f"{N_FLA}"),
        ("SKU × store observations", f"{meta[0]:,}"),
        ("Avg flavors listed / store", f"{avg_listed:.1f} of {N_FLA}"),
        ("Avg flavors in-stock / store", f"{avg_instk:.1f} of {N_FLA}"),
        ("Flavor availability rate", f"{100*avg_listed/N_FLA:.1f}%"),
        ("Stores carrying all 14 flavors", f"{full_range:,} ({100*full_range/n_store:.1f}%)"),
        ("Stores missing 3+ flavors", f"{gap3:,} ({100*gap3/n_store:.1f}%)"),
        ("Out-of-stock SKU×store instances", f"{oos_pairs:,}"),
        ("Avg discount off MRP", f"{avg_disc:.1f}%"),
        ("Avg buyable stock depth", f"{avg_inv:.1f} units"),
        ("Avg product rating", f"{avg_rating:.2f} / 5"),
    ]
    r0 = 4
    ws.cell(r0, 1, "Snapshot").font = SUB_FONT
    for i, (k, v) in enumerate(kpis):
        rr = r0 + 1 + i
        a = ws.cell(rr, 1, k); a.font = LABEL_FONT; a.border = BORDER
        b = ws.cell(rr, 2, v); b.border = BORDER; b.font = Font(bold=True)
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 32

    # flavor coverage for insights
    fla_cov = sorted(((fl, len(fla_listed[fl])) for fl in FLAVORS), key=lambda x: x[1])
    worst = fla_cov[0]; best = fla_cov[-1]
    # worst OOS flavor (share of listing stores that are OOS)
    oos_rate = sorted(
        ((fl, (len(fla_listed[fl]) - len(fla_instk[fl])) / max(1, len(fla_listed[fl])))
         for fl in FLAVORS), key=lambda x: -x[1])
    chip_listed = sum(len(fla_listed[fl]) for fl in FLAVORS if FLAVOR_CAT[fl] == CAT_CHIP) / 5
    soda_listed = sum(len(fla_listed[fl]) for fl in FLAVORS if FLAVOR_CAT[fl] == CAT_SODA) / 7

    insights = [
        f"The catalog has largely shifted from single units to MULTIPACKS: e.g. the "
        f"'Nimbu Masala' single is listed in only {len(sku_stores['554784'])} stores, but its "
        f"Pack-of-2 is in {len(sku_stores['771095'])}. Measured by flavor (all pack sizes rolled "
        f"up), Nimbu Masala reaches {len(fla_listed['Nimbu Masala Soda'])} stores — so distribution "
        f"is strong; it is the pack format that changed.",
        f"Across {n_store:,} stores, Dobra lists on average {avg_listed:.1f} of {N_FLA} flavors "
        f"({100*avg_listed/N_FLA:.0f}%); {full_range:,} stores ({100*full_range/n_store:.0f}%) carry "
        f"the complete range and {gap3:,} ({100*gap3/n_store:.0f}%) are missing 3+.",
        f"Widest flavor: '{best[0]}' ({best[1]:,} stores). Narrowest: '{worst[0]}' "
        f"({worst[1]:,} stores). By category, Goli Soda averages {soda_listed:.0f} stores/flavor vs "
        f"Tapioca Chips {chip_listed:.0f} — chips remain the distribution gap.",
        f"Biggest stock-availability risk: '{oos_rate[0][0]}' is out of stock in "
        f"{100*oos_rate[0][1]:.0f}% of the stores that list it, then '{oos_rate[1][0]}' "
        f"({100*oos_rate[1][1]:.0f}%). Avg buyable depth is only {avg_inv:.1f} units — thin.",
        f"Average discount off MRP is {avg_disc:.1f}% and the range rates {avg_rating:.2f}/5.",
        f"See 'Dark Store Coverage' for the per-store missing/out-of-stock flavor list, "
        f"'Flavor Coverage' for the per-flavor view and 'Stock Risk' for OOS hotspots.",
    ]
    ri = r0 + len(kpis) + 3
    ws.cell(ri, 1, "Key insights").font = SUB_FONT
    for j, tx in enumerate(insights):
        c = ws.cell(ri + 1 + j, 1, f"•  {tx}"); c.alignment = WRAP
        ws.merge_cells(start_row=ri + 1 + j, start_column=1, end_row=ri + 1 + j, end_column=7)
        ws.row_dimensions[ri + 1 + j].height = 46
    note = ws.cell(ri + len(insights) + 2, 1,
        "Method: 'listed' = the SKU appears in the store's Dobra catalog; 'in-stock' = buyable now "
        "(inventory > 0). A flavor is available at a store if any of its pack variants (single / "
        "pack-of-2 / pack-of-3 / promo) is present. The 4 multi-flavor combo bundles are excluded "
        "from the 14 flavors and shown in 'SKU Detail'. One targeted scrape, "
        f"{meta[4]:%d %b %Y}.")
    note.alignment = WRAP; note.font = Font(italic=True, color="808080", size=9)
    ws.merge_cells(start_row=ri + len(insights) + 2, start_column=1,
                   end_row=ri + len(insights) + 2, end_column=7)
    ws.row_dimensions[ri + len(insights) + 2].height = 60

    # ── 2. FLAVOR COVERAGE (14) ───────────────────────────────────────────────
    ws = wb.create_sheet("Flavor Coverage")
    headers = ["Flavor", "Category", "Primary Format", "Stores Listed", "Listed %",
               "Stores In-Stock", "In-Stock %", "Stores Missing", "Out-of-Stock %",
               "Avg Price (₹)", "Avg MRP (₹)", "Avg Discount %", "Avg Stock Depth", "Avg Rating"]
    rows = []
    for fl in sorted(FLAVORS, key=lambda x: -len(fla_listed[x])):
        pids = [p for p, v in CATALOG.items() if v[0] == fl]
        # primary format = variant with widest listing
        prim = max(pids, key=lambda p: len(sku_stores[p]))
        prices = [x for p in pids for x in sku_price[p]]
        mrps = [x for p in pids for x in sku_mrp[p]]
        discs = [x for p in pids for x in sku_disc[p]]
        invs = [x for p in pids for x in sku_inv[p]]
        rats = [x for p in pids for x in sku_rating[p]]
        lst = len(fla_listed[fl]); ins = len(fla_instk[fl])
        rows.append([
            fl, FLAVOR_CAT[fl], CATALOG[prim][2], lst, lst / n_store,
            ins, ins / n_store, n_store - lst,
            (lst - ins) / lst if lst else 0,
            round(sum(prices) / len(prices)) if prices else None,
            round(sum(mrps) / len(mrps)) if mrps else None,
            round(sum(discs) / len(discs), 1) if discs else None,
            round(sum(invs) / len(invs), 1) if invs else None,
            round(sum(rats) / len(rats), 2) if rats else None,
        ])
    write_table(ws, headers, rows, num_formats={
        4: "0.0%", 6: "0.0%", 8: "0.0%", 9: "₹#,##0", 10: "₹#,##0", 11: "0.0"})
    ws.conditional_formatting.add(f"E2:E{len(rows)+1}", ColorScaleRule(
        start_type="num", start_value=0, start_color="F8696B",
        mid_type="num", mid_value=0.85, mid_color="FFEB84",
        end_type="num", end_value=1, end_color="63BE7B"))
    ws.conditional_formatting.add(f"I2:I{len(rows)+1}", ColorScaleRule(
        start_type="num", start_value=0, start_color="63BE7B",
        mid_type="percentile", mid_value=50, mid_color="FFEB84",
        end_type="num", end_value=0.2, end_color="F8696B"))
    autofit(ws, headers, [[str(x) for x in r] for r in rows])

    # ── 3. SKU DETAIL (all 29) ────────────────────────────────────────────────
    ws = wb.create_sheet("SKU Detail")
    headers = ["Product", "Flavor / Bundle", "Category", "Format", "Product ID",
               "Stores Listed", "Listed %", "In-Stock %", "Avg Price (₹)",
               "Avg MRP (₹)", "Avg Discount %", "Avg Stock Depth", "Avg Rating"]
    catrank = {CAT_SODA: 0, CAT_CC: 1, CAT_CHIP: 2, "Combo": 3}
    rows = []
    for pid in sorted(CATALOG, key=lambda p: (catrank[CATALOG[p][1]], -len(sku_stores[p]))):
        fl, cat, fmt = CATALOG[pid]
        lst = len(sku_stores[pid]); ins = len(sku_instk[pid])
        rows.append([
            prod_name.get(pid, fl), fl, cat, fmt, pid, lst, lst / n_store,
            ins / lst if lst else 0,
            round(sum(sku_price[pid]) / len(sku_price[pid])) if sku_price[pid] else None,
            round(sum(sku_mrp[pid]) / len(sku_mrp[pid])) if sku_mrp[pid] else None,
            round(sum(sku_disc[pid]) / len(sku_disc[pid]), 1) if sku_disc[pid] else None,
            round(sum(sku_inv[pid]) / len(sku_inv[pid]), 1) if sku_inv[pid] else None,
            round(sum(sku_rating[pid]) / len(sku_rating[pid]), 2) if sku_rating[pid] else None,
        ])
    write_table(ws, headers, rows, num_formats={
        6: "0.0%", 7: "0.0%", 8: "₹#,##0", 9: "₹#,##0", 10: "0.0", 4: "@"})
    autofit(ws, headers, [[str(x) for x in r] for r in rows])
    ws.column_dimensions["A"].width = 46

    # ── 4. DARK STORE COVERAGE ────────────────────────────────────────────────
    ws = wb.create_sheet("Dark Store Coverage")
    headers = ["Merchant ID", "City", "State", "Region", "Flavors Listed",
               "Flavors In-Stock", "Missing", "Availability %", "Distinct SKUs",
               "Avg Stock Depth", "Missing Flavors", "Out-of-Stock Flavors"]
    rows = []
    for m in sorted(stores, key=lambda x: (-(N_FLA - len(store_listed[x])),
                    store_dim[x]["city"])):
        d = store_dim[m]
        lst = store_listed[m]; ins = store_instk[m]
        miss = [fl for fl in FLAVORS if fl not in lst]
        oos = [fl for fl in FLAVORS if fl in store_oos[m]]
        inv = round(sum(store_inv[m]) / len(store_inv[m]), 1) if store_inv[m] else None
        rows.append([m, d["city"].title(), d["state"], d["region"], len(lst), len(ins),
                     N_FLA - len(lst), len(lst) / N_FLA, len(store_skus[m]), inv,
                     ", ".join(miss), ", ".join(oos)])
    write_table(ws, headers, rows, num_formats={7: "0.0%", 0: "@"})
    ws.conditional_formatting.add(f"G2:G{len(rows)+1}", ColorScaleRule(
        start_type="num", start_value=0, start_color="63BE7B",
        mid_type="num", mid_value=3, mid_color="FFEB84",
        end_type="num", end_value=N_FLA, end_color="F8696B"))
    for col, w in {"A": 13, "B": 16, "C": 14, "D": 14, "K": 55, "L": 55}.items():
        ws.column_dimensions[col].width = w

    # ── 5. STORE × FLAVOR MATRIX ──────────────────────────────────────────────
    ws = wb.create_sheet("Store x Flavor Matrix")
    order = sorted(FLAVORS, key=lambda x: -len(fla_listed[x]))
    headers = ["Merchant ID", "City"] + order + ["Missing"]
    rows = []
    for m in sorted(stores, key=lambda x: (store_dim[x]["city"], -(N_FLA - len(store_listed[x])))):
        cells = []
        for fl in order:
            if fl in store_instk[m]:
                cells.append("Yes")
            elif fl in store_listed[m]:
                cells.append("OOS")
            else:
                cells.append("—")
        rows.append([m, store_dim[m]["city"].title()] + cells + [N_FLA - len(store_listed[m])])
    write_table(ws, headers, rows, num_formats={0: "@"})
    for c in range(3, 3 + len(order)):
        ws.cell(1, c).alignment = Alignment(textRotation=90, horizontal="center", vertical="bottom")
        ws.column_dimensions[get_column_letter(c)].width = 5
    ws.row_dimensions[1].height = 130
    ws.column_dimensions["A"].width = 13; ws.column_dimensions["B"].width = 16

    # ── 6. CITY COVERAGE ──────────────────────────────────────────────────────
    ws = wb.create_sheet("City Coverage")
    headers = ["City", "State", "Stores", "Avg Flavors Listed", "Avg In-Stock",
               "Availability %", "Stores w/ Full Range", "Stores Missing 3+"]
    bycity = defaultdict(list)
    for m in stores:
        bycity[store_dim[m]["city"]].append(m)
    rows = []
    for city, ms in bycity.items():
        n = len(ms)
        al = sum(len(store_listed[m]) for m in ms) / n
        ai = sum(len(store_instk[m]) for m in ms) / n
        full = sum(1 for m in ms if len(store_listed[m]) == N_FLA)
        g3 = sum(1 for m in ms if N_FLA - len(store_listed[m]) >= 3)
        st = next((store_dim[m]["state"] for m in ms if store_dim[m]["state"]), "")
        rows.append([city.title(), st, n, round(al, 1), round(ai, 1), al / N_FLA, full, g3])
    rows.sort(key=lambda r: (-r[2], -r[5]))
    write_table(ws, headers, rows, num_formats={5: "0.0%"})
    ws.conditional_formatting.add(f"F2:F{len(rows)+1}", ColorScaleRule(
        start_type="num", start_value=0, start_color="F8696B",
        mid_type="num", mid_value=0.85, mid_color="FFEB84",
        end_type="num", end_value=1, end_color="63BE7B"))
    autofit(ws, headers, [[str(x) for x in r] for r in rows])

    # ── 7. STOCK RISK ─────────────────────────────────────────────────────────
    ws = wb.create_sheet("Stock Risk")
    headers = ["Flavor", "Category", "Stores Listed", "Stores Out-of-Stock",
               "Out-of-Stock %", "Stores Low (1-2 units)", "Low-Stock %", "Avg Stock Depth"]
    rows = []
    for fl in FLAVORS:
        pids = [p for p, v in CATALOG.items() if v[0] == fl]
        lst = len(fla_listed[fl])
        oos = lst - len(fla_instk[fl])
        # low = in-stock (depth>=1) but best variant depth <= 2 units
        low = sum(1 for m in fla_instk[fl] if 1 <= sf_depth[(m, fl)] <= 2)
        invs = [x for p in pids for x in sku_inv[p]]
        rows.append([fl, FLAVOR_CAT[fl], lst, oos, oos / lst if lst else 0,
                     low, low / lst if lst else 0,
                     round(sum(invs) / len(invs), 1) if invs else None])
    rows.sort(key=lambda r: -r[4])
    write_table(ws, headers, rows, num_formats={4: "0.0%", 6: "0.0%"})
    ws.conditional_formatting.add(f"E2:E{len(rows)+1}", ColorScaleRule(
        start_type="num", start_value=0, start_color="63BE7B",
        mid_type="percentile", mid_value=50, mid_color="FFEB84",
        end_type="num", end_value=0.2, end_color="F8696B"))
    autofit(ws, headers, [[str(x) for x in r] for r in rows])

    wb.save(output)
    print(f"Saved {output}  ({len(wb.sheetnames)} sheets, {n_store:,} stores, {meta[2]} SKUs)")


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--output", "-o",
                    default=f"Dobra_SKU_Scrape_Analysis_{date.today():%Y%m%d}.xlsx")
    args = ap.parse_args()
    asyncio.run(main(args.output))
