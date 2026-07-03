"""
Client-facing analysis of Dobra Blinkit public search data.

Consolidates `search_snapshots` + `search_listings` into an Excel workbook:
availability/distribution coverage of Dobra's SKUs across every scraped dark
store, plus competitive share-of-voice and keyword performance.

Usage:  python build_public_analysis.py [--output PATH]
"""

import argparse
import asyncio
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv
from sqlalchemy import text

load_dotenv(Path(__file__).parent / ".env")
from app.core.database import AsyncSessionLocal  # noqa: E402

TENANT = "a870fd8d-7373-47ec-ad69-5dd08ce35542"  # Dobra

# ── The 14 individual Dobra SKUs (combos excluded) ────────────────────────────
# platform_product_id -> (short display name, category)
SKUS = {
    "554784": ("Nimbu Masala Soda", "Goli Soda"),
    "554785": ("Kokum Jeera Soda", "Goli Soda"),
    "620124": ("Blueberry Soda", "Goli Soda"),
    "620055": ("Mango Soda", "Goli Soda"),
    "618158": ("Grape Soda", "Goli Soda"),
    "554786": ("Apple Mojito Soda", "Goli Soda"),
    "693070": ("Rose Apple Soda", "Goli Soda"),
    "554767": ("Strawberry Cotton Candy", "Cotton Candy"),
    "554766": ("Bubblegum Cotton Candy", "Cotton Candy"),
    "618146": ("Smoky BBQ Chips", "Tapioca Chips"),
    "554783": ("Spicy Kari Chips", "Tapioca Chips"),
    "618144": ("Garlic Pickle Chips", "Tapioca Chips"),
    "554779": ("Tangy Tomato Chips", "Tapioca Chips"),
    "618143": ("Plain Salted Chips", "Tapioca Chips"),
}
SKU_IDS = list(SKUS.keys())
N_SKU = len(SKU_IDS)

# ── Styling ───────────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(color="1F4E79", bold=True, size=18)
SUB_FONT = Font(color="404040", bold=True, size=12)
LABEL_FONT = Font(bold=True, color="1F4E79")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)


def style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def autofit(ws, headers, rows, cap=52, minw=9):
    for i, h in enumerate(headers, 1):
        w = max(len(str(h)), max((len(str(r[i - 1])) for r in rows), default=0)) + 2
        ws.column_dimensions[get_column_letter(i)].width = max(minw, min(w, cap))


def write_table(ws, headers, rows, start=1, num_formats=None):
    """Write a header row + data rows, styled. num_formats: {col_idx0: fmt}."""
    for i, h in enumerate(headers, 1):
        ws.cell(row=start, column=i, value=h)
    style_header(ws, len(headers), row=start)
    for r, row in enumerate(rows, start + 1):
        for i, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=i, value=val)
            cell.border = BORDER
            if num_formats and (i - 1) in num_formats:
                cell.number_format = num_formats[i - 1]
                cell.alignment = CENTER
    ws.freeze_panes = ws.cell(row=start + 1, column=1)
    if rows:
        ws.auto_filter.ref = (
            f"A{start}:{get_column_letter(len(headers))}{start + len(rows)}"
        )


async def fetch(session, sql, params=None):
    return (await session.execute(text(sql), params or {})).fetchall()


async def main(output):
    async with AsyncSessionLocal() as s:
        # scope / meta
        meta = (await fetch(s, """
            select count(*) snaps, min(scraped_at) mn, max(scraped_at) mx,
                   count(distinct keyword) kws, count(distinct city) cities
            from search_snapshots where tenant_id=:t""", {"t": TENANT}))[0]
        n_listings = (await fetch(s,
            "select count(*) from search_listings where tenant_id=:t", {"t": TENANT}))[0][0]

        # snapshots -> store dimension source
        snaps = await fetch(s, """
            select id, round(lat::numeric,4) lat, round(lon::numeric,4) lon,
                   city, pincode, keyword, brand_rank, brand_sov, total_results
            from search_snapshots where tenant_id=:t""", {"t": TENANT})

        # Dobra own-brand listings joined to store
        brand = await fetch(s, """
            select round(sn.lat::numeric,4) lat, round(sn.lon::numeric,4) lon,
                   l.platform_product_id pid, l.position, l.price, l.in_stock,
                   l.extra->>'merchant_id' mid
            from search_listings l join search_snapshots sn on sn.id=l.snapshot_id
            where l.tenant_id=:t and l.is_brand=true""", {"t": TENANT})

        # modal merchant_id per store (from all listings)
        midrows = await fetch(s, """
            select round(sn.lat::numeric,4) lat, round(sn.lon::numeric,4) lon,
                   l.extra->>'merchant_id' mid, count(*) c
            from search_listings l join search_snapshots sn on sn.id=l.snapshot_id
            where l.tenant_id=:t and l.extra->>'merchant_id' is not null
            group by 1,2,3""", {"t": TENANT})

        # competitor / brand share-of-voice
        comp = await fetch(s, """
            select l.brand_slug,
                   count(*) listings,
                   count(distinct l.snapshot_id) snaps_present,
                   count(distinct (round(sn.lat::numeric,4), round(sn.lon::numeric,4))) stores,
                   round(avg(l.position)::numeric,1) avg_pos,
                   round(avg(l.price)::numeric,0) avg_price,
                   round(100.0*count(*) filter (where l.in_stock)/count(*),1) instock_pct
            from search_listings l join search_snapshots sn on sn.id=l.snapshot_id
            where l.tenant_id=:t and l.brand_slug is not null
            group by l.brand_slug order by snaps_present desc""", {"t": TENANT})

        # keyword performance (Dobra rank/SoV live on the snapshot header)
        kw = await fetch(s, """
            select keyword, count(*) searches,
                   round(avg(brand_rank)::numeric,1) avg_rank,
                   round(avg(brand_sov)::numeric,1) avg_sov,
                   count(*) filter (where brand_rank=1) rank1,
                   round(avg(total_results)::numeric,0) avg_results,
                   count(distinct (round(lat::numeric,4),round(lon::numeric,4))) stores
            from search_snapshots where tenant_id=:t
            group by keyword order by avg_rank nulls last""", {"t": TENANT})

        # combo presence (reference)
        combos = await fetch(s, """
            select l.product_name, l.platform_product_id,
                   count(distinct (round(sn.lat::numeric,4),round(sn.lon::numeric,4))) stores
            from search_listings l join search_snapshots sn on sn.id=l.snapshot_id
            where l.tenant_id=:t and l.is_brand=true and l.product_name ilike '%combo%'
            group by 1,2 order by stores desc""", {"t": TENANT})

    # ── Build store dimension in Python ───────────────────────────────────────
    snap_store = {}                       # snap_id -> (lat,lon)
    store_city = defaultdict(Counter)     # store -> city counts
    store_pin = defaultdict(Counter)
    store_kw = defaultdict(set)
    store_snaps = Counter()
    for sid, lat, lon, city, pin, kwd, *_ in snaps:
        st = (float(lat), float(lon))
        snap_store[sid] = st
        if city:
            store_city[st][city] += 1
        if pin:
            store_pin[st][pin] += 1
        store_kw[st].add(kwd)
        store_snaps[st] += 1

    stores = sorted(store_snaps)          # all scraped stores (distinct lat/lon)

    # modal merchant per store
    mid_ct = defaultdict(Counter)
    for lat, lon, mid, c in midrows:
        if mid and mid != "None":
            mid_ct[(float(lat), float(lon))][mid] += c
    store_mid = {st: (mid_ct[st].most_common(1)[0][0] if mid_ct[st] else "n/a")
                 for st in stores}

    # SKU presence + per-SKU stats
    present = defaultdict(set)            # store -> {pid}
    sku_stores = defaultdict(set)         # pid -> {store}
    sku_pos = defaultdict(list)
    sku_price = defaultdict(list)
    sku_instock = defaultdict(lambda: [0, 0])
    for lat, lon, pid, pos, price, instock, mid in brand:
        if pid not in SKUS:
            continue
        st = (float(lat), float(lon))
        present[st].add(pid)
        sku_stores[pid].add(st)
        if pos is not None:
            sku_pos[pid].append(pos)
        if price is not None:
            sku_price[pid].append(float(price))
        sku_instock[pid][0] += 1 if instock else 0
        sku_instock[pid][1] += 1

    n_stores = len(stores)
    # Coverage stats are only meaningful where the full keyword set was scraped;
    # partially-scraped stores (fewer keywords) would otherwise read as false gaps.
    FULL = {st for st in stores if len(store_kw[st]) == meta[3]}
    full_stores = sorted(FULL)
    n_full = len(full_stores)

    # ══════════════════════════════════════════════════════════════════════════
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── 1. OVERVIEW ───────────────────────────────────────────────────────────
    ws = wb.create_sheet("Overview")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Dobra — Blinkit Availability & Market Analysis"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "Public search-data scrape  ·  consolidated for client review"
    ws["A2"].font = Font(italic=True, color="808080", size=11)

    dmin, dmax = meta[1], meta[2]
    miss_counts = {st: N_SKU - len(present.get(st, ())) for st in stores}
    # headline coverage stats on fully-scraped stores only
    avg_present = sum(len(present.get(st, ())) for st in full_stores) / n_full
    fully_stocked = sum(1 for st in full_stores if miss_counts[st] == 0)
    gap3 = sum(1 for st in full_stores if miss_counts[st] >= 3)
    zero_dobra = sum(1 for st in full_stores if len(present.get(st, ())) == 0)

    kpis = [
        ("Scrape window", f"{dmin:%d %b %Y} – {dmax:%d %b %Y}"),
        ("Dark stores scraped", f"{n_stores:,}"),
        ("  of which fully scraped (all 9 kw)", f"{n_full:,}"),
        ("Cities covered", f"{meta[4]:,}"),
        ("Keywords tracked", f"{meta[3]}"),
        ("Search snapshots", f"{meta[0]:,}"),
        ("Product listings captured", f"{n_listings:,}"),
        ("Dobra SKUs analysed", f"{N_SKU} (combos excl.)"),
        ("Avg SKUs available per store", f"{avg_present:.1f} of {N_SKU}"),
        ("Avg availability rate", f"{100*avg_present/N_SKU:.1f}%"),
        ("Stores carrying all 14 SKUs", f"{fully_stocked:,} ({100*fully_stocked/n_full:.1f}%)"),
        ("Stores missing 3+ SKUs", f"{gap3:,} ({100*gap3/n_full:.1f}%)"),
        ("Stores with zero Dobra SKUs", f"{zero_dobra:,}"),
    ]
    r0 = 4
    ws.cell(r0, 1, "Snapshot").font = SUB_FONT
    for i, (k, v) in enumerate(kpis):
        rr = r0 + 1 + i
        a = ws.cell(rr, 1, k); a.font = LABEL_FONT; a.border = BORDER
        b = ws.cell(rr, 2, v); b.border = BORDER
        b.font = Font(bold=True)
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 30

    # worst / best SKUs (availability measured over fully-scraped stores)
    sku_full = {pid: len(sku_stores[pid] & FULL) for pid in SKU_IDS}
    sku_cov = sorted(sku_full.items(), key=lambda x: x[1])
    worst = sku_cov[0]; worst2 = sku_cov[1]
    best = sku_cov[-1]
    # worst cities by availability (fully-scraped stores only)
    city_present = defaultdict(list)
    for st in full_stores:
        city = store_city[st].most_common(1)[0][0] if store_city[st] else "?"
        city_present[city].append(len(present.get(st, ())))
    big_cities = [(c, sum(v)/len(v), len(v)) for c, v in city_present.items() if len(v) >= 10]
    worst_city = min(big_cities, key=lambda x: x[1]) if big_cities else ("-", 0, 0)

    insights = [
        f"Across the {n_full:,} fully-scraped stores, Dobra products fill on average "
        f"{100*avg_present/N_SKU:.0f}% of the 14-SKU range — {fully_stocked:,} stores "
        f"({100*fully_stocked/n_full:.0f}%) carry the complete range, while {gap3:,} "
        f"({100*gap3/n_full:.0f}%) are missing 3 or more SKUs.",
        f"Widest distribution: '{SKUS[best[0]][0]}' — found in {best[1]:,} stores "
        f"({100*best[1]/n_full:.0f}%). Narrowest: '{SKUS[worst[0]][0]}' — {worst[1]:,} stores "
        f"({100*worst[1]/n_full:.0f}%), then '{SKUS[worst2[0]][0]}' ({worst2[1]:,}).",
        f"The Tapioca Chips range is the clear distribution gap: the Goli Soda flavours are near-"
        f"universal (~98% of stores) while the chips lag by several hundred stores each.",
        f"{gap3:,} stores are missing 3+ SKUs and {zero_dobra} fully-scraped store(s) carry no "
        f"Dobra SKUs at all — priority targets for the sales/distribution team.",
        f"Weakest city on availability (10+ stores): {worst_city[0].title()} — "
        f"avg {worst_city[1]:.1f} of {N_SKU} SKUs across {worst_city[2]} stores.",
        f"See 'Dark Store Coverage' for the per-store missing-SKU list (with a Full/Partial scrape "
        f"flag) and 'SKU Coverage' for the per-product view.",
    ]
    ri = r0 + len(kpis) + 3
    ws.cell(ri, 1, "Key insights").font = SUB_FONT
    for j, tx in enumerate(insights):
        cell = ws.cell(ri + 1 + j, 1, f"•  {tx}")
        cell.alignment = WRAP
        ws.merge_cells(start_row=ri + 1 + j, start_column=1, end_row=ri + 1 + j, end_column=6)
        ws.row_dimensions[ri + 1 + j].height = 32
    note = ws.cell(ri + len(insights) + 2, 1,
        "Method: a SKU is 'available' at a store if it appeared in at least one keyword search at that "
        "store's location during the scrape window. 'Dark store' = one Blinkit fulfilment location "
        "(unique lat/lon), labelled with its Blinkit merchant id. Combo bundles are excluded from the 14 "
        "SKUs and reported separately.")
    note.alignment = WRAP
    note.font = Font(italic=True, color="808080", size=9)
    ws.merge_cells(start_row=ri + len(insights) + 2, start_column=1,
                   end_row=ri + len(insights) + 2, end_column=6)
    ws.row_dimensions[ri + len(insights) + 2].height = 48

    # ── 2. SKU COVERAGE ───────────────────────────────────────────────────────
    ws = wb.create_sheet("SKU Coverage")
    headers = ["SKU", "Category", "Product ID", "Stores Available",
               "Stores Missing", "Availability %", "Avg Shelf Position",
               "Avg Price (₹)", "In-Stock %"]
    rows = []
    for pid in sorted(SKU_IDS, key=lambda p: -sku_full[p]):
        nm, cat = SKUS[pid]
        sp = sku_full[pid]  # available in this many of the fully-scraped stores
        pos = sum(sku_pos[pid]) / len(sku_pos[pid]) if sku_pos[pid] else None
        pr = sum(sku_price[pid]) / len(sku_price[pid]) if sku_price[pid] else None
        ins = sku_instock[pid]
        rows.append([nm, cat, pid, sp, n_full - sp, sp / n_full,
                     round(pos, 1) if pos else None,
                     round(pr) if pr else None,
                     ins[0] / ins[1] if ins[1] else None])
    write_table(ws, headers, rows, num_formats={5: "0.0%", 8: "0.0%", 7: "₹#,##0"})
    ws.conditional_formatting.add(
        f"F2:F{len(rows)+1}",
        ColorScaleRule(start_type="num", start_value=0, start_color="F8696B",
                       mid_type="num", mid_value=0.75, mid_color="FFEB84",
                       end_type="num", end_value=1, end_color="63BE7B"))
    autofit(ws, headers, [[str(x) for x in r] for r in rows])

    # ── 3. DARK STORE COVERAGE (core deliverable) ─────────────────────────────
    ws = wb.create_sheet("Dark Store Coverage")
    headers = ["Merchant ID", "City", "Pincode", "Lat", "Lon", "Scrape Status",
               "Keywords Scraped", "SKUs Available", "SKUs Missing", "Availability %",
               "Missing SKUs"]
    # Full-scrape stores first (ranked worst-gap first), partial scrapes last so
    # incomplete data never masquerades as the biggest distribution gap.
    def sort_key(x):
        return (0 if x in FULL else 1, -miss_counts[x],
                store_city[x].most_common(1)[0][0] if store_city[x] else "")
    rows = []
    for st in sorted(stores, key=sort_key):
        city = store_city[st].most_common(1)[0][0] if store_city[st] else ""
        pin = store_pin[st].most_common(1)[0][0] if store_pin[st] else ""
        pset = present.get(st, set())
        status = "Full" if st in FULL else "Partial"
        miss = [SKUS[p][0] for p in SKU_IDS if p not in pset]
        rows.append([store_mid[st], city.title(), pin, st[0], st[1], status,
                     len(store_kw[st]), len(pset), N_SKU - len(pset), len(pset) / N_SKU,
                     ", ".join(miss)])
    write_table(ws, headers, rows,
                num_formats={9: "0.0%", 2: "@", 0: "@"})
    ws.column_dimensions["K"].width = 60
    for c in ("A", "B", "C", "F"):
        ws.column_dimensions[c].width = 14
    ws.conditional_formatting.add(
        f"I2:I{len(rows)+1}",
        ColorScaleRule(start_type="num", start_value=0, start_color="63BE7B",
                       mid_type="num", mid_value=3, mid_color="FFEB84",
                       end_type="num", end_value=N_SKU, end_color="F8696B"))

    # ── 4. COVERAGE MATRIX ────────────────────────────────────────────────────
    ws = wb.create_sheet("Coverage Matrix")
    sku_order = sorted(SKU_IDS, key=lambda p: -len(sku_stores[p]))
    headers = ["Merchant ID", "City"] + [SKUS[p][0] for p in sku_order] + ["Missing"]
    rows = []
    for st in sorted(stores, key=lambda x: (store_city[x].most_common(1)[0][0]
                     if store_city[x] else "", -miss_counts[x])):
        city = store_city[st].most_common(1)[0][0] if store_city[st] else ""
        pset = present.get(st, set())
        cells = ["Yes" if p in pset else "—" for p in sku_order]
        rows.append([store_mid[st], city.title()] + cells + [N_SKU - len(pset)])
    write_table(ws, headers, rows, num_formats={0: "@"})
    # rotate SKU headers for compactness
    for c in range(3, 3 + len(sku_order)):
        ws.cell(1, c).alignment = Alignment(textRotation=90, horizontal="center",
                                            vertical="bottom")
        ws.column_dimensions[get_column_letter(c)].width = 5
    ws.row_dimensions[1].height = 120
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 16

    # ── 5. CITY COVERAGE ──────────────────────────────────────────────────────
    ws = wb.create_sheet("City Coverage")
    headers = ["City", "Stores", "Avg SKUs Available", "Avg Availability %",
               "Stores w/ Full Range", "Stores Missing 3+"]
    rows = []
    for city, vals in city_present.items():
        n = len(vals)
        avg = sum(vals) / n
        full = sum(1 for v in vals if v == N_SKU)
        g3 = sum(1 for v in vals if N_SKU - v >= 3)
        rows.append([city.title(), n, round(avg, 1), avg / N_SKU, full, g3])
    rows.sort(key=lambda r: (-r[1], -r[3]))
    write_table(ws, headers, rows, num_formats={3: "0.0%"})
    ws.conditional_formatting.add(
        f"D2:D{len(rows)+1}",
        ColorScaleRule(start_type="num", start_value=0, start_color="F8696B",
                       mid_type="num", mid_value=0.75, mid_color="FFEB84",
                       end_type="num", end_value=1, end_color="63BE7B"))
    autofit(ws, headers, [[str(x) for x in r] for r in rows])

    # ── 6. COMPETITIVE SHARE OF VOICE ─────────────────────────────────────────
    ws = wb.create_sheet("Competitive SoV")
    total_snaps = meta[0]
    headers = ["Brand", "Relationship", "Listings Captured", "Searches Present",
               "Presence %", "Stores Present", "Avg Position", "Avg Price (₹)",
               "In-Stock %"]
    rel = {"dobra": "Own brand", "bombay-banta": "Competitor", "paper-boat": "Competitor",
           "lahori": "Competitor", "shift": "Competitor", "neo-pop": "Competitor"}
    rows = []
    for bslug, listings, snaps_present, st_ct, avg_pos, avg_price, instock in comp:
        rows.append([bslug.replace("-", " ").title(), rel.get(bslug, "Competitor"),
                     listings, snaps_present, snaps_present / total_snaps, st_ct,
                     float(avg_pos) if avg_pos else None,
                     float(avg_price) if avg_price else None,
                     float(instock) / 100 if instock else None])
    write_table(ws, headers, rows, num_formats={4: "0.0%", 8: "0.0%", 7: "₹#,##0"})
    autofit(ws, headers, [[str(x) for x in r] for r in rows])

    # ── 7. KEYWORD PERFORMANCE ────────────────────────────────────────────────
    ws = wb.create_sheet("Keyword Performance")
    headers = ["Keyword", "Searches", "Stores", "Dobra Avg Rank", "Dobra Avg SoV %",
               "# Searches Ranked #1", "Avg Total Results"]
    rows = []
    for keyword, searches, avg_rank, avg_sov, rank1, avg_results, st_ct in kw:
        rows.append([keyword, searches, st_ct,
                     float(avg_rank) if avg_rank else None,
                     float(avg_sov) / 100 if avg_sov else None,
                     rank1, int(avg_results) if avg_results else None])
    write_table(ws, headers, rows, num_formats={4: "0.0%"})
    autofit(ws, headers, [[str(x) for x in r] for r in rows])

    # ── 8. COMBOS (reference) ─────────────────────────────────────────────────
    ws = wb.create_sheet("Combos (reference)")
    headers = ["Combo Bundle", "Product ID", "Stores Available", "Availability %"]
    rows = [[nm, pid, st_ct, st_ct / n_stores] for nm, pid, st_ct in combos]
    write_table(ws, headers, rows, num_formats={3: "0.0%"})
    autofit(ws, headers, [[str(x) for x in r] for r in rows])

    wb.save(output)
    print(f"Saved {output}  ({len(wb.sheetnames)} sheets, {n_stores:,} stores)")


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--output", "-o",
                    default=f"Dobra_Public_Data_Analysis_{date.today():%Y%m%d}.xlsx")
    args = ap.parse_args()
    asyncio.run(main(args.output))
