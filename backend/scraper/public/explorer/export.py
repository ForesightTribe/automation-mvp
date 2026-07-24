"""Explorer Excel export — the multi-sheet workbook.

`write_workbook(insights, result, path)` renders the typed `ExplorerInsights`
(from `build_insights`) into an .xlsx: INSIGHT sheets first (formatted, with
colour scales + a couple of native bar charts), RAW sheets last (full field set).
The writer only formats what `build_insights` computed — no aggregation here.
"""
from collections import OrderedDict

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.schemas.explorer import ExplorerInsights
from scraper.public.explorer.orchestrator import ExplorerResult

_HEADER_FILL = PatternFill("solid", fgColor="1F2937")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_TITLE_FONT = Font(bold=True, size=16)
_LABEL_FONT = Font(bold=True)

# Number formats
_RUP = "₹#,##0"
_RUP2 = "₹#,##0.00"   # per-unit prices — small, need decimals (₹6.25 / 100 ml)
_PCT = "0.0"
_NUM = "#,##0"
_DEC = "0.0"


def _style(ws: Worksheet, columns: list[dict]) -> None:
    for j, c in enumerate(columns, 1):
        cell = ws.cell(row=1, column=j)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        letter = get_column_letter(j)
        width = len(str(c["header"]))
        for row in range(2, ws.max_row + 1):
            v = ws.cell(row=row, column=j).value
            if v is not None:
                width = max(width, len(str(v)))
        ws.column_dimensions[letter].width = min(max(width + 2, 10), 46)
        if c.get("fmt"):
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=j).number_format = c["fmt"]
        if c.get("cf") and ws.max_row >= 2:
            _color_scale(ws, j, c["cf"])
    ws.freeze_panes = "A2"


def _color_scale(ws: Worksheet, col: int, kind: str) -> None:
    letter = get_column_letter(col)
    rng = f"{letter}2:{letter}{ws.max_row}"
    hi, lo = "63BE7B", "F8696B"
    start, end = (lo, hi) if kind == "good_high" else (hi, lo)
    ws.conditional_formatting.add(rng, ColorScaleRule(
        start_type="min", start_color=start,
        mid_type="percentile", mid_value=50, mid_color="FFEB84",
        end_type="max", end_color=end,
    ))


def _table(wb: Workbook, title: str, columns: list[dict], rows: list[dict]) -> Worksheet:
    ws = wb.create_sheet(title[:31])
    ws.append([c["header"] for c in columns])
    for r in rows:
        ws.append([r.get(c["key"]) for c in columns])
    _style(ws, columns)
    return ws


def _bar(ws: Worksheet, title: str, cat_col: int, val_col: int, anchor: str, max_rows: int = 15) -> None:
    if ws.max_row < 2:
        return
    chart = BarChart()
    chart.type = "col"
    chart.title = title
    chart.height = 8
    chart.width = 20
    chart.legend = None
    last = min(ws.max_row, 1 + max_rows)
    chart.add_data(Reference(ws, min_col=val_col, min_row=1, max_row=last), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=cat_col, min_row=2, max_row=last))
    ws.add_chart(chart, anchor)


# ── Sheets ────────────────────────────────────────────────────────────────────

def _overview_sheet(wb: Workbook, ov) -> None:
    ws = wb.create_sheet("Run Overview")
    ws["A1"] = "Explorer Report"
    ws["A1"].font = _TITLE_FONT
    pairs = [
        ("Marketplace", ov.marketplace),
        ("Brand", ov.brand),
        ("Mode", ov.mode),
        ("Label", ov.label or "—"),
        ("Keywords", ", ".join(ov.keywords) or "—"),
        ("Cities", ", ".join(ov.cities) or "all catalog cities"),
        ("Locations scraped", ov.locations_scraped),
        ("Sampling", "full census" if ov.full else f"{ov.sample}/city"),
        ("Generated", ov.generated_at.strftime("%Y-%m-%d %H:%M")),
        ("", ""),
        ("Overall SoV %", ov.overall_sov_pct),
        ("Average rank", ov.avg_rank),
        ("In-stock % (own)", ov.in_stock_pct),
        ("Keywords in top-3", ov.keywords_top3),
        ("Strongest keyword", ov.strongest_keyword or "—"),
        ("Weakest keyword", ov.weakest_keyword or "—"),
        ("Strongest city", ov.strongest_city or "—"),
        ("Weakest city", ov.weakest_city or "—"),
        ("Total listings captured", ov.total_listings),
        ("Competitors discovered", ov.total_competitors),
        ("Fetch errors", ov.errors),
    ]
    r = 3
    for k, v in pairs:
        if k:
            ws.cell(r, 1, k).font = _LABEL_FONT
            ws.cell(r, 2, v)
        r += 1
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 44


def _insight_sheets(wb: Workbook, ins: ExplorerInsights) -> None:
    kw = _table(wb, "Keyword Scorecard", [
        {"header": "Keyword", "key": "keyword"},
        {"header": "Locations", "key": "locations", "fmt": _NUM},
        {"header": "Avg Rank", "key": "avg_rank", "fmt": _DEC, "cf": "good_low"},
        {"header": "Best Rank", "key": "best_rank", "fmt": _NUM},
        {"header": "SoV %", "key": "sov_pct", "fmt": _PCT, "cf": "good_high"},
        {"header": "Presence %", "key": "presence_pct", "fmt": _PCT, "cf": "good_high"},
        {"header": "In-stock %", "key": "in_stock_pct", "fmt": _PCT, "cf": "good_high"},
        {"header": "Competitors", "key": "competitors", "fmt": _NUM},
        {"header": "Top Competitor", "key": "top_competitor"},
    ], [r.model_dump() for r in ins.keywords])
    _bar(kw, "SoV % by keyword", cat_col=1, val_col=5, anchor="K2")

    _table(wb, "Geography", [
        {"header": "City", "key": "city"},
        {"header": "Locations", "key": "locations", "fmt": _NUM},
        {"header": "Avg Rank", "key": "avg_rank", "fmt": _DEC, "cf": "good_low"},
        {"header": "SoV %", "key": "sov_pct", "fmt": _PCT, "cf": "good_high"},
        {"header": "In-stock %", "key": "in_stock_pct", "fmt": _PCT, "cf": "good_high"},
        {"header": "Keywords", "key": "keywords", "fmt": _NUM},
    ], [r.model_dump() for r in ins.geography])

    comp = _table(wb, "Competitor Landscape", [
        {"header": "Competitor", "key": "competitor"},
        {"header": "Locations", "key": "locations", "fmt": _NUM, "cf": "good_high"},
        {"header": "Keywords", "key": "keywords", "fmt": _NUM},
        {"header": "Appearances", "key": "appearances", "fmt": _NUM},
        {"header": "Avg Position", "key": "avg_position", "fmt": _DEC},
        {"header": "Avg Price", "key": "avg_price", "fmt": _RUP},
        {"header": "Share %", "key": "share_pct", "fmt": _PCT},
    ], [r.model_dump() for r in ins.competitors])
    _bar(comp, "Top competitors by locations", cat_col=1, val_col=2, anchor="I2", max_rows=10)

    _table(wb, "Price & Discount", [
        {"header": "Keyword", "key": "keyword"},
        {"header": "Own Avg", "key": "own_avg", "fmt": _RUP},
        {"header": "Own Min", "key": "own_min", "fmt": _RUP},
        {"header": "Own Max", "key": "own_max", "fmt": _RUP},
        {"header": "Own Disc %", "key": "own_discount_pct", "fmt": _PCT},
        {"header": "Comp Avg", "key": "comp_avg", "fmt": _RUP},
        {"header": "Comp Min", "key": "comp_min", "fmt": _RUP},
        {"header": "Comp Median", "key": "comp_median", "fmt": _RUP},
        {"header": "Comp Max", "key": "comp_max", "fmt": _RUP},
        # Per-unit band — the fair cross-pack-size comparison (see Basis column).
        {"header": "Basis", "key": "unit_uom"},
        {"header": "Own Avg /u", "key": "own_avg_unit", "fmt": _RUP2},
        {"header": "Own Min /u", "key": "own_min_unit", "fmt": _RUP2},
        {"header": "Own Max /u", "key": "own_max_unit", "fmt": _RUP2},
        {"header": "Comp Avg /u", "key": "comp_avg_unit", "fmt": _RUP2},
        {"header": "Comp Median /u", "key": "comp_median_unit", "fmt": _RUP2},
        {"header": "Comp Max /u", "key": "comp_max_unit", "fmt": _RUP2},
    ], [r.model_dump() for r in ins.pricing])

    _table(wb, "Availability", [
        {"header": "Keyword", "key": "keyword"},
        {"header": "City", "key": "city"},
        {"header": "Own Found", "key": "own_found", "fmt": _NUM},
        {"header": "Own In-stock", "key": "own_in_stock", "fmt": _NUM},
        {"header": "In-stock %", "key": "in_stock_pct", "fmt": _PCT, "cf": "good_high"},
    ], [r.model_dump() for r in ins.availability])

    if ins.catalog:
        _table(wb, "Own Catalog", [
            {"header": "Product ID", "key": "product_id"},
            {"header": "Name", "key": "name"},
            {"header": "Found Locations", "key": "found_locations", "fmt": _NUM},
            {"header": "Reach %", "key": "reach_pct", "fmt": _PCT, "cf": "good_high"},
            {"header": "Distribution %", "key": "distribution_pct", "fmt": _PCT, "cf": "good_high"},
            {"header": "Pack", "key": "pack_size", "fmt": _DEC},
            {"header": "Unit", "key": "pack_uom"},
            {"header": "Price Min", "key": "price_min", "fmt": _RUP},
            {"header": "Price Median", "key": "price_median", "fmt": _RUP},
            {"header": "Price Max", "key": "price_max", "fmt": _RUP},
            {"header": "Min /u", "key": "unit_price_min", "fmt": _RUP2},
            {"header": "Median /u", "key": "unit_price_median", "fmt": _RUP2},
            {"header": "Max /u", "key": "unit_price_max", "fmt": _RUP2},
            {"header": "Disc %", "key": "discount_pct", "fmt": _PCT},
            {"header": "Rating", "key": "rating", "fmt": _DEC},
            {"header": "Combo", "key": "is_combo"},
        ], [r.model_dump() for r in ins.catalog])


def _raw_sheets(wb: Workbook, result: ExplorerResult) -> None:
    _table(wb, "Raw - Snapshots", [
        {"header": h, "key": k, "fmt": f} for h, k, f in [
            ("Keyword", "keyword", None), ("City", "city", None), ("Zone", "zone", None),
            ("Lat", "lat", None), ("Lon", "lon", None), ("Merchant", "merchant_id", None),
            ("Total Results", "total_results", _NUM), ("Brand Rank", "brand_rank", _NUM),
            ("Brand SoV %", "brand_sov_pct", _PCT), ("Brand Products", "brand_product_count", _NUM),
        ]
    ], result.snapshots)

    listing_cols = [
        ("Keyword", "keyword", None), ("City", "city", None), ("Lat", "lat", None),
        ("Lon", "lon", None), ("Position", "position", _NUM),
        ("Merchant", "merchant_id", None), ("Name", "name", None),
        ("Brand", "brand", None), ("Own?", "is_brand", None), ("Brand Slug", "brand_slug", None),
        ("Price", "price", _RUP), ("MRP", "mrp", _RUP), ("Discount %", "discount_pct", _PCT),
        ("In Stock", "in_stock", None), ("Inventory", "inventory", _NUM),
        ("Product ID", "product_id", None), ("Unit", "unit", None),
        ("Pack Size", "pack_size", _DEC), ("Pack UOM", "pack_uom", None),
        ("Pack Count", "pack_count", _NUM), ("Rating", "rating", _DEC),
        ("State", "product_state", None), ("L0", "l0", None), ("L1", "l1", None),
        ("L2", "l2", None), ("Merchant Type", "merchant_type", None), ("Combo", "is_combo", None),
    ]
    _table(wb, "Raw - Listings",
           [{"header": h, "key": k, "fmt": f} for h, k, f in listing_cols],
           result.listings)

    if result.sku_rows:
        _table(wb, "Raw - Catalog SKUs",
               [{"header": h, "key": k, "fmt": f} for h, k, f in [
                   ("Product ID", "product_id", None), ("Name", "name", None),
                   ("City", "city", None), ("Lat", "lat", None), ("Lon", "lon", None),
                   ("Merchant", "merchant_id", None), ("Merchant Type", "merchant_type", None),
                   ("Price", "price", _RUP), ("MRP", "mrp", _RUP),
                   ("Discount %", "discount_pct", _PCT),
                   ("Unit", "unit", None), ("Pack Size", "pack_size", _DEC),
                   ("Pack UOM", "pack_uom", None), ("Pack Count", "pack_count", _NUM),
                   ("In Stock", "in_stock", None),
                   ("Inventory", "inventory", _NUM), ("Rating", "rating", _DEC),
                   ("Combo", "is_combo", None),
               ]], result.sku_rows)

    # Distinct locations actually used
    locs = OrderedDict()
    for r in result.snapshots + result.sku_rows:
        locs[(r.get("lat"), r.get("lon"))] = {
            "city": r.get("city"), "zone": r.get("zone"),
            "pincode": r.get("pincode"), "lat": r.get("lat"), "lon": r.get("lon"),
        }
    _table(wb, "Raw - Locations", [
        {"header": "City", "key": "city"}, {"header": "Zone", "key": "zone"},
        {"header": "Pincode", "key": "pincode"}, {"header": "Lat", "key": "lat"},
        {"header": "Lon", "key": "lon"},
    ], list(locs.values()))


def write_workbook(insights: ExplorerInsights, result: ExplorerResult, path: str) -> str:
    """Render the insights + raw rows to an .xlsx at `path`. Returns `path`."""
    wb = Workbook()
    wb.remove(wb.active)  # drop the default empty sheet
    _overview_sheet(wb, insights.overview)
    _insight_sheets(wb, insights)
    _raw_sheets(wb, result)
    wb.save(path)
    return path
