"""
Export Blinkit scraped data for a single tenant to an Excel workbook.

Usage:
    python export_to_excel.py <tenant_id>
    python export_to_excel.py <tenant_id> --output report.xlsx
"""

import asyncio
import argparse
import json
import uuid
from datetime import date, datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv
from sqlmodel import select

load_dotenv(Path(__file__).parent / ".env")

from app.core.database import AsyncSessionLocal
from app.models.blinkit_marketing import (
    AdPerformanceSummary,
    AdCampaign,
    SponsoredSOV,
    BrandCollection,
    VisibilityPlan,
)
from app.models.blinkit_seller import (
    BlinkitSellerSale,
    BlinkitSellerSalesSummary,
    BlinkitPO,
    BlinkitPOSnapshot,
    BlinkitSOH,
    BlinkitScorecardWeekly,
    BlinkitScorecardFacility,
    BlinkitScorecardKeySku,
)

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_STRIP = {"id", "upsert_key", "tenant_id", "scrape_job_id", "platform"}


def _to_header(key: str) -> str:
    return key.replace("_", " ").title()


def _clean_value(val):
    if isinstance(val, datetime):
        return val.replace(tzinfo=None)
    if isinstance(val, (dict, list)):
        return json.dumps(val, default=str)
    return val


def _write_sheet(ws, rows: list[dict]) -> None:
    if not rows:
        ws.append(["(no data)"])
        return

    all_keys = list(dict.fromkeys(k for row in rows for k in row))

    for col_idx, key in enumerate(all_keys, start=1):
        cell = ws.cell(row=1, column=col_idx, value=_to_header(key))
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, key in enumerate(all_keys, start=1):
            ws.cell(row=row_idx, column=col_idx, value=_clean_value(row.get(key)))

    for col_idx, key in enumerate(all_keys, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(_to_header(key)),
            max((len(str(row.get(key) or "")) for row in rows), default=0),
        )
        ws.column_dimensions[col_letter].width = min(max_len + 2, 55)


def _strip(obj) -> dict:
    return {k: v for k, v in obj.__dict__.items() if k not in _STRIP and not k.startswith("_")}


async def export(tenant_id: str, output_path: str) -> None:
    tid = uuid.UUID(tenant_id)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    async with AsyncSessionLocal() as session:

        async def fetch(model):
            result = await session.execute(select(model).where(model.tenant_id == tid))
            return result.scalars().all()

        def add_sheet(name: str, rows: list[dict]) -> None:
            _write_sheet(wb.create_sheet(name), rows)

        # ── Ad Performance Summary ────────────────────────────────────────────
        raw_perf = await fetch(AdPerformanceSummary)
        flat_perf = []
        for obj in raw_perf:
            row: dict = {}
            for k, v in _strip(obj).items():
                if k == "budget_distribution" and isinstance(v, dict):
                    for bk, bv in v.items():
                        row[f"budget_dist_{bk}"] = bv
                else:
                    row[k] = v
            flat_perf.append(row)
        add_sheet("Ad Performance Summary", flat_perf)

        # ── Ad Campaigns ──────────────────────────────────────────────────────
        add_sheet("Ad Campaigns", [_strip(o) for o in await fetch(AdCampaign)])

        # ── Sponsored SOV ─────────────────────────────────────────────────────
        add_sheet("Sponsored SOV", [_strip(o) for o in await fetch(SponsoredSOV)])

        # ── Brand Collections ─────────────────────────────────────────────────
        add_sheet("Brand Collections", [_strip(o) for o in await fetch(BrandCollection)])

        # ── Visibility Plans ──────────────────────────────────────────────────
        add_sheet("Visibility Plans", [_strip(o) for o in await fetch(VisibilityPlan)])

        # ── Seller Sales ──────────────────────────────────────────────────────
        add_sheet("Seller Sales", [_strip(o) for o in await fetch(BlinkitSellerSale)])

        # ── Sales Summary ─────────────────────────────────────────────────────
        add_sheet("Sales Summary", [_strip(o) for o in await fetch(BlinkitSellerSalesSummary)])

        # ── Purchase Orders ───────────────────────────────────────────────────
        raw_pos = await fetch(BlinkitPO)
        po_line_items: list[dict] = []
        flat_pos = []
        for obj in raw_pos:
            items = obj.raw.get("items") or [] if isinstance(obj.raw, dict) else []
            for item in items:
                po_line_items.append({"po_number": obj.po_number, **item})
            d = _strip(obj)
            d.pop("raw", None)
            flat_pos.append(d)
        add_sheet("Purchase Orders", flat_pos)
        add_sheet("PO Line Items", po_line_items)

        # ── PO Snapshots ──────────────────────────────────────────────────────
        add_sheet("PO Snapshots", [_strip(o) for o in await fetch(BlinkitPOSnapshot)])

        # ── Stock On Hand ─────────────────────────────────────────────────────
        add_sheet("Stock On Hand", [_strip(o) for o in await fetch(BlinkitSOH)])

        # ── Scorecard Weekly ──────────────────────────────────────────────────
        raw_weekly = await fetch(BlinkitScorecardWeekly)
        flat_weekly = []
        all_categories: list[dict] = []

        for obj in raw_weekly:
            row = {}
            d = _strip(obj)
            for k, v in d.items():
                if k == "overall" and isinstance(v, dict):
                    for sk, sv in v.items():
                        row[f"overall_{sk}"] = sv
                elif k == "best_category" and isinstance(v, dict):
                    for sk, sv in v.items():
                        row[f"best_cat_{sk}"] = sv
                elif k == "categories" and isinstance(v, list):
                    for cat in v:
                        all_categories.append({
                            "manufacturer_id": obj.manufacturer_id,
                            "from_date_ist": obj.from_date_ist,
                            **cat,
                        })
                else:
                    row[k] = v
            flat_weekly.append(row)

        add_sheet("Scorecard Weekly", flat_weekly)
        add_sheet("Scorecard Categories", all_categories)

        # ── Scorecard Facilities ──────────────────────────────────────────────
        add_sheet("Scorecard Facilities", [_strip(o) for o in await fetch(BlinkitScorecardFacility)])

        # ── Scorecard Key SKUs ────────────────────────────────────────────────
        add_sheet("Scorecard Key SKUs", [_strip(o) for o in await fetch(BlinkitScorecardKeySku)])

    wb.save(output_path)
    print(f"Exported {len(wb.sheetnames)} sheets → {output_path}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Export Blinkit tenant data to Excel")
    parser.add_argument("tenant_id", help="Tenant ID to export")
    parser.add_argument(
        "--output", "-o",
        default=None,
        help="Output file path (default: blinkit_<tenant_id>_<date>.xlsx)",
    )
    args = parser.parse_args()

    output = args.output or f"blinkit_{args.tenant_id}_{date.today().isoformat()}.xlsx"
    asyncio.run(export(args.tenant_id, output))
